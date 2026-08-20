"""
ventana_reporte_presupuestos.py
Pestaña "Presupuestos" del módulo Reportes: registro completo de todos
los presupuestos con filtros por rango de fechas, estado y búsqueda.
Incluye tarjetas de resumen (con tasa de conversión a venta), tabla con
panel lateral de detalle, y exportación en los 7 formatos vía
BotonReporteGeneral (PDF con dashboard, PDF simple, Word, ODT, Excel,
CSV y JSON).
"""
import tkinter as tk
from tkinter import ttk
import datetime

from utilidades_ui import formatear_gs, formatear_cantidad, habilitar_deseleccion_treeview
from widget_calendario import abrir_selector_fecha, abrir_selector_rango_fechas
from menu_reporte_general import BotonReporteGeneral
from models_presupuestos import (
    ESTADOS, listar_presupuestos_en_rango, obtener_presupuesto,
)

AZUL        = "#1d5fd6"
AZUL_OSC    = "#163d8c"
GRIS_FONDO  = "#f4f5f7"
GRIS_BORDE  = "#e2e8f0"
BLANCO      = "#ffffff"
VERDE       = "#16a34a"
NARANJA     = "#d97706"
GRIS_TEXTO  = "#6b7280"
NEGRO       = "#1e293b"

COLOR_ESTADO = {
    "Pendiente": "#1d5fd6", "Aprobado": "#16a34a", "Rechazado": "#dc2626",
    "Vencido": "#6b7280", "Convertido": "#7c3aed",
}


class PanelReportePresupuestos(tk.Frame):
    def __init__(self, parent, usuario_actual):
        super().__init__(parent, bg=BLANCO)
        self.usuario_actual = usuario_actual

        hoy = datetime.date.today()
        self.fecha_desde = hoy.replace(day=1)
        self.fecha_hasta = hoy

        self.presupuestos_actuales: list[dict] = []
        self.presupuesto_detalle_id: int | None = None

        self._construir_encabezado()
        self._construir_filtros()
        self._construir_tarjetas()
        self._construir_cuerpo()
        self._cargar()

    # ── Encabezado ───────────────────────────────────────────
    def _construir_encabezado(self):
        enc = tk.Frame(self, bg=AZUL, height=54)
        enc.pack(fill="x")
        enc.pack_propagate(False)
        tk.Label(enc, text="📊  Reportes de Presupuestos",
                 font=("Segoe UI", 15, "bold"), bg=AZUL, fg=BLANCO
                 ).pack(side="left", padx=20, pady=12)

        self.boton_reporte = BotonReporteGeneral(
            enc, obtener_datos_callback=self._obtener_datos_reporte,
            nombre_archivo_base="Reporte_General_Presupuestos",
        )
        self.boton_reporte.generador_pdf_dashboard = self._generar_pdf_dashboard
        self.boton_reporte.generador_excel = self._generar_excel
        self.boton_reporte.pack(side="right", padx=(0, 12), pady=12)

        tk.Button(enc, text="🔄 Actualizar", font=("Segoe UI", 9),
                  bg=AZUL_OSC, fg=BLANCO, relief="flat", padx=12, pady=4,
                  cursor="hand2", command=self._cargar
                  ).pack(side="right", padx=16, pady=12)

    # ── Barra de filtros ─────────────────────────────────────
    def _construir_filtros(self):
        barra = tk.Frame(self, bg=GRIS_FONDO, pady=10)
        barra.pack(fill="x", padx=0)
        interior = tk.Frame(barra, bg=GRIS_FONDO)
        interior.pack(fill="x", padx=16)

        tk.Label(interior, text="Desde:", font=("Segoe UI", 9, "bold"), bg=GRIS_FONDO).pack(side="left")
        self.lbl_desde = tk.Label(interior, text=self.fecha_desde.strftime("%d/%m/%Y"),
                                   font=("Segoe UI", 9), bg=BLANCO, fg=AZUL,
                                   relief="solid", bd=1, padx=8, pady=4, cursor="hand2")
        self.lbl_desde.pack(side="left", padx=(4, 12))
        self.lbl_desde.bind("<Button-1>", lambda e: self._abrir_cal_desde())

        tk.Label(interior, text="Hasta:", font=("Segoe UI", 9, "bold"), bg=GRIS_FONDO).pack(side="left")
        self.lbl_hasta = tk.Label(interior, text=self.fecha_hasta.strftime("%d/%m/%Y"),
                                   font=("Segoe UI", 9), bg=BLANCO, fg=AZUL,
                                   relief="solid", bd=1, padx=8, pady=4, cursor="hand2")
        self.lbl_hasta.pack(side="left", padx=(4, 6))
        self.lbl_hasta.bind("<Button-1>", lambda e: self._abrir_cal_hasta())

        tk.Button(interior, text="📅 Elegir Rango", font=("Segoe UI", 8), bg=BLANCO,
                  relief="solid", bd=1, padx=8, pady=3, cursor="hand2",
                  command=self._abrir_selector_rango).pack(side="left", padx=(0, 10))

        for texto, accion in [("Hoy", self._periodo_hoy), ("Esta semana", self._periodo_semana),
                               ("Este mes", self._periodo_mes), ("Este año", self._periodo_anio)]:
            tk.Button(interior, text=texto, font=("Segoe UI", 8), bg=BLANCO, relief="solid",
                      bd=1, padx=8, pady=3, cursor="hand2", command=accion).pack(side="left", padx=2)

        tk.Frame(interior, bg=GRIS_BORDE, width=1).pack(side="left", fill="y", padx=12, pady=2)

        tk.Label(interior, text="Estado:", font=("Segoe UI", 9, "bold"), bg=GRIS_FONDO).pack(side="left")
        self.var_estado = tk.StringVar(value="Todos")
        combo = ttk.Combobox(interior, textvariable=self.var_estado,
                             values=["Todos"] + ESTADOS + ["Vencido"], state="readonly",
                             width=14, font=("Segoe UI", 9))
        combo.pack(side="left", padx=(4, 16))
        combo.bind("<<ComboboxSelected>>", lambda e: self.after(50, self._cargar))

        tk.Label(interior, text="🔍", font=("Segoe UI", 11), bg=GRIS_FONDO).pack(side="left")
        self.var_busqueda = tk.StringVar()
        entry_busq = tk.Entry(interior, textvariable=self.var_busqueda, font=("Segoe UI", 9),
                              relief="solid", bd=1, width=20)
        entry_busq.pack(side="left", padx=(2, 4), ipady=3)
        entry_busq.bind("<Return>", lambda e: self._cargar())

        tk.Button(interior, text="Buscar", font=("Segoe UI", 9), bg=AZUL, fg=BLANCO,
                  relief="flat", padx=10, pady=3, cursor="hand2",
                  command=self._cargar).pack(side="left")

    # ── Tarjetas de resumen ──────────────────────────────────
    def _construir_tarjetas(self):
        self.frame_tarjetas = tk.Frame(self, bg=BLANCO)
        self.frame_tarjetas.pack(fill="x", padx=16, pady=(12, 4))

        definiciones = [
            ("cantidad",    "Presupuestos",        AZUL,    "📝"),
            ("total",       "Total Cotizado",      VERDE,   "💰"),
            ("convertidos", "Convertidos en Venta", "#7c3aed", "🔄"),
            ("tasa",        "Tasa de Conversión",  NARANJA, "📈"),
        ]
        self.labels_tarjetas: dict[str, tk.Label] = {}
        for clave, titulo, color, icono in definiciones:
            card = tk.Frame(self.frame_tarjetas, bg=BLANCO, highlightthickness=1,
                            highlightbackground=GRIS_BORDE)
            card.pack(side="left", fill="both", expand=True, padx=(0, 8), ipady=8)
            tk.Label(card, text=icono, font=("Segoe UI", 16), bg=BLANCO).pack(
                anchor="w", padx=12, pady=(10, 0))
            lbl_val = tk.Label(card, text="—", font=("Segoe UI", 14, "bold"), bg=BLANCO, fg=color)
            lbl_val.pack(anchor="w", padx=12)
            tk.Label(card, text=titulo, font=("Segoe UI", 8), bg=BLANCO, fg=GRIS_TEXTO).pack(
                anchor="w", padx=12, pady=(0, 8))
            self.labels_tarjetas[clave] = lbl_val

    # ── Cuerpo: tabla + panel detalle ───────────────────────
    def _construir_cuerpo(self):
        self.frame_cuerpo = tk.Frame(self, bg=BLANCO)
        self.frame_cuerpo.pack(fill="both", expand=True, padx=16, pady=(4, 12))
        self.frame_cuerpo.grid_columnconfigure(0, weight=1)
        self.frame_cuerpo.grid_columnconfigure(1, weight=0)
        self.frame_cuerpo.grid_rowconfigure(0, weight=1)
        self._construir_tabla()
        self._construir_panel_detalle()

    def _construir_tabla(self):
        frame_tabla = tk.Frame(self.frame_cuerpo, bg=BLANCO)
        frame_tabla.grid(row=0, column=0, sticky="nsew")

        cols = ("id", "fecha", "cliente", "validez", "estado", "total")
        encs = ("ID", "Fecha", "Cliente", "Válido Hasta", "Estado", "Total")
        anchos = (50, 130, 220, 110, 110, 120)

        self.tabla = ttk.Treeview(frame_tabla, columns=cols, show="headings", selectmode="browse")
        habilitar_deseleccion_treeview(self.tabla)
        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        style.configure("Treeview", font=("Segoe UI", 9), rowheight=26)
        for col, enc, ancho in zip(cols, encs, anchos):
            self.tabla.heading(col, text=enc)
            self.tabla.column(col, width=ancho, anchor="center", minwidth=40)
        self.tabla.column("cliente", anchor="w")
        for estado, color in COLOR_ESTADO.items():
            self.tabla.tag_configure(estado, foreground=color)

        sb_y = ttk.Scrollbar(frame_tabla, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=sb_y.set)
        self.tabla.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_y_h = ttk.Scrollbar(frame_tabla, orient="horizontal", command=self.tabla.xview)
        self.tabla.configure(xscrollcommand=sb_y_h.set)
        sb_y_h.grid(row=1, column=0, sticky="ew")
        frame_tabla.grid_rowconfigure(0, weight=1)
        frame_tabla.grid_columnconfigure(0, weight=1)

        self.tabla.bind("<<TreeviewSelect>>", self._al_seleccionar)
        self.lbl_contador = tk.Label(frame_tabla, text="", font=("Segoe UI", 8), bg=GRIS_FONDO,
                                     fg=GRIS_TEXTO, anchor="w")
        self.lbl_contador.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(2, 0))

    def _construir_panel_detalle(self):
        self.frame_detalle = tk.Frame(self.frame_cuerpo, bg=GRIS_FONDO, width=320)
        self.frame_detalle.grid(row=0, column=1, sticky="ns", padx=(10, 0))
        self.frame_detalle.grid_propagate(False)
        self.frame_detalle.grid_remove()

    # ── Accesos rápidos de período ──────────────────────────
    def _abrir_cal_desde(self):
        abrir_selector_fecha(self, self.fecha_desde, self._al_elegir_desde)

    def _abrir_cal_hasta(self):
        abrir_selector_fecha(self, self.fecha_hasta, self._al_elegir_hasta)

    def _abrir_selector_rango(self):
        abrir_selector_rango_fechas(self, self.fecha_desde, self.fecha_hasta, self._al_elegir_rango)

    def _al_elegir_rango(self, fecha_desde, fecha_hasta):
        self.fecha_desde = fecha_desde
        self.fecha_hasta = fecha_hasta
        self._sync_labels()
        self._cargar()

    def _al_elegir_desde(self, fecha):
        self.fecha_desde = fecha
        self._sync_labels()
        self._cargar()

    def _al_elegir_hasta(self, fecha):
        self.fecha_hasta = fecha
        self._sync_labels()
        self._cargar()

    def _periodo_hoy(self):
        hoy = datetime.date.today()
        self.fecha_desde = self.fecha_hasta = hoy
        self._sync_labels(); self._cargar()

    def _periodo_semana(self):
        hoy = datetime.date.today()
        self.fecha_desde = hoy - datetime.timedelta(days=hoy.weekday())
        self.fecha_hasta = hoy
        self._sync_labels(); self._cargar()

    def _periodo_mes(self):
        hoy = datetime.date.today()
        self.fecha_desde = hoy.replace(day=1)
        self.fecha_hasta = hoy
        self._sync_labels(); self._cargar()

    def _periodo_anio(self):
        hoy = datetime.date.today()
        self.fecha_desde = hoy.replace(month=1, day=1)
        self.fecha_hasta = hoy
        self._sync_labels(); self._cargar()

    def _sync_labels(self):
        self.lbl_desde.config(text=self.fecha_desde.strftime("%d/%m/%Y"))
        self.lbl_hasta.config(text=self.fecha_hasta.strftime("%d/%m/%Y"))

    # ── Carga y refresco ────────────────────────────────────
    def _cargar(self):
        estado = self.var_estado.get()
        estado_filtro = None if estado == "Todos" else estado
        self.presupuestos_actuales = listar_presupuestos_en_rango(
            self.fecha_desde.isoformat(), self.fecha_hasta.isoformat(),
            estado=estado_filtro, busqueda=self.var_busqueda.get(),
        )
        self._poblar_tabla()
        self._actualizar_tarjetas()

    def _poblar_tabla(self):
        for fila in self.tabla.get_children():
            self.tabla.delete(fila)
        for p in self.presupuestos_actuales:
            self.tabla.insert("", "end", iid=str(p["id"]), values=(
                p["id"], _formatear_fecha_hora(p["fecha"]), p["cliente"],
                _formatear_fecha(p["fecha_validez"]), p["estado_efectivo"], formatear_gs(p["total"]),
            ), tags=(p["estado_efectivo"],))
        n = len(self.presupuestos_actuales)
        self.lbl_contador.config(
            text=f"  {n} presupuesto{'s' if n != 1 else ''} encontrado{'s' if n != 1 else ''}"
                 f"  —  {self.fecha_desde.strftime('%d/%m/%Y')} al {self.fecha_hasta.strftime('%d/%m/%Y')}"
        )

    def _actualizar_tarjetas(self):
        total = sum(p["total"] for p in self.presupuestos_actuales)
        cantidad = len(self.presupuestos_actuales)
        convertidos = [p for p in self.presupuestos_actuales if p["estado_efectivo"] == "Convertido"]
        tasa = (len(convertidos) / cantidad * 100) if cantidad else 0

        self.labels_tarjetas["cantidad"].config(text=str(cantidad))
        self.labels_tarjetas["total"].config(text=formatear_gs(total))
        self.labels_tarjetas["convertidos"].config(text=str(len(convertidos)))
        self.labels_tarjetas["tasa"].config(text=f"{tasa:.1f}%")

    # ── Selección → panel detalle ────────────────────────────
    def _al_seleccionar(self, event=None):
        sel = self.tabla.selection()
        if not sel:
            return
        presupuesto_id = int(sel[0])
        if presupuesto_id == self.presupuesto_detalle_id:
            return
        self.presupuesto_detalle_id = presupuesto_id
        self._mostrar_detalle(presupuesto_id)

    def _mostrar_detalle(self, presupuesto_id: int):
        for w in self.frame_detalle.winfo_children():
            w.destroy()
        self.frame_detalle.grid()

        det = obtener_presupuesto(presupuesto_id)
        if not det:
            tk.Label(self.frame_detalle, text="No se pudo cargar el detalle.",
                     bg=GRIS_FONDO, font=("Segoe UI", 9)).pack(pady=20)
            return

        hdr = tk.Frame(self.frame_detalle, bg=AZUL)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"Presupuesto #{det['id']}", font=("Segoe UI", 11, "bold"),
                bg=AZUL, fg=BLANCO).pack(side="left", padx=12, pady=8)
        tk.Button(hdr, text="✕", font=("Segoe UI", 9), bg=AZUL, fg=BLANCO, relief="flat",
                 cursor="hand2", command=self._cerrar_detalle).pack(side="right", padx=8)

        canvas = tk.Canvas(self.frame_detalle, bg=GRIS_FONDO, highlightthickness=0)
        sb = tk.Scrollbar(self.frame_detalle, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)

        inner = tk.Frame(canvas, bg=GRIS_FONDO)
        canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        def _scroll(event):
            if event.num == 4:
                canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                canvas.yview_scroll(1, "units")
            else:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<MouseWheel>", _scroll); canvas.bind("<Button-4>", _scroll)
        canvas.bind("<Button-5>", _scroll); inner.bind("<MouseWheel>", _scroll)

        pad = {"padx": 12, "pady": 3}

        def fila_info(etiqueta, valor, color_val=NEGRO):
            f = tk.Frame(inner, bg=GRIS_FONDO)
            f.pack(fill="x", **pad)
            tk.Label(f, text=etiqueta, font=("Segoe UI", 8), bg=GRIS_FONDO, fg=GRIS_TEXTO,
                    width=13, anchor="w").pack(side="left")
            tk.Label(f, text=valor, font=("Segoe UI", 9, "bold"), bg=GRIS_FONDO, fg=color_val,
                    wraplength=180, justify="left").pack(side="left")

        tk.Label(inner, text="", bg=GRIS_FONDO).pack(pady=(6, 0))
        fila_info("Cliente:", det["cliente_nombre"])
        fila_info("Fecha:", _formatear_fecha_hora(det["fecha"]))
        fila_info("Válido hasta:", _formatear_fecha(det["fecha_validez"]))
        fila_info("Estado:", det["estado_efectivo"], COLOR_ESTADO.get(det["estado_efectivo"], NEGRO))
        fila_info("Vendedor:", det["vendedor"] or "—")

        tk.Frame(inner, bg=GRIS_BORDE, height=1).pack(fill="x", padx=12, pady=(8, 8))
        tk.Label(inner, text="ARTÍCULOS", font=("Segoe UI", 8, "bold"), bg=GRIS_FONDO,
                fg=GRIS_TEXTO).pack(anchor="w", padx=12)

        for item in det["items"]:
            f = tk.Frame(inner, bg=BLANCO, highlightthickness=1, highlightbackground=GRIS_BORDE)
            f.pack(fill="x", padx=12, pady=4)
            tk.Label(f, text=item["nombre"], font=("Segoe UI", 9, "bold"), bg=BLANCO,
                    wraplength=260, justify="left").pack(anchor="w", padx=8, pady=(6, 0))
            sub = tk.Frame(f, bg=BLANCO)
            sub.pack(fill="x", padx=8, pady=(2, 6))
            cant_txt = formatear_cantidad(item["cantidad"], item["unidad_medida"])
            tk.Label(sub, text=f"{cant_txt} x {formatear_gs(item['precio_unitario'])}",
                    font=("Segoe UI", 8), bg=BLANCO, fg=GRIS_TEXTO).pack(side="left")
            tk.Label(sub, text=formatear_gs(item["importe"]), font=("Segoe UI", 9, "bold"),
                    bg=BLANCO, fg=VERDE).pack(side="right")

        tk.Frame(inner, bg=GRIS_BORDE, height=1).pack(fill="x", padx=12, pady=(8, 8))
        f_total = tk.Frame(inner, bg=GRIS_FONDO)
        f_total.pack(fill="x", padx=12, pady=(0, 16))
        tk.Label(f_total, text="TOTAL", font=("Segoe UI", 9, "bold"), bg=GRIS_FONDO, fg=NEGRO).pack(
            side="left")
        tk.Label(f_total, text=formatear_gs(det["total"]), font=("Segoe UI", 12, "bold"),
                bg=GRIS_FONDO, fg=AZUL).pack(side="right")

    def _cerrar_detalle(self):
        self.frame_detalle.grid_remove()
        self.presupuesto_detalle_id = None
        if self.tabla.selection():
            self.tabla.selection_remove(*self.tabla.selection())

    # ---------------- REPORTE GENERAL ----------------
    def _obtener_datos_reporte(self) -> dict:
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""
        total = sum(p["total"] for p in self.presupuestos_actuales)
        cantidad = len(self.presupuestos_actuales)
        convertidos = [p for p in self.presupuestos_actuales if p["estado_efectivo"] == "Convertido"]
        tasa = (len(convertidos) / cantidad * 100) if cantidad else 0

        filtros_aplicados = []
        if self.var_estado.get() != "Todos":
            filtros_aplicados.append(f"Estado: {self.var_estado.get()}")
        busq = self.var_busqueda.get()
        if busq:
            filtros_aplicados.append(f"Búsqueda: \"{busq}\"")
        texto_filtros = "  •  ".join(filtros_aplicados) if filtros_aplicados else "Sin filtros adicionales"
        periodo_texto = (f"Período: {self.fecha_desde.strftime('%d/%m/%Y')} al "
                        f"{self.fecha_hasta.strftime('%d/%m/%Y')}   ({texto_filtros})")

        secciones = [
            {
                "tipo": "resumen", "titulo": "RESUMEN GENERAL",
                "filas": [
                    ("Cantidad de Presupuestos", str(cantidad)),
                    ("Total Cotizado", formatear_gs(total)),
                    ("Convertidos en Venta", str(len(convertidos))),
                    ("Tasa de Conversión", f"{tasa:.1f}%"),
                ],
            },
            {
                "tipo": "tabla", "titulo": "DETALLE DE PRESUPUESTOS",
                "encabezados": ["ID", "Fecha", "Cliente", "Válido Hasta", "Estado", "Total"],
                "filas": [
                    [str(p["id"]), _formatear_fecha_hora(p["fecha"]), p["cliente"],
                     _formatear_fecha(p["fecha_validez"]), p["estado_efectivo"], formatear_gs(p["total"])]
                    for p in self.presupuestos_actuales
                ],
            },
        ]
        return {"titulo": "Reporte General de Presupuestos", "subtitulo": periodo_texto,
                "generado_por": nombre_usuario, "secciones": secciones}

    def _generar_pdf_dashboard(self, ruta: str):
        from reporte_presupuestos_pdf import generar_reporte_presupuestos_pdf
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""
        generar_reporte_presupuestos_pdf(
            ruta, self.fecha_desde.isoformat(), self.fecha_hasta.isoformat(),
            generado_por=nombre_usuario,
        )

    def _generar_excel(self, ruta: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from models_presupuestos import obtener_presupuesto as _detalle

        total = sum(p["total"] for p in self.presupuestos_actuales)
        cantidad = len(self.presupuestos_actuales)
        convertidos = [p for p in self.presupuestos_actuales if p["estado_efectivo"] == "Convertido"]
        tasa = (len(convertidos) / cantidad * 100) if cantidad else 0
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen"
        ws.append(["Reporte General de Presupuestos"])
        ws.append([f"Período: {self.fecha_desde.strftime('%d/%m/%Y')} al "
                    f"{self.fecha_hasta.strftime('%d/%m/%Y')}"])
        ws.append([f"Generado por: {nombre_usuario}"])
        ws.append([])
        ws.append(["Indicador", "Valor"])
        for etiqueta, valor in [("Cantidad de Presupuestos", cantidad), ("Total Cotizado", total),
                                 ("Convertidos en Venta", len(convertidos)),
                                 ("Tasa de Conversión (%)", round(tasa, 1))]:
            ws.append([etiqueta, valor])
        for cell in ws[5]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D5FD6")
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 20

        ws2 = wb.create_sheet("Detalle de Presupuestos")
        cols_det = ["ID", "Fecha", "Cliente", "Válido Hasta", "Estado", "Total"]
        ws2.append(cols_det)
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D5FD6")
            cell.alignment = Alignment(horizontal="center")
        for p in self.presupuestos_actuales:
            ws2.append([p["id"], _formatear_fecha_hora(p["fecha"]), p["cliente"],
                       _formatear_fecha(p["fecha_validez"]), p["estado_efectivo"], p["total"]])
        for col in ws2.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        ws3 = wb.create_sheet("Detalle de Artículos")
        ws3.append(["Presupuesto Nro.", "Producto", "Cantidad", "Precio Unit.", "Importe"])
        for cell in ws3[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D5FD6")
            cell.alignment = Alignment(horizontal="center")
        for p in self.presupuestos_actuales:
            det = _detalle(p["id"])
            if not det:
                continue
            for item in det["items"]:
                ws3.append([p["id"], item["nombre"], item["cantidad"], item["precio_unitario"],
                           item["importe"]])
        for col in ws3.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws3.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(ruta)


def _formatear_fecha(fecha_iso: str) -> str:
    if not fecha_iso:
        return "—"
    try:
        return datetime.date.fromisoformat(fecha_iso[:10]).strftime("%d/%m/%Y")
    except ValueError:
        return fecha_iso


def _formatear_fecha_hora(fecha_hora: str) -> str:
    if not fecha_hora:
        return "—"
    try:
        return datetime.datetime.fromisoformat(fecha_hora).strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return fecha_hora
