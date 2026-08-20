"""
ventana_licencia.py
Tres ventanas modales relacionadas con licenciamiento y uso:
- VentanaActivarLicencia: pide el serial cuando no hay licencia vigente.
- VentanaGeneradorLicencias: genera nuevos seriales (protegido con admin/admin).
- VentanaEstadisticasUso: muestra el cronómetro detallado y hora pico de uso.
"""
import sys
import subprocess
import threading
import tkinter as tk
from tkinter import ttk, messagebox

from models_licencia import (
    activar_licencia, generar_y_guardar_serial,
    generar_seriales_en_lote, listar_seriales_generados, normalizar_serial,
    verificar_credenciales_admin,
    ETIQUETAS_UNIDADES, etiqueta_a_clave, describir_duracion,
    describir_duracion_componentes,
    eliminar_seriales,
)
from models_sesion import (
    tiempo_total_hoy, tiempo_total_general, hora_pico,
    actividad_por_hora, ultimas_sesiones, formatear_duracion,
)
from menu_reporte_general import BotonReporteGeneral
from utilidades_ui import habilitar_deseleccion_treeview
from traducciones import t

AZUL_RIBBON = "#1d5fd6"
AZUL_OSCURO = "#163d8c"
GRIS_FONDO = "#f4f5f7"
VERDE = "#16a34a"
ROJO = "#dc2626"


def copiar_al_portapapeles(texto: str, widget_tk: tk.Misc | None = None) -> bool:
    """Copia texto al portapapeles. PARTE 1 (instantánea): pone el texto en
    el clipboard de Tk para que esté disponible AHORA. PARTE 2 (en hilo de
    fondo): lo replica al portapapeles del sistema operativo para que
    sobreviva al cierre de la app. La UI nunca se bloquea.
    """
    texto = (texto or "").strip()
    if not texto:
        return False

    # ---- 1) Tk clipboard (instantáneo) ----
    if widget_tk is not None:
        try:
            widget_tk.clipboard_clear()
            widget_tk.clipboard_append(texto)
            widget_tk.update_idletasks()
        except Exception:
            pass

    # ---- 2) Portapapeles del SO (en hilo de fondo, no bloquea) ----
    def _copiar_so():
        try:
            if sys.platform == "win32":
                # CREATE_NO_WINDOW = 0x08000000 evita parpadeo de consola
                proc = subprocess.Popen(
                    "clip", stdin=subprocess.PIPE, shell=True,
                    creationflags=0x08000000,
                )
                proc.communicate(input=texto.encode("utf-16le"))
            elif sys.platform == "darwin":
                proc = subprocess.Popen(["pbcopy"], stdin=subprocess.PIPE)
                proc.communicate(input=texto.encode("utf-8"))
            else:
                for cmd in (["xclip", "-selection", "clipboard"], ["xsel", "--clipboard"]):
                    try:
                        proc = subprocess.Popen(cmd, stdin=subprocess.PIPE)
                        proc.communicate(input=texto.encode("utf-8"))
                        if proc.returncode == 0:
                            return
                    except FileNotFoundError:
                        continue
        except Exception:
            pass

    threading.Thread(target=_copiar_so, daemon=True).start()
    return True


# ============================================================
# 1) ACTIVAR LICENCIA (ventana raíz obligatoria al inicio)
# ============================================================
class VentanaActivarLicencia(tk.Tk):
    """Ventana raíz que pide el serial de 16 dígitos antes de poder
    iniciar sesión. No se puede cerrar sin activar una licencia válida
    (excepto cerrando la app entera)."""

    def __init__(self, on_activada):
        super().__init__()
        self.on_activada = on_activada

        self.title("Activación de Licencia")
        self.geometry("520x340")
        self.minsize(460, 320)
        self.configure(bg="white")
        self.protocol("WM_DELETE_WINDOW", self._intentar_cerrar)
        self._centrar()

        # Forzar que aparezca al frente al iniciar (importante en Windows/IDLE)
        self.lift()
        self.attributes("-topmost", True)
        self.after(200, lambda: self.attributes("-topmost", False))

        # Barra superior
        barra = tk.Frame(self, bg=AZUL_RIBBON, height=40)
        barra.pack(fill="x")
        barra.pack_propagate(False)
        tk.Label(barra, text="🔐  Activación de Licencia", font=("Segoe UI", 12, "bold"),
                 bg=AZUL_RIBBON, fg="white").pack(side="left", padx=20, pady=10)

        cuerpo = tk.Frame(self, bg="white")
        cuerpo.pack(fill="both", expand=True, padx=30, pady=20)

        tk.Label(cuerpo,
                 text="Para usar el sistema necesitás ingresar un serial de licencia válido.\n"
                      "Formato: XXXX-XXXX-XXXX-XXXX  (16 caracteres)",
                 font=("Segoe UI", 10), bg="white", fg="#374151", justify="left",
                 wraplength=440).pack(anchor="w", pady=(0, 12))

        tk.Label(cuerpo, text="Serial:", font=("Segoe UI", 10, "bold"),
                 bg="white").pack(anchor="w")
        self.var_serial = tk.StringVar()
        entry = tk.Entry(cuerpo, textvariable=self.var_serial,
                         font=("Consolas", 14, "bold"), justify="center")
        entry.pack(fill="x", ipady=8, pady=(4, 0))
        entry.focus()
        entry.bind("<KeyRelease>", lambda e: self._auto_formatear())
        entry.bind("<Return>", lambda e: self._activar())

        # Estado
        self.lbl_estado = tk.Label(cuerpo, text="", font=("Segoe UI", 9),
                                    bg="white", fg=ROJO, wraplength=440, justify="left")
        self.lbl_estado.pack(anchor="w", pady=(10, 0))

        # Botones
        botones = tk.Frame(cuerpo, bg="white")
        botones.pack(pady=(20, 0), fill="x")
        tk.Button(botones, text="✔ Activar", font=("Segoe UI", 10, "bold"),
                  bg=VERDE, fg="white", relief="flat", padx=20, pady=8,
                  cursor="hand2", command=self._activar).pack(side="left", padx=(0, 8))
        tk.Button(botones, text="🔑 Generar licencia (admin)",
                  font=("Segoe UI", 9), bg="white", fg=AZUL_OSCURO,
                  relief="solid", bd=1, padx=14, pady=7, cursor="hand2",
                  command=self._abrir_generador).pack(side="left")
        tk.Button(botones, text="Salir del sistema", font=("Segoe UI", 9),
                  bg="white", fg=ROJO, relief="solid", bd=1, padx=14, pady=7,
                  cursor="hand2", command=self._salir_sistema).pack(side="right")

    def _centrar(self):
        self.update_idletasks()
        w, h = 520, 340
        x = (self.winfo_screenwidth() - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _auto_formatear(self):
        valor = self.var_serial.get()
        limpio = "".join(c for c in valor.upper() if c.isalnum())[:16]
        with_guiones = "-".join(limpio[i:i+4] for i in range(0, len(limpio), 4))
        if with_guiones != valor:
            self.var_serial.set(with_guiones)

    def _activar(self):
        serial = normalizar_serial(self.var_serial.get())
        if len(serial.replace("-", "")) != 16:
            self.lbl_estado.config(text="⚠ El serial debe tener 16 caracteres alfanuméricos.")
            return
        # Reconfirmación de seguridad: antes de activar de verdad, se
        # vuelve a pedir usuario/contraseña de administrador.
        _solicitar_credenciales_admin(
            self, "🔐 Confirmar activación",
            "Para activar este serial, confirmá la contraseña de administrador.",
            on_exito=lambda: self._activar_confirmado(serial),
        )

    def _activar_confirmado(self, serial: str):
        ok, msg = activar_licencia(serial)
        if ok:
            self.lbl_estado.config(text="✓ " + msg, fg=VERDE)
            self.update()
            self.after(800, self._cerrar_ok)
        else:
            self.lbl_estado.config(text="✗ " + msg, fg=ROJO)

    def _cerrar_ok(self):
        self.destroy()
        self.on_activada()

    def _abrir_generador(self):
        """Abre el generador de licencias pidiendo credenciales admin/admin."""
        _solicitar_credenciales_y_abrir_generador(self)

    def _intentar_cerrar(self):
        messagebox.showinfo(
            "Licencia requerida",
            "Necesitás activar una licencia para usar el sistema.\n\n"
            "Si querés salir, usá el botón 'Salir del sistema'.",
            parent=self,
        )

    def _salir_sistema(self):
        if messagebox.askyesno("Salir", "¿Cerrar el sistema sin activar licencia?", parent=self):
            self.destroy()


# ============================================================
# 2) GENERADOR DE LICENCIAS (admin/admin)
# ============================================================
class VentanaGeneradorLicencias(tk.Toplevel):
    """Permite a un administrador generar seriales nuevos en cantidad ilimitada,
    eligiendo el tipo (mensual/anual) y la duración en meses."""

    def __init__(self, parent):
        super().__init__(parent)
        from utilidades_ui import ajustar_tamaño_ventana

        self.title("Generador de Licencias")
        self.resizable(True, True)
        self.configure(bg="white")
        self.grab_set()
        # NOTA: no usamos transient(parent) a propósito. En Windows, una
        # ventana Toplevel marcada como "transient" de su padre pierde los
        # botones nativos de minimizar y maximizar (queda solo con cerrar).
        try:
            self.attributes("-toolwindow", False)
        except tk.TclError:
            pass

        # Un grab_set() activo puede hacer que Windows ignore el click en
        # "Minimizar" (la ventana no puede ocultarse mientras retiene el
        # grab). Lo liberamos apenas se minimiza y lo recuperamos al
        # restaurarla, para que minimizar funcione sin perder el
        # comportamiento modal mientras la ventana está visible.
        def _al_cambiar_estado_ventana(event=None):
            try:
                estado = self.state()
            except tk.TclError:
                return
            if estado == "iconic":
                self.grab_release()
            elif estado == "normal":
                try:
                    self.grab_set()
                except tk.TclError:
                    pass

        self.bind("<Unmap>", _al_cambiar_estado_ventana)
        self.bind("<Map>", _al_cambiar_estado_ventana)

        # Barra superior
        barra = tk.Frame(self, bg=AZUL_RIBBON, height=40)
        barra.pack(fill="x")
        barra.pack_propagate(False)
        tk.Label(barra, text=t("licencia_generador_titulo"), font=("Segoe UI", 12, "bold"),
                 bg=AZUL_RIBBON, fg="white").pack(side="left", padx=20, pady=10)

        # --- Panel de generación ---
        panel = tk.Frame(self, bg=GRIS_FONDO, padx=20, pady=15)
        panel.pack(fill="x")

        # Ilimitado / Permanente (excluyente: si se marca, se ignoran los
        # campos de duración combinada de abajo)
        fila_ilimitado = tk.Frame(panel, bg=GRIS_FONDO)
        fila_ilimitado.pack(fill="x", pady=(0, 8))
        self.var_ilimitado = tk.BooleanVar(value=False)
        tk.Checkbutton(fila_ilimitado, text="Ilimitado / Permanente (sin vencimiento)",
                       variable=self.var_ilimitado, font=("Segoe UI", 9, "bold"),
                       bg=GRIS_FONDO, command=self._on_ilimitado_cambiado).pack(side="left")

        # Duración combinada: se puede cargar más de una unidad a la vez,
        # por ejemplo Meses=1 + Días=15 = "1 mes y 15 días".
        tk.Label(panel, text="Duración (combiná las unidades que necesites):",
                 font=("Segoe UI", 9, "bold"), bg=GRIS_FONDO).pack(anchor="w", pady=(4, 4))

        self.frame_duracion = tk.Frame(panel, bg=GRIS_FONDO)
        self.frame_duracion.pack(fill="x", pady=(0, 8))

        self.vars_componentes = {}
        self.spins_componentes = {}
        unidades_ui = [
            ("anios", "Años"), ("meses", "Meses"), ("semanas", "Semanas"),
            ("dias", "Días"), ("horas", "Horas"), ("minutos", "Minutos"),
        ]
        for i, (clave, etiqueta) in enumerate(unidades_ui):
            fila, col = divmod(i, 3)
            celda = tk.Frame(self.frame_duracion, bg=GRIS_FONDO)
            celda.grid(row=fila, column=col, sticky="w", padx=(0, 22), pady=4)
            tk.Label(celda, text=f"{etiqueta}:", font=("Segoe UI", 9), bg=GRIS_FONDO,
                     width=8, anchor="e").pack(side="left")
            var = tk.IntVar(value=0)
            spin = tk.Spinbox(celda, from_=0, to=99999, width=6, textvariable=var,
                              font=("Segoe UI", 10))
            spin.pack(side="left", padx=(6, 0))
            self.vars_componentes[clave] = var
            self.spins_componentes[clave] = spin

        # Cantidad de seriales a generar
        fila_cantidad = tk.Frame(panel, bg=GRIS_FONDO)
        fila_cantidad.pack(fill="x", pady=(4, 8))
        tk.Label(fila_cantidad, text="Cantidad de seriales:",
                 font=("Segoe UI", 9, "bold"), bg=GRIS_FONDO).pack(side="left")
        self.var_cantidad = tk.IntVar(value=1)
        tk.Spinbox(fila_cantidad, from_=1, to=500, width=6, textvariable=self.var_cantidad,
                   font=("Segoe UI", 10)).pack(side="left", padx=(10, 0))

        # Tip de uso
        tk.Label(panel,
                 text="Tip: elegí 'Ilimitado / Permanente' para una licencia sin vencimiento.",
                 font=("Segoe UI", 8, "italic"), bg=GRIS_FONDO,
                 fg="#6b7280").pack(anchor="w", pady=(4, 6))

        # Botón generar
        tk.Button(panel, text="＋ Generar seriales", font=("Segoe UI", 10, "bold"),
                  bg=VERDE, fg="white", relief="flat", padx=20, pady=8,
                  cursor="hand2", command=self._generar).pack(anchor="w", pady=(2, 0))

        # --- Barra inferior (se empaqueta ANTES que la tabla, con
        # side="bottom", para garantizarle su espacio siempre visible;
        # si no, al no entrar todo el contenido, la tabla —que se expande—
        # se quedaba con todo el lugar y esta barra quedaba invisible) ---
        inferior = tk.Frame(self, bg=GRIS_FONDO, pady=10)
        inferior.pack(fill="x", side="bottom")
        tk.Frame(inferior, bg="#e2e8f0", height=1).pack(fill="x", side="top")

        fila_botones = tk.Frame(inferior, bg=GRIS_FONDO)
        fila_botones.pack(fill="x", pady=(10, 4), padx=15)
        tk.Button(fila_botones, text="📋 Copiar serial",
                  font=("Segoe UI", 9, "bold"), bg=AZUL_RIBBON, fg="white",
                  relief="flat", padx=16, pady=8, cursor="hand2",
                  activebackground=AZUL_OSCURO, activeforeground="white",
                  command=self._copiar_serial).pack(side="left")
        tk.Button(fila_botones, text="🗑 Eliminar serial",
                  font=("Segoe UI", 9, "bold"), bg=ROJO, fg="white",
                  relief="flat", padx=16, pady=8, cursor="hand2",
                  activebackground="#b91c1c", activeforeground="white",
                  command=self._eliminar_serial).pack(side="left", padx=(10, 0))
        tk.Button(fila_botones, text="Cerrar", font=("Segoe UI", 9, "bold"),
                  bg="white", fg=AZUL_OSCURO, relief="solid", bd=1, padx=18, pady=8,
                  cursor="hand2", command=self.destroy).pack(side="right")

        tk.Label(inferior,
                  text="Doble clic = copiar   •   Suprimir = eliminar   •   Ctrl+clic = selección múltiple",
                  font=("Segoe UI", 8, "italic"), bg=GRIS_FONDO,
                  fg="#6b7280").pack(anchor="w", padx=15, pady=(0, 2))

        # --- Tabla de seriales ---
        tk.Label(self, text="Seriales generados (últimos 200)",
                 font=("Segoe UI", 10, "bold"), bg="white").pack(anchor="w", padx=15, pady=(0, 6))

        contenedor = tk.Frame(self, bg="white", padx=15, pady=10)
        contenedor.pack(fill="both", expand=True)

        cols = ("serial", "duracion", "generado", "usada", "fecha_uso")
        nombres = (t("licencia_col_serial"), t("licencia_col_duracion"), t("licencia_col_generado"), t("licencia_col_usada"), t("licencia_col_fecha_uso"))
        anchos = (180, 140, 140, 90, 140)
        self.tabla = ttk.Treeview(contenedor, columns=cols, show="headings",
                                   height=10, selectmode="extended")
        habilitar_deseleccion_treeview(self.tabla)
        for c, n, a in zip(cols, nombres, anchos):
            self.tabla.heading(c, text=n)
            self.tabla.column(c, width=a, anchor="center")
        self.tabla.column("serial", anchor="w")
        self.tabla.column("duracion", anchor="w")
        contenedor.grid_rowconfigure(0, weight=1)
        contenedor.grid_columnconfigure(0, weight=1)
        scroll = ttk.Scrollbar(contenedor, orient="vertical", command=self.tabla.yview)
        scroll_h = ttk.Scrollbar(contenedor, orient="horizontal", command=self.tabla.xview)
        self.tabla.configure(yscrollcommand=scroll.set, xscrollcommand=scroll_h.set)
        self.tabla.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")
        scroll_h.grid(row=1, column=0, sticky="ew")

        self._cargar_tabla()
        # Atajos sobre la tabla
        self.tabla.bind("<Double-1>", lambda e: self._copiar_serial())
        self.tabla.bind("<Control-c>", lambda e: self._copiar_serial())
        self.tabla.bind("<Control-C>", lambda e: self._copiar_serial())
        self.tabla.bind("<Delete>", lambda e: self._eliminar_serial())

        # Menú contextual con clic derecho
        self.menu_contextual = tk.Menu(self.tabla, tearoff=0, font=("Segoe UI", 9))
        self.menu_contextual.add_command(label="📋 Copiar serial", command=self._copiar_serial)
        self.menu_contextual.add_separator()
        self.menu_contextual.add_command(label="🗑 Eliminar serial", command=self._eliminar_serial)
        self.tabla.bind("<Button-3>", self._abrir_menu_contextual)

        ajustar_tamaño_ventana(self, ancho_min=660, alto_min=480,
                               margen_alto=20, ancho_max=1000, alto_max=760)

    def _on_ilimitado_cambiado(self):
        """Si se marca 'Ilimitado', deshabilitamos los campos de duración
        combinada porque no aplican (son excluyentes entre sí)."""
        estado = "disabled" if self.var_ilimitado.get() else "normal"
        for spin in self.spins_componentes.values():
            spin.config(state=estado)

    def _generar(self):
        cantidad = self.var_cantidad.get()
        if cantidad <= 0:
            messagebox.showwarning("Cantidad inválida",
                                    "La cantidad de seriales debe ser mayor a 0.", parent=self)
            return

        if self.var_ilimitado.get():
            componentes = {"ilimitado": True}
        else:
            componentes = {clave: var.get() for clave, var in self.vars_componentes.items()}
            if not any(v > 0 for v in componentes.values()):
                messagebox.showwarning(
                    "Duración vacía",
                    "Cargá al menos una unidad de tiempo (Años, Meses, Semanas, Días, "
                    "Horas o Minutos), o marcá 'Ilimitado / Permanente'.", parent=self)
                return

        duracion_legible = describir_duracion_componentes(componentes)
        if cantidad == 1:
            ok, msg, serial = generar_y_guardar_serial(0, "meses", componentes=componentes)
            if ok:
                # Insertar la nueva fila al principio sin recargar toda la tabla
                self._insertar_fila_nueva(serial, duracion_legible)
                # Copiar al portapapeles INMEDIATAMENTE para que el usuario
                # pueda pegar de una vez si ya hizo clic en otro lado
                copiar_al_portapapeles(serial, self)
                _mostrar_serial_copiable(self, serial, duracion_legible)
            else:
                messagebox.showerror("Error", msg, parent=self)
        else:
            generados = generar_seriales_en_lote(cantidad, 0, "meses", componentes=componentes)
            if generados:
                # Recarga simple solo cuando hay lote (varias filas)
                self._cargar_tabla()
                # Copiar el bloque entero al portapapeles
                copiar_al_portapapeles("\n".join(generados), self)
                _mostrar_seriales_lote(self, generados, duracion_legible)
            else:
                messagebox.showerror("Error", "No se pudo generar ningún serial.",
                                     parent=self)

    def _insertar_fila_nueva(self, serial: str, duracion_legible: str):
        """Inserta solo la fila recién creada al principio (más rápido que
        recargar toda la tabla)."""
        # Recuperar el id real del último insertado para que el iid sea correcto
        seriales = listar_seriales_generados(1)
        if not seriales:
            return
        lic = seriales[0]
        # Insertar al inicio (índice 0)
        self.tabla.insert("", 0, iid=str(lic["id"]), values=(
            lic["serial"], lic["duracion_legible"],
            lic["fecha_generacion"], "Disponible", "—",
        ))
        self.tabla.selection_set(str(lic["id"]))
        self.tabla.see(str(lic["id"]))

    def _cargar_tabla(self):
        for f in self.tabla.get_children():
            self.tabla.delete(f)
        for lic in listar_seriales_generados(200):
            estado = "✓ Usada" if lic["usada"] else "Disponible"
            self.tabla.insert("", "end", iid=str(lic["id"]), values=(
                lic["serial"], lic["duracion_legible"],
                lic["fecha_generacion"], estado, lic["fecha_uso"] or "—"
            ))

    def _abrir_menu_contextual(self, event):
        # Seleccionar la fila bajo el cursor si no estaba ya seleccionada
        fila = self.tabla.identify_row(event.y)
        if not fila:
            return
        if fila not in self.tabla.selection():
            self.tabla.selection_set(fila)
        self.menu_contextual.tk_popup(event.x_root, event.y_root)

    def _copiar_serial(self):
        sel = self.tabla.selection()
        if not sel:
            self._mostrar_toast("Seleccioná un serial primero", ROJO)
            return
        # Soportar múltiples seriales (uno por línea)
        seriales = [str(self.tabla.item(s)["values"][0]) for s in sel]
        texto = "\n".join(seriales)
        copiar_al_portapapeles(texto, self)
        cantidad = len(seriales)
        msg = (f"✓ {cantidad} seriales copiados" if cantidad > 1
               else f"✓ Copiado: {seriales[0]}")
        self._mostrar_toast(msg, VERDE)

    def _eliminar_serial(self):
        sel = self.tabla.selection()
        if not sel:
            self._mostrar_toast("Seleccioná un serial primero", ROJO)
            return
        ids = [int(s) for s in sel]
        # Mostrar resumen de qué se va a borrar
        if len(ids) == 1:
            serial_txt = self.tabla.item(sel[0])["values"][0]
            estado = self.tabla.item(sel[0])["values"][3]
            advertencia = ""
            if "Usada" in str(estado):
                # Verificar si es la licencia activa
                from models_licencia import obtener_licencia_activa
                activa = obtener_licencia_activa()
                if activa and activa["serial"] == serial_txt:
                    advertencia = ("\n\n⚠ Este serial es la LICENCIA ACTIVA del sistema. "
                                    "Al eliminarlo, el sistema te pedirá una nueva licencia "
                                    "en el próximo minuto.")
            pregunta = f"¿Eliminar el serial?\n\n{serial_txt}{advertencia}"
        else:
            pregunta = (f"¿Eliminar los {len(ids)} seriales seleccionados?\n\n"
                        "⚠ Si alguno es la licencia activa, el sistema te pedirá "
                        "una nueva licencia en el próximo minuto.")

        if not messagebox.askyesno("Confirmar eliminación", pregunta, parent=self):
            return

        eliminados, alguna_activa = eliminar_seriales(ids)

        # Borrar las filas del tree sin recargar todo
        for iid in sel:
            try:
                self.tabla.delete(iid)
            except tk.TclError:
                pass

        if alguna_activa:
            self._mostrar_toast(f"✓ {eliminados} eliminado(s). Licencia desvinculada.", "#f59e0b")
        else:
            self._mostrar_toast(f"✓ {eliminados} serial(es) eliminado(s)", VERDE)

    def _mostrar_toast(self, mensaje: str, color: str = VERDE, duracion_ms: int = 2500):
        """Muestra un mensaje no-bloqueante por unos segundos en la esquina
        de la ventana. Mucho más rápido que un messagebox."""
        if hasattr(self, "_toast_actual") and self._toast_actual:
            try:
                self._toast_actual.destroy()
            except tk.TclError:
                pass
        toast = tk.Toplevel(self)
        toast.overrideredirect(True)
        toast.configure(bg=color)
        try:
            toast.attributes("-topmost", True)
        except tk.TclError:
            pass
        tk.Label(toast, text=" " + mensaje + " ", font=("Segoe UI", 10, "bold"),
                  bg=color, fg="white", padx=14, pady=8).pack()
        # Posicionar abajo-derecha de la ventana del generador
        self.update_idletasks()
        x = self.winfo_rootx() + self.winfo_width() - 360
        y = self.winfo_rooty() + self.winfo_height() - 70
        toast.geometry(f"+{x}+{y}")
        self._toast_actual = toast
        toast.after(duracion_ms, lambda: toast.destroy() if toast.winfo_exists() else None)


def _mostrar_serial_copiable(parent, serial: str, duracion_legible: str):
    """Diálogo que muestra UN serial recién generado en un Entry seleccionable
    + botón 'Copiar al portapapeles' que sobrevive al cierre de la app."""
    from utilidades_ui import ajustar_tamaño_ventana

    dlg = tk.Toplevel(parent)
    dlg.title("Serial generado")
    dlg.resizable(True, True)
    dlg.configure(bg="white")
    dlg.grab_set()
    # NOTA: no usamos transient(parent) a propósito. En Windows, una
    # ventana Toplevel marcada como "transient" de su padre pierde los
    # botones nativos de minimizar y maximizar (queda solo con cerrar).
    try:
        dlg.attributes("-toolwindow", False)
    except tk.TclError:
        pass
    dlg.lift()
    try:
        dlg.attributes("-topmost", True)
        dlg.after(50, lambda: dlg.attributes("-topmost", False))
    except tk.TclError:
        pass

    barra = tk.Frame(dlg, bg=AZUL_RIBBON, height=36)
    barra.pack(fill="x")
    barra.pack_propagate(False)
    tk.Label(barra, text="🔑  Nuevo serial generado", font=("Segoe UI", 11, "bold"),
             bg=AZUL_RIBBON, fg="white").pack(side="left", padx=18, pady=8)

    cont = tk.Frame(dlg, bg="white", padx=24, pady=18)
    cont.pack(fill="both", expand=True)

    tk.Label(cont, text=f"Duración: {duracion_legible}",
             font=("Segoe UI", 10, "bold"), bg="white",
             fg=AZUL_OSCURO).pack(anchor="w")

    tk.Label(cont, text="Serial:", font=("Segoe UI", 9, "bold"),
             bg="white").pack(anchor="w", pady=(10, 4))

    # Entry seleccionable y de solo lectura
    var = tk.StringVar(value=serial)
    entry = tk.Entry(cont, textvariable=var, font=("Consolas", 16, "bold"),
                     justify="center", relief="solid", bd=1)
    entry.pack(fill="x", ipady=10)
    entry.config(state="readonly", readonlybackground="white",
                  fg=AZUL_OSCURO)
    # Pre-seleccionar todo el texto para que el usuario pueda copiar con Ctrl+C
    entry.config(state="normal")
    entry.select_range(0, "end")
    entry.icursor("end")
    entry.focus_set()
    entry.config(state="readonly")
    # Permitir selección incluso en estado readonly: re-habilitamos para seleccionar
    def _seleccionar_todo(_e=None):
        entry.config(state="normal")
        entry.select_range(0, "end")
        entry.icursor("end")
        entry.focus_set()
        entry.config(state="readonly")
    entry.bind("<Button-1>", lambda e: dlg.after(10, _seleccionar_todo))

    estado = tk.Label(cont, text="", font=("Segoe UI", 9), bg="white", fg=VERDE)
    estado.pack(anchor="w", pady=(8, 0))

    def copiar():
        ok = copiar_al_portapapeles(serial, dlg)
        if ok:
            estado.config(text="✓ Copiado al portapapeles. Podés pegarlo con Ctrl+V "
                                "incluso después de cerrar.", fg=VERDE)
        else:
            estado.config(text="⚠ No se pudo acceder al portapapeles del sistema. "
                                "Seleccioná el texto y copialo con Ctrl+C.", fg=ROJO)

    botones = tk.Frame(cont, bg="white")
    botones.pack(fill="x", pady=(14, 0))
    tk.Button(botones, text="📋 Copiar al portapapeles",
              font=("Segoe UI", 10, "bold"), bg=VERDE, fg="white",
              relief="flat", padx=18, pady=8, cursor="hand2",
              command=copiar).pack(side="left")
    tk.Button(botones, text="Cerrar", font=("Segoe UI", 9),
              bg="white", relief="solid", bd=1, padx=16, pady=6,
              cursor="hand2", command=dlg.destroy).pack(side="right")

    # Copiar al portapapeles automáticamente al abrir (UX cómoda)
    dlg.after_idle(copiar)

    ajustar_tamaño_ventana(dlg, ancho_min=480, alto_min=280,
                           margen_alto=20, ancho_max=640, alto_max=420)


def _mostrar_seriales_lote(parent, generados: list[str], duracion_legible: str):
    """Diálogo para lotes: muestra todos los seriales en un Text seleccionable
    y permite copiarlos todos al portapapeles."""
    dlg = tk.Toplevel(parent)
    dlg.title(f"{len(generados)} seriales generados")
    dlg.geometry("560x440")
    dlg.minsize(520, 400)
    dlg.configure(bg="white")
    dlg.grab_set()
    dlg.transient(parent)
    dlg.lift()
    dlg.attributes("-topmost", True)
    dlg.after(50, lambda: dlg.attributes("-topmost", False))

    barra = tk.Frame(dlg, bg=AZUL_RIBBON, height=36)
    barra.pack(fill="x")
    barra.pack_propagate(False)
    tk.Label(barra, text=f"🔑  {len(generados)} seriales generados",
             font=("Segoe UI", 11, "bold"),
             bg=AZUL_RIBBON, fg="white").pack(side="left", padx=18, pady=8)

    cont = tk.Frame(dlg, bg="white", padx=20, pady=15)
    cont.pack(fill="both", expand=True)

    tk.Label(cont, text=f"Duración: {duracion_legible}",
             font=("Segoe UI", 10, "bold"), bg="white",
             fg=AZUL_OSCURO).pack(anchor="w")

    tk.Label(cont, text="Seriales (uno por línea):", font=("Segoe UI", 9, "bold"),
             bg="white").pack(anchor="w", pady=(10, 4))

    frame_txt = tk.Frame(cont, bg="white")
    frame_txt.pack(fill="both", expand=True)
    texto = tk.Text(frame_txt, font=("Consolas", 11), wrap="none",
                    relief="solid", bd=1, height=12)
    texto.pack(side="left", fill="both", expand=True)
    sb = ttk.Scrollbar(frame_txt, orient="vertical", command=texto.yview)
    sb.pack(side="right", fill="y")
    texto.config(yscrollcommand=sb.set)

    contenido = "\n".join(generados)
    texto.insert("1.0", contenido)
    texto.tag_add("sel", "1.0", "end")
    texto.focus_set()

    estado = tk.Label(cont, text="", font=("Segoe UI", 9), bg="white", fg=VERDE)
    estado.pack(anchor="w", pady=(8, 0))

    def copiar_todos():
        ok = copiar_al_portapapeles(contenido, dlg)
        if ok:
            estado.config(text=f"✓ {len(generados)} seriales copiados al portapapeles. "
                                "Sobreviven al cierre de esta ventana.", fg=VERDE)
        else:
            estado.config(text="⚠ No se pudo copiar automáticamente. "
                                "Seleccioná el texto y usá Ctrl+C.", fg=ROJO)

    botones = tk.Frame(cont, bg="white")
    botones.pack(fill="x", pady=(12, 0))
    tk.Button(botones, text="📋 Copiar TODOS al portapapeles",
              font=("Segoe UI", 10, "bold"), bg=VERDE, fg="white",
              relief="flat", padx=18, pady=8, cursor="hand2",
              command=copiar_todos).pack(side="left")
    tk.Button(botones, text="Cerrar", font=("Segoe UI", 9),
              bg="white", relief="solid", bd=1, padx=16, pady=6,
              cursor="hand2", command=dlg.destroy).pack(side="right")

    # Copiar automáticamente al abrir
    dlg.after_idle(copiar_todos)


def _solicitar_credenciales_admin(parent, titulo: str, subtitulo: str, on_exito):
    """Pide usuario y contraseña de administrador; si son correctas,
    ejecuta on_exito() y cierra el diálogo. Reutilizable para cualquier
    acción que necesite una reconfirmación de credenciales admin (abrir
    el generador, confirmar una activación, etc.).

    Usa la barra de título NATIVA del sistema operativo (con sus botones
    reales de minimizar, maximizar y cerrar) y un layout responsive con
    grid, para que la ventana se pueda redimensionar o maximizar sin que
    ningún campo o botón quede cortado fuera del área visible."""
    from utilidades_ui import ajustar_tamaño_ventana

    dialog = tk.Toplevel(parent)
    dialog.title("Credenciales de administrador")
    dialog.resizable(True, True)
    dialog.configure(bg="white")
    dialog.grab_set()
    # NOTA: no usamos transient(parent) a propósito. En Windows, una
    # ventana Toplevel marcada como "transient" de su padre pierde los
    # botones nativos de minimizar y maximizar (queda solo con cerrar).
    try:
        dialog.attributes("-toolwindow", False)
    except tk.TclError:
        pass

    dialog.grid_rowconfigure(0, weight=1)
    dialog.grid_columnconfigure(0, weight=1)

    # ── Cuerpo scrolleable (por si la ventana queda muy chica) ──────
    contenedor_scroll = tk.Frame(dialog, bg="white")
    contenedor_scroll.grid(row=0, column=0, sticky="nsew")
    contenedor_scroll.grid_rowconfigure(0, weight=1)
    contenedor_scroll.grid_columnconfigure(0, weight=1)

    canvas = tk.Canvas(contenedor_scroll, bg="white", highlightthickness=0)
    scrollbar = ttk.Scrollbar(contenedor_scroll, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.grid(row=0, column=0, sticky="nsew")
    scrollbar.grid(row=0, column=1, sticky="ns")

    cont = tk.Frame(canvas, bg="white")
    ventana_canvas = canvas.create_window((0, 0), window=cont, anchor="nw")

    def _actualizar_scrollregion(event=None):
        canvas.configure(scrollregion=canvas.bbox("all"))

    def _ajustar_ancho_cuerpo(event):
        canvas.itemconfig(ventana_canvas, width=event.width)

    cont.bind("<Configure>", _actualizar_scrollregion)
    canvas.bind("<Configure>", _ajustar_ancho_cuerpo)

    def _scroll(event):
        if event.num == 4:
            canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            canvas.yview_scroll(1, "units")
        else:
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    for w in (canvas, cont):
        w.bind("<MouseWheel>", _scroll)
        w.bind("<Button-4>", _scroll)
        w.bind("<Button-5>", _scroll)

    # ── Barra de color (NO es la barra de título del SO) ─────────────
    barra = tk.Frame(cont, bg=AZUL_RIBBON, height=34)
    barra.pack(fill="x")
    barra.pack_propagate(False)
    tk.Label(barra, text=titulo, font=("Segoe UI", 11, "bold"),
             bg=AZUL_RIBBON, fg="white").pack(side="left", padx=15, pady=6)

    cuerpo_campos = tk.Frame(cont, bg="white", padx=20, pady=15)
    cuerpo_campos.pack(fill="x")

    if subtitulo:
        tk.Label(cuerpo_campos, text=subtitulo, font=("Segoe UI", 9), bg="white",
                 fg="#374151", wraplength=320, justify="left").pack(anchor="w", pady=(0, 10))

    tk.Label(cuerpo_campos, text="Usuario:", font=("Segoe UI", 9, "bold"),
             bg="white").pack(anchor="w")
    var_usuario = tk.StringVar()
    e_user = tk.Entry(cuerpo_campos, textvariable=var_usuario, font=("Segoe UI", 10))
    e_user.pack(fill="x", pady=(2, 8))
    e_user.focus()

    tk.Label(cuerpo_campos, text="Contraseña:", font=("Segoe UI", 9, "bold"),
             bg="white").pack(anchor="w")
    var_pass = tk.StringVar()
    e_pass = tk.Entry(cuerpo_campos, textvariable=var_pass, font=("Segoe UI", 10), show="•")
    e_pass.pack(fill="x", pady=(2, 8))

    lbl_msg = tk.Label(cuerpo_campos, text="", font=("Segoe UI", 9), bg="white",
                       fg=ROJO, wraplength=320, justify="left")
    lbl_msg.pack(anchor="w", fill="x")

    def _intentar():
        if verificar_credenciales_admin(var_usuario.get(), var_pass.get()):
            dialog.destroy()
            on_exito()
        else:
            lbl_msg.config(text="✗ Credenciales incorrectas.")
            var_pass.set("")
            e_pass.focus()

    botones = tk.Frame(cuerpo_campos, bg="white")
    botones.pack(fill="x", pady=(10, 0))
    btn_entrar = tk.Button(botones, text="Entrar", font=("Segoe UI", 9, "bold"),
              bg=VERDE, fg="white", relief="flat", padx=16, pady=6,
              cursor="hand2", command=_intentar)
    btn_entrar.pack(side="left")
    btn_entrar.bind("<Return>", lambda e: _intentar())
    tk.Button(botones, text="Cancelar", font=("Segoe UI", 9),
              bg="white", relief="solid", bd=1, padx=16, pady=5,
              cursor="hand2", command=dialog.destroy).pack(side="right")

    # Enter en cualquiera de los dos campos confirma el acceso.
    e_pass.bind("<Return>", lambda e: _intentar())
    e_user.bind("<Return>", lambda e: _intentar())

    ajustar_tamaño_ventana(dialog, ancho_min=380, alto_min=260,
                           margen_alto=20, ancho_max=480, alto_max=560)


def _solicitar_credenciales_y_abrir_generador(parent):
    """Pide usuario y contraseña; si son admin/admin abre el generador."""
    _solicitar_credenciales_admin(
        parent, "🔐 Acceso al generador", "",
        on_exito=lambda: VentanaGeneradorLicencias(parent),
    )


# ============================================================
# 3) ESTADÍSTICAS DE USO (cronómetro y hora pico)
# ============================================================
class VentanaEstadisticasUso(tk.Toplevel):
    """Muestra el tiempo de uso del sistema, la hora pico y un historial
    de sesiones."""

    def __init__(self, parent, usuario_actual=None, tiempo_sesion_actual=0):
        super().__init__(parent)
        from utilidades_ui import ajustar_tamaño_ventana

        self.title("Estadísticas de Uso del Sistema")
        self.resizable(True, True)
        self.configure(bg="white")
        self.grab_set()
        # NOTA: no usamos transient(parent) a propósito. En Windows, una
        # ventana Toplevel marcada como "transient" de su padre pierde los
        # botones nativos de minimizar y maximizar (queda solo con cerrar).
        try:
            self.attributes("-toolwindow", False)
        except tk.TclError:
            pass

        self.usuario_actual = usuario_actual

        barra = tk.Frame(self, bg=AZUL_RIBBON, height=40)
        barra.pack(fill="x")
        barra.pack_propagate(False)
        tk.Label(barra, text=t("uso_titulo"), font=("Segoe UI", 12, "bold"),
                 bg=AZUL_RIBBON, fg="white").pack(side="left", padx=20, pady=10)

        self.boton_reporte = BotonReporteGeneral(
            barra, obtener_datos_callback=self._obtener_datos_reporte,
            nombre_archivo_base="Reporte_Uso_del_Sistema",
        )
        self.boton_reporte.generador_pdf_dashboard = self._generar_pdf_dashboard
        self.boton_reporte.generador_excel = self._generar_excel
        self.boton_reporte.pack(side="right", padx=15, pady=6)

        # --- KPIs ---
        kpis = tk.Frame(self, bg="white", padx=15, pady=15)
        kpis.pack(fill="x")

        self.tiempo_sesion_actual = tiempo_sesion_actual
        self.uso_hoy = tiempo_total_hoy()
        self.uso_general = tiempo_total_general()
        self.hora_pico_datos = hora_pico()
        self.sesiones_listadas = ultimas_sesiones(30)

        hoy = self.uso_hoy
        general = self.uso_general
        pico = self.hora_pico_datos

        # Tarjetas KPI
        self._tarjeta(kpis, "Sesión actual", formatear_duracion(tiempo_sesion_actual),
                      AZUL_RIBBON).pack(side="left", padx=5, fill="x", expand=True)
        self._tarjeta(kpis, "Uso de hoy", formatear_duracion(hoy),
                      VERDE).pack(side="left", padx=5, fill="x", expand=True)
        self._tarjeta(kpis, "Uso total acumulado", formatear_duracion(general),
                      "#7c3aed").pack(side="left", padx=5, fill="x", expand=True)
        if pico:
            h, segs = pico
            texto_pico = f"{h:02d}:00 a {h:02d}:59"
            sub = f"({formatear_duracion(segs)} acumulado)"
        else:
            texto_pico = "Sin datos"
            sub = ""
        self._tarjeta(kpis, "Hora pico de uso", texto_pico, "#f59e0b", sub=sub
                      ).pack(side="left", padx=5, fill="x", expand=True)

        # --- Distribución por hora ---
        tk.Label(self, text=t("uso_actividad_por_hora"),
                 font=("Segoe UI", 10, "bold"), bg="white").pack(anchor="w", padx=20, pady=(8, 4))
        frame_grafico = tk.Frame(self, bg="white", padx=20)
        frame_grafico.pack(fill="x")
        self._dibujar_grafico_horas(frame_grafico)

        # --- Historial de sesiones ---
        tk.Label(self, text=t("uso_ultimas_sesiones"),
                 font=("Segoe UI", 10, "bold"), bg="white").pack(anchor="w", padx=20, pady=(12, 4))
        cont = tk.Frame(self, bg="white", padx=15)
        cont.pack(fill="both", expand=True)
        cols = ("usuario", "inicio", "fin", "duracion")
        nombres = (t("uso_col_usuario"), t("uso_col_inicio"), t("uso_col_fin"), t("licencia_col_duracion"))
        anchos = (140, 160, 160, 100)
        tabla = ttk.Treeview(cont, columns=cols, show="headings", height=8)
        habilitar_deseleccion_treeview(tabla)
        for c, n, a in zip(cols, nombres, anchos):
            tabla.heading(c, text=n)
            tabla.column(c, width=a, anchor="center")
        tabla.column("usuario", anchor="w")
        cont.grid_rowconfigure(0, weight=1)
        cont.grid_columnconfigure(0, weight=1)
        scroll = ttk.Scrollbar(cont, orient="vertical", command=tabla.yview)
        scroll_h = ttk.Scrollbar(cont, orient="horizontal", command=tabla.xview)
        tabla.configure(yscrollcommand=scroll.set, xscrollcommand=scroll_h.set)
        tabla.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")
        scroll_h.grid(row=1, column=0, sticky="ew")

        for s in self.sesiones_listadas:
            tabla.insert("", "end", values=(
                s["usuario_nombre"], s["fecha_inicio"], s["fecha_fin"],
                formatear_duracion(s["duracion_segundos"]),
            ))

        tk.Button(self, text=t("cerrar"), font=("Segoe UI", 9, "bold"),
                  bg="white", fg=AZUL_OSCURO, relief="solid", bd=1, padx=20, pady=6,
                  cursor="hand2", command=self.destroy).pack(pady=10)

        ajustar_tamaño_ventana(self, ancho_min=640, alto_min=480,
                               margen_alto=20, ancho_max=1100, alto_max=800)

    def _tarjeta(self, parent, label, valor, color, sub=""):
        card = tk.Frame(parent, bg="white", relief="solid", bd=1, padx=10, pady=10)
        tk.Label(card, text=label, font=("Segoe UI", 9), bg="white",
                 fg="#6b7280").pack(anchor="w")
        tk.Label(card, text=valor, font=("Segoe UI", 14, "bold"),
                 bg="white", fg=color).pack(anchor="w", pady=(4, 0))
        if sub:
            tk.Label(card, text=sub, font=("Segoe UI", 8),
                     bg="white", fg="#9ca3af").pack(anchor="w")
        return card

    def _dibujar_grafico_horas(self, parent):
        datos = actividad_por_hora()
        self.actividad_por_hora_datos = datos  # para incluir en la exportación
        if not datos:
            tk.Label(parent, text=t("uso_sin_datos"),
                     font=("Segoe UI", 9, "italic"), bg="white",
                     fg="#9ca3af").pack(anchor="w")
            return
        mapa = {h: s for h, s in datos}
        max_seg = max(mapa.values()) if mapa else 1

        canvas = tk.Canvas(parent, height=140, bg="white", highlightthickness=0)
        canvas.pack(fill="x")

        def redibujar(event=None):
            canvas.delete("all")
            w = canvas.winfo_width() or 700
            altura_grafico = 100
            margen = 10
            ancho_barra = (w - 2 * margen) / 24
            for hora in range(24):
                segs = mapa.get(hora, 0)
                alto = (segs / max_seg) * altura_grafico if max_seg else 0
                x = margen + hora * ancho_barra
                y2 = margen + altura_grafico
                y1 = y2 - alto
                color = AZUL_RIBBON if segs == max_seg else "#93c5fd"
                canvas.create_rectangle(x + 2, y1, x + ancho_barra - 2, y2,
                                         fill=color, outline="")
                canvas.create_text(x + ancho_barra / 2, y2 + 12,
                                    text=f"{hora:02d}", font=("Segoe UI", 7),
                                    fill="#6b7280")

        canvas.bind("<Configure>", redibujar)
        canvas.after(50, redibujar)
    # ---------------- REPORTE GENERAL ----------------
    def _generar_pdf_dashboard(self, ruta: str):
        """PDF con dashboard: portada con KPIs, gráfico de barras de
        actividad por hora y tabla de historial de sesiones."""
        from reporte_uso_pdf import generar_pdf_uso_sistema
        nombre_usuario = (
            self.usuario_actual.get("nombre_completo", "")
            if self.usuario_actual else ""
        )
        generar_pdf_uso_sistema(
            ruta_destino=ruta,
            tiempo_sesion_actual=self.tiempo_sesion_actual,
            uso_hoy=self.uso_hoy,
            uso_general=self.uso_general,
            hora_pico_datos=self.hora_pico_datos,
            actividad_por_hora_datos=getattr(self, "actividad_por_hora_datos", []),
            sesiones=self.sesiones_listadas or [],
            generado_por=nombre_usuario,
        )

    def _obtener_datos_reporte(self) -> dict:
        """Estructura neutral consumida por todos los formatos de exportación
        (PDF simple, Word, ODT, CSV, JSON). Refleja los mismos datos que se
        ven en pantalla: KPIs de la sesión actual, uso de hoy y total,
        hora pico, actividad por hora y las últimas 30 sesiones."""
        nombre_usuario = (
            self.usuario_actual.get("nombre_completo", "")
            if self.usuario_actual else ""
        )

        # ── Hora pico ──────────────────────────────────────────
        if self.hora_pico_datos:
            h, segs = self.hora_pico_datos
            texto_pico = f"{h:02d}:00 a {h:02d}:59  ({formatear_duracion(segs)} acumulado)"
        else:
            texto_pico = "Sin datos suficientes"

        # ── Actividad por hora (tabla de 24 horas) ─────────────
        mapa_horas = dict(getattr(self, "actividad_por_hora_datos", []) or [])
        filas_horas = [
            [f"{h:02d}:00", formatear_duracion(mapa_horas.get(h, 0)),
             str(mapa_horas.get(h, 0)) + " seg"]
            for h in range(24)
        ]

        # ── Historial de sesiones ──────────────────────────────
        filas_sesiones = []
        for s in (self.sesiones_listadas or []):
            filas_sesiones.append([
                s.get("usuario_nombre", ""),
                s.get("fecha_inicio", ""),
                s.get("fecha_fin", "") or "—",
                formatear_duracion(s.get("duracion_segundos", 0)),
            ])

        import datetime
        secciones = [
            {
                "tipo": "resumen",
                "titulo": "RESUMEN DE USO",
                "filas": [
                    ("Sesión actual", formatear_duracion(self.tiempo_sesion_actual)),
                    ("Uso de hoy", formatear_duracion(self.uso_hoy)),
                    ("Uso total acumulado", formatear_duracion(self.uso_general)),
                    ("Hora pico de uso", texto_pico),
                ],
            },
            {
                "tipo": "tabla",
                "titulo": "ACTIVIDAD POR HORA DEL DÍA",
                "encabezados": ["Hora", "Duración", "Segundos"],
                "filas": filas_horas,
            },
            {
                "tipo": "tabla",
                "titulo": "ÚLTIMAS SESIONES (30 más recientes)",
                "encabezados": ["Usuario", "Inicio", "Fin", "Duración"],
                "filas": filas_sesiones,
            },
        ]

        return {
            "titulo": "Reporte de Uso del Sistema",
            "subtitulo": f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}",
            "generado_por": nombre_usuario,
            "secciones": secciones,
        }

    def _generar_excel(self, ruta: str):
        """Excel con tres hojas: Resumen KPIs, Actividad por Hora y
        Historial de Sesiones."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        def _enc(ws, fila_num):
            for cell in ws[fila_num]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1D5FD6")
                cell.alignment = Alignment(horizontal="center")

        nombre_usuario = (
            self.usuario_actual.get("nombre_completo", "")
            if self.usuario_actual else ""
        )
        if self.hora_pico_datos:
            h, segs = self.hora_pico_datos
            texto_pico = f"{h:02d}:00 a {h:02d}:59  ({formatear_duracion(segs)} acumulado)"
        else:
            texto_pico = "Sin datos"

        wb = openpyxl.Workbook()

        # ── Hoja 1: Resumen KPIs ──────────────────────────────
        ws1 = wb.active
        ws1.title = "Resumen KPIs"
        ws1.append(["Reporte de Uso del Sistema"])
        ws1["A1"].font = Font(bold=True, size=14)
        ws1.append([f"Generado por: {nombre_usuario}"])
        ws1.append([])
        ws1.append(["Indicador", "Valor"])
        _enc(ws1, 4)
        ws1.append(["Sesión actual",          formatear_duracion(self.tiempo_sesion_actual)])
        ws1.append(["Uso de hoy",             formatear_duracion(self.uso_hoy)])
        ws1.append(["Uso total acumulado",    formatear_duracion(self.uso_general)])
        ws1.append(["Hora pico de uso",       texto_pico])
        ws1.column_dimensions["A"].width = 30
        ws1.column_dimensions["B"].width = 35

        # ── Hoja 2: Actividad por Hora ────────────────────────
        ws2 = wb.create_sheet("Actividad por Hora")
        ws2.append(["Hora", "Duración", "Segundos acumulados"])
        _enc(ws2, 1)
        mapa_horas = dict(getattr(self, "actividad_por_hora_datos", []) or [])
        for h in range(24):
            segs = mapa_horas.get(h, 0)
            ws2.append([f"{h:02d}:00", formatear_duracion(segs), segs])
        ws2.column_dimensions["A"].width = 10
        ws2.column_dimensions["B"].width = 22
        ws2.column_dimensions["C"].width = 24

        # ── Hoja 3: Historial de Sesiones ─────────────────────
        ws3 = wb.create_sheet("Historial de Sesiones")
        ws3.append(["Usuario", "Inicio", "Fin", "Duración"])
        _enc(ws3, 1)
        for s in (self.sesiones_listadas or []):
            ws3.append([
                s.get("usuario_nombre", ""),
                s.get("fecha_inicio", ""),
                s.get("fecha_fin", "") or "—",
                formatear_duracion(s.get("duracion_segundos", 0)),
            ])
        for col in ws3.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws3.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(ruta)
