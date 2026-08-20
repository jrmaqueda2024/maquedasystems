"""
ventana_rrhh.py
Módulo de Recursos Humanos: personal, asistencia, adelantos y reportes.
Solo accesible para Administradores.
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import datetime
import os
import shutil

from models_rrhh import (
    inicializar_tablas_rrhh, listar_empleados, obtener_empleado,
    crear_empleado, editar_empleado, desactivar_empleado, activar_empleado,
    registrar_asistencia, obtener_asistencia_dia, obtener_resumen_periodo,
    calcular_liquidacion, listar_adelantos, registrar_adelanto,
    marcar_adelanto_descontado, eliminar_adelanto, sum_adelantos_pendientes,
    sum_adelantos_pendientes_por_empleado,
    resumen_rrhh, carpeta_adjuntos_rrhh, ESTADOS_ASISTENCIA,
)
from utilidades_ui import forzar_mayusculas, ajustar_tamaño_ventana, formatear_gs, habilitar_deseleccion_treeview
from traducciones import t
from menu_reporte_general import BotonReporteGeneral
from widget_calendario import abrir_selector_fecha


def _fecha_o_hoy(texto: str) -> datetime.date:
    """Convierte un texto ISO (YYYY-MM-DD) a date; si está vacío o mal
    formado, devuelve la fecha de hoy en vez de romper el selector."""
    try:
        return datetime.date.fromisoformat(texto)
    except (ValueError, TypeError):
        return datetime.date.today()

AZUL       = "#1d5fd6"
AZUL_OSC   = "#163d8c"
GRIS_FONDO = "#f4f5f7"
GRIS_BORDE = "#e2e8f0"
BLANCO     = "#ffffff"
VERDE      = "#16a34a"
ROJO       = "#dc2626"
NARANJA    = "#d97706"
NEGRO      = "#1e293b"
GRIS_TEXT  = "#6b7280"


# ═════════════════════════════════════════════════════════════
#  PANEL PRINCIPAL
# ═════════════════════════════════════════════════════════════
class PanelRRHH(tk.Frame):
    def __init__(self, parent, usuario_actual):
        super().__init__(parent, bg=BLANCO)
        self.usuario_actual = usuario_actual
        inicializar_tablas_rrhh()
        self._construir_ui()

    def _construir_ui(self):
        # ── Encabezado ────────────────────────────────────────
        enc = tk.Frame(self, bg=AZUL, height=54)
        enc.pack(fill="x")
        enc.pack_propagate(False)
        tk.Label(enc, text=t("rrhh_titulo"),
                 font=("Segoe UI", 15, "bold"), bg=AZUL, fg=BLANCO
                 ).pack(side="left", padx=20, pady=12)

        self.boton_reporte = BotonReporteGeneral(
            enc, obtener_datos_callback=self._obtener_datos_reporte,
            nombre_archivo_base="Reporte_RRHH",
        )
        self.boton_reporte.generador_excel = self._generar_excel
        self.boton_reporte.pack(side="right", padx=(0, 12), pady=12)

        # ── Notebook con 3 pestañas ───────────────────────────
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=8, pady=6)

        tab_personal   = tk.Frame(self.nb, bg=BLANCO)
        tab_asistencia = tk.Frame(self.nb, bg=BLANCO)
        tab_adelantos  = tk.Frame(self.nb, bg=BLANCO)

        self.nb.add(tab_personal,   text=t("rrhh_tab_personal"))
        self.nb.add(tab_asistencia, text=t("rrhh_tab_asistencia"))
        self.nb.add(tab_adelantos,  text=t("rrhh_tab_adelantos"))

        cargar_personal   = _construir_tab_personal(tab_personal,   self.usuario_actual)
        cargar_asistencia = _construir_tab_asistencia(tab_asistencia, self.usuario_actual)
        # Adelantos avisa a Personal (on_cambio_externo) apenas se registra,
        # marca o elimina un vale, para que la columna "ADELANTOS PEND." y
        # el resumen inferior de Personal se vean actualizados al instante,
        # sin tener que cambiar de pestaña ni salir y volver a entrar al
        # módulo (antes había que hacer eso para que se reflejara).
        cargar_adelantos = _construir_tab_adelantos(
            tab_adelantos, self.usuario_actual, on_cambio_externo=lambda: cargar_personal())

        # Además, cambiar de pestaña siempre refresca TODO (por si se editó
        # algo desde otra ventana, como Asistencia Técnica o el reporte),
        # igual que ya hacen Streaming e Importaciones.
        self._cargas_pestañas = [cargar_personal, cargar_asistencia, cargar_adelantos]
        self.nb.bind("<<NotebookTabChanged>>", lambda e: self._refrescar_pestaña_visible())

    def _refrescar_pestaña_visible(self):
        try:
            idx = self.nb.index(self.nb.select())
        except tk.TclError:
            return
        if 0 <= idx < len(self._cargas_pestañas):
            self._cargas_pestañas[idx]()

    # ── Datos para el Reporte General ─────────────────────────
    def _obtener_datos_reporte(self) -> dict:
        r = resumen_rrhh()
        empleados = listar_empleados(solo_activos=False)
        adelantos = listar_adelantos()
        nombre_usuario = self.usuario_actual.get("nombre_completo", "")

        secciones = [
            {
                "tipo": "resumen",
                "titulo": "RESUMEN DE RECURSOS HUMANOS",
                "filas": [
                    ("Empleados activos",     str(r["cantidad_empleados"])),
                    ("Total sueldos/mes",     formatear_gs(r["total_sueldos"])),
                    ("Adelantos pendientes",  formatear_gs(r["total_adelantos"])),
                    ("Resta del sueldo",      formatear_gs(r["total_resta_sueldo"])),
                ],
            },
            {
                "tipo": "tabla",
                "titulo": "NÓMINA DE PERSONAL",
                "encabezados": ["ID", "Nombre", "Cargo", "Departamento",
                                "Ingreso", "Sueldo", "Estado"],
                "filas": [
                    [str(e["id"]), e["nombre"], e["cargo"], e["departamento"],
                     e["fecha_ingreso"], formatear_gs(e["sueldo_mensual"]),
                     "Activo" if e["activo"] else "Inactivo"]
                    for e in empleados
                ],
            },
            {
                "tipo": "tabla",
                "titulo": "ADELANTOS / VALES",
                "encabezados": ["ID", "Empleado", "Fecha", "Monto",
                                "Descripción", "Estado"],
                "filas": [
                    [str(a["id"]), a["empleado_nombre"], a["fecha"],
                     formatear_gs(a["monto"]), a["descripcion"], a["estado"]]
                    for a in adelantos
                ],
            },
        ]
        return {
            "titulo":        "Reporte de Recursos Humanos",
            "subtitulo":     f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}",
            "generado_por":  nombre_usuario,
            "secciones":     secciones,
        }

    def _generar_excel(self, ruta: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        def _enc(ws):
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1D5FD6")
                cell.alignment = Alignment(horizontal="center")

        r = resumen_rrhh()
        wb = openpyxl.Workbook()

        # Hoja resumen
        ws1 = wb.active
        ws1.title = "Resumen"
        ws1.append(["Indicador", "Valor"])
        _enc(ws1)
        ws1.append(["Empleados activos",    r["cantidad_empleados"]])
        ws1.append(["Total sueldos/mes",    r["total_sueldos"]])
        ws1.append(["Adelantos pendientes", r["total_adelantos"]])
        ws1.append(["Resta del sueldo",     r["total_resta_sueldo"]])

        # Hoja nómina
        ws2 = wb.create_sheet("Nómina")
        ws2.append(["ID", "Nombre", "Cargo", "Departamento",
                    "Ingreso", "Sueldo", "Horas/día", "Estado"])
        _enc(ws2)
        for e in listar_empleados(solo_activos=False):
            ws2.append([e["id"], e["nombre"], e["cargo"], e["departamento"],
                        e["fecha_ingreso"], e["sueldo_mensual"], e["horas_dia"],
                        "Activo" if e["activo"] else "Inactivo"])

        # Hoja adelantos
        ws3 = wb.create_sheet("Adelantos")
        ws3.append(["ID", "Empleado", "Fecha", "Monto", "Descripción", "Estado"])
        _enc(ws3)
        for a in listar_adelantos():
            ws3.append([a["id"], a["empleado_nombre"], a["fecha"],
                        a["monto"], a["descripcion"], a["estado"]])

        wb.save(ruta)


# ═════════════════════════════════════════════════════════════
#  TAB PERSONAL
# ═════════════════════════════════════════════════════════════

def _agregar_deseleccion(tabla, extra_callback=None):
    """Click en área vacía o Escape deselecciona la fila actual."""
    def _deseleccionar(event=None):
        tabla.selection_remove(*tabla.selection())
        if extra_callback:
            extra_callback()

    def _click_vacio(event):
        if not tabla.identify_row(event.y):
            _deseleccionar()

    tabla.bind("<Button-1>", _click_vacio, add="+")
    tabla.bind("<Escape>",   lambda e: _deseleccionar())

def _construir_tab_personal(parent, usuario_actual):
    frame = tk.Frame(parent, bg=BLANCO)
    frame.pack(fill="both", expand=True)
    frame.grid_rowconfigure(1, weight=1)
    frame.grid_columnconfigure(0, weight=1)

    # Barra superior
    barra = tk.Frame(frame, bg=GRIS_FONDO)
    barra.grid(row=0, column=0, sticky="ew", padx=8, pady=6)

    var_busqueda = tk.StringVar()
    var_activos  = tk.BooleanVar(value=True)

    tk.Label(barra, text="Buscar:", font=("Segoe UI", 9),
             bg=GRIS_FONDO).pack(side="left")
    entry_busq = tk.Entry(barra, textvariable=var_busqueda,
                          font=("Segoe UI", 9), width=30)
    entry_busq.pack(side="left", padx=(4, 12), ipady=3)

    tk.Checkbutton(barra, text="Solo activos", variable=var_activos,
                   bg=GRIS_FONDO, font=("Segoe UI", 9),
                   command=lambda: _cargar_personal()).pack(side="left")

    tk.Button(barra, text="+ Nuevo Empleado",
              font=("Segoe UI", 9, "bold"), bg=VERDE, fg=BLANCO,
              relief="flat", padx=10, pady=4, cursor="hand2",
              command=lambda: _abrir_formulario_empleado(
                  parent, None, _cargar_personal)
              ).pack(side="right", padx=4)

    entry_busq.bind("<KeyRelease>", lambda e: _cargar_personal())
    var_busqueda.trace_add("write", lambda *_: _cargar_personal())

    # Grilla
    contenedor = tk.Frame(frame, bg=BLANCO)
    contenedor.grid(row=1, column=0, sticky="nsew", padx=8)
    contenedor.grid_rowconfigure(0, weight=1)
    contenedor.grid_columnconfigure(0, weight=1)

    cols   = ("id", "nombre", "cargo", "departamento", "ingreso",
              "antiguedad", "sueldo", "horas_dia", "adelantos")
    encabs = ("ID", "NOMBRE", "CARGO", "DEPARTAMENTO", "INGRESO",
              "ANTIGÜEDAD", "SUELDO/MES", "HS/DÍA", "ADELANTOS PEND.")
    anchos = (40, 200, 130, 120, 90, 90, 110, 70, 110)

    tabla = ttk.Treeview(contenedor, columns=cols, show="headings",
                          selectmode="browse")
    habilitar_deseleccion_treeview(tabla)
    for col, enc, ancho in zip(cols, encabs, anchos):
        tabla.heading(col, text=enc)
        tabla.column(col, width=ancho,
                     anchor="w" if col == "nombre" else "center")

    tabla.tag_configure("inactivo", foreground="#9ca3af",
                        background="#f3f4f6")

    sb_y = ttk.Scrollbar(contenedor, orient="vertical",   command=tabla.yview)
    sb_x = ttk.Scrollbar(contenedor, orient="horizontal", command=tabla.xview)
    tabla.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
    tabla.grid(row=0, column=0, sticky="nsew")
    sb_y.grid(row=0, column=1, sticky="ns")
    sb_x.grid(row=1, column=0, sticky="ew")

    # Panel resumen inferior
    panel_res = tk.Frame(frame, bg=GRIS_FONDO)
    panel_res.grid(row=2, column=0, sticky="ew", padx=8, pady=4)

    lbl_cant     = tk.Label(panel_res, text="", font=("Segoe UI", 10, "bold"),
                            bg=GRIS_FONDO)
    lbl_sueldo   = tk.Label(panel_res, text="", font=("Segoe UI", 10, "bold"),
                            bg=GRIS_FONDO, fg=AZUL)
    lbl_adelant  = tk.Label(panel_res, text="", font=("Segoe UI", 10, "bold"),
                            bg=GRIS_FONDO, fg=NARANJA)
    lbl_resta    = tk.Label(panel_res, text="", font=("Segoe UI", 10, "bold"),
                            bg=GRIS_FONDO, fg=VERDE)

    for lbl in (lbl_cant, lbl_sueldo, lbl_adelant, lbl_resta):
        lbl.pack(side="left", padx=20, pady=6)

    def _calcular_antiguedad(fecha_ingreso: str) -> str:
        try:
            fi = datetime.date.fromisoformat(fecha_ingreso)
            delta = datetime.date.today() - fi
            años  = delta.days // 365
            meses = (delta.days % 365) // 30
            if años > 0:
                return f"{años}a {meses}m"
            return f"{meses} mes{'es' if meses != 1 else ''}"
        except Exception:
            return "—"

    def _cargar_personal():
        for r in tabla.get_children():
            tabla.delete(r)
        texto = var_busqueda.get().lower()
        empleados = listar_empleados(solo_activos=var_activos.get())
        # Una sola consulta agregada para TODOS los empleados, en vez de
        # una consulta (con su propia conexión a la BD) por cada fila de
        # la grilla: antes esto era lo que hacía que agregar un empleado
        # o un adelanto se sintiera lento, sobre todo con varios empleados
        # cargados.
        adelantos_por_empleado = sum_adelantos_pendientes_por_empleado()
        total_sueldos = 0
        total_adelantos = 0
        for e in empleados:
            if texto and texto not in e["nombre"].lower() \
                    and texto not in e["cargo"].lower():
                continue
            adel = adelantos_por_empleado.get(e["id"], 0)
            total_sueldos  += e["sueldo_mensual"]
            total_adelantos += adel
            tag = () if e["activo"] else ("inactivo",)
            tabla.insert("", "end", iid=str(e["id"]), values=(
                e["id"], e["nombre"], e["cargo"], e["departamento"],
                e["fecha_ingreso"], _calcular_antiguedad(e["fecha_ingreso"]),
                formatear_gs(e["sueldo_mensual"]),
                f"{e['horas_dia']:g} hs",
                formatear_gs(adel) if adel > 0 else "—",
            ), tags=tag)

        r = resumen_rrhh()
        lbl_cant.config(text=f"Empleados activos: {r['cantidad_empleados']}")
        lbl_sueldo.config(text=f"Total sueldos/mes: {formatear_gs(r['total_sueldos'])}")
        lbl_adelant.config(text=f"Adelantos pendientes: {formatear_gs(r['total_adelantos'])}")
        lbl_resta.config(text=f"Resta del sueldo: {formatear_gs(r['total_resta_sueldo'])}")

    def _abrir_menu_contextual(event):
        iid = tabla.identify_row(event.y)
        if not iid:
            return
        tabla.selection_set(iid)
        emp = obtener_empleado(int(iid))
        menu = tk.Menu(tabla, tearoff=0)
        menu.add_command(label="✏ Editar empleado",
                         command=lambda: _abrir_formulario_empleado(
                             parent, emp, _cargar_personal))
        menu.add_command(label="💵 Ver adelantos",
                         command=lambda: _ver_adelantos_empleado(parent, emp))
        menu.add_separator()
        if emp["activo"]:
            menu.add_command(label="🚫 Desactivar empleado",
                             command=lambda: _toggle_activo(emp, False))
        else:
            menu.add_command(label="✅ Activar empleado",
                             command=lambda: _toggle_activo(emp, True))
        menu.tk_popup(event.x_root, event.y_root)

    def _toggle_activo(emp, activar):
        accion = "activar" if activar else "desactivar"
        if messagebox.askyesno("Confirmar", f"¿Deseas {accion} a {emp['nombre']}?"):
            if activar:
                activar_empleado(emp["id"])
            else:
                desactivar_empleado(emp["id"])
            _cargar_personal()

    tabla.bind("<Button-3>", _abrir_menu_contextual)
    tabla.bind("<Double-1>", lambda e: (
        _abrir_formulario_empleado(
            parent, obtener_empleado(int(tabla.selection()[0])), _cargar_personal)
        if tabla.selection() else None
    ))
    _agregar_deseleccion(tabla)

    _cargar_personal()
    return _cargar_personal


# ═════════════════════════════════════════════════════════════
#  FORMULARIO EMPLEADO
# ═════════════════════════════════════════════════════════════
def _abrir_formulario_empleado(parent, empleado, on_guardado):
    es_edicion = empleado is not None
    win = tk.Toplevel(parent)
    win.title("Editar Empleado" if es_edicion else "Nuevo Empleado")
    win.configure(bg=BLANCO)
    win.grab_set()
    try:
        win.attributes("-toolwindow", False)
    except Exception:
        pass

    campos = {}

    def _campo(frame, label, key, row, col=0, ancho=30,
               mayusculas=False, placeholder=""):
        tk.Label(frame, text=label, font=("Segoe UI", 9, "bold"),
                 bg=BLANCO).grid(row=row, column=col*2, sticky="w",
                                 padx=(12, 4), pady=4)
        var = tk.StringVar(value=str(empleado.get(key, "") if es_edicion else ""))

        COLOR_PH = "#9ca3af"
        COLOR_OK = "#1e293b"

        entry = tk.Entry(frame, font=("Segoe UI", 9), width=ancho, fg=COLOR_OK)
        entry.grid(row=row, column=col*2+1, sticky="ew", padx=(0, 12), pady=4)

        # Mostrar valor real o placeholder (si no hay ninguno de los dos,
        # el campo queda vacío con el color normal, listo para escribir)
        if es_edicion and var.get():
            entry.insert(0, var.get())
            entry.config(fg=COLOR_OK)
        elif placeholder:
            entry.insert(0, placeholder)
            entry.config(fg=COLOR_PH)

        def _focus_in(e, _e=entry, _ph=placeholder, _var=var):
            if placeholder and _e.get() == _ph:
                _e.delete(0, "end")
                _e.config(fg=COLOR_OK)

        def _focus_out(e, _e=entry, _ph=placeholder, _var=var):
            val = _e.get().strip()
            if placeholder and (not val or val == _ph):
                _e.delete(0, "end")
                _e.insert(0, _ph)
                _e.config(fg=COLOR_PH)
                _var.set("")
            else:
                _e.config(fg=COLOR_OK)
                _var.set(val)

        def _key(e, _e=entry, _var=var, _ph=placeholder):
            val = _e.get()
            if val != _ph:
                _e.config(fg=COLOR_OK)
                _var.set(val)

        entry.bind("<FocusIn>",    _focus_in)
        entry.bind("<FocusOut>",   _focus_out)
        entry.bind("<KeyRelease>", _key)

        if mayusculas:
            forzar_mayusculas(entry, var)

        campos[key] = var
        return entry

    barra = tk.Frame(win, bg=AZUL, height=36)
    barra.pack(fill="x")
    barra.pack_propagate(False)
    tk.Label(barra, text="Datos del Empleado",
             font=("Segoe UI", 11, "bold"), bg=AZUL, fg=BLANCO
             ).pack(side="left", padx=12, pady=8)

    cuerpo = tk.Frame(win, bg=BLANCO)
    cuerpo.pack(fill="both", expand=True)
    cuerpo.grid_columnconfigure(1, weight=1)
    cuerpo.grid_columnconfigure(3, weight=1)

    _campo(cuerpo, "Nombre completo:", "nombre",       0, 0, 35, mayusculas=True)
    _campo(cuerpo, "Cargo:",           "cargo",        1, 0, 35, mayusculas=True)
    _campo(cuerpo, "Departamento:",    "departamento", 2, 0, 35, mayusculas=True)
    _campo(cuerpo, "Teléfono:",        "telefono",     3, 0, 35, mayusculas=True)
    _campo(cuerpo, "Email:",           "email",        4, 0, 35)

    # Forzar minúsculas en el campo Email
    def _forzar_minusculas_email(*args):
        val = campos["email"].get()
        val_lower = val.lower()
        if val != val_lower:
            campos["email"].set(val_lower)
    campos["email"].trace_add("write", _forzar_minusculas_email)

    tk.Label(cuerpo, text="Fecha de ingreso:", font=("Segoe UI", 9, "bold"),
             bg=BLANCO).grid(row=0, column=2, sticky="w", padx=(12, 4), pady=4)
    var_fi = tk.StringVar(value=empleado.get("fecha_ingreso", "") if es_edicion
                          else datetime.date.today().isoformat())
    frame_fi = tk.Frame(cuerpo, bg=BLANCO)
    frame_fi.grid(row=0, column=3, sticky="w", padx=(0, 12))
    lbl_fi = tk.Label(frame_fi, textvariable=var_fi, font=("Segoe UI", 9),
                      bg=BLANCO, width=14, relief="solid", bd=1)
    lbl_fi.pack(side="left")
    tk.Button(frame_fi, text="📅", font=("Segoe UI", 9), relief="flat",
              cursor="hand2",
              command=lambda: abrir_selector_fecha(
                  win, _fecha_o_hoy(var_fi.get()), lambda d: var_fi.set(d.isoformat()))
              ).pack(side="left", padx=2)
    campos["fecha_ingreso"] = var_fi

    _campo(cuerpo, "Sueldo mensual (Gs):", "sueldo_mensual", 1, 1, 18, mayusculas=True)
    _campo(cuerpo, "Horas por día:",       "horas_dia",      2, 1, 18, mayusculas=True)

    tk.Label(cuerpo, text="Observaciones:", font=("Segoe UI", 9, "bold"),
             bg=BLANCO).grid(row=3, column=2, sticky="nw", padx=(12, 4), pady=4)
    var_obs = tk.StringVar(value=empleado.get("observaciones", "") if es_edicion else "")
    txt_obs = tk.Text(cuerpo, font=("Segoe UI", 9), width=28, height=3, fg="#1e293b")
    txt_obs.grid(row=3, column=3, rowspan=2, sticky="ew", padx=(0, 12), pady=4)
    if es_edicion and empleado.get("observaciones", ""):
        txt_obs.insert("1.0", empleado.get("observaciones", ""))

    def _mayus_text(event):
        if event.char and event.char.isalpha():
            txt_obs.insert(tk.INSERT, event.char.upper())
            return "break"
    txt_obs.bind("<Key>", _mayus_text)

    # Botones
    pie = tk.Frame(win, bg=BLANCO)
    pie.pack(fill="x", pady=8)

    def _guardar():
        nombre = campos["nombre"].get().strip()
        if not nombre:
            messagebox.showerror("Error", "El nombre es obligatorio.")
            return
        try:
            sueldo = float(campos["sueldo_mensual"].get().replace(".", "").replace(",", ".") or 0)
            horas  = float(campos["horas_dia"].get().replace(",", ".") or 8)
        except ValueError:
            messagebox.showerror("Error", "Sueldo y horas deben ser números.")
            return

        obs = txt_obs.get("1.0", "end").strip()
        args = (
            nombre, campos["cargo"].get().strip(),
            campos["departamento"].get().strip(),
            campos["telefono"].get().strip(),
            campos["email"].get().strip(),
            campos["fecha_ingreso"].get(),
            sueldo, horas, obs,
        )
        if es_edicion:
            editar_empleado(empleado["id"], *args)
        else:
            crear_empleado(*args)
        on_guardado()
        win.destroy()
        messagebox.showinfo("Listo", "Empleado guardado correctamente.")

    btn_g_emp = tk.Button(pie, text="✔ Guardar", font=("Segoe UI", 10, "bold"),
              bg=VERDE, fg=BLANCO, relief="flat", padx=16, pady=6,
              cursor="hand2", command=_guardar)
    btn_g_emp.pack(side="left", padx=16)
    btn_g_emp.bind("<Return>", lambda e: _guardar())
    btn_c_emp = tk.Button(pie, text="✖ Cancelar", font=("Segoe UI", 10),
              bg=BLANCO, relief="solid", bd=1, padx=16, pady=5,
              cursor="hand2", command=win.destroy)
    btn_c_emp.pack(side="right", padx=16)
    btn_c_emp.bind("<Return>", lambda e: win.destroy())

    ajustar_tamaño_ventana(win, ancho_min=620, alto_min=340, margen_alto=20)


def _ver_adelantos_empleado(parent, empleado):
    """Abre la pestaña de adelantos filtrada por ese empleado."""
    messagebox.showinfo("Adelantos",
                        f"Para ver los adelantos de {empleado['nombre']}, "
                        "usá la pestaña 'Adelantos / Vales' y filtrá por nombre.")


# ═════════════════════════════════════════════════════════════
#  TAB ASISTENCIA
# ═════════════════════════════════════════════════════════════
def _construir_tab_asistencia(parent, usuario_actual):
    frame = tk.Frame(parent, bg=BLANCO)
    frame.pack(fill="both", expand=True)
    frame.grid_rowconfigure(2, weight=1)
    frame.grid_columnconfigure(0, weight=1)

    hoy = datetime.date.today()
    var_fecha = tk.StringVar(value=hoy.isoformat())

    # ── Controles de fecha ─────────────────────────────────
    barra = tk.Frame(frame, bg=GRIS_FONDO)
    barra.grid(row=0, column=0, sticky="ew", padx=8, pady=6)

    tk.Label(barra, text="Fecha:", font=("Segoe UI", 9, "bold"),
             bg=GRIS_FONDO).pack(side="left", padx=(8, 4))

    def _ir_dia(delta):
        try:
            d = datetime.date.fromisoformat(var_fecha.get())
            var_fecha.set((d + datetime.timedelta(days=delta)).isoformat())
            _cargar_asistencia()
        except Exception:
            pass

    tk.Button(barra, text="◀", font=("Segoe UI", 10), bg=GRIS_FONDO,
              relief="flat", cursor="hand2",
              command=lambda: _ir_dia(-1)).pack(side="left")

    lbl_fecha = tk.Label(barra, textvariable=var_fecha,
                         font=("Segoe UI", 10, "bold"), bg=GRIS_FONDO,
                         width=12, relief="solid", bd=1, cursor="hand2")
    lbl_fecha.pack(side="left", padx=2)
    lbl_fecha.bind("<Button-1>", lambda e: abrir_selector_fecha(
        frame, _fecha_o_hoy(var_fecha.get()),
        lambda d: (var_fecha.set(d.isoformat()), _cargar_asistencia())))

    tk.Button(barra, text="▶", font=("Segoe UI", 10), bg=GRIS_FONDO,
              relief="flat", cursor="hand2",
              command=lambda: _ir_dia(1)).pack(side="left")

    tk.Button(barra, text="Hoy", font=("Segoe UI", 9),
              bg=AZUL, fg=BLANCO, relief="flat", padx=8, cursor="hand2",
              command=lambda: (var_fecha.set(hoy.isoformat()),
                               _cargar_asistencia())
              ).pack(side="left", padx=8)

    # Rango para resumen
    tk.Label(barra, text="  Resumen Desde:", font=("Segoe UI", 9),
             bg=GRIS_FONDO).pack(side="left")
    var_desde = tk.StringVar(value=hoy.replace(day=1).isoformat())
    var_hasta = tk.StringVar(value=hoy.isoformat())

    tk.Label(barra, text="Desde:", font=("Segoe UI", 9),
             bg=GRIS_FONDO).pack(side="left", padx=(8, 2))
    lbl_desde = tk.Label(barra, textvariable=var_desde, font=("Segoe UI", 9),
                         bg=GRIS_FONDO, width=11, relief="solid", bd=1, cursor="hand2")
    lbl_desde.pack(side="left")
    lbl_desde.bind("<Button-1>", lambda e: abrir_selector_fecha(
        frame, _fecha_o_hoy(var_desde.get()), lambda d: var_desde.set(d.isoformat())))

    tk.Label(barra, text="Hasta:", font=("Segoe UI", 9),
             bg=GRIS_FONDO).pack(side="left", padx=(8, 2))
    lbl_hasta = tk.Label(barra, textvariable=var_hasta, font=("Segoe UI", 9),
                         bg=GRIS_FONDO, width=11, relief="solid", bd=1, cursor="hand2")
    lbl_hasta.pack(side="left")
    lbl_hasta.bind("<Button-1>", lambda e: abrir_selector_fecha(
        frame, _fecha_o_hoy(var_hasta.get()), lambda d: var_hasta.set(d.isoformat())))

    tk.Button(barra, text="Ver resumen", font=("Segoe UI", 9),
              bg=NARANJA, fg=BLANCO, relief="flat", padx=8, cursor="hand2",
              command=lambda: _mostrar_resumen_periodo(
                  frame, var_desde.get(), var_hasta.get())
              ).pack(side="left", padx=8)

    # ── Grilla de asistencia ───────────────────────────────
    contenedor = tk.Frame(frame, bg=BLANCO)
    contenedor.grid(row=2, column=0, sticky="nsew", padx=8)
    contenedor.grid_rowconfigure(0, weight=1)
    contenedor.grid_columnconfigure(0, weight=1)

    cols   = ("nombre", "cargo", "estado", "hora_entrada", "hora_salida", "obs")
    encabs = ("NOMBRE", "CARGO", "ESTADO", "ENTRADA", "SALIDA", "OBSERVACIONES")
    anchos = (200, 130, 100, 80, 80, 200)

    tabla = ttk.Treeview(contenedor, columns=cols, show="headings",
                          selectmode="browse")
    habilitar_deseleccion_treeview(tabla)
    for col, enc, ancho in zip(cols, encabs, anchos):
        tabla.heading(col, text=enc)
        tabla.column(col, width=ancho,
                     anchor="w" if col in ("nombre", "obs") else "center")

    tabla.tag_configure("ausente",  background="#fef2f2", foreground=ROJO)
    tabla.tag_configure("tardanza", background="#fef3c7", foreground=NARANJA)
    tabla.tag_configure("licencia", background="#dbeafe", foreground=AZUL)
    tabla.tag_configure("feriado",  background="#f3f4f6", foreground=GRIS_TEXT)

    sb_y = ttk.Scrollbar(contenedor, orient="vertical",   command=tabla.yview)
    sb_x = ttk.Scrollbar(contenedor, orient="horizontal", command=tabla.xview)
    tabla.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
    tabla.grid(row=0, column=0, sticky="nsew")
    sb_y.grid(row=0, column=1, sticky="ns")
    sb_x.grid(row=1, column=0, sticky="ew")

    # Panel de edición rápida
    panel_edit = tk.Frame(frame, bg=GRIS_FONDO)
    panel_edit.grid(row=3, column=0, sticky="ew", padx=8, pady=4)

    # Indicador del empleado seleccionado actualmente
    lbl_emp_sel = tk.Label(panel_edit, text="Ningún empleado seleccionado",
                            font=("Segoe UI", 9, "italic"), bg=GRIS_FONDO,
                            fg=GRIS_TEXT)
    lbl_emp_sel.pack(side="left", padx=(6, 12))

    tk.Frame(panel_edit, bg=GRIS_BORDE, width=1).pack(side="left", fill="y", pady=2)

    tk.Label(panel_edit, text="Estado:", font=("Segoe UI", 9),
             bg=GRIS_FONDO).pack(side="left", padx=(8, 2))
    var_estado = tk.StringVar(value="Presente")
    combo_estado = ttk.Combobox(panel_edit, textvariable=var_estado,
                                 values=ESTADOS_ASISTENCIA, width=12,
                                 state="readonly")
    combo_estado.pack(side="left", padx=4)

    for lbl_txt, placeholder in [("Entrada:", "08:00"), ("Salida:", "17:00")]:
        tk.Label(panel_edit, text=lbl_txt, font=("Segoe UI", 9),
                 bg=GRIS_FONDO).pack(side="left", padx=(8, 2))
        var_hora = tk.StringVar()

        def _hacer_entry_placeholder(var=var_hora, ph=placeholder):
            e = tk.Entry(panel_edit, font=("Segoe UI", 9), width=7,
                         fg="#9ca3af")
            e.insert(0, ph)

            def _focus_in(event, _e=e, _ph=ph, _var=var):
                if _e.get() == _ph:
                    _e.delete(0, "end")
                    _e.config(fg="#1e293b")
                    _var.set("")

            def _focus_out(event, _e=e, _ph=ph, _var=var):
                val = _e.get().strip()
                if not val or val == _ph:
                    _e.delete(0, "end")
                    _e.insert(0, _ph)
                    _e.config(fg="#9ca3af")
                    _var.set("")
                else:
                    _var.set(val)

            def _key_release(event, _e=e, _var=var):
                _var.set(_e.get())

            e.bind("<FocusIn>",   _focus_in)
            e.bind("<FocusOut>",  _focus_out)
            e.bind("<KeyRelease>",_key_release)
            e.pack(side="left")
            return var

        if lbl_txt == "Entrada:":
            var_hora_entrada = _hacer_entry_placeholder()
        else:
            var_hora_salida  = _hacer_entry_placeholder()

    var_obs_asist = tk.StringVar()
    tk.Label(panel_edit, text="Obs:", font=("Segoe UI", 9),
             bg=GRIS_FONDO).pack(side="left", padx=(8, 2))

    entry_obs = tk.Entry(panel_edit, font=("Segoe UI", 9),
                         width=24, fg="#9ca3af")
    entry_obs.insert(0, "Ej: Llegó tarde por tráfico")
    entry_obs.pack(side="left")

    def _obs_focus_in(e):
        if entry_obs.get() == "Ej: Llegó tarde por tráfico":
            entry_obs.delete(0, "end")
            entry_obs.config(fg="#1e293b")
            var_obs_asist.set("")

    def _obs_focus_out(e):
        val = entry_obs.get().strip()
        if not val or val == "Ej: Llegó tarde por tráfico":
            entry_obs.delete(0, "end")
            entry_obs.insert(0, "Ej: Llegó tarde por tráfico")
            entry_obs.config(fg="#9ca3af")
            var_obs_asist.set("")
        else:
            var_obs_asist.set(val)

    def _obs_key(e):
        var_obs_asist.set(entry_obs.get())

    entry_obs.bind("<FocusIn>",   _obs_focus_in)
    entry_obs.bind("<FocusOut>",  _obs_focus_out)
    entry_obs.bind("<KeyRelease>", _obs_key)

    def _registrar():
        sel = tabla.selection()
        if not sel:
            messagebox.showinfo("Seleccionar",
                                "Hacé click en un empleado de la lista para seleccionarlo,\n"
                                "luego presioná ✔ Guardar.\n\n"
                                "Para marcar a todos a la vez usá el botón 'Marcar todos'.")
            return
        emp_id = int(sel[0])
        registrar_asistencia(
            emp_id, var_fecha.get(), var_estado.get(),
            var_hora_entrada.get(), var_hora_salida.get(),
            var_obs_asist.get(),
        )
        _cargar_asistencia()

    def _registrar_todos():
        if not messagebox.askyesno("Confirmar",
                                   f"¿Marcar a TODOS como '{var_estado.get()}' "
                                   f"el {var_fecha.get()}?"):
            return
        for iid in tabla.get_children():
            registrar_asistencia(
                int(iid), var_fecha.get(), var_estado.get(),
                var_hora_entrada.get(), var_hora_salida.get(),
                var_obs_asist.get(),
            )
        _cargar_asistencia()

    tk.Button(panel_edit, text="✔ Guardar",
              font=("Segoe UI", 9, "bold"), bg=VERDE, fg=BLANCO,
              relief="flat", padx=8, cursor="hand2",
              command=_registrar).pack(side="left", padx=8)
    tk.Button(panel_edit, text="Marcar todos",
              font=("Segoe UI", 9), bg=AZUL, fg=BLANCO,
              relief="flat", padx=8, cursor="hand2",
              command=_registrar_todos).pack(side="left", padx=2)

    def _cargar_asistencia():
        for r in tabla.get_children():
            tabla.delete(r)
        registros = obtener_asistencia_dia(var_fecha.get())
        for reg in registros:
            estado = reg["estado"]
            tag = {
                "Ausente":  "ausente",
                "Tardanza": "tardanza",
                "Licencia": "licencia",
                "Feriado":  "feriado",
            }.get(estado, "")
            iid = str(reg["id"])
            try:
                tabla.insert("", "end", iid=iid, values=(
                    reg["nombre"], reg["cargo"], estado,
                    reg["hora_entrada"], reg["hora_salida"], reg["observaciones"],
                ), tags=(tag,) if tag else ())
            except Exception:
                # Si el iid ya existe, actualizar los valores
                tabla.item(iid, values=(
                    reg["nombre"], reg["cargo"], estado,
                    reg["hora_entrada"], reg["hora_salida"], reg["observaciones"],
                ), tags=(tag,) if tag else ())

        # Al seleccionar una fila, precargar estado actual en los controles
        def _al_seleccionar(event):
            sel = tabla.selection()
            if not sel:
                lbl_emp_sel.config(text="Ningún empleado seleccionado", fg=GRIS_TEXT)
                return
            vals = tabla.item(sel[0])["values"]
            nombre = vals[0] if vals else ""
            estado_actual = vals[2] if vals and len(vals) > 2 else "Presente"
            lbl_emp_sel.config(
                text=f"✔ Seleccionado: {nombre}", fg=VERDE,
                font=("Segoe UI", 9, "bold"))
            if estado_actual != "Sin registro":
                var_estado.set(estado_actual)
            else:
                var_estado.set("Presente")

        tabla.bind("<<TreeviewSelect>>", _al_seleccionar)

    def _reset_panel():
        """Resetea el indicador de empleado al deseleccionar."""
        lbl_emp_sel.config(text="Ningún empleado seleccionado",
                           fg=GRIS_TEXT, font=("Segoe UI", 9, "italic"))

    _agregar_deseleccion(tabla, extra_callback=_reset_panel)

    _cargar_asistencia()
    return _cargar_asistencia


def _mostrar_resumen_periodo(parent, desde, hasta):
    """Ventana con resumen de asistencia de todos los empleados en el período."""
    win = tk.Toplevel(parent)
    win.title(f"Resumen de Asistencia: {desde} al {hasta}")
    win.configure(bg=BLANCO)
    win.grab_set()
    try:
        win.attributes("-toolwindow", False)
    except Exception:
        pass

    barra = tk.Frame(win, bg=AZUL, height=36)
    barra.pack(fill="x")
    barra.pack_propagate(False)
    tk.Label(barra, text=f"Asistencia {desde} → {hasta}",
             font=("Segoe UI", 11, "bold"), bg=AZUL, fg=BLANCO
             ).pack(side="left", padx=12, pady=8)

    cont = tk.Frame(win, bg=BLANCO)
    cont.pack(fill="both", expand=True, padx=8, pady=8)
    cont.grid_rowconfigure(0, weight=1)
    cont.grid_columnconfigure(0, weight=1)

    cols   = ("nombre", "cargo", "presentes", "ausentes", "tardanzas",
              "licencias", "descuento", "adelantos", "neto")
    encabs = ("NOMBRE", "CARGO", "PRESENTES", "AUSENTES", "TARDANZAS",
              "LICENCIAS", "DESCUENTO", "ADELANTOS PEND.", "SUELDO NETO EST.")
    anchos = (180, 110, 80, 80, 80, 80, 110, 120, 130)

    tabla = ttk.Treeview(cont, columns=cols, show="headings", selectmode="browse")
    habilitar_deseleccion_treeview(tabla)
    for col, enc, ancho in zip(cols, encabs, anchos):
        tabla.heading(col, text=enc)
        tabla.column(col, width=ancho, anchor="center")
    tabla.column("nombre", anchor="w")

    sb_y = ttk.Scrollbar(cont, orient="vertical",   command=tabla.yview)
    sb_x = ttk.Scrollbar(cont, orient="horizontal", command=tabla.xview)
    tabla.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
    tabla.grid(row=0, column=0, sticky="nsew")
    sb_y.grid(row=0, column=1, sticky="ns")
    sb_x.grid(row=1, column=0, sticky="ew")

    for e in listar_empleados(solo_activos=True):
        liq = calcular_liquidacion(e["id"], desde, hasta)
        tabla.insert("", "end", values=(
            e["nombre"], e["cargo"],
            liq["presentes"], liq["ausentes"], liq["tardanzas"], liq["licencias"],
            formatear_gs(liq["descuento_ausencias"]),
            formatear_gs(liq["adelantos_pendientes"]),
            formatear_gs(liq["sueldo_neto"]),
        ))

    tk.Button(win, text="Cerrar", font=("Segoe UI", 9), bg=BLANCO,
              relief="solid", bd=1, padx=14, pady=4,
              command=win.destroy).pack(pady=8)

    ajustar_tamaño_ventana(win, ancho_min=900, alto_min=400, margen_alto=20,
                           ancho_max=1200, alto_max=700)


# ═════════════════════════════════════════════════════════════
#  TAB ADELANTOS
# ═════════════════════════════════════════════════════════════
def _construir_tab_adelantos(parent, usuario_actual, on_cambio_externo=None):
    frame = tk.Frame(parent, bg=BLANCO)
    frame.pack(fill="both", expand=True)
    frame.grid_rowconfigure(1, weight=1)
    frame.grid_columnconfigure(0, weight=1)

    # Barra superior
    barra = tk.Frame(frame, bg=GRIS_FONDO)
    barra.grid(row=0, column=0, sticky="ew", padx=8, pady=6)

    var_filtro_emp = tk.StringVar()
    var_filtro_est = tk.StringVar(value="Todos")

    tk.Label(barra, text="Empleado:", font=("Segoe UI", 9),
             bg=GRIS_FONDO).pack(side="left", padx=(8, 2))
    tk.Entry(barra, textvariable=var_filtro_emp,
             font=("Segoe UI", 9), width=22).pack(side="left", padx=(0, 10))

    tk.Label(barra, text="Estado:", font=("Segoe UI", 9),
             bg=GRIS_FONDO).pack(side="left")
    combo_est = ttk.Combobox(barra, textvariable=var_filtro_est,
                              values=["Todos", "Pendiente", "Descontado"],
                              width=12, state="readonly")
    combo_est.pack(side="left", padx=4)

    tk.Button(barra, text="🔍 Filtrar", font=("Segoe UI", 9),
              bg=AZUL_OSC, fg=BLANCO, relief="flat", padx=8, cursor="hand2",
              command=lambda: _cargar_adelantos()
              ).pack(side="left", padx=8)

    tk.Button(barra, text="+ Nuevo Vale",
              font=("Segoe UI", 9, "bold"), bg=VERDE, fg=BLANCO,
              relief="flat", padx=10, pady=4, cursor="hand2",
              command=lambda: _abrir_formulario_adelanto(
                  parent, None, usuario_actual, _cargar_adelantos_y_notificar)
              ).pack(side="right", padx=4)

    var_filtro_emp.trace_add("write", lambda *_: _cargar_adelantos())
    combo_est.bind("<<ComboboxSelected>>", lambda e: _cargar_adelantos())

    # Grilla
    contenedor = tk.Frame(frame, bg=BLANCO)
    contenedor.grid(row=1, column=0, sticky="nsew", padx=8)
    contenedor.grid_rowconfigure(0, weight=1)
    contenedor.grid_columnconfigure(0, weight=1)

    cols   = ("id", "empleado", "fecha", "monto", "descripcion",
              "estado", "registrado_por", "adjunto")
    encabs = ("ID", "EMPLEADO", "FECHA", "MONTO", "DESCRIPCIÓN",
              "ESTADO", "REGISTRADO POR", "ADJUNTO")
    anchos = (40, 180, 90, 110, 220, 90, 130, 80)

    tabla = ttk.Treeview(contenedor, columns=cols, show="headings",
                          selectmode="browse")
    habilitar_deseleccion_treeview(tabla)
    for col, enc, ancho in zip(cols, encabs, anchos):
        tabla.heading(col, text=enc)
        tabla.column(col, width=ancho,
                     anchor="w" if col in ("descripcion", "empleado") else "center")

    tabla.tag_configure("pendiente",   foreground=NARANJA)
    tabla.tag_configure("descontado",  foreground=GRIS_TEXT)

    sb_y = ttk.Scrollbar(contenedor, orient="vertical",   command=tabla.yview)
    sb_x = ttk.Scrollbar(contenedor, orient="horizontal", command=tabla.xview)
    tabla.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
    tabla.grid(row=0, column=0, sticky="nsew")
    sb_y.grid(row=0, column=1, sticky="ns")
    sb_x.grid(row=1, column=0, sticky="ew")

    # Panel resumen inferior
    panel_res = tk.Frame(frame, bg=GRIS_FONDO)
    panel_res.grid(row=2, column=0, sticky="ew", padx=8, pady=4)

    lbl_total_pend = tk.Label(panel_res, text="", font=("Segoe UI", 10, "bold"),
                               bg=GRIS_FONDO, fg=NARANJA)
    lbl_total_pend.pack(side="left", padx=20, pady=6)
    lbl_total_desc = tk.Label(panel_res, text="", font=("Segoe UI", 10, "bold"),
                               bg=GRIS_FONDO, fg=GRIS_TEXT)
    lbl_total_desc.pack(side="left", padx=20, pady=6)

    adelantos_cache = []

    def _cargar_adelantos():
        nonlocal adelantos_cache
        for r in tabla.get_children():
            tabla.delete(r)
        estado_filtro = None if var_filtro_est.get() == "Todos" else var_filtro_est.get()
        todos = listar_adelantos(estado=estado_filtro)
        texto = var_filtro_emp.get().lower()
        adelantos_cache = [a for a in todos
                           if not texto or texto in a["empleado_nombre"].lower()]
        total_pend = 0
        total_desc = 0
        for a in adelantos_cache:
            tag = "pendiente" if a["estado"] == "Pendiente" else "descontado"
            adj = "📎 Ver" if a.get("archivo_adjunto") else "—"
            tabla.insert("", "end", iid=str(a["id"]), values=(
                a["id"], a["empleado_nombre"], a["fecha"],
                formatear_gs(a["monto"]), a["descripcion"],
                a["estado"], a.get("registrado_por", ""), adj,
            ), tags=(tag,))
            if a["estado"] == "Pendiente":
                total_pend += a["monto"]
            else:
                total_desc += a["monto"]
        lbl_total_pend.config(text=f"Pendientes: {formatear_gs(total_pend)}")
        lbl_total_desc.config(text=f"Descontados: {formatear_gs(total_desc)}")

    def _cargar_adelantos_y_notificar():
        """Recarga esta pestaña Y avisa a Personal al instante (sin esperar
        a que el usuario cambie de pestaña), porque la grilla de Personal
        muestra el adelanto pendiente de cada empleado y el resumen
        inferior también depende de estos datos."""
        _cargar_adelantos()
        if on_cambio_externo:
            on_cambio_externo()

    def _abrir_menu_contextual(event):
        iid = tabla.identify_row(event.y)
        if not iid:
            return
        tabla.selection_set(iid)
        adelanto = next((a for a in adelantos_cache if str(a["id"]) == iid), None)
        if not adelanto:
            return
        menu = tk.Menu(tabla, tearoff=0)
        if adelanto["estado"] == "Pendiente":
            menu.add_command(
                label="✔ Marcar como descontado",
                command=lambda: (
                    marcar_adelanto_descontado(
                        adelanto["id"], datetime.date.today().isoformat()),
                    _cargar_adelantos_y_notificar(),
                ))
        if adelanto.get("archivo_adjunto"):
            menu.add_command(
                label="📎 Ver archivo adjunto",
                command=lambda: _abrir_adjunto(adelanto["archivo_adjunto"]))
        menu.add_separator()
        menu.add_command(
            label="🗑 Eliminar vale",
            command=lambda: _eliminar(adelanto))
        menu.tk_popup(event.x_root, event.y_root)

    def _eliminar(adelanto):
        if messagebox.askyesno("Confirmar",
                               f"¿Eliminar el vale de {formatear_gs(adelanto['monto'])} "
                               f"de {adelanto['empleado_nombre']}?"):
            eliminar_adelanto(adelanto["id"])
            _cargar_adelantos_y_notificar()

    tabla.bind("<Button-3>", _abrir_menu_contextual)
    tabla.bind("<Double-1>", lambda e: (
        _abrir_adjunto(
            next((a for a in adelantos_cache
                  if str(a["id"]) == tabla.selection()[0]), {}).get("archivo_adjunto", "")
        ) if tabla.selection() else None
    ))
    _agregar_deseleccion(tabla)

    _cargar_adelantos()
    return _cargar_adelantos


def _abrir_adjunto(ruta):
    if not ruta or not os.path.exists(ruta):
        messagebox.showinfo("Sin archivo", "No hay archivo adjunto o no se encontró.")
        return
    import subprocess, sys
    try:
        if sys.platform == "win32":
            os.startfile(ruta)
        elif sys.platform == "darwin":
            subprocess.run(["open", ruta])
        else:
            subprocess.run(["xdg-open", ruta])
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{e}")


# ═════════════════════════════════════════════════════════════
#  FORMULARIO ADELANTO
# ═════════════════════════════════════════════════════════════
def _abrir_formulario_adelanto(parent, adelanto, usuario_actual, on_guardado):
    win = tk.Toplevel(parent)
    win.title("Nuevo Vale / Adelanto")
    win.configure(bg=BLANCO)
    win.grab_set()
    try:
        win.attributes("-toolwindow", False)
    except Exception:
        pass

    barra = tk.Frame(win, bg=AZUL, height=36)
    barra.pack(fill="x")
    barra.pack_propagate(False)
    tk.Label(barra, text="💵 Registrar Vale / Adelanto",
             font=("Segoe UI", 11, "bold"), bg=AZUL, fg=BLANCO
             ).pack(side="left", padx=12, pady=8)

    cuerpo = tk.Frame(win, bg=BLANCO, padx=20, pady=12)
    cuerpo.pack(fill="both", expand=True)

    # Empleado
    tk.Label(cuerpo, text="Empleado:", font=("Segoe UI", 9, "bold"),
             bg=BLANCO).grid(row=0, column=0, sticky="w", pady=4)
    empleados = listar_empleados(solo_activos=True)
    nombres   = [f"{e['id']} – {e['nombre']}" for e in empleados]
    var_emp   = tk.StringVar()
    combo_emp = ttk.Combobox(cuerpo, textvariable=var_emp, values=nombres,
                              width=35, state="readonly")
    combo_emp.grid(row=0, column=1, sticky="ew", padx=(8, 0), pady=4)

    # Fecha
    tk.Label(cuerpo, text="Fecha:", font=("Segoe UI", 9, "bold"),
             bg=BLANCO).grid(row=1, column=0, sticky="w", pady=4)
    var_fecha = tk.StringVar(value=datetime.date.today().isoformat())
    frame_f = tk.Frame(cuerpo, bg=BLANCO)
    frame_f.grid(row=1, column=1, sticky="w", padx=(8, 0))
    tk.Label(frame_f, textvariable=var_fecha, font=("Segoe UI", 9),
             bg=BLANCO, width=12, relief="solid", bd=1).pack(side="left")
    tk.Button(frame_f, text="📅", font=("Segoe UI", 9), relief="flat",
              cursor="hand2",
              command=lambda: abrir_selector_fecha(
                  win, _fecha_o_hoy(var_fecha.get()), lambda d: var_fecha.set(d.isoformat()))
              ).pack(side="left", padx=2)

    # Monto
    tk.Label(cuerpo, text="Monto (Gs):", font=("Segoe UI", 9, "bold"),
             bg=BLANCO).grid(row=2, column=0, sticky="w", pady=4)
    var_monto = tk.StringVar()
    tk.Entry(cuerpo, textvariable=var_monto, font=("Segoe UI", 9),
             width=20).grid(row=2, column=1, sticky="w", padx=(8, 0), pady=4)

    # Descripción
    tk.Label(cuerpo, text="Descripción:", font=("Segoe UI", 9, "bold"),
             bg=BLANCO).grid(row=3, column=0, sticky="nw", pady=4)
    txt_desc = tk.Text(cuerpo, font=("Segoe UI", 9), width=38, height=3)
    txt_desc.grid(row=3, column=1, sticky="ew", padx=(8, 0), pady=4)
    def _mayus_desc(event):
        if event.char and event.char.isalpha():
            txt_desc.insert(tk.INSERT, event.char.upper())
            return "break"
    txt_desc.bind("<Key>", _mayus_desc)

    # Archivo adjunto
    tk.Label(cuerpo, text="Adjunto (foto/PDF):", font=("Segoe UI", 9, "bold"),
             bg=BLANCO).grid(row=4, column=0, sticky="w", pady=4)
    var_adjunto = tk.StringVar()
    frame_adj = tk.Frame(cuerpo, bg=BLANCO)
    frame_adj.grid(row=4, column=1, sticky="ew", padx=(8, 0))
    lbl_adj = tk.Label(frame_adj, text="Sin archivo", font=("Segoe UI", 8),
                        bg=BLANCO, fg=GRIS_TEXT, width=28, anchor="w")
    lbl_adj.pack(side="left")

    def _elegir_adjunto():
        ruta = filedialog.askopenfilename(
            title="Seleccionar aval firmado",
            filetypes=[("Imágenes y PDF", "*.pdf *.jpg *.jpeg *.png *.webp"),
                       ("Todos los archivos", "*.*")])
        if ruta:
            var_adjunto.set(ruta)
            lbl_adj.config(text=os.path.basename(ruta), fg=AZUL)

    tk.Button(frame_adj, text="📂 Elegir archivo", font=("Segoe UI", 8),
              bg=GRIS_FONDO, relief="solid", bd=1, cursor="hand2",
              command=_elegir_adjunto).pack(side="left", padx=4)

    cuerpo.grid_columnconfigure(1, weight=1)

    # Botones
    pie = tk.Frame(win, bg=BLANCO)
    pie.pack(fill="x", pady=8)

    def _guardar():
        if not var_emp.get():
            messagebox.showerror("Error", "Seleccioná un empleado.")
            return
        try:
            monto = float(var_monto.get().replace(".", "").replace(",", ".") or 0)
            if monto <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Error", "El monto debe ser un número mayor a 0.")
            return

        emp_id = int(var_emp.get().split("–")[0].strip())

        # Copiar el adjunto a la carpeta RRHH
        ruta_destino = ""
        if var_adjunto.get():
            carpeta = carpeta_adjuntos_rrhh()
            ext      = os.path.splitext(var_adjunto.get())[1]
            nombre_archivo = (f"adelanto_{emp_id}_"
                              f"{var_fecha.get().replace('-','')}_{monto:.0f}{ext}")
            ruta_destino = os.path.join(carpeta, nombre_archivo)
            shutil.copy2(var_adjunto.get(), ruta_destino)

        registrado_por = usuario_actual.get("nombre_completo", "")
        registrar_adelanto(emp_id, var_fecha.get(), monto,
                           txt_desc.get("1.0", "end").strip(),
                           ruta_destino, registrado_por)
        on_guardado()
        win.destroy()
        messagebox.showinfo("Listo", "Vale registrado correctamente.")

    btn_g_adel = tk.Button(pie, text="✔ Guardar", font=("Segoe UI", 10, "bold"),
              bg=VERDE, fg=BLANCO, relief="flat", padx=16, pady=6,
              cursor="hand2", command=_guardar)
    btn_g_adel.pack(side="left", padx=16)
    btn_g_adel.bind("<Return>", lambda e: _guardar())
    btn_c_adel = tk.Button(pie, text="✖ Cancelar", font=("Segoe UI", 10),
              bg=BLANCO, relief="solid", bd=1, padx=16, pady=5,
              cursor="hand2", command=win.destroy)
    btn_c_adel.pack(side="right", padx=16)
    btn_c_adel.bind("<Return>", lambda e: win.destroy())

    ajustar_tamaño_ventana(win, ancho_min=500, alto_min=340, margen_alto=20)
