"""
ventana_importacion.py
Módulo Importaciones: registrar compras hechas en plataformas del exterior
(eBay, AliExpress, Temu, Shein, Alibaba, Made in China) que llegan en caja
a un casillero (Miami o Shenzhen), calcular automáticamente cuánto le
corresponde de flete a cada unidad según el peso de la caja y la tarifa
por kilo del courier, cargar el precio de venta al público para ver la
ganancia, y un Dashboard de inversión / rentabilidad. Sigue el mismo
patrón visual que los módulos Restaurante/Comedor y Alquiler de Streaming.
"""
import datetime
import tkinter as tk
from tkinter import ttk, messagebox

from models_importacion import (
    TIPOS_ENVIO, CASILLEROS, ESTADOS_COMPRA, COLOR_ESTADO,
    listar_couriers, obtener_courier, crear_courier, editar_courier,
    cambiar_estado_courier, eliminar_courier,
    listar_plataformas, crear_plataforma, cambiar_estado_plataforma, eliminar_plataforma,
    obtener_tasa_cambio, obtener_info_tasa_cambio, guardar_tasa_cambio, actualizar_tasa_cambio_automatica,
    listar_compras, obtener_compra_detalle, crear_compra, editar_compra,
    cambiar_estado_compra, eliminar_compra, set_precio_venta,
    previsualizar_costos, enviar_a_inventario, buscar_productos_similares,
    conteos_dashboard, rentabilidad_por_plataforma, rentabilidad_por_producto,
)
from utilidades_ui import ajustar_tamaño_ventana, forzar_mayusculas, habilitar_deseleccion_treeview, formatear_gs
from widget_calendario import abrir_selector_fecha
from menu_reporte_general import BotonReporteGeneral
import threading

AZUL_RIBBON = "#1d5fd6"
GRIS_FONDO = "#f4f5f7"
VERDE = "#16a34a"
ROJO = "#dc2626"
NARANJA = "#d97706"
GRIS_TEXTO = "#6b7280"
MORADO = "#7c3aed"


def _usd(valor) -> str:
    """Formatea un monto en dólares (moneda habitual de estas plataformas
    y de las tarifas de courier), con separador de miles."""
    try:
        return f"US$ {float(valor):,.2f}"
    except (TypeError, ValueError):
        return "US$ 0.00"


def _dual(valor_usd) -> str:
    """Formatea un monto en dólares junto con su equivalente en guaraníes,
    usando el tipo de cambio cargado en el módulo (editable en el
    Dashboard). Ej: 'US$ 4.30  (Gs. 31.390)'."""
    try:
        valor_usd = float(valor_usd)
    except (TypeError, ValueError):
        valor_usd = 0.0
    tasa = obtener_tasa_cambio()
    return f"{_usd(valor_usd)}  ({formatear_gs(valor_usd * tasa)})"


def _crear_seccion_scrollable(parent):
    frame_exterior = tk.Frame(parent, bg="white")
    canvas = tk.Canvas(frame_exterior, bg="white", highlightthickness=0)
    scrollbar = ttk.Scrollbar(frame_exterior, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    frame_interior = tk.Frame(canvas, bg="white")
    id_ventana_interior = canvas.create_window((0, 0), window=frame_interior, anchor="nw")

    def _actualizar_scrollregion(_e=None):
        canvas.configure(scrollregion=canvas.bbox("all"))
    frame_interior.bind("<Configure>", _actualizar_scrollregion)

    def _ajustar_ancho_interior(event):
        canvas.itemconfig(id_ventana_interior, width=event.width)
    canvas.bind("<Configure>", _ajustar_ancho_interior)

    def _rueda_mouse(event):
        if event.num == 5 or event.delta < 0:
            canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0:
            canvas.yview_scroll(-1, "units")

    def _activar_scroll(_e=None):
        canvas.bind_all("<MouseWheel>", _rueda_mouse)
        canvas.bind_all("<Button-4>", _rueda_mouse)
        canvas.bind_all("<Button-5>", _rueda_mouse)

    def _desactivar_scroll(_e=None):
        canvas.unbind_all("<MouseWheel>")
        canvas.unbind_all("<Button-4>")
        canvas.unbind_all("<Button-5>")

    frame_exterior.bind("<Enter>", _activar_scroll)
    frame_exterior.bind("<Leave>", _desactivar_scroll)

    return frame_exterior, frame_interior


def _campo_fecha(parent, variable: tk.StringVar):
    frame = tk.Frame(parent, bg="white")
    entry = tk.Entry(frame, textvariable=variable, font=("Segoe UI", 10), state="readonly", width=12)
    entry.pack(side="left")

    def _actual():
        try:
            return datetime.date.fromisoformat(variable.get())
        except (ValueError, TypeError):
            return datetime.date.today()

    tk.Button(frame, text="📅", font=("Segoe UI", 9), bg="white", relief="solid", bd=1, cursor="hand2",
              command=lambda: abrir_selector_fecha(
                  parent.winfo_toplevel(), _actual(), lambda d: variable.set(d.isoformat()))).pack(
                  side="left", padx=(4, 0))
    tk.Button(frame, text="✕", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
              cursor="hand2", command=lambda: variable.set("")).pack(side="left", padx=(2, 0))
    return frame


# ============================================================
# PANEL PRINCIPAL
# ============================================================
class PanelImportacion(tk.Frame):
    def __init__(self, parent, usuario_actual):
        super().__init__(parent, bg="white")
        self.usuario_actual = usuario_actual

        encabezado = tk.Frame(self, bg=AZUL_RIBBON, height=54)
        encabezado.pack(fill="x")
        encabezado.pack_propagate(False)
        tk.Label(encabezado, text="📦 Importaciones", font=("Segoe UI", 15, "bold"),
                 bg=AZUL_RIBBON, fg="white").pack(side="left", padx=20, pady=12)

        self.boton_reporte = BotonReporteGeneral(
            encabezado, obtener_datos_callback=self._obtener_datos_reporte,
            nombre_archivo_base="Reporte_Importaciones",
        )
        self.boton_reporte.generador_excel = self._generar_excel
        self.boton_reporte.pack(side="right", padx=(0, 12), pady=12)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        self.tab_dashboard = _TabDashboard(self.notebook, self)
        self.tab_compras = _TabCompras(self.notebook, self)
        self.tab_couriers = _TabCouriers(self.notebook, self)

        self.notebook.add(self.tab_dashboard, text="📊 Dashboard")
        self.notebook.add(self.tab_compras, text="📦 Compras")
        self.notebook.add(self.tab_couriers, text="🚚 Couriers / Casilleros")

        self.notebook.bind("<<NotebookTabChanged>>", lambda e: self.refrescar_todo())

    def refrescar_todo(self):
        self.tab_dashboard.cargar()
        self.tab_compras.cargar()
        self.tab_couriers.cargar()

    # ── Datos para el Reporte General ─────────────────────────
    def _obtener_datos_reporte(self) -> dict:
        d = conteos_dashboard()
        compras = listar_compras()
        couriers = listar_couriers(solo_activos=False)
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if isinstance(self.usuario_actual, dict) else ""

        tiempo_txt = f"{d['tiempo_promedio_dias']} días" if d["tiempo_promedio_dias"] is not None else "—"
        margen_txt = f"{d['margen_promedio_pct']}%" if d["margen_promedio_pct"] is not None else "—"

        secciones = [
            {
                "tipo": "resumen",
                "titulo": "RESUMEN DE IMPORTACIONES",
                "filas": [
                    ("Compras registradas", str(d["total_compras"])),
                    ("En proceso", str(d["en_proceso"])),
                    ("Unidades importadas", str(int(d["total_unidades"]))),
                    ("Unidades en inventario", str(int(d["unidades_en_inventario"]))),
                    ("Tiempo prom. compra→recibido", tiempo_txt),
                    ("Costo productos", _dual(d["total_costo_productos"])),
                    ("Costo envíos a casillero", _dual(d["total_envios"])),
                    ("Inversión total", _dual(d["inversion_total"])),
                    ("Ganancia potencial", _dual(d["ganancia_potencial"])),
                    ("Margen promedio", margen_txt),
                ],
            },
            {
                "tipo": "tabla",
                "titulo": "COMPRAS DE IMPORTACIÓN",
                "encabezados": ["ID", "Plataforma", "Referencia", "Courier", "Tipo Envío",
                                "Peso Caja", "Unidades", "Inversión", "Estado", "Fecha Compra"],
                "filas": [
                    [str(c["id"]), c["plataforma"], c["referencia"], c["courier"], c["tipo_envio"],
                     f'{c["peso_caja_kg"]:.3f} kg', str(int(c["total_unidades"])), _dual(c["inversion_total"]),
                     c["estado"], c["fecha_compra"]]
                    for c in compras
                ],
            },
            {
                "tipo": "tabla",
                "titulo": "RENTABILIDAD POR PLATAFORMA",
                "encabezados": ["Plataforma", "Unidades", "Inversión", "Venta Potencial", "Ganancia Potencial"],
                "filas": [
                    [r["plataforma"], str(int(r["unidades"])), _dual(r["costo"]),
                     _dual(r["ingreso_potencial"]), _dual(r["margen"])]
                    for r in rentabilidad_por_plataforma()
                ],
            },
            {
                "tipo": "tabla",
                "titulo": "RENTABILIDAD POR PRODUCTO",
                "encabezados": ["Producto", "Unidades", "Costo Unit. Prom.", "Venta Unit. Prom.", "Ganancia Potencial"],
                "filas": [
                    [r["producto"], str(int(r["unidades"])), _dual(r["costo_unitario_promedio"]),
                     _dual(r["precio_venta_promedio"]) if r["precio_venta_promedio"] else "Sin precio",
                     _dual(r["ganancia"])]
                    for r in rentabilidad_por_producto()
                ],
            },
            {
                "tipo": "tabla",
                "titulo": "COURIERS / CASILLEROS",
                "encabezados": ["Courier", "Email", "RUC", "US$/kg Aéreo", "US$/kg Marítimo", "Estado"],
                "filas": [
                    [c["nombre"], c["email"], c["ruc"], _usd(c["costo_kg_aereo"]),
                     _usd(c["costo_kg_maritimo"]), "Activo" if c["activo"] else "Inactivo"]
                    for c in couriers
                ],
            },
        ]
        return {
            "titulo": "Reporte de Importaciones",
            "subtitulo": f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}",
            "generado_por": nombre_usuario,
            "secciones": secciones,
        }

    def _generar_excel(self, ruta: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        def _enc(ws):
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1D5FD6")
                cell.alignment = Alignment(horizontal="center")

        d = conteos_dashboard()
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Resumen"
        ws1.append(["Indicador", "Valor US$", "Valor Gs."])
        _enc(ws1)
        tasa = obtener_tasa_cambio()
        ws1.append(["Compras registradas", d["total_compras"], ""])
        ws1.append(["En proceso", d["en_proceso"], ""])
        ws1.append(["Unidades importadas", int(d["total_unidades"]), ""])
        ws1.append(["Unidades en inventario", int(d["unidades_en_inventario"]), ""])
        ws1.append(["Costo productos", d["total_costo_productos"], round(d["total_costo_productos"] * tasa)])
        ws1.append(["Costo envíos a casillero", d["total_envios"], round(d["total_envios"] * tasa)])
        ws1.append(["Inversión total", d["inversion_total"], round(d["inversion_total"] * tasa)])
        ws1.append(["Ganancia potencial", d["ganancia_potencial"], round(d["ganancia_potencial"] * tasa)])

        ws2 = wb.create_sheet("Compras")
        ws2.append(["ID", "Plataforma", "Referencia", "Courier", "Tipo Envío", "Peso Caja (kg)",
                    "Unidades", "Inversión (US$)", "Estado", "Fecha Compra"])
        _enc(ws2)
        for c in listar_compras():
            ws2.append([c["id"], c["plataforma"], c["referencia"], c["courier"], c["tipo_envio"],
                       c["peso_caja_kg"], int(c["total_unidades"]), c["inversion_total"],
                       c["estado"], c["fecha_compra"]])

        ws3 = wb.create_sheet("Rentabilidad por Plataforma")
        ws3.append(["Plataforma", "Unidades", "Inversión (US$)", "Venta Potencial (US$)", "Ganancia (US$)"])
        _enc(ws3)
        for r in rentabilidad_por_plataforma():
            ws3.append([r["plataforma"], int(r["unidades"]), r["costo"], r["ingreso_potencial"], r["margen"]])

        ws4 = wb.create_sheet("Rentabilidad por Producto")
        ws4.append(["Producto", "Unidades", "Costo Unit. Prom. (US$)", "Venta Unit. Prom. (US$)", "Ganancia (US$)"])
        _enc(ws4)
        for r in rentabilidad_por_producto():
            ws4.append([r["producto"], int(r["unidades"]), r["costo_unitario_promedio"],
                       r["precio_venta_promedio"] or 0, r["ganancia"]])

        ws5 = wb.create_sheet("Couriers")
        ws5.append(["Courier", "Email", "RUC", "Teléfono", "US$/kg Aéreo", "US$/kg Marítimo", "Estado"])
        _enc(ws5)
        for c in listar_couriers(solo_activos=False):
            ws5.append([c["nombre"], c["email"], c["ruc"], c["telefono"], c["costo_kg_aereo"],
                       c["costo_kg_maritimo"], "Activo" if c["activo"] else "Inactivo"])

        wb.save(ruta)


# ============================================================
# PESTAÑA: DASHBOARD
# ============================================================
class _TabDashboard(tk.Frame):
    def __init__(self, parent, panel_padre: PanelImportacion):
        super().__init__(parent, bg="white")
        self.panel_padre = panel_padre
        self._construir()
        self.cargar()

    def _construir(self):
        frame_ext, self.contenedor = _crear_seccion_scrollable(self)
        frame_ext.pack(fill="both", expand=True)

        self.frame_tasa = tk.Frame(self.contenedor, bg="#eef2ff", relief="solid", bd=1)
        self.frame_tasa.pack(fill="x", padx=4, pady=(6, 8))
        self.lbl_tasa = tk.Label(self.frame_tasa, text="", font=("Segoe UI", 9, "bold"),
                                  bg="#eef2ff", fg=AZUL_RIBBON)
        self.lbl_tasa.pack(side="left", padx=10, pady=6)
        tk.Button(self.frame_tasa, text="✏ Editar", font=("Segoe UI", 8), bg="white", relief="solid", bd=1,
                  cursor="hand2", command=self._editar_tasa_manual).pack(side="left", padx=(0, 6), pady=6)
        self.btn_actualizar_tasa = tk.Button(self.frame_tasa, text="🔄 Actualizar automático (Internet)",
                                              font=("Segoe UI", 8), bg="white", relief="solid", bd=1,
                                              cursor="hand2", command=self._actualizar_tasa_automatica)
        self.btn_actualizar_tasa.pack(side="left", pady=6)

        # Contenedor "de flujo": las tarjetas se acomodan en columnas que se
        # recalculan solas según el ancho disponible, así ninguna tarjeta
        # queda cortada fuera de la vista (antes se amontonaban todas en
        # una sola fila con pack(side="left") y las últimas no entraban).
        self.frame_tarjetas = tk.Frame(self.contenedor, bg="white")
        self.frame_tarjetas.pack(fill="x", pady=(6, 8), padx=4)
        self._tarjetas_widgets = []
        self._ancho_tarjetas_previo = None
        self.frame_tarjetas.bind("<Configure>", self._on_resize_tarjetas)

        self.frame_alertas = tk.Frame(self.contenedor, bg="white")
        self.frame_alertas.pack(fill="x", pady=(0, 10), padx=4)

        tk.Label(self.contenedor, text="Rentabilidad por plataforma", font=("Segoe UI", 10, "bold"),
                 bg="white").pack(anchor="w", padx=4)
        cont1 = tk.Frame(self.contenedor, bg="white")
        cont1.pack(fill="x", pady=(4, 12), padx=4)
        cont1.grid_rowconfigure(0, weight=1)
        cont1.grid_columnconfigure(0, weight=1)
        columnas1 = ("plataforma", "unidades", "costo", "ingreso", "margen")
        self.tabla_plataformas = ttk.Treeview(cont1, columns=columnas1, show="headings", height=6)
        for col, enc, w in zip(columnas1, ("PLATAFORMA", "UNIDADES", "INVERSIÓN", "VENTA POTENCIAL", "GANANCIA POTENCIAL"),
                                (140, 90, 210, 210, 210)):
            self.tabla_plataformas.heading(col, text=enc)
            self.tabla_plataformas.column(col, width=w, anchor="center" if col != "plataforma" else "w")
        sb1_v = ttk.Scrollbar(cont1, orient="vertical", command=self.tabla_plataformas.yview)
        sb1_h = ttk.Scrollbar(cont1, orient="horizontal", command=self.tabla_plataformas.xview)
        self.tabla_plataformas.configure(yscrollcommand=sb1_v.set, xscrollcommand=sb1_h.set)
        self.tabla_plataformas.grid(row=0, column=0, sticky="nsew")
        sb1_v.grid(row=0, column=1, sticky="ns")
        sb1_h.grid(row=1, column=0, sticky="ew")
        self.tabla_plataformas.tag_configure("positivo", foreground=VERDE)
        self.tabla_plataformas.tag_configure("negativo", foreground=ROJO)
        habilitar_deseleccion_treeview(self.tabla_plataformas)

        tk.Label(self.contenedor, text="Rentabilidad por producto", font=("Segoe UI", 10, "bold"),
                 bg="white").pack(anchor="w", padx=4)
        cont2 = tk.Frame(self.contenedor, bg="white")
        cont2.pack(fill="x", pady=(4, 12), padx=4)
        cont2.grid_rowconfigure(0, weight=1)
        cont2.grid_columnconfigure(0, weight=1)
        columnas2 = ("producto", "unidades", "costo_unit", "venta_unit", "ganancia")
        self.tabla_productos = ttk.Treeview(cont2, columns=columnas2, show="headings", height=8)
        for col, enc, w in zip(columnas2, ("PRODUCTO", "UNIDADES", "COSTO UNIT. PROM.", "VENTA UNIT. PROM.",
                                            "GANANCIA POTENCIAL"), (160, 80, 210, 210, 210)):
            self.tabla_productos.heading(col, text=enc)
            self.tabla_productos.column(col, width=w, anchor="center" if col != "producto" else "w")
        sb2_v = ttk.Scrollbar(cont2, orient="vertical", command=self.tabla_productos.yview)
        sb2_h = ttk.Scrollbar(cont2, orient="horizontal", command=self.tabla_productos.xview)
        self.tabla_productos.configure(yscrollcommand=sb2_v.set, xscrollcommand=sb2_h.set)
        self.tabla_productos.grid(row=0, column=0, sticky="nsew")
        sb2_v.grid(row=0, column=1, sticky="ns")
        sb2_h.grid(row=1, column=0, sticky="ew")
        self.tabla_productos.tag_configure("positivo", foreground=VERDE)
        self.tabla_productos.tag_configure("negativo", foreground=ROJO)
        habilitar_deseleccion_treeview(self.tabla_productos)

    ANCHO_TARJETA_MIN = 230

    def _tarjeta(self, titulo, valor, color):
        """Crea una tarjeta y la agrega a la lista de flujo (todavía sin
        ubicarla en la grilla: eso lo hace _reflow_tarjetas, que decide
        cuántas entran por fila según el ancho disponible)."""
        marco = tk.Frame(self.frame_tarjetas, bg=color, padx=14, pady=8)
        tk.Label(marco, text=str(valor), font=("Segoe UI", 14, "bold"), bg=color, fg="white").pack(anchor="w")
        tk.Label(marco, text=titulo, font=("Segoe UI", 8), bg=color, fg="white").pack(anchor="w")
        self._tarjetas_widgets.append(marco)

    def _on_resize_tarjetas(self, event):
        # Solo re-acomodamos si el ANCHO cambió de verdad (el widget también
        # dispara <Configure> cuando cambia el alto al agregar tarjetas, y
        # ahí no hace falta recalcular columnas).
        if event.width == self._ancho_tarjetas_previo:
            return
        self._ancho_tarjetas_previo = event.width
        self._reflow_tarjetas(event.width)

    def _reflow_tarjetas(self, ancho_disponible: int = None):
        if not self._tarjetas_widgets:
            return
        if ancho_disponible is None:
            ancho_disponible = self.frame_tarjetas.winfo_width()
        columnas = max(1, ancho_disponible // self.ANCHO_TARJETA_MIN)
        for i in range(columnas):
            self.frame_tarjetas.grid_columnconfigure(i, weight=1, uniform="tarjetas")
        for idx, marco in enumerate(self._tarjetas_widgets):
            fila, col = divmod(idx, columnas)
            marco.grid(row=fila, column=col, sticky="ew", padx=(0, 8), pady=(0, 8))

    def cargar(self):
        self._refrescar_texto_tasa()
        for w in self._tarjetas_widgets:
            w.destroy()
        self._tarjetas_widgets = []
        for w in self.frame_alertas.winfo_children():
            w.destroy()

        d = conteos_dashboard()
        self._tarjeta("Compras Registradas", d["total_compras"], AZUL_RIBBON)
        self._tarjeta("En Proceso", d["en_proceso"], NARANJA)
        self._tarjeta("Unidades Importadas", int(d["total_unidades"]), "#0f766e")
        self._tarjeta("Unidades en Inventario", int(d["unidades_en_inventario"]), MORADO)
        tiempo = f"{d['tiempo_promedio_dias']} días" if d["tiempo_promedio_dias"] is not None else "—"
        self._tarjeta("Tiempo Prom. Compra→Recibido", tiempo, GRIS_TEXTO)

        self._tarjeta("Costo Productos", _dual(d["total_costo_productos"]), NARANJA)
        self._tarjeta("Costo Envíos a Casillero", _dual(d["total_envios"]), NARANJA)
        self._tarjeta("Inversión Total", _dual(d["inversion_total"]), AZUL_RIBBON)
        color_ganancia = VERDE if d["ganancia_potencial"] >= 0 else ROJO
        self._tarjeta("Ganancia Potencial", _dual(d["ganancia_potencial"]), color_ganancia)
        margen_txt = f"{d['margen_promedio_pct']}%" if d["margen_promedio_pct"] is not None else "—"
        self._tarjeta("Margen Promedio", margen_txt, color_ganancia)

        self._reflow_tarjetas()

        if d["unidades_sin_precio"] > 0:
            aviso = tk.Frame(self.frame_alertas, bg="#fef3c7", relief="solid", bd=1)
            aviso.pack(fill="x")
            tk.Label(aviso, text=f"⚠ {int(d['unidades_sin_precio'])} unidad(es) todavía no tienen precio de "
                                 "venta al público cargado, así que no entran en la ganancia potencial.",
                     font=("Segoe UI", 9, "bold"), bg="#fef3c7", fg="#92400e").pack(anchor="w", padx=10, pady=6)

        for f in self.tabla_plataformas.get_children():
            self.tabla_plataformas.delete(f)
        for r in rentabilidad_por_plataforma():
            tag = "positivo" if r["margen"] >= 0 else "negativo"
            self.tabla_plataformas.insert("", "end", tags=(tag,), values=(
                r["plataforma"], int(r["unidades"]), _dual(r["costo"]), _dual(r["ingreso_potencial"]), _dual(r["margen"])))

        for f in self.tabla_productos.get_children():
            self.tabla_productos.delete(f)
        for r in rentabilidad_por_producto():
            tag = "positivo" if r["ganancia"] >= 0 else "negativo"
            venta_txt = _dual(r["precio_venta_promedio"]) if r["precio_venta_promedio"] else "Sin precio"
            self.tabla_productos.insert("", "end", tags=(tag,), values=(
                r["producto"], int(r["unidades"]), _dual(r["costo_unitario_promedio"]), venta_txt, _dual(r["ganancia"])))

    def _refrescar_texto_tasa(self):
        info = obtener_info_tasa_cambio()
        texto = f"💱 Tipo de cambio del módulo: 1 US$ = {formatear_gs(info['tasa'])}"
        if info["fecha_actualizacion"]:
            texto += f"   (actualizado: {info['fecha_actualizacion'][:16]})"
        else:
            texto += "   (valor de referencia, todavía no actualizado)"
        self.lbl_tasa.config(text=texto)

    def _editar_tasa_manual(self):
        tasa_actual = obtener_tasa_cambio()
        ventana = tk.Toplevel(self)
        ventana.title("Editar tipo de cambio")
        ventana.configure(bg="white")
        ventana.resizable(False, False)
        ventana.grab_set()
        cont = tk.Frame(ventana, bg="white", padx=18, pady=16)
        cont.pack(fill="both", expand=True)
        tk.Label(cont, text="1 US$ equivale a cuántos guaraníes:", font=("Segoe UI", 9, "bold"),
                 bg="white").pack(anchor="w")
        var_tasa = tk.StringVar(value=str(tasa_actual))
        entry = tk.Entry(cont, textvariable=var_tasa, font=("Segoe UI", 11), width=18)
        entry.pack(anchor="w", pady=(6, 12))
        entry.focus()
        entry.select_range(0, "end")

        def _guardar():
            try:
                valor = float(var_tasa.get().replace(",", "."))
            except ValueError:
                messagebox.showerror("Importaciones", "Ingresá un número válido.", parent=ventana)
                return
            ok, msg = guardar_tasa_cambio(valor)
            if not ok:
                messagebox.showerror("Importaciones", msg, parent=ventana)
                return
            ventana.destroy()
            self.panel_padre.refrescar_todo()

        barra = tk.Frame(cont, bg="white")
        barra.pack(fill="x")
        tk.Button(barra, text="Guardar", font=("Segoe UI", 9, "bold"), bg=AZUL_RIBBON, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=4, command=_guardar).pack(side="right")
        tk.Button(barra, text="Cancelar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=14, pady=4, command=ventana.destroy).pack(side="right", padx=(0, 8))
        entry.bind("<Return>", lambda e: _guardar())
        ajustar_tamaño_ventana(ventana, ancho_min=320, alto_min=170)

    def _actualizar_tasa_automatica(self):
        self.btn_actualizar_tasa.config(state="disabled", text="Consultando...")

        def _tarea():
            ok, msg, _tasa = actualizar_tasa_cambio_automatica()

            def _ui():
                if not self.winfo_exists():
                    return
                self.btn_actualizar_tasa.config(state="normal", text="🔄 Actualizar automático (Internet)")
                (messagebox.showinfo if ok else messagebox.showerror)("Importaciones", msg)
                if ok:
                    self.panel_padre.refrescar_todo()
            self.after(0, _ui)

        threading.Thread(target=_tarea, daemon=True).start()


# ============================================================
# PESTAÑA: COMPRAS
# ============================================================
class _TabCompras(tk.Frame):
    def __init__(self, parent, panel_padre: PanelImportacion):
        super().__init__(parent, bg="white")
        self.panel_padre = panel_padre
        self._construir()
        self.cargar()

    def _construir(self):
        barra = tk.Frame(self, bg="white")
        barra.pack(fill="x", pady=(8, 4), padx=4)

        tk.Button(barra, text="➕ Nueva Compra", font=("Segoe UI", 9, "bold"), bg=AZUL_RIBBON, fg="white",
                  relief="flat", cursor="hand2", padx=10, pady=4,
                  command=self._nueva_compra).pack(side="left")
        tk.Button(barra, text="✏ Ver / Editar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=10, command=self._editar_compra).pack(side="left", padx=(8, 0))
        tk.Button(barra, text="🗑 Eliminar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=10, command=self._eliminar_compra).pack(side="left", padx=(8, 0))
        tk.Button(barra, text="🏪 Tiendas", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=10, command=self._gestionar_tiendas).pack(side="left", padx=(8, 0))

        tk.Label(barra, text="Estado:", font=("Segoe UI", 9), bg="white").pack(side="left", padx=(20, 4))
        self.var_filtro_estado = tk.StringVar(value="Todos")
        combo_estado = ttk.Combobox(barra, textvariable=self.var_filtro_estado, state="readonly",
                                     values=["Todos"] + ESTADOS_COMPRA, width=18)
        combo_estado.pack(side="left")
        combo_estado.bind("<<ComboboxSelected>>", lambda e: self.cargar())

        tk.Label(barra, text="Plataforma:", font=("Segoe UI", 9), bg="white").pack(side="left", padx=(12, 4))
        self.var_filtro_plataforma = tk.StringVar(value="Todas")
        self.combo_plat = ttk.Combobox(barra, textvariable=self.var_filtro_plataforma, state="readonly", width=14)
        self.combo_plat.pack(side="left")
        self.combo_plat.bind("<<ComboboxSelected>>", lambda e: self.cargar())

        self.var_busqueda = tk.StringVar()
        entry_busq = tk.Entry(barra, textvariable=self.var_busqueda, font=("Segoe UI", 9), width=18)
        entry_busq.pack(side="left", padx=(12, 0))
        entry_busq.bind("<Return>", lambda e: self.cargar())
        tk.Button(barra, text="🔍", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", command=self.cargar).pack(side="left", padx=(4, 0))

        contenedor = tk.Frame(self, bg="white")
        contenedor.pack(fill="both", expand=True, padx=4, pady=(4, 8))
        contenedor.grid_rowconfigure(0, weight=1)
        contenedor.grid_columnconfigure(0, weight=1)

        columnas = ("id", "plataforma", "referencia", "courier", "tipo_envio", "peso", "unidades",
                    "inversion", "estado", "fecha_compra", "dias")
        self.tabla = ttk.Treeview(contenedor, columns=columnas, show="headings", selectmode="browse")
        habilitar_deseleccion_treeview(self.tabla)
        encabezados = {"id": "#", "plataforma": "PLATAFORMA", "referencia": "REFERENCIA", "courier": "COURIER",
                       "tipo_envio": "ENVÍO", "peso": "PESO CAJA", "unidades": "UNID.", "inversion": "INVERSIÓN",
                       "estado": "ESTADO", "fecha_compra": "F. COMPRA", "dias": "DÍAS"}
        anchos = {"id": 40, "plataforma": 100, "referencia": 110, "courier": 110, "tipo_envio": 70,
                  "peso": 80, "unidades": 55, "inversion": 190, "estado": 150, "fecha_compra": 90, "dias": 55}
        for col in columnas:
            self.tabla.heading(col, text=encabezados[col])
            self.tabla.column(col, width=anchos[col], anchor="center" if col != "referencia" else "w")
        sb = ttk.Scrollbar(contenedor, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=sb.set)
        self.tabla.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")
        self.tabla.bind("<Double-1>", lambda e: self._editar_compra())

        for estado, color in COLOR_ESTADO.items():
            self.tabla.tag_configure(estado, foreground=color)

    def _id_seleccionado(self):
        sel = self.tabla.selection()
        if not sel:
            return None
        return int(self.tabla.item(sel[0], "values")[0])

    def cargar(self):
        seleccion_previa = self.var_filtro_plataforma.get()
        nombres = [p["nombre"] for p in listar_plataformas(solo_activas=False)]
        self.combo_plat["values"] = ["Todas"] + nombres
        if seleccion_previa not in (["Todas"] + nombres):
            self.var_filtro_plataforma.set("Todas")

        for f in self.tabla.get_children():
            self.tabla.delete(f)
        estado = "" if self.var_filtro_estado.get() in ("", "Todos") else self.var_filtro_estado.get()
        plataforma = "" if self.var_filtro_plataforma.get() in ("", "Todas") else self.var_filtro_plataforma.get()
        for c in listar_compras(self.var_busqueda.get(), estado, plataforma):
            dias_txt = c["dias_transcurridos"] if c["dias_transcurridos"] is not None else "—"
            self.tabla.insert("", "end", tags=(c["estado"],), values=(
                c["id"], c["plataforma"], c["referencia"], c["courier"], c["tipo_envio"],
                f'{c["peso_caja_kg"]:.3f} kg', int(c["total_unidades"]), _dual(c["inversion_total"]),
                c["estado"], c["fecha_compra"], dias_txt))

    def _gestionar_tiendas(self):
        VentanaGestionTiendas(self.winfo_toplevel(), on_cambio=self.cargar)

    def _nueva_compra(self):
        VentanaFichaCompra(self.winfo_toplevel(), compra_id=None, on_guardado=self.cargar)

    def _editar_compra(self):
        compra_id = self._id_seleccionado()
        if compra_id is None:
            messagebox.showinfo("Importaciones", "Seleccioná una compra de la lista.")
            return
        VentanaFichaCompra(self.winfo_toplevel(), compra_id=compra_id, on_guardado=self.cargar)

    def _eliminar_compra(self):
        compra_id = self._id_seleccionado()
        if compra_id is None:
            messagebox.showinfo("Importaciones", "Seleccioná una compra de la lista.")
            return
        if not messagebox.askyesno("Confirmar", "¿Eliminar esta compra de importación? Esta acción no se puede deshacer."):
            return
        ok, msg = eliminar_compra(compra_id)
        (messagebox.showinfo if ok else messagebox.showerror)("Importaciones", msg)
        if ok:
            self.cargar()


# ============================================================
# FICHA DE COMPRA (cabecera + productos de la caja)
# ============================================================
class VentanaFichaCompra(tk.Toplevel):
    def __init__(self, parent, compra_id: int = None, on_guardado=None):
        super().__init__(parent)
        self.compra_id = compra_id
        self.es_nueva = compra_id is None
        self.on_guardado = on_guardado
        self.items = []  # lista de dicts en memoria hasta guardar

        self.title("Nueva Compra de Importación" if self.es_nueva else f"Compra de Importación #{compra_id}")
        self.configure(bg="white")
        self.resizable(True, True)
        self.grab_set()

        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._construir_titulo()
        self._construir_cuerpo()
        self._construir_barra_botones()

        if not self.es_nueva:
            self._cargar_datos_existentes()
        else:
            self._recalcular_preview()

        self.minsize(860, 620)
        ajustar_tamaño_ventana(self, ancho_min=860, alto_min=640, alto_max=self.winfo_screenheight() - 60)

    def _construir_titulo(self):
        barra = tk.Frame(self, bg=AZUL_RIBBON, height=34)
        barra.grid(row=0, column=0, sticky="ew")
        barra.grid_propagate(False)
        titulo = "📦 Nueva Compra de Importación" if self.es_nueva else f"📦 Compra de Importación #{self.compra_id}"
        tk.Label(barra, text=titulo, font=("Segoe UI", 11, "bold"), bg=AZUL_RIBBON, fg="white").pack(
            side="left", padx=15, pady=6)

    def _construir_cuerpo(self):
        frame_ext, contenedor_scroll = _crear_seccion_scrollable(self)
        frame_ext.grid(row=1, column=0, sticky="nsew")
        contenedor = tk.Frame(contenedor_scroll, bg="white")
        contenedor.pack(fill="both", expand=True, padx=16, pady=(12, 6))

        # ---- Datos de la compra ----
        marco_datos = tk.LabelFrame(contenedor, text=" Datos de la compra ", font=("Segoe UI", 9, "bold"),
                                     bg="white", fg=GRIS_TEXTO, padx=10, pady=8)
        marco_datos.pack(fill="x", pady=(0, 10))
        for col in (1, 3, 5):
            marco_datos.grid_columnconfigure(col, weight=1)

        tk.Label(marco_datos, text="Plataforma:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=0, column=0, sticky="e", padx=(0, 6), pady=4)
        frame_plataforma = tk.Frame(marco_datos, bg="white")
        frame_plataforma.grid(row=0, column=1, sticky="w", pady=4)
        self.var_plataforma = tk.StringVar()
        self.combo_plataforma = ttk.Combobox(frame_plataforma, textvariable=self.var_plataforma, state="readonly",
                                              width=14)
        self.combo_plataforma.pack(side="left")
        tk.Button(frame_plataforma, text="+", font=("Segoe UI", 9, "bold"), bg="white", relief="solid", bd=1,
                  width=2, cursor="hand2", command=self._agregar_tienda_rapido).pack(side="left", padx=(4, 0))
        self._refrescar_combo_plataformas()

        tk.Label(marco_datos, text="Referencia / N° Pedido:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=0, column=2, sticky="e", padx=(12, 6), pady=4)
        self.var_referencia = tk.StringVar()
        tk.Entry(marco_datos, textvariable=self.var_referencia, font=("Segoe UI", 10), width=20).grid(
            row=0, column=3, sticky="w", pady=4)

        tk.Label(marco_datos, text="Estado:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=0, column=4, sticky="e", padx=(12, 6), pady=4)
        self.var_estado = tk.StringVar(value=ESTADOS_COMPRA[0])
        ttk.Combobox(marco_datos, textvariable=self.var_estado, state="readonly",
                     values=ESTADOS_COMPRA, width=18).grid(row=0, column=5, sticky="w", pady=4)

        tk.Label(marco_datos, text="Courier:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=1, column=0, sticky="e", padx=(0, 6), pady=4)
        self._couriers = {c["nombre"]: c["id"] for c in listar_couriers(solo_activos=True)}
        self.var_courier = tk.StringVar()
        combo_courier = ttk.Combobox(marco_datos, textvariable=self.var_courier, state="readonly",
                                      values=list(self._couriers.keys()), width=16)
        combo_courier.grid(row=1, column=1, sticky="w", pady=4)
        combo_courier.bind("<<ComboboxSelected>>", lambda e: self._recalcular_preview())

        tk.Label(marco_datos, text="Casillero:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=1, column=2, sticky="e", padx=(12, 6), pady=4)
        self.var_casillero = tk.StringVar(value=CASILLEROS[0])
        ttk.Combobox(marco_datos, textvariable=self.var_casillero, state="readonly",
                     values=CASILLEROS, width=20).grid(row=1, column=3, sticky="w", pady=4)

        tk.Label(marco_datos, text="Tipo de envío:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=1, column=4, sticky="e", padx=(12, 6), pady=4)
        self.var_tipo_envio = tk.StringVar(value=TIPOS_ENVIO[0])
        combo_tipo = ttk.Combobox(marco_datos, textvariable=self.var_tipo_envio, state="readonly",
                                   values=TIPOS_ENVIO, width=18)
        combo_tipo.grid(row=1, column=5, sticky="w", pady=4)
        combo_tipo.bind("<<ComboboxSelected>>", lambda e: self._recalcular_preview())

        tk.Label(marco_datos, text="Peso de la caja (kg):", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=2, column=0, sticky="e", padx=(0, 6), pady=4)
        self.var_peso = tk.StringVar(value="0")
        entry_peso = tk.Entry(marco_datos, textvariable=self.var_peso, font=("Segoe UI", 10), width=16)
        entry_peso.grid(row=2, column=1, sticky="w", pady=4)
        entry_peso.bind("<KeyRelease>", lambda e: self._recalcular_preview())

        tk.Label(marco_datos, text="Costo envío manual (US$):", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=2, column=2, sticky="e", padx=(12, 6), pady=4)
        self.var_envio_manual = tk.StringVar()
        entry_manual = tk.Entry(marco_datos, textvariable=self.var_envio_manual, font=("Segoe UI", 10), width=16)
        entry_manual.grid(row=2, column=3, sticky="w", pady=4)
        entry_manual.bind("<KeyRelease>", lambda e: self._recalcular_preview())
        tk.Label(marco_datos, text="(dejar vacío para calcular automático: peso × tarifa/kg)",
                 font=("Segoe UI", 7), bg="white", fg=GRIS_TEXTO).grid(row=2, column=4, columnspan=2, sticky="w")

        self.lbl_costo_envio_calc = tk.Label(marco_datos, text="Costo de envío total: US$ 0.00",
                                              font=("Segoe UI", 10, "bold"), bg="white", fg=AZUL_RIBBON)
        self.lbl_costo_envio_calc.grid(row=3, column=0, columnspan=3, sticky="w", pady=(4, 0))

        tk.Label(marco_datos, text="Fecha de compra:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=4, column=0, sticky="e", padx=(0, 6), pady=(8, 4))
        self.var_fecha_compra = tk.StringVar(value=datetime.date.today().isoformat())
        _campo_fecha(marco_datos, self.var_fecha_compra).grid(row=4, column=1, sticky="w", pady=(8, 4))

        tk.Label(marco_datos, text="Fecha envío al casillero:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=4, column=2, sticky="e", padx=(12, 6), pady=(8, 4))
        self.var_fecha_envio = tk.StringVar()
        _campo_fecha(marco_datos, self.var_fecha_envio).grid(row=4, column=3, sticky="w", pady=(8, 4))

        tk.Label(marco_datos, text="Fecha de recepción final:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=4, column=4, sticky="e", padx=(12, 6), pady=(8, 4))
        self.var_fecha_recepcion = tk.StringVar()
        _campo_fecha(marco_datos, self.var_fecha_recepcion).grid(row=4, column=5, sticky="w", pady=(8, 4))

        tk.Label(marco_datos, text="Notas:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=5, column=0, sticky="ne", padx=(0, 6), pady=4)
        self.var_notas = tk.StringVar()
        tk.Entry(marco_datos, textvariable=self.var_notas, font=("Segoe UI", 10)).grid(
            row=5, column=1, columnspan=5, sticky="ew", pady=4)

        # ---- Productos de la caja ----
        marco_items = tk.LabelFrame(contenedor, text=" Productos dentro de la caja ", font=("Segoe UI", 9, "bold"),
                                     bg="white", fg=GRIS_TEXTO, padx=10, pady=8)
        marco_items.pack(fill="both", expand=True, pady=(0, 10))

        barra_items = tk.Frame(marco_items, bg="white")
        barra_items.pack(fill="x", pady=(0, 6))
        tk.Button(barra_items, text="➕ Agregar producto", font=("Segoe UI", 9, "bold"), bg=VERDE, fg="white",
                  relief="flat", cursor="hand2", padx=8, command=self._agregar_item).pack(side="left")
        tk.Button(barra_items, text="✏ Editar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=8, command=self._editar_item).pack(side="left", padx=(6, 0))
        tk.Button(barra_items, text="🗑 Quitar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=8, command=self._quitar_item).pack(side="left", padx=(6, 0))
        if not self.es_nueva:
            tk.Button(barra_items, text="📥 Enviar a Inventario", font=("Segoe UI", 9, "bold"), bg=MORADO,
                      fg="white", relief="flat", cursor="hand2", padx=8,
                      command=self._enviar_item_inventario).pack(side="left", padx=(16, 0))

        cont_tabla = tk.Frame(marco_items, bg="white")
        cont_tabla.pack(fill="both", expand=True)
        columnas = ("producto", "cantidad", "costo_unit", "envio_unit", "costo_total", "venta", "ganancia", "margen", "inventario")
        self.tabla_items = ttk.Treeview(cont_tabla, columns=columnas, show="headings", height=8)
        encabezados = {"producto": "PRODUCTO", "cantidad": "CANT.", "costo_unit": "COSTO COMPRA/U",
                       "envio_unit": "ENVÍO/U", "costo_total": "COSTO TOTAL/U", "venta": "VENTA PÚBLICO/U",
                       "ganancia": "GANANCIA/U", "margen": "MARGEN", "inventario": "INVENTARIO"}
        anchos = {"producto": 150, "cantidad": 55, "costo_unit": 95, "envio_unit": 80, "costo_total": 180,
                  "venta": 180, "ganancia": 180, "margen": 70, "inventario": 90}
        for col in columnas:
            self.tabla_items.heading(col, text=encabezados[col])
            self.tabla_items.column(col, width=anchos[col], anchor="center" if col != "producto" else "w")
        sb2 = ttk.Scrollbar(cont_tabla, orient="vertical", command=self.tabla_items.yview)
        self.tabla_items.configure(yscrollcommand=sb2.set)
        self.tabla_items.pack(side="left", fill="both", expand=True)
        sb2.pack(side="right", fill="y")
        self.tabla_items.bind("<Double-1>", lambda e: self._editar_item())
        self.tabla_items.tag_configure("positivo", foreground=VERDE)
        self.tabla_items.tag_configure("negativo", foreground=ROJO)
        self.tabla_items.tag_configure("enviado", foreground=MORADO)

        self.lbl_totales = tk.Label(contenedor, text="", font=("Segoe UI", 10, "bold"), bg="white", fg=AZUL_RIBBON)
        self.lbl_totales.pack(anchor="w", pady=(4, 0))

    def _refrescar_combo_plataformas(self, seleccionar: str = None):
        nombres = [p["nombre"] for p in listar_plataformas(solo_activas=False)]
        self.combo_plataforma["values"] = nombres
        if seleccionar and seleccionar not in nombres:
            nombres = nombres + [seleccionar]
            self.combo_plataforma["values"] = nombres
        if seleccionar:
            self.var_plataforma.set(seleccionar)
        elif not self.var_plataforma.get() and nombres:
            self.var_plataforma.set(nombres[0])

    def _agregar_tienda_rapido(self):
        ventana = tk.Toplevel(self)
        ventana.title("Nueva tienda")
        ventana.configure(bg="white")
        ventana.resizable(False, False)
        ventana.grab_set()
        cont = tk.Frame(ventana, bg="white", padx=18, pady=16)
        cont.pack(fill="both", expand=True)
        tk.Label(cont, text="Nombre de la tienda/plataforma:", font=("Segoe UI", 9, "bold"),
                 bg="white").pack(anchor="w")
        var_nombre = tk.StringVar()
        entry = tk.Entry(cont, textvariable=var_nombre, font=("Segoe UI", 10), width=26)
        entry.pack(anchor="w", pady=(6, 12))
        entry.focus()

        def _guardar():
            nombre = var_nombre.get().strip()
            if not nombre:
                messagebox.showerror("Importaciones", "El nombre es obligatorio.", parent=ventana)
                return
            ok, msg = crear_plataforma(nombre)
            if not ok:
                messagebox.showerror("Importaciones", msg, parent=ventana)
                return
            ventana.destroy()
            self._refrescar_combo_plataformas(seleccionar=nombre)

        barra = tk.Frame(cont, bg="white")
        barra.pack(fill="x")
        tk.Button(barra, text="Agregar", font=("Segoe UI", 9, "bold"), bg=AZUL_RIBBON, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=4, command=_guardar).pack(side="right")
        tk.Button(barra, text="Cancelar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=14, pady=4, command=ventana.destroy).pack(side="right", padx=(0, 8))
        entry.bind("<Return>", lambda e: _guardar())
        ajustar_tamaño_ventana(ventana, ancho_min=320, alto_min=150)

    def _construir_barra_botones(self):
        barra = tk.Frame(self, bg=GRIS_FONDO)
        barra.grid(row=2, column=0, sticky="ew")
        tk.Button(barra, text="💾 Guardar", font=("Segoe UI", 10, "bold"), bg=AZUL_RIBBON, fg="white",
                  relief="flat", cursor="hand2", padx=16, pady=6, command=self._guardar).pack(
                  side="right", padx=10, pady=8)
        tk.Button(barra, text="Cancelar", font=("Segoe UI", 10), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=16, pady=6, command=self.destroy).pack(side="right", pady=8)

    # ---------- datos ----------
    def _peso_actual(self) -> float:
        try:
            return float(self.var_peso.get().replace(",", "."))
        except ValueError:
            return 0.0

    def _envio_manual_actual(self):
        txt = self.var_envio_manual.get().strip()
        if not txt:
            return None
        try:
            return float(txt.replace(",", "."))
        except ValueError:
            return None

    def _courier_id_actual(self):
        return self._couriers.get(self.var_courier.get())

    def _cargar_datos_existentes(self):
        d = obtener_compra_detalle(self.compra_id)
        if not d:
            messagebox.showerror("Importaciones", "No se encontró la compra.")
            self.destroy()
            return
        self.var_plataforma.set(d["plataforma"])
        self.var_referencia.set(d["referencia"])
        self.var_estado.set(d["estado"])
        if d["courier"] and d["courier"] != "Sin courier":
            self.var_courier.set(d["courier"])
        self.var_casillero.set(d["casillero"])
        self.var_tipo_envio.set(d["tipo_envio"])
        self.var_peso.set(str(d["peso_caja_kg"]))
        if d["costo_envio_manual"] is not None:
            self.var_envio_manual.set(str(d["costo_envio_manual"]))
        self.var_fecha_compra.set(d["fecha_compra"] or "")
        self.var_fecha_envio.set(d["fecha_envio_casillero"] or "")
        self.var_fecha_recepcion.set(d["fecha_recepcion"] or "")
        self.var_notas.set(d["notas"])
        self.items = [dict(it) for it in d["items"]]
        self._recalcular_preview()

    def _recalcular_preview(self):
        items_simplificados = [{"producto_nombre": it["producto_nombre"], "cantidad": it["cantidad"],
                                 "costo_unitario_compra": it["costo_unitario_compra"],
                                 "precio_venta_publico": it.get("precio_venta_publico")} for it in self.items]
        costo_envio_total, items_recalc = previsualizar_costos(
            self._peso_actual(), self._courier_id_actual(), self.var_tipo_envio.get(),
            self._envio_manual_actual(), items_simplificados)

        # conserva metadata (id, enviado_inventario) al reescribir self.items
        for original, recalc in zip(self.items, items_recalc):
            original["costo_envio_unitario"] = recalc["costo_envio_unitario"]
            original["costo_total_unitario"] = recalc["costo_total_unitario"]
            original["ganancia_unitaria"] = recalc["ganancia_unitaria"]
            original["margen_pct"] = recalc["margen_pct"]

        self.lbl_costo_envio_calc.config(text=f"Costo de envío total de la caja: {_dual(costo_envio_total)}"
                                               f"  (se reparte entre {int(sum(it['cantidad'] for it in self.items)) if self.items else 0} unidad(es))")
        self._refrescar_tabla_items()

    def _refrescar_tabla_items(self):
        for f in self.tabla_items.get_children():
            self.tabla_items.delete(f)
        costo_total_inv = 0.0
        venta_total = 0.0
        for it in self.items:
            venta = it.get("precio_venta_publico")
            ganancia = it.get("ganancia_unitaria")
            margen = it.get("margen_pct")
            tag = "enviado" if it.get("enviado_inventario") else ("positivo" if (ganancia or 0) >= 0 else "negativo")
            costo_total_inv += it["cantidad"] * it["costo_total_unitario"]
            if venta:
                venta_total += it["cantidad"] * venta
            self.tabla_items.insert("", "end", tags=(tag,), values=(
                it["producto_nombre"], it["cantidad"], _usd(it["costo_unitario_compra"]),
                _usd(it["costo_envio_unitario"]), _dual(it["costo_total_unitario"]),
                _dual(venta) if venta else "—", _dual(ganancia) if ganancia is not None else "—",
                f"{margen}%" if margen is not None else "—",
                "✔ Enviado" if it.get("enviado_inventario") else "—",
            ))
        ganancia_total = venta_total - costo_total_inv if venta_total else None
        texto = f"Inversión total de la caja: {_dual(costo_total_inv)}"
        if ganancia_total is not None:
            texto += f"   |   Venta potencial: {_dual(venta_total)}   |   Ganancia potencial: {_dual(ganancia_total)}"
        self.lbl_totales.config(text=texto)

    # ---------- items ----------
    def _agregar_item(self):
        _DialogoItem(self, on_guardar=self._agregar_item_callback)

    def _agregar_item_callback(self, datos):
        self.items.append({
            "id": None, "producto_nombre": datos["producto_nombre"], "cantidad": datos["cantidad"],
            "costo_unitario_compra": datos["costo_unitario_compra"],
            "precio_venta_publico": datos["precio_venta_publico"],
            "costo_envio_unitario": 0, "costo_total_unitario": 0, "ganancia_unitaria": None,
            "margen_pct": None, "enviado_inventario": False, "producto_id_generado": None,
        })
        self._recalcular_preview()

    def _item_seleccionado_idx(self):
        sel = self.tabla_items.selection()
        if not sel:
            return None
        return self.tabla_items.index(sel[0])

    def _editar_item(self):
        idx = self._item_seleccionado_idx()
        if idx is None:
            messagebox.showinfo("Importaciones", "Seleccioná un producto de la lista.")
            return
        item = self.items[idx]
        if item.get("enviado_inventario"):
            messagebox.showinfo("Importaciones", "Este producto ya fue enviado a Inventario y no se puede editar.")
            return

        def _guardar(datos):
            item["producto_nombre"] = datos["producto_nombre"]
            item["cantidad"] = datos["cantidad"]
            item["costo_unitario_compra"] = datos["costo_unitario_compra"]
            item["precio_venta_publico"] = datos["precio_venta_publico"]
            self._recalcular_preview()

        _DialogoItem(self, on_guardar=_guardar, datos_iniciales=item)

    def _quitar_item(self):
        idx = self._item_seleccionado_idx()
        if idx is None:
            messagebox.showinfo("Importaciones", "Seleccioná un producto de la lista.")
            return
        if self.items[idx].get("enviado_inventario"):
            messagebox.showinfo("Importaciones", "Este producto ya fue enviado a Inventario y no se puede quitar.")
            return
        del self.items[idx]
        self._recalcular_preview()

    def _enviar_item_inventario(self):
        idx = self._item_seleccionado_idx()
        if idx is None:
            messagebox.showinfo("Importaciones", "Seleccioná un producto de la lista.")
            return
        item = self.items[idx]
        if item.get("enviado_inventario"):
            messagebox.showinfo("Importaciones", "Este producto ya fue enviado a Inventario.")
            return
        if item.get("id") is None:
            messagebox.showinfo("Importaciones", "Guardá la compra primero antes de enviar sus productos a Inventario.")
            return
        _DialogoEnviarInventario(self, item, on_enviado=self._recargar_tras_envio)

    def _recargar_tras_envio(self):
        self._cargar_datos_existentes()

    # ---------- guardar ----------
    def _guardar(self):
        try:
            peso = self._peso_actual()
        except ValueError:
            messagebox.showerror("Importaciones", "El peso de la caja no es válido.")
            return
        if not self.items:
            messagebox.showerror("Importaciones", "Agregá al menos un producto a la caja.")
            return

        items_para_guardar = [{"producto_nombre": it["producto_nombre"], "cantidad": it["cantidad"],
                                "costo_unitario_compra": it["costo_unitario_compra"],
                                "precio_venta_publico": it.get("precio_venta_publico")} for it in self.items]

        kwargs = dict(
            plataforma=self.var_plataforma.get(), referencia=self.var_referencia.get(),
            courier_id=self._courier_id_actual(), casillero=self.var_casillero.get(),
            tipo_envio=self.var_tipo_envio.get(), peso_caja_kg=peso,
            costo_envio_manual=self._envio_manual_actual(), estado=self.var_estado.get(),
            fecha_compra=self.var_fecha_compra.get() or datetime.date.today().isoformat(),
            fecha_envio_casillero=self.var_fecha_envio.get() or None,
            fecha_recepcion=self.var_fecha_recepcion.get() or None,
            notas=self.var_notas.get(), items=items_para_guardar,
        )

        # Todo el guardado va blindado: si pasa CUALQUIER cosa inesperada acá
        # adentro, el usuario tiene que ver el error sí o sí (antes, un error
        # en este bloque podía quedar silencioso en el .exe compilado y la
        # compra parecía "guardada" sin estarlo realmente).
        try:
            usuario = getattr(self.master, "usuario_actual", None)
            usuario_id = usuario.get("id") if isinstance(usuario, dict) else None
            if self.es_nueva:
                ok, msg, nuevo_id = crear_compra(usuario_id=usuario_id, **kwargs)
            else:
                ok, msg = editar_compra(self.compra_id, **kwargs)
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Importaciones", f"Ocurrió un error inesperado al guardar:\n\n{e}")
            return

        if not ok:
            messagebox.showerror("Importaciones", msg)
            return
        messagebox.showinfo("Importaciones", msg)
        if self.on_guardado:
            self.on_guardado()
        self.destroy()


class _DialogoItem(tk.Toplevel):
    """Alta/edición de un producto dentro de la caja."""
    def __init__(self, parent, on_guardar, datos_iniciales: dict = None):
        super().__init__(parent)
        self.on_guardar = on_guardar
        self.title("Producto de la caja")
        self.configure(bg="white")
        self.resizable(False, False)
        self.grab_set()

        cont = tk.Frame(self, bg="white", padx=16, pady=14)
        cont.pack(fill="both", expand=True)

        tk.Label(cont, text="Nombre del producto:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=0, column=0, sticky="e", pady=4, padx=(0, 8))
        self.var_nombre = tk.StringVar(value=(datos_iniciales or {}).get("producto_nombre", ""))
        tk.Entry(cont, textvariable=self.var_nombre, font=("Segoe UI", 10), width=28).grid(
            row=0, column=1, sticky="w", pady=4)

        tk.Label(cont, text="Cantidad (unidades):", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=1, column=0, sticky="e", pady=4, padx=(0, 8))
        self.var_cantidad = tk.StringVar(value=str((datos_iniciales or {}).get("cantidad", "1")))
        tk.Entry(cont, textvariable=self.var_cantidad, font=("Segoe UI", 10), width=28).grid(
            row=1, column=1, sticky="w", pady=4)

        tk.Label(cont, text="Costo pagado en la plataforma\n(por unidad, US$):", font=("Segoe UI", 9, "bold"),
                 bg="white", justify="right").grid(row=2, column=0, sticky="e", pady=4, padx=(0, 8))
        self.var_costo = tk.StringVar(value=str((datos_iniciales or {}).get("costo_unitario_compra", "0")))
        tk.Entry(cont, textvariable=self.var_costo, font=("Segoe UI", 10), width=28).grid(
            row=2, column=1, sticky="w", pady=4)

        tk.Label(cont, text="Precio de venta al público\n(por unidad, opcional):", font=("Segoe UI", 9, "bold"),
                 bg="white", justify="right").grid(row=3, column=0, sticky="e", pady=4, padx=(0, 8))
        precio_ini = (datos_iniciales or {}).get("precio_venta_publico")
        self.var_precio = tk.StringVar(value=str(precio_ini) if precio_ini else "")
        tk.Entry(cont, textvariable=self.var_precio, font=("Segoe UI", 10), width=28).grid(
            row=3, column=1, sticky="w", pady=4)

        barra = tk.Frame(cont, bg="white")
        barra.grid(row=4, column=0, columnspan=2, pady=(12, 0), sticky="e")
        tk.Button(barra, text="Guardar", font=("Segoe UI", 9, "bold"), bg=AZUL_RIBBON, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=4, command=self._guardar).pack(side="right")
        tk.Button(barra, text="Cancelar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=14, pady=4, command=self.destroy).pack(side="right", padx=(0, 8))

        self.bind("<Return>", lambda e: self._guardar())
        ajustar_tamaño_ventana(self, ancho_min=380, alto_min=280)

    def _guardar(self):
        nombre = self.var_nombre.get().strip()
        if not nombre:
            messagebox.showerror("Importaciones", "El nombre del producto es obligatorio.")
            return
        try:
            cantidad = float(self.var_cantidad.get().replace(",", "."))
            if cantidad <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Importaciones", "La cantidad debe ser un número mayor a cero.")
            return
        try:
            costo = float(self.var_costo.get().replace(",", "."))
        except ValueError:
            messagebox.showerror("Importaciones", "El costo unitario no es válido.")
            return
        precio_txt = self.var_precio.get().strip()
        precio = None
        if precio_txt:
            try:
                precio = float(precio_txt.replace(",", "."))
            except ValueError:
                messagebox.showerror("Importaciones", "El precio de venta no es válido.")
                return

        self.on_guardar({"producto_nombre": nombre, "cantidad": cantidad,
                          "costo_unitario_compra": costo, "precio_venta_publico": precio})
        self.destroy()


class _DialogoEnviarInventario(tk.Toplevel):
    """Envía las unidades de un ítem ya recibido al stock del módulo
    Productos/Inventario, para que se puedan vender desde Ventas."""
    def __init__(self, parent, item: dict, on_enviado):
        super().__init__(parent)
        self.item = item
        self.on_enviado = on_enviado
        self.producto_elegido_id = None

        self.title("Enviar a Inventario")
        self.configure(bg="white")
        self.resizable(False, False)
        self.grab_set()

        cont = tk.Frame(self, bg="white", padx=16, pady=14)
        cont.pack(fill="both", expand=True)

        tk.Label(cont, text=f"{item['producto_nombre']}  ·  {item['cantidad']} unidad(es)  ·  "
                            f"costo {_dual(item['costo_total_unitario'])}/u",
                 font=("Segoe UI", 10, "bold"), bg="white", wraplength=360, justify="left").pack(anchor="w", pady=(0, 10))

        tk.Label(cont, text="¿Es un producto nuevo o ya existe en tu catálogo de Productos?",
                 font=("Segoe UI", 9), bg="white", wraplength=360, justify="left").pack(anchor="w")

        self.var_modo = tk.StringVar(value="nuevo")
        tk.Radiobutton(cont, text="Crear como producto nuevo", variable=self.var_modo, value="nuevo",
                       bg="white", font=("Segoe UI", 9), command=self._actualizar_modo).pack(anchor="w", pady=(6, 0))
        tk.Radiobutton(cont, text="Sumar stock a un producto que ya existe", variable=self.var_modo,
                       value="existente", bg="white", font=("Segoe UI", 9), command=self._actualizar_modo).pack(
                       anchor="w")

        self.frame_buscar = tk.Frame(cont, bg="white")
        self.frame_buscar.pack(fill="x", pady=(6, 0))
        self.var_buscar = tk.StringVar(value=item["producto_nombre"])
        entry_buscar = tk.Entry(self.frame_buscar, textvariable=self.var_buscar, font=("Segoe UI", 9), width=26)
        entry_buscar.pack(side="left")
        tk.Button(self.frame_buscar, text="Buscar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", command=self._buscar).pack(side="left", padx=(6, 0))

        self.lista_resultados = tk.Listbox(cont, font=("Segoe UI", 9), height=5)
        self.lista_resultados.pack(fill="x", pady=(6, 0))
        self._resultados = []

        self._actualizar_modo()

        barra = tk.Frame(cont, bg="white")
        barra.pack(fill="x", pady=(12, 0))
        tk.Button(barra, text="📥 Enviar a Inventario", font=("Segoe UI", 9, "bold"), bg=MORADO, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=4, command=self._confirmar).pack(side="right")
        tk.Button(barra, text="Cancelar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=14, pady=4, command=self.destroy).pack(side="right", padx=(0, 8))

        ajustar_tamaño_ventana(self, ancho_min=400, alto_min=380)

    def _actualizar_modo(self):
        estado = "normal" if self.var_modo.get() == "existente" else "disabled"
        for w in self.frame_buscar.winfo_children():
            w.config(state=estado)
        self.lista_resultados.config(state=estado)
        if estado == "normal":
            self._buscar()

    def _buscar(self):
        self.lista_resultados.delete(0, "end")
        self._resultados = buscar_productos_similares(self.var_buscar.get())
        for p in self._resultados:
            self.lista_resultados.insert("end", f"{p['nombre']}  (stock: {p['stock']})")

    def _confirmar(self):
        producto_id = None
        if self.var_modo.get() == "existente":
            sel = self.lista_resultados.curselection()
            if not sel:
                messagebox.showerror("Importaciones", "Elegí un producto de la lista, o cambiá a 'Crear nuevo'.")
                return
            producto_id = self._resultados[sel[0]]["id"]
        ok, msg = enviar_a_inventario(self.item["id"], producto_existente_id=producto_id)
        (messagebox.showinfo if ok else messagebox.showerror)("Importaciones", msg)
        if ok:
            if self.on_enviado:
                self.on_enviado()
            self.destroy()


# ============================================================
# PESTAÑA: COURIERS / CASILLEROS
# ============================================================
class _TabCouriers(tk.Frame):
    def __init__(self, parent, panel_padre: PanelImportacion):
        super().__init__(parent, bg="white")
        self.panel_padre = panel_padre
        self._construir()
        self.cargar()

    def _construir(self):
        barra = tk.Frame(self, bg="white")
        barra.pack(fill="x", pady=(8, 4), padx=4)
        tk.Button(barra, text="➕ Nuevo Courier", font=("Segoe UI", 9, "bold"), bg=AZUL_RIBBON, fg="white",
                  relief="flat", cursor="hand2", padx=10, pady=4, command=self._nuevo).pack(side="left")
        tk.Button(barra, text="✏ Editar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=10, command=self._editar).pack(side="left", padx=(8, 0))
        tk.Button(barra, text="🔁 Activar/Desactivar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=10, command=self._alternar_estado).pack(side="left", padx=(8, 0))
        tk.Button(barra, text="🗑 Eliminar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=10, command=self._eliminar).pack(side="left", padx=(8, 0))

        tk.Label(self, text="El costo por kg (aéreo o marítimo) que cargues acá se usa automáticamente para "
                            "calcular el flete de cada compra según el peso de la caja.",
                 font=("Segoe UI", 8), bg="white", fg=GRIS_TEXTO).pack(anchor="w", padx=6, pady=(0, 4))

        contenedor = tk.Frame(self, bg="white")
        contenedor.pack(fill="both", expand=True, padx=4, pady=(0, 8))
        contenedor.grid_rowconfigure(0, weight=1)
        contenedor.grid_columnconfigure(0, weight=1)

        columnas = ("id", "nombre", "email", "ruc", "aereo", "maritimo", "activo")
        self.tabla = ttk.Treeview(contenedor, columns=columnas, show="headings", selectmode="browse")
        habilitar_deseleccion_treeview(self.tabla)
        encabezados = {"id": "#", "nombre": "COURIER", "email": "EMAIL", "ruc": "RUC",
                       "aereo": "US$/KG AÉREO", "maritimo": "US$/KG MARÍTIMO", "activo": "ESTADO"}
        anchos = {"id": 40, "nombre": 150, "email": 180, "ruc": 100, "aereo": 190, "maritimo": 190, "activo": 90}
        for col in columnas:
            self.tabla.heading(col, text=encabezados[col])
            self.tabla.column(col, width=anchos[col], anchor="center" if col not in ("nombre", "email") else "w")
        sb = ttk.Scrollbar(contenedor, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=sb.set)
        self.tabla.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")
        self.tabla.bind("<Double-1>", lambda e: self._editar())
        self.tabla.tag_configure("inactivo", foreground=GRIS_TEXTO)

    def _id_seleccionado(self):
        sel = self.tabla.selection()
        if not sel:
            return None
        return int(self.tabla.item(sel[0], "values")[0])

    def cargar(self):
        for f in self.tabla.get_children():
            self.tabla.delete(f)
        for c in listar_couriers(solo_activos=False):
            tag = "" if c["activo"] else "inactivo"
            self.tabla.insert("", "end", tags=(tag,), values=(
                c["id"], c["nombre"], c["email"], c["ruc"], _dual(c["costo_kg_aereo"]),
                _dual(c["costo_kg_maritimo"]), "Activo" if c["activo"] else "Inactivo"))

    def _nuevo(self):
        VentanaFichaCourier(self.winfo_toplevel(), courier_id=None, on_guardado=self.cargar)

    def _editar(self):
        courier_id = self._id_seleccionado()
        if courier_id is None:
            messagebox.showinfo("Importaciones", "Seleccioná un courier de la lista.")
            return
        VentanaFichaCourier(self.winfo_toplevel(), courier_id=courier_id, on_guardado=self.cargar)

    def _alternar_estado(self):
        courier_id = self._id_seleccionado()
        if courier_id is None:
            messagebox.showinfo("Importaciones", "Seleccioná un courier de la lista.")
            return
        c = obtener_courier(courier_id)
        cambiar_estado_courier(courier_id, not c["activo"])
        self.cargar()

    def _eliminar(self):
        courier_id = self._id_seleccionado()
        if courier_id is None:
            messagebox.showinfo("Importaciones", "Seleccioná un courier de la lista.")
            return
        if not messagebox.askyesno("Confirmar", "¿Eliminar este courier?"):
            return
        ok, msg = eliminar_courier(courier_id)
        (messagebox.showinfo if ok else messagebox.showerror)("Importaciones", msg)
        if ok:
            self.cargar()


class VentanaFichaCourier(tk.Toplevel):
    def __init__(self, parent, courier_id: int = None, on_guardado=None):
        super().__init__(parent)
        self.courier_id = courier_id
        self.es_nuevo = courier_id is None
        self.on_guardado = on_guardado

        self.title("Nuevo Courier" if self.es_nuevo else "Editar Courier")
        self.configure(bg="white")
        self.resizable(False, False)
        self.grab_set()

        cont = tk.Frame(self, bg="white", padx=18, pady=16)
        cont.pack(fill="both", expand=True)
        cont.grid_columnconfigure(1, weight=1)

        campos = [
            ("Nombre del courier:", "nombre", 30),
            ("Email de contacto:", "email", 30),
            ("RUC:", "ruc", 20),
            ("Teléfono:", "telefono", 20),
        ]
        self.vars = {}
        fila = 0
        for etiqueta, clave, ancho in campos:
            tk.Label(cont, text=etiqueta, font=("Segoe UI", 9, "bold"), bg="white").grid(
                row=fila, column=0, sticky="e", pady=4, padx=(0, 8))
            var = tk.StringVar()
            tk.Entry(cont, textvariable=var, font=("Segoe UI", 10), width=ancho).grid(
                row=fila, column=1, sticky="w", pady=4)
            self.vars[clave] = var
            fila += 1

        tk.Label(cont, text="Costo por kg vía AÉREA (US$):", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=fila, column=0, sticky="e", pady=4, padx=(0, 8))
        self.vars["costo_kg_aereo"] = tk.StringVar(value="0")
        tk.Entry(cont, textvariable=self.vars["costo_kg_aereo"], font=("Segoe UI", 10), width=30).grid(
            row=fila, column=1, sticky="w", pady=4)
        fila += 1

        tk.Label(cont, text="Costo por kg vía MARÍTIMA (US$):", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=fila, column=0, sticky="e", pady=4, padx=(0, 8))
        self.vars["costo_kg_maritimo"] = tk.StringVar(value="0")
        tk.Entry(cont, textvariable=self.vars["costo_kg_maritimo"], font=("Segoe UI", 10), width=30).grid(
            row=fila, column=1, sticky="w", pady=4)
        fila += 1

        tk.Label(cont, text="Dirección casillero Miami:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=fila, column=0, sticky="e", pady=4, padx=(0, 8))
        self.vars["direccion_casillero_miami"] = tk.StringVar()
        tk.Entry(cont, textvariable=self.vars["direccion_casillero_miami"], font=("Segoe UI", 10), width=30).grid(
            row=fila, column=1, sticky="w", pady=4)
        fila += 1

        tk.Label(cont, text="Dirección casillero Shenzhen:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=fila, column=0, sticky="e", pady=4, padx=(0, 8))
        self.vars["direccion_casillero_shenzhen"] = tk.StringVar()
        tk.Entry(cont, textvariable=self.vars["direccion_casillero_shenzhen"], font=("Segoe UI", 10), width=30).grid(
            row=fila, column=1, sticky="w", pady=4)
        fila += 1

        tk.Label(cont, text="Notas:", font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=fila, column=0, sticky="ne", pady=4, padx=(0, 8))
        self.vars["notas"] = tk.StringVar()
        tk.Entry(cont, textvariable=self.vars["notas"], font=("Segoe UI", 10), width=30).grid(
            row=fila, column=1, sticky="w", pady=4)
        fila += 1

        barra = tk.Frame(cont, bg="white")
        barra.grid(row=fila, column=0, columnspan=2, pady=(14, 0), sticky="e")
        tk.Button(barra, text="💾 Guardar", font=("Segoe UI", 9, "bold"), bg=AZUL_RIBBON, fg="white",
                  relief="flat", cursor="hand2", padx=14, pady=4, command=self._guardar).pack(side="right")
        tk.Button(barra, text="Cancelar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", padx=14, pady=4, command=self.destroy).pack(side="right", padx=(0, 8))

        if not self.es_nuevo:
            self._cargar_datos()

        ajustar_tamaño_ventana(self, ancho_min=420, alto_min=420)

    def _cargar_datos(self):
        c = obtener_courier(self.courier_id)
        if not c:
            messagebox.showerror("Importaciones", "No se encontró el courier.")
            self.destroy()
            return
        for clave, var in self.vars.items():
            var.set(str(c.get(clave, "")))

    def _guardar(self):
        nombre = self.vars["nombre"].get().strip()
        if not nombre:
            messagebox.showerror("Importaciones", "El nombre del courier es obligatorio.")
            return
        try:
            costo_aereo = float(self.vars["costo_kg_aereo"].get().replace(",", ".") or 0)
            costo_maritimo = float(self.vars["costo_kg_maritimo"].get().replace(",", ".") or 0)
        except ValueError:
            messagebox.showerror("Importaciones", "Las tarifas por kg deben ser números.")
            return

        kwargs = dict(
            nombre=nombre, email=self.vars["email"].get(), ruc=self.vars["ruc"].get(),
            telefono=self.vars["telefono"].get(), costo_kg_aereo=costo_aereo, costo_kg_maritimo=costo_maritimo,
            direccion_casillero_miami=self.vars["direccion_casillero_miami"].get(),
            direccion_casillero_shenzhen=self.vars["direccion_casillero_shenzhen"].get(),
            notas=self.vars["notas"].get(),
        )
        if self.es_nuevo:
            ok, msg = crear_courier(**kwargs)
        else:
            ok, msg = editar_courier(self.courier_id, **kwargs)

        (messagebox.showinfo if ok else messagebox.showerror)("Importaciones", msg)
        if ok:
            if self.on_guardado:
                self.on_guardado()
            self.destroy()


# ============================================================
# GESTIÓN DE TIENDAS / PLATAFORMAS
# ============================================================
class VentanaGestionTiendas(tk.Toplevel):
    """Alta rápida de tiendas nuevas (ej. Amazon, o cualquier otra que uses
    para importar) y activar/desactivar/eliminar las existentes. Las
    tiendas desactivadas dejan de aparecer para elegir en compras nuevas,
    pero las compras ya registradas con ellas se siguen viendo normalmente."""
    def __init__(self, parent, on_cambio=None):
        super().__init__(parent)
        self.on_cambio = on_cambio
        self.title("Tiendas / Plataformas de Importación")
        self.configure(bg="white")
        self.resizable(False, False)
        self.grab_set()

        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)

        barra = tk.Frame(self, bg=AZUL_RIBBON, height=34)
        barra.grid(row=0, column=0, sticky="ew")
        barra.grid_propagate(False)
        tk.Label(barra, text="🏪 Tiendas / Plataformas", font=("Segoe UI", 11, "bold"), bg=AZUL_RIBBON,
                 fg="white").pack(side="left", padx=15, pady=6)

        formulario = tk.Frame(self, bg="white")
        formulario.grid(row=1, column=0, sticky="ew", padx=16, pady=(12, 6))
        self.var_nombre = tk.StringVar()
        entry = tk.Entry(formulario, textvariable=self.var_nombre, font=("Segoe UI", 10))
        entry.pack(side="left", fill="x", expand=True)
        entry.bind("<Return>", lambda e: self._agregar())
        tk.Button(formulario, text="➕ Agregar", font=("Segoe UI", 9, "bold"), bg=AZUL_RIBBON, fg="white",
                  relief="flat", padx=10, pady=4, cursor="hand2", command=self._agregar).pack(
                  side="left", padx=(8, 0))

        contenedor = tk.Frame(self, bg="white")
        contenedor.grid(row=2, column=0, sticky="nsew", padx=16, pady=(6, 6))
        contenedor.grid_rowconfigure(0, weight=1)
        contenedor.grid_columnconfigure(0, weight=1)
        columnas = ("nombre", "estado")
        self.tabla = ttk.Treeview(contenedor, columns=columnas, show="headings", selectmode="browse")
        habilitar_deseleccion_treeview(self.tabla)
        for col, enc in zip(columnas, ("TIENDA", "ESTADO")):
            self.tabla.heading(col, text=enc)
            self.tabla.column(col, width=220 if col == "nombre" else 100)
        sb = ttk.Scrollbar(contenedor, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=sb.set)
        self.tabla.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")
        self.tabla.tag_configure("inactiva", foreground=GRIS_TEXTO)

        botones = tk.Frame(self, bg="white")
        botones.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 8))
        tk.Button(botones, text="🔁 Activar/Desactivar", font=("Segoe UI", 9, "bold"), bg="white",
                  relief="solid", bd=1, cursor="hand2", command=self._alternar_estado).pack(side="left")
        tk.Button(botones, text="🗑 Eliminar", font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                  cursor="hand2", command=self._eliminar).pack(side="left", padx=(8, 0))

        barra_final = tk.Frame(self, bg="white")
        barra_final.grid(row=4, column=0, sticky="ew", padx=16, pady=(0, 16))
        tk.Button(barra_final, text="Cerrar", font=("Segoe UI", 9, "bold"), bg="white",
                  relief="solid", bd=1, cursor="hand2", command=self._cerrar).pack(side="right")

        self.minsize(400, 440)
        ajustar_tamaño_ventana(self, ancho_min=400, alto_min=460)
        entry.focus()
        self._cargar()

    def _id_seleccionado(self):
        seleccion = self.tabla.selection()
        if not seleccion:
            return None
        return int(seleccion[0])

    def _cargar(self):
        for f in self.tabla.get_children():
            self.tabla.delete(f)
        for p in listar_plataformas(solo_activas=False):
            tag = "" if p["activa"] else "inactiva"
            self.tabla.insert("", "end", iid=str(p["id"]), tags=(tag,),
                               values=(p["nombre"], "Activa" if p["activa"] else "Inactiva"))

    def _agregar(self):
        ok, msg = crear_plataforma(self.var_nombre.get())
        if not ok:
            messagebox.showerror("Importaciones", msg, parent=self)
            return
        self.var_nombre.set("")
        self._cargar()
        if self.on_cambio:
            self.on_cambio()

    def _alternar_estado(self):
        pid = self._id_seleccionado()
        if pid is None:
            messagebox.showinfo("Importaciones", "Elegí una tienda de la lista.", parent=self)
            return
        plataforma = next((p for p in listar_plataformas(solo_activas=False) if p["id"] == pid), None)
        if not plataforma:
            return
        cambiar_estado_plataforma(pid, not plataforma["activa"])
        self._cargar()
        if self.on_cambio:
            self.on_cambio()

    def _eliminar(self):
        pid = self._id_seleccionado()
        if pid is None:
            messagebox.showinfo("Importaciones", "Elegí una tienda de la lista.", parent=self)
            return
        if not messagebox.askyesno("Confirmar", "¿Eliminar esta tienda?", parent=self):
            return
        ok, msg = eliminar_plataforma(pid)
        if not ok:
            messagebox.showerror("Importaciones", msg, parent=self)
            return
        self._cargar()
        if self.on_cambio:
            self.on_cambio()

    def _cerrar(self):
        if self.on_cambio:
            self.on_cambio()
        self.destroy()
