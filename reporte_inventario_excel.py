"""
reporte_inventario_excel.py
Genera el Reporte General de Inventario en formato Excel (.xlsx): resumen
general, productos bajo stock mínimo, top productos por valor de
inventario, y el listado completo de productos. Usa fórmulas reales de
Excel (no valores fijos) para que el archivo se recalcule si se edita.

Requiere la librería 'openpyxl' (pip install openpyxl).
"""
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models_catalogo import listar_productos, productos_bajo_stock_minimo, top_productos_por_valor_inventario

AZUL_RIBBON = "1D5FD6"
GRIS_CLARO = "F4F5F7"
BLANCO = "FFFFFF"
ROJO_BAJO_STOCK = "FDE8E8"

FUENTE = "Arial"


def _estilo_encabezado():
    return {
        "font": Font(name=FUENTE, bold=True, color=BLANCO, size=10),
        "fill": PatternFill("solid", start_color=AZUL_RIBBON),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def _aplicar_estilo_encabezado(celda):
    estilo = _estilo_encabezado()
    celda.font = estilo["font"]
    celda.fill = estilo["fill"]
    celda.alignment = estilo["alignment"]


def _autoancho_columnas(hoja, anchos: dict):
    for col, ancho in anchos.items():
        hoja.column_dimensions[col].width = ancho


def generar_reporte_inventario_excel(ruta_destino: str, generado_por: str = "") -> str:
    """Genera el archivo Excel en ruta_destino. Devuelve la ruta generada."""
    productos = listar_productos()
    bajo_stock = productos_bajo_stock_minimo()
    top_valor = top_productos_por_valor_inventario(limite=10)

    wb = Workbook()

    # ---------------- HOJA 1: RESUMEN ----------------
    hoja_resumen = wb.active
    hoja_resumen.title = "Resumen"

    hoja_resumen["A1"] = "Reporte General de Inventario"
    hoja_resumen["A1"].font = Font(name=FUENTE, bold=True, size=16, color=AZUL_RIBBON)
    hoja_resumen["A2"] = f"Generado el {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}" + (
        f" por {generado_por}" if generado_por else "")
    hoja_resumen["A2"].font = Font(name=FUENTE, italic=True, size=9, color="666666")

    hoja_resumen["A4"] = "RESUMEN GENERAL"
    _aplicar_estilo_encabezado(hoja_resumen["A4"])
    hoja_resumen.merge_cells("A4:B4")

    # Fórmulas reales referenciando la hoja "Listado" (se llena más abajo)
    filas_listado = len(productos) + 1  # +1 por el encabezado
    hoja_resumen["A5"] = "Cantidad de Productos Distintos"
    hoja_resumen["B5"] = f"=COUNTA(Listado!A2:A{filas_listado})"
    hoja_resumen["A6"] = "Cantidad Total de Unidades en Stock"
    hoja_resumen["B6"] = f"=SUM(Listado!G2:G{filas_listado})"
    hoja_resumen["A7"] = "Valor del Inventario (a Precio de Compra)"
    hoja_resumen["B7"] = f"=SUMPRODUCT(Listado!D2:D{filas_listado},Listado!G2:G{filas_listado})"
    hoja_resumen["A8"] = "Valor Comprometido (a Precio de Compra)"
    hoja_resumen["B8"] = f"=SUMPRODUCT(Listado!D2:D{filas_listado},Listado!H2:H{filas_listado})"
    hoja_resumen["A9"] = "Productos Bajo Stock Mínimo"
    hoja_resumen["B9"] = len(bajo_stock)

    for fila in range(5, 10):
        hoja_resumen[f"A{fila}"].font = Font(name=FUENTE, bold=True, size=10)
        hoja_resumen[f"B{fila}"].font = Font(name=FUENTE, size=10)
        hoja_resumen[f"B{fila}"].alignment = Alignment(horizontal="right")
    for fila in (7, 8):
        hoja_resumen[f"B{fila}"].number_format = '"Gs. "#,##0'

    _autoancho_columnas(hoja_resumen, {"A": 38, "B": 22})

    # ---------------- HOJA 2: PRODUCTOS BAJO STOCK MÍNIMO ----------------
    hoja_bajo_stock = wb.create_sheet("Bajo Stock Mínimo")
    encabezados_bajo_stock = ["Código", "Descripción", "Stock Actual", "Stock Mínimo", "Faltante"]
    for col, texto in enumerate(encabezados_bajo_stock, start=1):
        celda = hoja_bajo_stock.cell(row=1, column=col, value=texto)
        _aplicar_estilo_encabezado(celda)

    for i, p in enumerate(bajo_stock, start=2):
        hoja_bajo_stock.cell(row=i, column=1, value=p["id"])
        hoja_bajo_stock.cell(row=i, column=2, value=p["nombre"])
        hoja_bajo_stock.cell(row=i, column=3, value=p["stock"])
        hoja_bajo_stock.cell(row=i, column=4, value=p["stock_minimo"])
        hoja_bajo_stock.cell(row=i, column=5, value=f"=D{i}-C{i}")
        for col in range(1, 6):
            hoja_bajo_stock.cell(row=i, column=col).fill = PatternFill("solid", start_color=ROJO_BAJO_STOCK)
            hoja_bajo_stock.cell(row=i, column=col).font = Font(name=FUENTE, size=9)

    if not bajo_stock:
        hoja_bajo_stock.cell(row=2, column=1, value="No hay productos bajo el stock mínimo.")
        hoja_bajo_stock.cell(row=2, column=1).font = Font(name=FUENTE, italic=True, size=9, color="777777")

    _autoancho_columnas(hoja_bajo_stock, {"A": 10, "B": 32, "C": 14, "D": 14, "E": 12})

    # ---------------- HOJA 3: TOP PRODUCTOS POR VALOR ----------------
    hoja_top = wb.create_sheet("Top Valor Inventario")
    encabezados_top = ["#", "Producto", "Stock", "Precio Compra", "Valor en Inventario"]
    for col, texto in enumerate(encabezados_top, start=1):
        celda = hoja_top.cell(row=1, column=col, value=texto)
        _aplicar_estilo_encabezado(celda)

    for i, p in enumerate(top_valor, start=2):
        hoja_top.cell(row=i, column=1, value=i - 1)
        hoja_top.cell(row=i, column=2, value=p["nombre"])
        hoja_top.cell(row=i, column=3, value=p["stock"])
        hoja_top.cell(row=i, column=4, value=p["precio_compra"]).number_format = '"Gs. "#,##0'
        celda_valor = hoja_top.cell(row=i, column=5, value=f"=C{i}*D{i}")
        celda_valor.number_format = '"Gs. "#,##0'
        for col in range(1, 6):
            hoja_top.cell(row=i, column=col).font = Font(name=FUENTE, size=9)
            if i % 2 == 0:
                hoja_top.cell(row=i, column=col).fill = PatternFill("solid", start_color=GRIS_CLARO)

    _autoancho_columnas(hoja_top, {"A": 5, "B": 32, "C": 12, "D": 16, "E": 18})

    # ---------------- HOJA 4: LISTADO COMPLETO (usada por las fórmulas de Resumen) ----------------
    hoja_listado = wb.create_sheet("Listado")
    encabezados_listado = [
        "Código", "Descripción", "Marca", "Precio Compra", "Precio Venta",
        "Precio Mayorista", "Stock", "Comprometido", "Disponible", "Stock Mínimo",
    ]
    for col, texto in enumerate(encabezados_listado, start=1):
        celda = hoja_listado.cell(row=1, column=col, value=texto)
        _aplicar_estilo_encabezado(celda)

    for i, p in enumerate(productos, start=2):
        disponible = p["disponible"] if isinstance(p["disponible"], (int, float)) else p["stock"]
        valores_fila = [
            p["id"], p["nombre"], p["marca"], p["precio_compra"], p["precio_venta"],
            p["precio_mayorista"], p["stock"], p["comprometido"], disponible, p["stock_minimo"],
        ]
        for col, valor in enumerate(valores_fila, start=1):
            celda = hoja_listado.cell(row=i, column=col, value=valor)
            celda.font = Font(name=FUENTE, size=9)
            if col in (4, 5, 6):
                celda.number_format = '"Gs. "#,##0'
            if i % 2 == 0:
                celda.fill = PatternFill("solid", start_color=GRIS_CLARO)
            es_bajo_stock = p["stock"] <= p["stock_minimo"] and p["tipo_producto"] != "Servicio"
            if es_bajo_stock:
                celda.fill = PatternFill("solid", start_color=ROJO_BAJO_STOCK)

    _autoancho_columnas(hoja_listado, {
        "A": 10, "B": 32, "C": 16, "D": 14, "E": 14, "F": 16, "G": 10, "H": 14, "I": 12, "J": 14,
    })
    hoja_listado.freeze_panes = "A2"

    wb.save(ruta_destino)
    return ruta_destino