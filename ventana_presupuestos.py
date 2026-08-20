"""
ventana_presupuestos.py
Módulo Presupuestos: armar cotizaciones para clientes (existentes o
walk-in), con seguimiento de estado (Pendiente → Aprobado/Rechazado →
Convertido en venta), impresión en PDF para entregar al cliente, y
conversión directa a una venta real cuando el cliente confirma.
"""
import datetime
import tkinter as tk
from tkinter import ttk, messagebox

from models_presupuestos import (
    ESTADOS, listar_presupuestos, obtener_presupuesto, crear_presupuesto,
    actualizar_presupuesto, cambiar_estado_presupuesto, marcar_convertido,
    eliminar_presupuesto,
)
from models_catalogo import buscar_producto_por_codigo
from utilidades_ui import (
    habilitar_deseleccion_treeview,
    ajustar_tamaño_ventana, formatear_gs, formatear_cantidad,
    unidad_es_fraccionable, parsear_cantidad,
)
from traducciones import t
from widget_calendario import abrir_selector_fecha

AZUL_RIBBON = "#1d5fd6"
GRIS_FONDO = "#f4f5f7"
GRIS_BORDE = "#e2e8f0"
VERDE = "#16a34a"
ROJO = "#dc2626"
NARANJA = "#d97706"
GRIS_TEXTO = "#6b7280"

COLOR_ESTADO = {
    "Pendiente": "#1d5fd6", "Aprobado": "#16a34a", "Rechazado": "#dc2626",
    "Vencido": "#6b7280", "Convertido": "#7c3aed",
}


def _formatear_fecha(fecha: str) -> str:
    if not fecha:
        return "—"
    try:
        return datetime.date.fromisoformat(fecha[:10]).strftime("%d/%m/%Y")
    except ValueError:
        return fecha


def _formatear_fecha_hora(fecha: str) -> str:
    if not fecha:
        return "—"
    try:
        return datetime.datetime.fromisoformat(fecha).strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return fecha


# ============================================================
# PANEL PRINCIPAL
# ============================================================
class PanelPresupuestos(tk.Frame):
    def __init__(self, parent, usuario_actual):
        super().__init__(parent, bg="white")
        self.usuario_actual = usuario_actual
        self.presupuestos_actuales: list[dict] = []

        self._construir_barra_superior()
        self._construir_tabla()
        self._cargar()

    def _construir_barra_superior(self):
        barra = tk.Frame(self, bg="white")
        barra.pack(fill="x", padx=10, pady=(10, 5))

        tk.Button(barra, text=t("presup_nuevo"), font=("Segoe UI", 9, "bold"),
                  bg=AZUL_RIBBON, fg="white", relief="flat", padx=12, pady=6,
                  cursor="hand2", command=self._abrir_nuevo).pack(side="left")

        from menu_reporte_general import BotonReporteGeneral
        self.boton_reporte = BotonReporteGeneral(
            barra, obtener_datos_callback=self._obtener_datos_reporte_lista,
            nombre_archivo_base="Reporte_General_Presupuestos",
        )
        self.boton_reporte.generador_pdf_dashboard = self._generar_pdf_dashboard_lista
        self.boton_reporte.generador_excel = self._generar_excel_lista
        self.boton_reporte.pack(side="left", padx=(8, 0))

        self.btn_generar_pdf = tk.Button(
            barra, text=t("presup_generar_pdf"), font=("Segoe UI", 9, "bold"),
            bg="#f3f4f6", fg="#9ca3af", relief="flat", padx=12, pady=6,
            cursor="hand2", command=self._generar_pdf_seleccionado,
            activebackground="#b91c1c", activeforeground="white",
        )
        self.btn_generar_pdf.pack(side="left", padx=(8, 0))

        tk.Label(barra, text=t("estado_label"), font=("Segoe UI", 9, "bold"), bg="white").pack(
            side="left", padx=(16, 6))
        self.var_estado = tk.StringVar(value="todos")
        combo = ttk.Combobox(barra, textvariable=self.var_estado, state="readonly", width=14,
                             values=["todos"] + ESTADOS + ["Vencido"], font=("Segoe UI", 9))
        combo.pack(side="left")
        combo.bind("<<ComboboxSelected>>", lambda e: self.after(50, self._cargar))

        tk.Label(barra, text=t("buscar") + ":", font=("Segoe UI", 9), bg="white").pack(side="right", padx=(0, 5))
        self.var_busqueda = tk.StringVar()
        entry = tk.Entry(barra, textvariable=self.var_busqueda, font=("Segoe UI", 9), width=25)
        entry.pack(side="right")
        entry.bind("<KeyRelease>", lambda e: self._cargar())

        tk.Button(barra, text="🔄", font=("Segoe UI", 10), bg="white", relief="flat",
                  cursor="hand2", command=self._cargar).pack(side="right", padx=(0, 10))

    def _construir_tabla(self):
        contenedor = tk.Frame(self, bg="white")
        contenedor.pack(fill="both", expand=True, padx=10, pady=(5, 10))
        contenedor.grid_rowconfigure(0, weight=1)
        contenedor.grid_columnconfigure(0, weight=1)

        columnas = ("codigo", "fecha", "cliente", "validez", "estado", "total", "vendedor")
        encabezados = (t("col_codigo_mayus"), t("col_fecha_mayus"), t("col_cliente_mayus"), t("presup_col_valido_hasta"), t("col_estado_mayus"), t("col_total_mayus"), t("col_vendedor_mayus"))
        anchos = (70, 130, 220, 110, 110, 120, 160)

        self.tabla = ttk.Treeview(contenedor, columns=columnas, show="headings", selectmode="browse")
        habilitar_deseleccion_treeview(self.tabla)
        for col, enc, ancho in zip(columnas, encabezados, anchos):
            self.tabla.heading(col, text=enc)
            self.tabla.column(col, width=ancho,
                              anchor="w" if col in ("cliente", "vendedor") else "center")

        for estado, color in COLOR_ESTADO.items():
            self.tabla.tag_configure(estado, foreground=color)

        sb_y = ttk.Scrollbar(contenedor, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=sb_y.set)
        self.tabla.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_y_h = ttk.Scrollbar(contenedor, orient="horizontal", command=self.tabla.xview)
        self.tabla.configure(xscrollcommand=sb_y_h.set)
        sb_y_h.grid(row=1, column=0, sticky="ew")

        self.tabla.bind("<Double-1>", lambda e: self._abrir_detalle_seleccionado())
        self.tabla.bind("<Button-3>", self._abrir_menu_contextual_fila)
        self.tabla.bind("<<TreeviewSelect>>", self._al_seleccionar_fila)

        self.menu_contextual_fila = tk.Menu(self, tearoff=0)
        self.menu_contextual_fila.add_command(label="👁 Ver Detalle", command=self._abrir_detalle_seleccionado)
        self.menu_contextual_fila.add_command(label="📄 Generar PDF para Cliente",
                                              command=self._generar_pdf_seleccionado)
        self.menu_contextual_fila.add_separator()
        self.menu_contextual_fila.add_command(label="📊 Reporte General de este Presupuesto",
                                              command=self._abrir_menu_reporte_individual)

    def _al_seleccionar_fila(self, event=None):
        hay_seleccion = bool(self.tabla.selection())
        self.btn_generar_pdf.config(
            bg="#dc2626" if hay_seleccion else "#f3f4f6",
            fg="white" if hay_seleccion else "#9ca3af",
        )

    def _generar_pdf_seleccionado(self):
        seleccion = self.tabla.selection()
        if not seleccion:
            messagebox.showinfo("Seleccioná un presupuesto",
                                "Primero hacé clic sobre un presupuesto de la lista "
                                "y después presioná \"Generar PDF\".", parent=self)
            return
        presupuesto_id = int(seleccion[0])
        det = obtener_presupuesto(presupuesto_id)
        if det is None:
            return

        from tkinter import filedialog
        try:
            from documento_presupuesto_pdf import generar_documento_presupuesto_pdf
        except ImportError:
            messagebox.showerror("Falta una librería",
                                 "Para generar el PDF se necesita instalar 'reportlab'.\n\n"
                                 "Abre una terminal y ejecutá:\n\npip install reportlab",
                                 parent=self)
            return
        ruta = filedialog.asksaveasfilename(
            title="Guardar Presupuesto", initialfile=f"presupuesto_{presupuesto_id}.pdf",
            defaultextension=".pdf", filetypes=[("Archivo PDF", "*.pdf")],
        )
        if not ruta:
            return
        try:
            generar_documento_presupuesto_pdf(ruta, det)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el PDF:\n{e}", parent=self)
            return
        if messagebox.askyesno("PDF generado", f"El presupuesto se guardó en:\n{ruta}\n\n"
                                               "¿Querés abrirlo ahora?", parent=self):
            import os, sys, subprocess
            try:
                if sys.platform == "win32":
                    os.startfile(ruta)
                elif sys.platform == "darwin":
                    subprocess.run(["open", ruta])
                else:
                    subprocess.run(["xdg-open", ruta])
            except Exception:
                pass

    def _abrir_menu_contextual_fila(self, event):
        fila = self.tabla.identify_row(event.y)
        if not fila:
            return
        self.tabla.selection_set(fila)
        try:
            self.menu_contextual_fila.tk_popup(event.x_root, event.y_root)
        finally:
            self.menu_contextual_fila.grab_release()

    def _cargar(self):
        for fila in self.tabla.get_children():
            self.tabla.delete(fila)
        estado = self.var_estado.get()
        self.presupuestos_actuales = listar_presupuestos(estado_filtro=estado,
                                                          busqueda=self.var_busqueda.get())
        for p in self.presupuestos_actuales:
            self.tabla.insert("", "end", iid=str(p["id"]), values=(
                p["id"], _formatear_fecha_hora(p["fecha"]), p["cliente"],
                _formatear_fecha(p["fecha_validez"]), p["estado_efectivo"],
                formatear_gs(p["total"]), p["vendedor"] or "—",
            ), tags=(p["estado_efectivo"],))

    def _abrir_detalle_seleccionado(self):
        seleccion = self.tabla.selection()
        if not seleccion:
            return
        VentanaDetallePresupuesto(self, int(seleccion[0]), self.usuario_actual, on_cambio=self._cargar)

    def _abrir_nuevo(self):
        VentanaEditarPresupuesto(self, self.usuario_actual, presupuesto_id=None,
                                 on_guardado=self._cargar)

    # ---------------- REPORTE GENERAL: LISTA COMPLETA ----------------
    def _obtener_datos_reporte_lista(self) -> dict:
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""
        cantidad = len(self.presupuestos_actuales)
        total = sum(p["total"] for p in self.presupuestos_actuales)
        convertidos = [p for p in self.presupuestos_actuales if p["estado_efectivo"] == "Convertido"]
        tasa = (len(convertidos) / cantidad * 100) if cantidad else 0

        filtros = []
        if self.var_estado.get() != "todos":
            filtros.append(f"Estado: {self.var_estado.get()}")
        if self.var_busqueda.get():
            filtros.append(f"Búsqueda: \"{self.var_busqueda.get()}\"")
        subtitulo = "  •  ".join(filtros) if filtros else "Todos los presupuestos registrados"

        return {
            "titulo": "Reporte General de Presupuestos",
            "subtitulo": subtitulo,
            "generado_por": nombre_usuario,
            "secciones": [
                {
                    "tipo": "resumen", "titulo": "RESUMEN GENERAL",
                    "filas": [
                        ("Cantidad de Presupuestos", str(cantidad)),
                        ("Total Cotizado", formatear_gs(total)),
                        ("Convertidos en Venta", str(len(convertidos))),
                        ("Tasa de Conversión", f"{tasa:.1f}%"),
                    ],
                },
                {
                    "tipo": "tabla", "titulo": "DETALLE DE PRESUPUESTOS",
                    "encabezados": ["Código", "Fecha", "Cliente", "Válido Hasta", "Estado", "Total"],
                    "filas": [
                        [str(p["id"]), _formatear_fecha_hora(p["fecha"]), p["cliente"],
                         _formatear_fecha(p["fecha_validez"]), p["estado_efectivo"],
                         formatear_gs(p["total"])]
                        for p in self.presupuestos_actuales
                    ],
                },
            ],
        }

    def _generar_pdf_dashboard_lista(self, ruta: str):
        from reporte_presupuestos_pdf import generar_reporte_presupuestos_pdf
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""
        if self.presupuestos_actuales:
            fechas = [p["fecha"][:10] for p in self.presupuestos_actuales]
            fecha_desde, fecha_hasta = min(fechas), max(fechas)
        else:
            hoy = datetime.date.today().isoformat()
            fecha_desde = fecha_hasta = hoy
        generar_reporte_presupuestos_pdf(ruta, fecha_desde, fecha_hasta, generado_por=nombre_usuario)

    def _generar_excel_lista(self, ruta: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        cantidad = len(self.presupuestos_actuales)
        total = sum(p["total"] for p in self.presupuestos_actuales)
        convertidos = [p for p in self.presupuestos_actuales if p["estado_efectivo"] == "Convertido"]
        tasa = (len(convertidos) / cantidad * 100) if cantidad else 0
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen"
        ws.append(["Reporte General de Presupuestos"])
        ws.append([f"Generado por: {nombre_usuario}"])
        ws.append([])
        ws.append(["Indicador", "Valor"])
        for etiqueta, valor in [("Cantidad de Presupuestos", cantidad), ("Total Cotizado", total),
                                 ("Convertidos en Venta", len(convertidos)),
                                 ("Tasa de Conversión (%)", round(tasa, 1))]:
            ws.append([etiqueta, valor])
        for cell in ws[4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D5FD6")
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 20

        ws2 = wb.create_sheet("Detalle de Presupuestos")
        ws2.append(["Código", "Fecha", "Cliente", "Válido Hasta", "Estado", "Total"])
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D5FD6")
            cell.alignment = Alignment(horizontal="center")
        for p in self.presupuestos_actuales:
            ws2.append([p["id"], _formatear_fecha_hora(p["fecha"]), p["cliente"],
                       _formatear_fecha(p["fecha_validez"]), p["estado_efectivo"], p["total"]])
        for col in ws2.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(ruta)

    # ---------------- REPORTE GENERAL: UN SOLO REGISTRO ----------------
    def _abrir_menu_reporte_individual(self):
        seleccion = self.tabla.selection()
        if not seleccion:
            return
        presupuesto_id = int(seleccion[0])

        from menu_reporte_general import BotonReporteGeneral
        boton_temporal = BotonReporteGeneral(
            self, obtener_datos_callback=lambda: self._obtener_datos_reporte_individual(presupuesto_id),
            nombre_archivo_base=f"Presupuesto_{presupuesto_id}",
        )
        boton_temporal.generador_excel = lambda ruta: self._generar_excel_individual(ruta, presupuesto_id)

        x = self.tabla.winfo_rootx() + 40
        y = self.tabla.winfo_rooty() + 40
        boton_temporal.menu.tk_popup(x, y)

    def _obtener_datos_reporte_individual(self, presupuesto_id: int) -> dict:
        det = obtener_presupuesto(presupuesto_id)
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""
        if det is None:
            return {"titulo": f"Presupuesto Nro. {presupuesto_id}", "subtitulo": "",
                    "generado_por": nombre_usuario, "secciones": []}
        return {
            "titulo": f"Presupuesto Nro. {det['id']}",
            "subtitulo": f"Cliente: {det['cliente_nombre']}  —  Estado: {det['estado_efectivo']}",
            "generado_por": nombre_usuario,
            "secciones": [
                {
                    "tipo": "resumen", "titulo": "DATOS DEL PRESUPUESTO",
                    "filas": [
                        ("Fecha", _formatear_fecha_hora(det["fecha"])),
                        ("Válido hasta", _formatear_fecha(det["fecha_validez"])),
                        ("Cliente", det["cliente_nombre"]),
                        ("Estado", det["estado_efectivo"]),
                        ("Vendedor", det["vendedor"] or "—"),
                        ("Total", formatear_gs(det["total"])),
                    ],
                },
                {
                    "tipo": "tabla", "titulo": "ARTÍCULOS COTIZADOS",
                    "encabezados": ["Código", "Descripción", "Cantidad", "Precio Unit.", "Importe"],
                    "filas": [
                        [str(it["producto_id"]) if it["producto_id"] else "Libre", it["nombre"],
                         formatear_cantidad(it["cantidad"], it["unidad_medida"]),
                         formatear_gs(it["precio_unitario"]), formatear_gs(it["importe"])]
                        for it in det["items"]
                    ],
                },
            ],
        }

    def _generar_excel_individual(self, ruta: str, presupuesto_id: int):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        det = obtener_presupuesto(presupuesto_id)
        if det is None:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Presupuesto"
        ws.append([f"Presupuesto Nro. {det['id']}"])
        ws.append([f"Cliente: {det['cliente_nombre']}"])
        ws.append([f"Estado: {det['estado_efectivo']}"])
        ws.append([f"Fecha: {_formatear_fecha_hora(det['fecha'])}"])
        ws.append([])
        ws.append(["Código", "Descripción", "Cantidad", "Precio Unit.", "Importe"])
        for cell in ws[6]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D5FD6")
            cell.alignment = Alignment(horizontal="center")
        for it in det["items"]:
            ws.append([it["producto_id"] or "Libre", it["nombre"], it["cantidad"],
                      it["precio_unitario"], it["importe"]])
        ws.append([])
        ws.append(["", "", "", "TOTAL", det["total"]])
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        wb.save(ruta)


# ============================================================
# EDITOR (crear / modificar mientras está Pendiente)
# ============================================================
class VentanaEditarPresupuesto(tk.Toplevel):
    def __init__(self, parent, usuario_actual, presupuesto_id=None, on_guardado=None):
        super().__init__(parent)
        self.usuario_actual = usuario_actual
        self.presupuesto_id = presupuesto_id
        self.on_guardado = on_guardado
        self.items = []  # {"producto": {...}, "cantidad":, "precio_unitario":}
        self.cliente_seleccionado = None
        self.fecha_validez = datetime.date.today() + datetime.timedelta(days=15)

        if presupuesto_id is not None:
            det = obtener_presupuesto(presupuesto_id)
            if det is None:
                self.destroy()
                return
            self._cargar_desde_detalle(det)

        self.title(f"Presupuesto Nro. {presupuesto_id}" if presupuesto_id else "Nuevo Presupuesto")
        self.configure(bg="white")
        self.resizable(True, True)
        self.grab_set()
        # Sin self.transient(parent): en Windows, una ventana "transient"
        # se trata como diálogo y pierde los botones de minimizar/
        # maximizar sin importar resizable(); al no marcarla como
        # transient, se comporta como una ventana normal con los tres
        # botones (minimizar, maximizar, cerrar) funcionando.
        self.lift()
        self.focus_force()

        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._construir_titulo()
        self._construir_barra_codigo()
        self._construir_grilla()
        self._construir_pie()
        self._registrar_atajos()

        self.minsize(780, 580)
        ajustar_tamaño_ventana(self, ancho_min=780, alto_min=580)
        self._actualizar_vista()
        self.entry_codigo.focus()

    def _cargar_desde_detalle(self, det):
        self.fecha_validez = None
        if det["fecha_validez"]:
            try:
                self.fecha_validez = datetime.date.fromisoformat(det["fecha_validez"][:10])
            except ValueError:
                pass
        if self.fecha_validez is None:
            self.fecha_validez = datetime.date.today() + datetime.timedelta(days=15)

        if det["cliente_id"]:
            from models_clientes import obtener_cliente
            self.cliente_seleccionado = obtener_cliente(det["cliente_id"])
        self._cliente_nombre_libre = det["cliente_nombre"]
        self._cliente_documento_libre = det["cliente_documento"]
        self._cliente_direccion_libre = det["cliente_direccion"]
        self._cliente_telefono_libre = det["cliente_telefono"]
        self._observaciones_previas = det["observaciones"]

        for item in det["items"]:
            if item["es_libre"]:
                producto = {"id": None, "nombre": item["nombre"], "es_libre": True,
                            "unidad_medida": "Unidad"}
            else:
                producto = {"id": item["producto_id"], "nombre": item["nombre"],
                            "unidad_medida": item["unidad_medida"]}
            self.items.append({"producto": producto, "cantidad": item["cantidad"],
                               "precio_unitario": item["precio_unitario"]})

    def _construir_titulo(self):
        barra = tk.Frame(self, bg=AZUL_RIBBON, height=34)
        barra.grid(row=0, column=0, sticky="ew")
        barra.grid_propagate(False)
        titulo = f"Presupuesto Nro. {self.presupuesto_id}" if self.presupuesto_id else "Nuevo Presupuesto"
        tk.Label(barra, text=titulo, font=("Segoe UI", 11, "bold"),
                 bg=AZUL_RIBBON, fg="white").pack(side="left", padx=15, pady=6)

    def _construir_barra_codigo(self):
        contenedor = tk.Frame(self, bg="white")
        contenedor.grid(row=1, column=0, sticky="ew", padx=10, pady=(10, 5))

        barra = tk.Frame(contenedor, bg="white")
        barra.pack(fill="x")
        tk.Label(barra, text=t("preventa_codigo_producto"), font=("Segoe UI", 10, "bold"),
                 bg="white").pack(side="left", padx=(0, 8))
        self.var_codigo = tk.StringVar()
        self.entry_codigo = tk.Entry(barra, textvariable=self.var_codigo, font=("Segoe UI", 11), width=26)
        self.entry_codigo.pack(side="left", padx=(0, 8))
        self.entry_codigo.bind("<Return>", lambda e: self._agregar_por_codigo())
        tk.Button(barra, text=t("ventas_agregar_producto"), font=("Segoe UI", 9, "bold"),
                  bg=AZUL_RIBBON, fg="white", command=self._agregar_por_codigo).pack(side="left")

        atajos = tk.Frame(contenedor, bg="white")
        atajos.pack(fill="x", pady=(8, 0))
        tk.Button(atajos, text=t("atajo_f1_cliente"), font=("Segoe UI", 8, "bold"),
                  bg="#7c3aed", fg="white", relief="flat", padx=10, pady=6,
                  cursor="hand2", command=self._abrir_asignar_cliente).pack(side="left", padx=(0, 6))
        tk.Button(atajos, text=t("atajo_f2_buscar"), font=("Segoe UI", 8, "bold"),
                  bg="#0891b2", fg="white", relief="flat", padx=10, pady=6,
                  cursor="hand2", command=self._abrir_buscar_producto).pack(side="left", padx=(0, 6))
        tk.Button(atajos, text=t("atajo_del_borrar"), font=("Segoe UI", 8, "bold"),
                  bg="#6b7280", fg="white", relief="flat", padx=10, pady=6,
                  cursor="hand2", command=self._borrar_articulo_seleccionado).pack(side="left", padx=(0, 6))

    def _construir_grilla(self):
        contenedor = tk.Frame(self, bg="white")
        contenedor.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 5))
        contenedor.grid_rowconfigure(0, weight=1)
        contenedor.grid_columnconfigure(0, weight=1)

        columnas = ("codigo", "descripcion", "cantidad", "precio_unitario", "importe")
        encabezados = ("Código", "Descripción", "Cant.", "Precio Unit.", "Importe")

        self.tabla = ttk.Treeview(contenedor, columns=columnas, show="headings", selectmode="browse")
        habilitar_deseleccion_treeview(self.tabla)
        for col, enc in zip(columnas, encabezados):
            self.tabla.heading(col, text=enc)
            ancho = 280 if col == "descripcion" else 110
            self.tabla.column(col, width=ancho, anchor="w" if col == "descripcion" else "center")
        self.tabla.tag_configure("importe", background="#dcfce7")

        sb_y = ttk.Scrollbar(contenedor, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=sb_y.set)
        self.tabla.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_y_h = ttk.Scrollbar(contenedor, orient="horizontal", command=self.tabla.xview)
        self.tabla.configure(xscrollcommand=sb_y_h.set)
        sb_y_h.grid(row=1, column=0, sticky="ew")

        self.tabla.bind("<Double-1>", lambda e: self._editar_item_seleccionado())
        self.tabla.bind("<Button-3>", self._abrir_menu_contextual)

        self.menu_contextual = tk.Menu(self, tearoff=0)
        self.menu_contextual.add_command(label="✏ Editar Artículo", command=self._editar_item_seleccionado)
        self.menu_contextual.add_separator()
        self.menu_contextual.add_command(
            label="🏷 Aplicar Precio Minorista",
            command=lambda: self._aplicar_tipo_precio_rapido(mayorista=False))
        self.menu_contextual.add_command(
            label="🏷 Aplicar Precio Mayorista",
            command=lambda: self._aplicar_tipo_precio_rapido(mayorista=True))
        self.menu_contextual.add_separator()
        self.menu_contextual.add_command(label="🗑 Borrar Artículo", command=self._borrar_articulo_seleccionado)

    def _abrir_menu_contextual(self, event):
        fila = self.tabla.identify_row(event.y)
        if not fila:
            return
        self.tabla.selection_set(fila)
        try:
            self.menu_contextual.tk_popup(event.x_root, event.y_root)
        finally:
            self.menu_contextual.grab_release()

    def _aplicar_tipo_precio_rapido(self, mayorista: bool):
        """Desde el menú contextual: aplica directo el precio Mayorista o
        Minorista del producto a la línea seleccionada, sin abrir el
        diálogo de edición completo."""
        seleccion = self.tabla.selection()
        if not seleccion:
            return
        indice = int(seleccion[0])
        item = self.items[indice]
        p = item["producto"]
        if p.get("es_libre") or p.get("precio_venta") is None or p.get("precio_mayorista") is None:
            messagebox.showinfo("No disponible",
                                "Este artículo no tiene precio Mayorista/Minorista configurado.",
                                parent=self)
            return
        item["es_mayoreo"] = mayorista
        item["precio_unitario"] = p["precio_mayorista"] if mayorista else p["precio_venta"]
        self._actualizar_vista()

    def _construir_pie(self):
        pie = tk.Frame(self, bg=GRIS_FONDO)
        pie.grid(row=3, column=0, sticky="ew")

        fila_sup = tk.Frame(pie, bg=GRIS_FONDO)
        fila_sup.pack(fill="x", padx=10, pady=(8, 4))

        tk.Label(fila_sup, text=t("presup_valido_hasta"), font=("Segoe UI", 9, "bold"), bg=GRIS_FONDO).pack(
            side="left")
        self.lbl_fecha_validez = tk.Label(fila_sup, text=self.fecha_validez.strftime("%d/%m/%Y"),
                                          font=("Segoe UI", 9), bg="white", relief="solid", bd=1,
                                          width=12, cursor="hand2")
        self.lbl_fecha_validez.pack(side="left", padx=(6, 20))
        self.lbl_fecha_validez.bind("<Button-1>", lambda e: self._elegir_fecha_validez())

        tk.Label(fila_sup, text=t("observaciones_label"), font=("Segoe UI", 9, "bold"), bg=GRIS_FONDO).pack(
            side="left")
        self.var_observaciones = tk.StringVar(value=getattr(self, "_observaciones_previas", ""))
        tk.Entry(fila_sup, textvariable=self.var_observaciones, font=("Segoe UI", 9), width=40).pack(
            side="left", padx=(6, 0))

        fila_inf = tk.Frame(pie, bg=GRIS_FONDO)
        fila_inf.pack(fill="x", padx=10, pady=(0, 8))

        self.label_cliente = tk.Label(fila_inf, text=f"{t('cliente_label')} {t('ocasional')}",
                                      font=("Segoe UI", 9, "bold"), bg=GRIS_FONDO)
        self.label_cliente.pack(side="left")

        self.label_contador = tk.Label(fila_inf, text=t("presup_productos_contador").format(n=0), font=("Segoe UI", 9),
                                       bg=GRIS_FONDO, fg=GRIS_TEXTO)
        self.label_contador.pack(side="left", padx=(20, 0))

        botones = tk.Frame(fila_inf, bg=GRIS_FONDO)
        botones.pack(side="right")
        tk.Button(botones, text=t("presup_cancelar_x"), font=("Segoe UI", 10, "bold"), bg="white",
                  relief="solid", bd=1, command=self.destroy).pack(side="right")
        tk.Button(botones, text=t("presup_guardar_presupuesto"), font=("Segoe UI", 10, "bold"),
                  bg="white", relief="solid", bd=1, command=self._guardar).pack(side="right", padx=(0, 8))

        self.label_total = tk.Label(fila_inf, text="Gs. 0", font=("Segoe UI", 18, "bold"),
                                    bg=GRIS_FONDO, fg="#1d4ed8")
        self.label_total.pack(side="right", padx=20)

    def _registrar_atajos(self):
        self.bind("<F1>", lambda e: self._abrir_asignar_cliente())
        self.bind("<F2>", lambda e: self._abrir_buscar_producto())
        self.bind("<Delete>", lambda e: self._borrar_articulo_seleccionado())

    def _abrir_asignar_cliente(self):
        from ventanas_auxiliares_venta import VentanaAsignarCliente
        VentanaAsignarCliente(self, on_seleccionado=self._al_asignar_cliente)

    def _al_asignar_cliente(self, cliente):
        self.cliente_seleccionado = cliente
        self._actualizar_vista()

    def _elegir_fecha_validez(self):
        def al_elegir(fecha):
            self.fecha_validez = fecha
            self.lbl_fecha_validez.config(text=fecha.strftime("%d/%m/%Y"))
        abrir_selector_fecha(self, self.fecha_validez, al_elegir)

    def _agregar_por_codigo(self):
        codigo = self.var_codigo.get().strip()
        if not codigo:
            return
        producto = buscar_producto_por_codigo(codigo)
        if producto is None or not producto["activo"]:
            messagebox.showerror("No encontrado", f"No existe un producto activo con el código '{codigo}'.",
                                 parent=self)
            self.var_codigo.set("")
            return
        self.var_codigo.set("")
        self._agregar_item(producto, 1)

    def _abrir_buscar_producto(self):
        from ventanas_auxiliares_venta import VentanaBuscarProducto
        VentanaBuscarProducto(self, on_seleccionado=lambda p: self._agregar_item(p, 1))

    def _agregar_item(self, producto, cantidad):
        for item in self.items:
            if item["producto"].get("id") == producto.get("id") and producto.get("id") is not None:
                item["cantidad"] += cantidad
                self._actualizar_vista()
                return
        precio = producto.get("precio_venta", producto.get("precio_unitario", 0))
        self.items.append({"producto": producto, "cantidad": cantidad, "precio_unitario": precio})
        self._actualizar_vista()
        self.entry_codigo.focus()

    def _borrar_articulo_seleccionado(self):
        seleccion = self.tabla.selection()
        if not seleccion:
            return
        indice = int(seleccion[0])
        del self.items[indice]
        self._actualizar_vista()

    def _editar_item_seleccionado(self):
        seleccion = self.tabla.selection()
        if not seleccion:
            return
        indice = int(seleccion[0])
        item = self.items[indice]
        VentanaEditarItemPresupuesto(self, item, on_confirmado=self._actualizar_vista)

    def _actualizar_vista(self):
        for fila in self.tabla.get_children():
            self.tabla.delete(fila)

        total = 0
        for i, item in enumerate(self.items):
            p = item["producto"]
            importe = item["cantidad"] * item["precio_unitario"]
            total += importe
            unidad = p.get("unidad_medida", "Unidad")
            cantidad_txt = formatear_cantidad(item["cantidad"], unidad)
            codigo_txt = p.get("id") if p.get("id") else "Libre"
            nombre_txt = f"{p['nombre']}  🏷 Mayoreo" if item.get("es_mayoreo") else p["nombre"]
            self.tabla.insert("", "end", iid=str(i), values=(
                codigo_txt, nombre_txt, cantidad_txt,
                formatear_gs(item["precio_unitario"]), formatear_gs(importe),
            ), tags=("importe",))

        self.label_contador.config(text=t("presup_productos_contador").format(n=len(self.items)))
        self.label_total.config(text=formatear_gs(total))

        if self.cliente_seleccionado:
            self.label_cliente.config(text=f"Cliente: {self.cliente_seleccionado['nombre']}")
        elif getattr(self, "_cliente_nombre_libre", ""):
            self.label_cliente.config(text=f"Cliente: {self._cliente_nombre_libre}")
        else:
            self.label_cliente.config(text=f"{t('cliente_label')} {t('ocasional')}")

    def _items_para_guardar(self):
        items = []
        for item in self.items:
            p = item["producto"]
            if p.get("es_libre") or p.get("id") is None:
                items.append({"producto_id": None, "descripcion_libre": p["nombre"],
                             "cantidad": item["cantidad"], "precio_unitario": item["precio_unitario"]})
            else:
                items.append({"producto_id": p["id"], "cantidad": item["cantidad"],
                             "precio_unitario": item["precio_unitario"]})
        return items

    def _datos_cliente(self):
        if self.cliente_seleccionado:
            return (self.cliente_seleccionado["id"], self.cliente_seleccionado["nombre"],
                    self.cliente_seleccionado.get("nro_documento", ""),
                    self.cliente_seleccionado.get("direccion", ""),
                    self.cliente_seleccionado.get("telefono", ""))
        return (None, getattr(self, "_cliente_nombre_libre", "") or "Ocasional",
                getattr(self, "_cliente_documento_libre", ""),
                getattr(self, "_cliente_direccion_libre", ""),
                getattr(self, "_cliente_telefono_libre", ""))

    def _guardar(self):
        if not self.items:
            messagebox.showwarning("Presupuesto vacío", "Agrega al menos un producto antes de guardar.",
                                   parent=self)
            return

        cliente_id, cliente_nombre, cliente_doc, cliente_dir, cliente_tel = self._datos_cliente()
        fecha_validez_iso = self.fecha_validez.isoformat() if self.fecha_validez else ""

        if self.presupuesto_id is None:
            ok, msg, presupuesto_id = crear_presupuesto(
                items=self._items_para_guardar(), usuario_id=self.usuario_actual.get("id"),
                cliente_id=cliente_id, cliente_nombre=cliente_nombre, cliente_documento=cliente_doc,
                cliente_direccion=cliente_dir, cliente_telefono=cliente_tel,
                fecha_validez=fecha_validez_iso, observaciones=self.var_observaciones.get(),
            )
        else:
            ok, msg = actualizar_presupuesto(
                self.presupuesto_id, self._items_para_guardar(), cliente_id, cliente_nombre,
                cliente_doc, cliente_dir, cliente_tel, fecha_validez_iso,
                self.var_observaciones.get(),
            )

        if not ok:
            messagebox.showerror("No se pudo guardar", msg, parent=self)
            return
        messagebox.showinfo("Presupuesto guardado", msg, parent=self)
        if self.on_guardado:
            self.on_guardado()
        self.destroy()


# ============================================================
# EDITAR CANTIDAD / PRECIO DE UN ITEM YA AGREGADO
# ============================================================
class VentanaEditarItemPresupuesto(tk.Toplevel):
    def __init__(self, parent, item, on_confirmado):
        super().__init__(parent)
        self.item = item
        self.on_confirmado = on_confirmado
        producto = item["producto"]

        self.title("Editar Artículo")
        self.configure(bg="white")
        self.resizable(True, True)
        self.grab_set()
        # Sin self.transient(parent): en Windows las ventanas "transient"
        # se tratan como diálogos y pierden minimizar/maximizar sin
        # importar resizable(); al no marcarla como transient, aparecen
        # los tres botones normales.
        self.lift()
        self.focus_force()

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        contenedor = tk.Frame(self, bg="white")
        contenedor.grid(row=0, column=0, sticky="nsew")
        contenedor.grid_rowconfigure(1, weight=1)
        contenedor.grid_columnconfigure(0, weight=1)

        tk.Label(contenedor, text=producto["nombre"], font=("Segoe UI", 10, "bold"),
                 bg="white", wraplength=280).grid(row=0, column=0, padx=16, pady=(16, 10))

        cuerpo = tk.Frame(contenedor, bg="white")
        cuerpo.grid(row=1, column=0, padx=16, sticky="n")
        cuerpo.grid_columnconfigure(1, weight=1)

        unidad = producto.get("unidad_medida", "Unidad")
        texto_cant = self._texto_editable(item["cantidad"], unidad)

        # ── Tipo de precio: Minorista / Mayorista, solo si el producto
        # tiene ambos precios cargados (no aplica a artículos libres) ──
        self.var_tipo_precio = tk.StringVar(value="mayorista" if item.get("es_mayoreo") else "minorista")
        precio_venta = producto.get("precio_venta")
        precio_mayorista = producto.get("precio_mayorista")
        fila = 0
        if not producto.get("es_libre") and precio_venta is not None and precio_mayorista is not None:
            tk.Label(cuerpo, text=t("presup_tipo_precio"), font=("Segoe UI", 9, "bold"), bg="white").grid(
                row=fila, column=0, sticky="ne", pady=6, padx=(0, 8))
            frame_tipo = tk.Frame(cuerpo, bg="white")
            frame_tipo.grid(row=fila, column=1, sticky="w", pady=6)
            tk.Radiobutton(frame_tipo, text=f"Minorista (Gs. {precio_venta:,.0f})",
                          variable=self.var_tipo_precio, value="minorista", bg="white",
                          font=("Segoe UI", 9),
                          command=lambda: self._aplicar_tipo_precio(precio_venta, precio_mayorista)
                          ).pack(anchor="w")
            tk.Radiobutton(frame_tipo, text=f"Mayorista (Gs. {precio_mayorista:,.0f})",
                          variable=self.var_tipo_precio, value="mayorista", bg="white",
                          font=("Segoe UI", 9),
                          command=lambda: self._aplicar_tipo_precio(precio_venta, precio_mayorista)
                          ).pack(anchor="w")
            fila += 1

        tk.Label(cuerpo, text=t("cantidad_label"), font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=fila, column=0, sticky="e", pady=6, padx=(0, 8))
        self.var_cantidad = tk.StringVar(value=texto_cant)
        entry_cant = tk.Entry(cuerpo, textvariable=self.var_cantidad, font=("Segoe UI", 10), width=14)
        entry_cant.grid(row=fila, column=1, sticky="ew", pady=6)
        fila += 1

        tk.Label(cuerpo, text=t("precio_unitario_label"), font=("Segoe UI", 9, "bold"), bg="white").grid(
            row=fila, column=0, sticky="e", pady=6, padx=(0, 8))
        self.var_precio = tk.StringVar(value=str(int(round(item["precio_unitario"]))))
        entry_precio = tk.Entry(cuerpo, textvariable=self.var_precio, font=("Segoe UI", 10), width=14)
        entry_precio.grid(row=fila, column=1, sticky="ew", pady=6)

        botones = tk.Frame(contenedor, bg="white")
        botones.grid(row=2, column=0, pady=16)
        tk.Button(botones, text=t("aceptar_boton"), font=("Segoe UI", 9, "bold"), bg=AZUL_RIBBON, fg="white",
                  relief="flat", padx=14, pady=6, cursor="hand2", command=self._confirmar).pack(
                      side="left", padx=6)
        tk.Button(botones, text=t("cancelar_x"), font=("Segoe UI", 9, "bold"), bg="#f3f4f6",
                  relief="flat", padx=14, pady=6, cursor="hand2", command=self.destroy).pack(
                      side="left", padx=6)

        entry_cant.focus()
        entry_cant.selection_range(0, "end")
        self.bind("<Return>", lambda e: self._confirmar())
        self.bind("<Escape>", lambda e: self.destroy())
        self.minsize(360, 260)
        ajustar_tamaño_ventana(self, ancho_min=360, alto_min=260)

    def _aplicar_tipo_precio(self, precio_venta, precio_mayorista):
        nuevo_precio = precio_mayorista if self.var_tipo_precio.get() == "mayorista" else precio_venta
        self.var_precio.set(str(int(round(nuevo_precio))))

    def _texto_editable(self, valor, unidad):
        if unidad_es_fraccionable(unidad):
            s = f"{valor:.10f}".rstrip("0").rstrip(".")
            return s.replace(".", ",")
        return str(int(round(valor)))

    def _confirmar(self):
        try:
            cantidad = parsear_cantidad(self.var_cantidad.get())
            if cantidad <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Cantidad inválida", "Ingresa una cantidad numérica mayor a cero.",
                                 parent=self)
            return
        try:
            precio = parsear_cantidad(self.var_precio.get())
            if precio < 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Precio inválido", "Ingresa un precio unitario numérico válido.",
                                 parent=self)
            return

        self.item["cantidad"] = cantidad
        self.item["precio_unitario"] = precio
        self.item["es_mayoreo"] = (self.var_tipo_precio.get() == "mayorista")
        self.destroy()
        self.on_confirmado()


# ============================================================
# DETALLE / GESTIÓN DE UN PRESUPUESTO
# ============================================================
class VentanaDetallePresupuesto(tk.Toplevel):
    def __init__(self, parent, presupuesto_id: int, usuario_actual, on_cambio=None):
        super().__init__(parent)
        self.presupuesto_id = presupuesto_id
        self.usuario_actual = usuario_actual
        self.on_cambio = on_cambio

        self.presupuesto = obtener_presupuesto(presupuesto_id)
        if self.presupuesto is None:
            self.destroy()
            return

        self.title(f"Presupuesto Nro. {presupuesto_id}")
        self.configure(bg="white")
        self.resizable(True, True)
        self.grab_set()
        self.transient(parent)

        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._construir_titulo()
        self._construir_info()
        self._construir_items()
        self._construir_pie()

        self.minsize(620, 560)
        ajustar_tamaño_ventana(self, ancho_min=620, alto_min=560)

    def _construir_titulo(self):
        p = self.presupuesto
        barra = tk.Frame(self, bg=AZUL_RIBBON, height=34)
        barra.grid(row=0, column=0, sticky="ew")
        barra.grid_propagate(False)
        tk.Label(barra, text=f"Presupuesto Nro. {self.presupuesto_id}", font=("Segoe UI", 11, "bold"),
                 bg=AZUL_RIBBON, fg="white").pack(side="left", padx=15, pady=6)
        color = COLOR_ESTADO.get(p["estado_efectivo"], "#6b7280")
        tk.Label(barra, text=f"  {p['estado_efectivo']}  ", font=("Segoe UI", 9, "bold"),
                bg=color, fg="white").pack(side="right", padx=15, pady=6)

    def _construir_info(self):
        p = self.presupuesto
        contenedor = tk.Frame(self, bg="white")
        contenedor.grid(row=1, column=0, sticky="ew", padx=16, pady=(12, 6))
        contenedor.grid_columnconfigure(1, weight=1)

        def fila(r, etiqueta, valor):
            tk.Label(contenedor, text=etiqueta, font=("Segoe UI", 9, "bold"), bg="white").grid(
                row=r, column=0, sticky="e", pady=3, padx=(0, 10))
            tk.Label(contenedor, text=valor, font=("Segoe UI", 9), bg="white",
                    wraplength=340, justify="left", anchor="w").grid(row=r, column=1, sticky="w", pady=3)

        fila(0, "Cliente:", p["cliente_nombre"])
        fila(1, "Fecha:", _formatear_fecha_hora(p["fecha"]))
        fila(2, "Válido hasta:", _formatear_fecha(p["fecha_validez"]))
        fila(3, "Vendedor:", p["vendedor"] or "—")
        if p["observaciones"]:
            fila(4, "Observaciones:", p["observaciones"])

    def _construir_items(self):
        contenedor = tk.Frame(self, bg="white")
        contenedor.grid(row=2, column=0, sticky="nsew", padx=16)
        contenedor.grid_rowconfigure(0, weight=1)
        contenedor.grid_columnconfigure(0, weight=1)

        cols = ("codigo", "descripcion", "cantidad", "precio", "importe")
        encs = ("Código", "Descripción", "Cant.", "Precio Unit.", "Importe")
        tabla = ttk.Treeview(contenedor, columns=cols, show="headings")
        habilitar_deseleccion_treeview(tabla)
        for col, enc in zip(cols, encs):
            tabla.heading(col, text=enc)
            ancho = 260 if col == "descripcion" else 100
            tabla.column(col, width=ancho, anchor="w" if col == "descripcion" else "center")
        tabla.grid(row=0, column=0, sticky="nsew")
        sb = ttk.Scrollbar(contenedor, orient="vertical", command=tabla.yview)
        tabla.configure(yscrollcommand=sb.set)
        sb.grid(row=0, column=1, sticky="ns")
        sb_h = ttk.Scrollbar(contenedor, orient="horizontal", command=tabla.xview)
        tabla.configure(xscrollcommand=sb_h.set)
        sb_h.grid(row=1, column=0, sticky="ew")

        for item in self.presupuesto["items"]:
            codigo = item["producto_id"] if item["producto_id"] else "Libre"
            cantidad_txt = formatear_cantidad(item["cantidad"], item["unidad_medida"])
            tabla.insert("", "end", values=(
                codigo, item["nombre"], cantidad_txt,
                formatear_gs(item["precio_unitario"]), formatear_gs(item["importe"]),
            ))

    def _construir_pie(self):
        p = self.presupuesto
        pie = tk.Frame(self, bg=GRIS_FONDO)
        pie.grid(row=3, column=0, sticky="ew")

        fila_total = tk.Frame(pie, bg=GRIS_FONDO)
        fila_total.pack(fill="x", padx=15, pady=(10, 4))
        tk.Label(fila_total, text=t("col_total_mayus3"), font=("Segoe UI", 10, "bold"), bg=GRIS_FONDO).pack(side="left")
        tk.Label(fila_total, text=formatear_gs(p["total"]), font=("Segoe UI", 16, "bold"),
                bg=GRIS_FONDO, fg="#1d4ed8").pack(side="right")

        botones = tk.Frame(pie, bg=GRIS_FONDO)
        botones.pack(fill="x", padx=15, pady=(4, 12))

        estado = p["estado_efectivo"]

        tk.Button(botones, text=t("presup_generar_pdf_boton"), font=("Segoe UI", 9, "bold"), bg="#dc2626", fg="white",
                 relief="flat", cursor="hand2", activebackground="#b91c1c", activeforeground="white",
                 command=self._generar_pdf).pack(side="left")

        if estado in ("Pendiente", "Vencido"):
            tk.Button(botones, text=t("presup_editar"), font=("Segoe UI", 9, "bold"), bg="white",
                     relief="solid", bd=1, cursor="hand2",
                     command=self._editar).pack(side="left", padx=(8, 0))
            tk.Button(botones, text=t("presup_aprobar"), font=("Segoe UI", 9, "bold"), bg=VERDE, fg="white",
                     relief="flat", cursor="hand2",
                     command=lambda: self._cambiar_estado("Aprobado")).pack(side="left", padx=(8, 0))
            tk.Button(botones, text=t("presup_rechazar"), font=("Segoe UI", 9, "bold"), bg=ROJO, fg="white",
                     relief="flat", cursor="hand2",
                     command=lambda: self._cambiar_estado("Rechazado")).pack(side="left", padx=(8, 0))
        elif estado == "Aprobado":
            tk.Button(botones, text=t("presup_convertir_venta"), font=("Segoe UI", 9, "bold"),
                     bg=AZUL_RIBBON, fg="white", relief="flat", cursor="hand2",
                     command=self._convertir_a_venta).pack(side="left", padx=(8, 0))

        if estado != "Convertido":
            tk.Button(botones, text=t("eliminar_boton"), font=("Segoe UI", 9, "bold"), bg="white",
                     fg=ROJO, relief="solid", bd=1, cursor="hand2",
                     command=self._eliminar).pack(side="left", padx=(8, 0))

        tk.Button(botones, text=t("cerrar"), font=("Segoe UI", 9, "bold"), bg="white",
                 relief="solid", bd=1, command=self.destroy).pack(side="right")

    def _generar_pdf(self):
        from tkinter import filedialog
        try:
            from documento_presupuesto_pdf import generar_documento_presupuesto_pdf
        except ImportError:
            messagebox.showerror("Falta una librería",
                                 "Para generar el PDF se necesita instalar 'reportlab'.\n\n"
                                 "Abre una terminal y ejecutá:\n\npip install reportlab",
                                 parent=self)
            return
        ruta = filedialog.asksaveasfilename(
            title="Guardar Presupuesto", initialfile=f"presupuesto_{self.presupuesto_id}.pdf",
            defaultextension=".pdf", filetypes=[("Archivo PDF", "*.pdf")],
        )
        if not ruta:
            return
        try:
            generar_documento_presupuesto_pdf(ruta, self.presupuesto)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el PDF:\n{e}", parent=self)
            return
        messagebox.showinfo("PDF generado", f"El presupuesto se guardó en:\n{ruta}", parent=self)

    def _editar(self):
        self.destroy()
        VentanaEditarPresupuesto(self.master, self.usuario_actual, presupuesto_id=self.presupuesto_id,
                                 on_guardado=self.on_cambio)

    def _cambiar_estado(self, nuevo_estado):
        ok, msg = cambiar_estado_presupuesto(self.presupuesto_id, nuevo_estado)
        if not ok:
            messagebox.showerror("No se pudo actualizar", msg, parent=self)
            return
        messagebox.showinfo("Estado actualizado", msg, parent=self)
        if self.on_cambio:
            self.on_cambio()
        self.destroy()

    def _convertir_a_venta(self):
        p = self.presupuesto
        items_venta = []
        for item in p["items"]:
            if item["es_libre"]:
                producto = {"id": None, "nombre": item["nombre"], "es_libre": True}
            else:
                from models_catalogo import buscar_producto_por_codigo
                producto_real = buscar_producto_por_codigo(str(item["producto_id"]))
                producto = producto_real or {"id": item["producto_id"], "nombre": item["nombre"]}
            items_venta.append({"producto": producto, "cantidad": item["cantidad"],
                               "precio_unitario": item["precio_unitario"]})

        cliente = None
        if p["cliente_id"]:
            from models_clientes import obtener_cliente
            cliente = obtener_cliente(p["cliente_id"])

        from ventana_cobrar import VentanaCobrar
        VentanaCobrar(
            self, items_venta=items_venta, cliente=cliente, usuario_actual=self.usuario_actual,
            condicion_inicial="contado", on_venta_procesada=self._al_finalizar_venta,
        )

    def _al_finalizar_venta(self):
        # Tomamos la venta recién creada (la más reciente) para vincularla.
        try:
            from database import conectar
            conn = conectar()
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM ventas ORDER BY id DESC LIMIT 1")
            fila = cursor.fetchone()
            conn.close()
            if fila:
                marcar_convertido(self.presupuesto_id, fila[0])
        except Exception:
            pass
        if self.on_cambio:
            self.on_cambio()
        self.destroy()

    def _eliminar(self):
        if not messagebox.askyesno("Eliminar Presupuesto",
                                   f"¿Eliminar el Presupuesto Nro. {self.presupuesto_id}?\n\n"
                                   "Esta acción no se puede deshacer.", parent=self):
            return
        ok, msg = eliminar_presupuesto(self.presupuesto_id)
        if not ok:
            messagebox.showerror("No se pudo eliminar", msg, parent=self)
            return
        if self.on_cambio:
            self.on_cambio()
        self.destroy()
