"""
ventana_datos.py
Módulo de Gestión de Datos: permite al administrador hacer copias de
seguridad de la base de datos (exportar), restaurarlas (importar) y
también exportar/importar datos en formato Excel (.xlsx).

Acciones disponibles:
  1. Exportar copia de seguridad  → copia el archivo .db a un destino elegido.
  2. Importar copia de seguridad  → reemplaza la BD actual con un .db externo
                                    (pide confirmación y hace backup previo).
  3. Exportar a Excel             → escribe las tablas de todos los módulos del
                                    negocio (ventas, préstamos, presupuestos,
                                    asistencia técnica, veterinaria, restaurante,
                                    streaming, RRHH, etc.) en hojas de un libro
                                    .xlsx descargable.
  4. Importar desde Excel         → lee filas de hojas con formato conocido y
                                    las inserta en la BD (productos y clientes).

Solo accesible para administradores.
"""
import tkinter as tk
from tkinter import messagebox, filedialog
import os
import shutil
import csv
import datetime

from database import conectar, obtener_ruta_bd
from models_rrhh import inicializar_tablas_rrhh
from utilidades_ui import habilitar_deseleccion_treeview
from traducciones import t

AZUL_RIBBON  = "#1d5fd6"
AZUL_OSCURO  = "#163d8c"
GRIS_FONDO   = "#f4f5f7"
GRIS_BORDE   = "#e2e8f0"
BLANCO       = "#ffffff"
VERDE        = "#16a34a"
ROJO         = "#dc2626"
NARANJA      = "#d97706"
GRIS_TEXTO   = "#6b7280"


# ─────────────────────────────────────────────────────────────
#  PANEL PRINCIPAL (embebido en frame_contenido de main.py)
# ─────────────────────────────────────────────────────────────
class PanelDatos(tk.Frame):
    def __init__(self, parent, usuario_actual, on_reiniciar_sesion=None):
        super().__init__(parent, bg=GRIS_FONDO)
        self.usuario_actual = usuario_actual
        # Callback opcional: si se provee, se llama tras una importación
        # exitosa de BD para que el sistema se reinicie automáticamente
        # (cierra la sesión y vuelve al login con la nueva BD ya cargada).
        self.on_reiniciar_sesion = on_reiniciar_sesion
        self._construir_ui()

    def _construir_ui(self):
        # ── Encabezado ──────────────────────────────────────────
        encabezado = tk.Frame(self, bg=AZUL_RIBBON, height=54)
        encabezado.pack(fill="x")
        encabezado.pack_propagate(False)
        tk.Label(encabezado, text=t("datos_titulo"),
                 font=("Segoe UI", 15, "bold"), bg=AZUL_RIBBON, fg=BLANCO
                 ).pack(side="left", padx=20, pady=12)

        # ── Área de tarjetas centrada ────────────────────────────
        canvas = tk.Canvas(self, bg=GRIS_FONDO, highlightthickness=0)
        scroll = tk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)

        contenedor = tk.Frame(canvas, bg=GRIS_FONDO)
        canvas.create_window((0, 0), window=contenedor, anchor="nw")
        contenedor.bind("<Configure>",
                        lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # ── Rueda del mouse: bindear al canvas y a todos los hijos ──
        def _scroll_rueda(event):
            # Windows: event.delta; Linux: Button-4/5
            if event.num == 4:
                canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                canvas.yview_scroll(1, "units")
            else:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _bind_scroll_recursivo(widget):
            widget.bind("<MouseWheel>", _scroll_rueda)   # Windows / macOS
            widget.bind("<Button-4>",   _scroll_rueda)   # Linux scroll arriba
            widget.bind("<Button-5>",   _scroll_rueda)   # Linux scroll abajo
            for hijo in widget.winfo_children():
                _bind_scroll_recursivo(hijo)

        # Guardamos la función para poder aplicarla a widgets que se añaden luego
        self._bind_scroll = _bind_scroll_recursivo
        self._canvas      = canvas
        _bind_scroll_recursivo(canvas)

        # Re-bindear cuando el contenedor añada nuevos hijos (tarjetas, labels, etc.)
        contenedor.bind("<Configure>",
                        lambda e: (_bind_scroll_recursivo(contenedor),
                                   canvas.configure(scrollregion=canvas.bbox("all"))))

        # ── Sección: Base de datos ───────────────────────────────
        self._seccion(contenedor, t("datos_seccion_bd"))

        fila1 = tk.Frame(contenedor, bg=GRIS_FONDO)
        fila1.pack(fill="x", padx=30, pady=(0, 10))

        self._tarjeta(
            fila1,
            icono="⬆", titulo="Exportar copia de seguridad",
            descripcion="Guarda una copia del archivo de base de datos (.db)\nen la ubicación que elijas.",
            color_boton=AZUL_RIBBON, texto_boton="Exportar .db",
            comando=self._exportar_bd,
        )
        self._tarjeta(
            fila1,
            icono="⬇", titulo="Importar copia de seguridad",
            descripcion="Restaura la base de datos desde un archivo .db\nprevio. ¡Se reemplazará la BD actual!",
            color_boton=NARANJA, texto_boton="Importar .db",
            comando=self._importar_bd,
        )

        # ── Sección: Excel ───────────────────────────────────────
        self._seccion(contenedor, t("datos_seccion_excel"))

        fila2 = tk.Frame(contenedor, bg=GRIS_FONDO)
        fila2.pack(fill="x", padx=30, pady=(0, 10))

        self._tarjeta(
            fila2,
            icono="📤", titulo="Exportar a Excel",
            descripcion="Genera un libro .xlsx con hojas para Ventas, Productos,\n"
                        "Créditos, Compras, Presupuestos, Préstamos, Asistencia\n"
                        "Técnica, Veterinaria, Restaurante, Streaming y RRHH.",
            color_boton=VERDE, texto_boton="Exportar .xlsx",
            comando=self._exportar_excel,
        )
        self._tarjeta(
            fila2,
            icono="📥", titulo="Importar desde Excel",
            descripcion="Lee un .xlsx con formato MAQUEDASYSTEMS\ne inserta Productos y Clientes en la BD.",
            color_boton=AZUL_OSCURO, texto_boton="Importar .xlsx",
            comando=self._importar_excel,
        )

        # ── Sección: CSV ─────────────────────────────────────────
        self._seccion(contenedor, t("datos_seccion_csv"))

        fila3 = tk.Frame(contenedor, bg=GRIS_FONDO)
        fila3.pack(fill="x", padx=30, pady=(0, 10))

        self._tarjeta(
            fila3,
            icono="📤", titulo="Exportar a CSV",
            descripcion="Exporta Productos, Clientes o Ventas a archivos\n.csv separados por comas, compatibles con cualquier planilla.",
            color_boton=VERDE, texto_boton="Exportar .csv",
            comando=self._exportar_csv,
        )
        self._tarjeta(
            fila3,
            icono="📥", titulo="Importar desde CSV",
            descripcion="Lee un archivo .csv con encabezados estándar\ne inserta Productos o Clientes en la BD.",
            color_boton=AZUL_OSCURO, texto_boton="Importar .csv",
            comando=self._importar_csv,
        )

        # ── Sección: Avanzado ────────────────────────────────────
        self._seccion(contenedor, t("datos_seccion_avanzado"))

        fila4 = tk.Frame(contenedor, bg=GRIS_FONDO)
        fila4.pack(fill="x", padx=30, pady=(0, 10))

        self._tarjeta(
            fila4,
            icono="🖥", titulo="Terminal SQL",
            descripcion="Ejecuta consultas SQL directamente sobre la base\nde datos. Solo para uso avanzado — hace backup\nautomático antes de cualquier cambio.",
            color_boton="#1e293b",
            texto_boton="Abrir Terminal SQL",
            comando=self._abrir_terminal_sql,
        )

        # ── Nota informativa ────────────────────────────────────
        nota = tk.Frame(contenedor, bg="#fefce8", relief="flat", bd=0)
        nota.pack(fill="x", padx=30, pady=(10, 20))
        tk.Frame(nota, bg="#fbbf24", width=4).pack(side="left", fill="y")
        tk.Label(nota,
                 text="⚠  Importante: antes de importar una copia de seguridad, el sistema\n"
                      "realizará un respaldo automático de los datos actuales para evitar pérdidas.",
                 font=("Segoe UI", 9), bg="#fefce8", fg="#92400e",
                 justify="left", padx=14, pady=10).pack(side="left", anchor="w")

    # ── Helpers de UI ────────────────────────────────────────────
    def _seccion(self, parent, titulo):
        f = tk.Frame(parent, bg=GRIS_FONDO)
        f.pack(fill="x", padx=30, pady=(20, 8))
        tk.Label(f, text=titulo, font=("Segoe UI", 12, "bold"),
                 bg=GRIS_FONDO, fg="#1e293b").pack(side="left")
        tk.Frame(f, bg=GRIS_BORDE, height=1).pack(side="left", fill="x", expand=True, padx=(12, 0), pady=7)

    def _tarjeta(self, parent, icono, titulo, descripcion,
                 color_boton, texto_boton, comando):
        card = tk.Frame(parent, bg=BLANCO, relief="flat", bd=0,
                        highlightthickness=1, highlightbackground=GRIS_BORDE)
        card.pack(side="left", fill="both", expand=True, padx=(0, 16), pady=4, ipadx=10, ipady=12)

        tk.Label(card, text=icono, font=("Segoe UI", 28),
                 bg=BLANCO).pack(anchor="w", padx=16, pady=(14, 4))
        tk.Label(card, text=titulo, font=("Segoe UI", 11, "bold"),
                 bg=BLANCO, fg="#1e293b").pack(anchor="w", padx=16)
        tk.Label(card, text=descripcion, font=("Segoe UI", 9),
                 bg=BLANCO, fg=GRIS_TEXTO, justify="left",
                 wraplength=260).pack(anchor="w", padx=16, pady=(4, 14))

        tk.Button(card, text=texto_boton, font=("Segoe UI", 10, "bold"),
                  bg=color_boton, fg=BLANCO, relief="flat", bd=0,
                  padx=18, pady=8, cursor="hand2",
                  activebackground=AZUL_OSCURO, activeforeground=BLANCO,
                  command=comando).pack(anchor="w", padx=16, pady=(0, 14))

    # ─────────────────────────────────────────────────────────────
    #  ACCIONES
    # ─────────────────────────────────────────────────────────────

    # ── 1. EXPORTAR BD ──────────────────────────────────────────
    def _abrir_terminal_sql(self):
        from ventana_terminal_sql import VentanaTerminalSQL
        VentanaTerminalSQL(self, self.usuario_actual)

    def _exportar_bd(self):
        ahora = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_sugerido = f"backup_maquedasystems_{ahora}.db"

        destino = filedialog.asksaveasfilename(
            title="Guardar copia de seguridad",
            defaultextension=".db",
            initialfile=nombre_sugerido,
            filetypes=[("Base de datos SQLite", "*.db"), ("Todos los archivos", "*.*")],
        )
        if not destino:
            return

        try:
            shutil.copy2(obtener_ruta_bd(), destino)
            messagebox.showinfo(
                "Exportación exitosa",
                f"✔ Copia de seguridad guardada en:\n{destino}",
            )
        except Exception as e:
            messagebox.showerror("Error al exportar", f"No se pudo guardar la copia:\n{e}", parent=self)

    # ── 2. IMPORTAR BD ──────────────────────────────────────────
    def _importar_bd(self):
        origen = filedialog.askopenfilename(
            title="Seleccionar copia de seguridad (.db)",
            filetypes=[("Base de datos SQLite", "*.db"), ("Todos los archivos", "*.*")],
        )
        if not origen:
            return

        # Validar que el archivo sea realmente una BD SQLite de MAQUEDASYSTEMS
        # antes de reemplazar la BD activa, para evitar corromperla con un
        # archivo equivocado.
        try:
            import sqlite3 as _sqlite3
            conn_test = _sqlite3.connect(origen)
            cursor_test = conn_test.cursor()
            # Verificar que existan las tablas principales del sistema
            cursor_test.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tablas_backup = {row[0] for row in cursor_test.fetchall()}
            conn_test.close()
            tablas_requeridas = {"productos", "ventas", "clientes", "usuarios"}
            faltantes = tablas_requeridas - tablas_backup
            if faltantes:
                messagebox.showerror(
                    "Archivo inválido",
                    f"El archivo seleccionado no parece ser una copia de seguridad "
                    f"de MAQUEDASYSTEMS.\n\n"
                    f"Tablas esperadas no encontradas: {', '.join(sorted(faltantes))}\n\n"
                    "Operación cancelada.", parent=self,
                )
                return
        except Exception as e:
            messagebox.showerror(
                "Archivo inválido",
                f"No se pudo leer el archivo seleccionado como base de datos SQLite:\n{e}\n\n"
                "Asegurate de seleccionar un archivo .db exportado desde MAQUEDASYSTEMS.",
                parent=self,
            )
            return

        confirmar = messagebox.askyesno(
            "⚠ Confirmar restauración",
            "Esto REEMPLAZARÁ toda la base de datos actual con el archivo seleccionado.\n\n"
            "Se guardará un respaldo automático de los datos actuales antes de continuar.\n\n"
            "Al finalizar, el sistema cerrará la sesión automáticamente para recargar "
            "los datos desde cero.\n\n"
            "¿Deseás continuar?",
        )
        if not confirmar:
            return

        ruta_bd = obtener_ruta_bd()

        # Backup automático previo
        ahora = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        carpeta_bd = os.path.dirname(ruta_bd)
        backup_auto = os.path.join(carpeta_bd, f"backup_auto_{ahora}.db")
        try:
            shutil.copy2(ruta_bd, backup_auto)
        except Exception as e:
            messagebox.showerror(
                "Error", f"No se pudo crear el respaldo automático:\n{e}\n\nOperación cancelada.",
                parent=self,
            )
            return

        try:
            shutil.copy2(origen, ruta_bd)
            messagebox.showinfo(
                "Restauración exitosa",
                f"✔ Base de datos restaurada correctamente.\n\n"
                f"Respaldo automático guardado en:\n{backup_auto}\n\n"
                "El sistema se reiniciará ahora para cargar los nuevos datos.",
                parent=self,
            )
            # Reiniciar la sesión automáticamente para que el sistema
            # vuelva al login ya con la nueva BD cargada desde cero, sin
            # que el usuario tenga que cerrar y abrir manualmente.
            if self.on_reiniciar_sesion:
                self.on_reiniciar_sesion()
        except Exception as e:
            # Intentar recuperar el backup si algo falló
            try:
                shutil.copy2(backup_auto, ruta_bd)
            except Exception:
                pass
            messagebox.showerror("Error al importar", f"No se pudo restaurar la BD:\n{e}", parent=self)

    # ── 3. EXPORTAR EXCEL ───────────────────────────────────────
    def _exportar_excel(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror(
                "Librería faltante",
                "Se necesita 'openpyxl' para exportar a Excel.\n\n"
                "Instálala con:\n  pip install openpyxl",
            )
            return

        ahora = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        destino = filedialog.asksaveasfilename(
            title="Guardar exportación Excel",
            defaultextension=".xlsx",
            initialfile=f"maquedasystems_export_{ahora}.xlsx",
            filetypes=[("Libro Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
        )
        if not destino:
            return

        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # quitar hoja vacía por defecto
            conn = conectar()
            cursor = conn.cursor()

            # ── Hoja: Productos ──────────────────────────────────
            ws_prod = wb.create_sheet("Productos")
            cursor.execute("""
                SELECT p.id, p.nombre, p.codigo_barras, p.precio_compra,
                       p.precio AS precio_venta, p.precio_mayorista, p.stock,
                       p.comprometido, p.tipo_producto, p.control_stock,
                       p.activo, c.nombre AS categoria, m.nombre AS marca
                FROM productos p
                LEFT JOIN categorias c ON p.categoria_id = c.id
                LEFT JOIN marcas m ON p.marca_id = m.id
                ORDER BY p.id
            """)
            cols_prod = ["ID", "Nombre", "Código de Barras", "Precio Compra",
                         "Precio Venta", "Precio Mayorista", "Stock",
                         "Comprometido", "Tipo", "Control Stock",
                         "Activo", "Categoría", "Marca"]
            _escribir_hoja(ws_prod, cols_prod, cursor.fetchall())

            # ── Hoja: Clientes ───────────────────────────────────
            ws_cli = wb.create_sheet("Clientes")
            cursor.execute("""
                SELECT id, nombre, razon_social, nro_documento, ruc,
                       telefono, email, direccion,
                       tipo_persona, nacionalidad,
                       credito_permitido, observaciones, fecha_creacion
                FROM clientes ORDER BY id
            """)
            cols_cli = ["ID", "Nombre", "Razón Social", "Nro. Documento", "RUC",
                        "Teléfono", "Email", "Dirección",
                        "Tipo Persona", "Nacionalidad",
                        "Crédito Permitido", "Observaciones", "Fecha Alta"]
            _escribir_hoja(ws_cli, cols_cli, cursor.fetchall())

            # ── Hoja: Ventas ─────────────────────────────────────
            ws_ventas = wb.create_sheet("Ventas")
            cursor.execute("""
                SELECT v.id, v.fecha,
                       COALESCE(cl.nombre, 'Ocasional') AS cliente,
                       u.nombre_completo AS vendedor,
                       v.total, v.condicion, v.forma_pago, v.estado
                FROM ventas v
                LEFT JOIN clientes cl ON v.cliente_id = cl.id
                LEFT JOIN usuarios u  ON v.usuario_id  = u.id
                ORDER BY v.id DESC
            """)
            cols_ventas = ["ID Venta", "Fecha", "Cliente", "Vendedor",
                           "Total", "Condición", "Forma de Pago", "Estado"]
            _escribir_hoja(ws_ventas, cols_ventas, cursor.fetchall())

            # ── Hoja: Detalle de Ventas ──────────────────────────
            ws_det = wb.create_sheet("Detalle de Ventas")
            cursor.execute("""
                SELECT dv.id, dv.venta_id,
                       COALESCE(p.nombre, dv.descripcion_libre) AS producto,
                       dv.cantidad, dv.precio_unitario,
                       (dv.cantidad * dv.precio_unitario) AS importe
                FROM detalle_ventas dv
                LEFT JOIN productos p ON dv.producto_id = p.id
                ORDER BY dv.venta_id DESC, dv.id
            """)
            cols_det = ["ID", "ID Venta", "Producto", "Cantidad",
                        "Precio Unitario", "Importe"]
            _escribir_hoja(ws_det, cols_det, cursor.fetchall())

            # ── Hoja: Inventario (stock actual) ──────────────────
            ws_inv = wb.create_sheet("Inventario")
            cursor.execute("""
                SELECT p.id,
                       p.nombre,
                       COALESCE(c.nombre, '—') AS categoria,
                       COALESCE(m.nombre, '—') AS marca,
                       p.stock,
                       p.comprometido,
                       (p.stock - p.comprometido) AS disponible,
                       p.stock_minimo,
                       CASE WHEN (p.stock - p.comprometido) < p.stock_minimo
                            THEN 'Bajo stock' ELSE 'OK' END AS estado_stock,
                       p.precio_compra,
                       p.precio AS precio_venta,
                       (p.stock * p.precio_compra) AS valor_inventario,
                       p.tipo_producto,
                       p.control_stock,
                       p.unidad_medida,
                       CASE WHEN p.activo = 1 THEN 'Activo' ELSE 'Inactivo' END AS estado
                FROM productos p
                LEFT JOIN categorias c ON p.categoria_id = c.id
                LEFT JOIN marcas m     ON p.marca_id     = m.id
                WHERE p.tipo_producto != 'Servicio'
                  AND p.control_stock != 'Ilimitado'
                ORDER BY categoria, p.nombre
            """)
            cols_inv = [
                "ID", "Producto", "Categoría", "Marca",
                "Stock", "Comprometido", "Disponible", "Stock Mínimo",
                "Estado Stock", "Precio Compra", "Precio Venta",
                "Valor Inventario", "Tipo", "Control Stock",
                "Unidad de Medida", "Estado"
            ]
            filas_inv = cursor.fetchall()
            _escribir_hoja(ws_inv, cols_inv, filas_inv)

            # Resaltar en rojo las filas con stock bajo
            try:
                from openpyxl.styles import PatternFill, Font
                rojo_claro = PatternFill("solid", fgColor="FFE5E5")
                font_rojo  = Font(color="CC0000", bold=True)
                col_estado = cols_inv.index("Estado Stock") + 1  # 1-indexed
                for row_idx, fila in enumerate(filas_inv, start=2):
                    if fila[8] == "Bajo stock":  # índice de estado_stock
                        for col_idx in range(1, len(cols_inv) + 1):
                            celda = ws_inv.cell(row=row_idx, column=col_idx)
                            celda.fill = rojo_claro
                        ws_inv.cell(row=row_idx, column=col_estado).font = font_rojo
            except Exception:
                pass  # el resaltado es opcional

            # ── Hoja: Movimientos de Inventario ──────────────────
            ws_mov = wb.create_sheet("Movimientos Inventario")
            cursor.execute("""
                SELECT mi.id,
                       mi.fecha,
                       COALESCE(p.nombre, mi.producto_nombre_historico, '(Producto eliminado)') AS producto,
                       mi.tipo,
                       CASE WHEN mi.es_ilimitado = 1 THEN 'Ilimitado'
                            ELSE mi.cantidad END AS cantidad,
                       CASE WHEN mi.es_ilimitado = 1 THEN 'Ilimitado'
                            ELSE mi.stock_resultante END AS stock_resultante,
                       mi.motivo,
                       mi.nro_comprobante,
                       mi.observaciones,
                       COALESCE(u.nombre_completo, '—') AS usuario
                FROM movimientos_inventario mi
                LEFT JOIN productos p ON mi.producto_id = p.id
                LEFT JOIN usuarios u ON mi.usuario_id = u.id
                ORDER BY mi.fecha DESC, mi.id DESC
            """)
            cols_mov = [
                "ID", "Fecha", "Producto", "Tipo",
                "Cantidad", "Stock Resultante", "Motivo",
                "Nro. Comprobante", "Observaciones", "Registrado por"
            ]
            _escribir_hoja(ws_mov, cols_mov, cursor.fetchall())

            # ── Hoja: Presupuestos ────────────────────────────────
            ws_pres = wb.create_sheet("Presupuestos")
            cursor.execute("""
                SELECT pr.id, pr.fecha, pr.cliente_nombre, pr.fecha_validez,
                       pr.estado, pr.total, COALESCE(u.nombre_completo, '—'),
                       pr.venta_id, pr.observaciones
                FROM presupuestos pr
                LEFT JOIN usuarios u ON pr.usuario_id = u.id
                ORDER BY pr.id DESC
            """)
            cols_pres = ["ID", "Fecha", "Cliente", "Fecha Validez", "Estado",
                         "Total", "Vendedor", "ID Venta (si se convirtió)", "Observaciones"]
            _escribir_hoja(ws_pres, cols_pres, cursor.fetchall())

            # ── Hoja: Detalle de Presupuestos ────────────────────
            ws_det_pres = wb.create_sheet("Detalle Presupuestos")
            cursor.execute("""
                SELECT dp.id, dp.presupuesto_id,
                       COALESCE(p.nombre, dp.descripcion_libre, '(Artículo libre)'),
                       dp.cantidad, dp.precio_unitario,
                       (dp.cantidad * dp.precio_unitario)
                FROM detalle_presupuestos dp
                LEFT JOIN productos p ON dp.producto_id = p.id
                ORDER BY dp.presupuesto_id DESC, dp.id
            """)
            cols_det_pres = ["ID", "ID Presupuesto", "Producto", "Cantidad",
                              "Precio Unitario", "Importe"]
            _escribir_hoja(ws_det_pres, cols_det_pres, cursor.fetchall())

            # ── Hoja: Créditos ────────────────────────────────────
            ws_cred = wb.create_sheet("Créditos")
            cursor.execute("""
                SELECT c.id, c.venta_id, COALESCE(cl.nombre, '—'), c.fecha,
                       c.fecha_vencimiento, c.descripcion, c.deuda_total,
                       c.pagado, (c.deuda_total - c.pagado)
                FROM creditos c
                LEFT JOIN clientes cl ON c.cliente_id = cl.id
                ORDER BY c.id DESC
            """)
            cols_cred = ["ID", "ID Venta", "Cliente", "Fecha", "Fecha Vencimiento",
                         "Descripción", "Deuda Total", "Pagado", "Saldo Pendiente"]
            _escribir_hoja(ws_cred, cols_cred, cursor.fetchall())

            # ── Hoja: Pagos de Créditos ───────────────────────────
            ws_pagos = wb.create_sheet("Pagos de Créditos")
            cursor.execute("""
                SELECT id, credito_id, fecha, monto
                FROM pagos_credito ORDER BY credito_id DESC, id
            """)
            cols_pagos = ["ID", "ID Crédito", "Fecha", "Monto Pagado"]
            _escribir_hoja(ws_pagos, cols_pagos, cursor.fetchall())

            # ── Hoja: Compras ─────────────────────────────────────
            ws_compras = wb.create_sheet("Compras")
            cursor.execute("""
                SELECT co.id, co.fecha_y_hora, co.fecha_compra,
                       COALESCE(pv.nombre, '—'), co.nro_comprobante, co.importe
                FROM compras co
                LEFT JOIN proveedores pv ON co.proveedor_id = pv.id
                ORDER BY co.id DESC
            """)
            cols_compras = ["ID", "Fecha y Hora", "Fecha Compra", "Proveedor",
                             "Nro. Comprobante", "Importe"]
            _escribir_hoja(ws_compras, cols_compras, cursor.fetchall())

            # ── Hoja: Detalle de Compras ──────────────────────────
            ws_det_compras = wb.create_sheet("Detalle de Compras")
            cursor.execute("""
                SELECT dc.id, dc.compra_id,
                       COALESCE(p.nombre, dc.producto_nombre_historico, '(Producto eliminado)'),
                       dc.cantidad, dc.precio_unitario,
                       (dc.cantidad * dc.precio_unitario)
                FROM detalle_compras dc
                LEFT JOIN productos p ON dc.producto_id = p.id
                ORDER BY dc.compra_id DESC, dc.id
            """)
            cols_det_compras = ["ID", "ID Compra", "Producto", "Cantidad",
                                 "Precio Unitario", "Importe"]
            _escribir_hoja(ws_det_compras, cols_det_compras, cursor.fetchall())

            # ── Hoja: Proveedores ──────────────────────────────────
            ws_prov = wb.create_sheet("Proveedores")
            cursor.execute("""
                SELECT id, nombre, contacto, telefono, direccion, ruc
                FROM proveedores ORDER BY nombre
            """)
            cols_prov = ["ID", "Nombre", "Contacto", "Teléfono", "Dirección", "RUC"]
            _escribir_hoja(ws_prov, cols_prov, cursor.fetchall())

            # ── Hoja: Movimientos de Caja (Entradas/Salidas) ──────
            ws_caja = wb.create_sheet("Movimientos de Caja")
            cursor.execute("""
                SELECT cm.id, cm.fecha, cm.tipo, cm.monto, cm.descripcion,
                       COALESCE(u.nombre_completo, '—')
                FROM caja_movimientos cm
                LEFT JOIN usuarios u ON cm.usuario_id = u.id
                ORDER BY cm.fecha DESC, cm.id DESC
            """)
            cols_caja = ["ID", "Fecha", "Tipo", "Monto", "Descripción", "Registrado por"]
            _escribir_hoja(ws_caja, cols_caja, cursor.fetchall())

            # ── Hoja: Devoluciones ─────────────────────────────────
            ws_devol = wb.create_sheet("Devoluciones")
            cursor.execute("""
                SELECT d.id, d.venta_id,
                       COALESCE(p.nombre, d.producto_nombre_historico, '(Producto eliminado)'),
                       d.cantidad, d.importe, d.fecha, COALESCE(u.nombre_completo, '—')
                FROM devoluciones d
                LEFT JOIN productos p ON d.producto_id = p.id
                LEFT JOIN usuarios u ON d.usuario_id = u.id
                ORDER BY d.fecha DESC, d.id DESC
            """)
            cols_devol = ["ID", "ID Venta", "Producto", "Cantidad", "Importe",
                          "Fecha", "Registrado por"]
            _escribir_hoja(ws_devol, cols_devol, cursor.fetchall())

            # ── Hoja: Facturas ─────────────────────────────────────
            ws_fact = wb.create_sheet("Facturas")
            cursor.execute("""
                SELECT id, venta_id, nro_factura, fecha, razon_social, ruc,
                       valor_total, estado
                FROM facturas ORDER BY id DESC
            """)
            cols_fact = ["ID", "ID Venta", "Nro. Factura", "Fecha", "Razón Social",
                         "RUC", "Valor Total", "Estado"]
            _escribir_hoja(ws_fact, cols_fact, cursor.fetchall())

            # ── Hoja: Asistencia Técnica ───────────────────────────
            ws_tecnico = wb.create_sheet("Asistencia Técnica")
            cursor.execute("""
                SELECT ct.id, ct.fecha_entrada, ct.cliente_nombre, ct.cliente_telefono,
                       COALESCE(te.nombre, ct.tipo_equipo_texto, '—'), ct.nro_serie,
                       ct.descripcion_equipo, ct.prioridad, ct.estado,
                       ct.fecha_estado, ct.fecha_retiro,
                       CASE WHEN ct.anulado = 1 THEN 'Sí' ELSE 'No' END,
                       COALESCE(u.nombre_completo, '—')
                FROM casos_tecnicos ct
                LEFT JOIN tipos_equipo te ON ct.tipo_equipo_id = te.id
                LEFT JOIN usuarios u ON ct.usuario_id = u.id
                ORDER BY ct.id DESC
            """)
            cols_tecnico = ["ID", "Fecha Entrada", "Cliente", "Teléfono", "Tipo de Equipo",
                            "Nro. Serie", "Descripción", "Prioridad", "Estado",
                            "Fecha de Estado", "Fecha de Retiro", "Anulado", "Registrado por"]
            _escribir_hoja(ws_tecnico, cols_tecnico, cursor.fetchall())

            # ── Hoja: Préstamos ─────────────────────────────────────
            ws_prestamos = wb.create_sheet("Préstamos")
            cursor.execute("""
                SELECT pr.id, pr.fecha_desembolso, COALESCE(cl.nombre, '—'),
                       pr.capital, pr.tasa_interes, pr.frecuencia, pr.cantidad_cuotas,
                       pr.sistema, pr.tasa_mora_diaria, pr.estado,
                       COALESCE(u.nombre_completo, '—'), pr.fecha_creacion
                FROM prestamos pr
                LEFT JOIN clientes cl ON pr.cliente_id = cl.id
                LEFT JOIN usuarios u ON pr.usuario_id = u.id
                ORDER BY pr.id DESC
            """)
            cols_prestamos = ["ID", "Fecha Desembolso", "Cliente", "Capital",
                              "Tasa Interés", "Frecuencia", "Cant. Cuotas", "Sistema",
                              "Tasa Mora Diaria", "Estado", "Registrado por", "Fecha Creación"]
            _escribir_hoja(ws_prestamos, cols_prestamos, cursor.fetchall())

            # ── Hoja: Cuotas de Préstamos ────────────────────────────
            ws_cuotas = wb.create_sheet("Cuotas de Préstamos")
            cursor.execute("""
                SELECT c.id, c.prestamo_id, c.nro_cuota, c.fecha_vencimiento,
                       c.capital, c.interes, c.pagado_capital, c.pagado_interes,
                       c.pagado_mora, c.fecha_pago_completo
                FROM cuotas_prestamo c
                ORDER BY c.prestamo_id DESC, c.nro_cuota
            """)
            cols_cuotas = ["ID", "ID Préstamo", "Nro. Cuota", "Fecha Vencimiento",
                          "Capital", "Interés", "Pagado Capital", "Pagado Interés",
                          "Pagado Mora", "Fecha Pago Completo"]
            _escribir_hoja(ws_cuotas, cols_cuotas, cursor.fetchall())

            # ── Hoja: Mascotas (Veterinaria) ─────────────────────────
            ws_mascotas = wb.create_sheet("Veterinaria - Mascotas")
            cursor.execute("""
                SELECT m.id, COALESCE(cl.nombre, m.dueño_nombre, '—'), m.dueño_telefono,
                       m.nombre, COALESCE(e.nombre, m.especie_texto, '—'), m.raza, m.sexo,
                       m.fecha_nacimiento, m.peso_kg,
                       CASE WHEN m.esterilizado = 1 THEN 'Sí' ELSE 'No' END,
                       CASE WHEN m.fallecido = 1 THEN 'Sí' ELSE 'No' END
                FROM mascotas m
                LEFT JOIN clientes cl ON m.cliente_id = cl.id
                LEFT JOIN especies_mascota e ON m.especie_id = e.id
                ORDER BY m.id DESC
            """)
            cols_mascotas = ["ID", "Dueño", "Teléfono", "Nombre Mascota", "Especie",
                             "Raza", "Sexo", "Fecha Nacimiento", "Peso (Kg)",
                             "Esterilizado", "Fallecido"]
            _escribir_hoja(ws_mascotas, cols_mascotas, cursor.fetchall())

            # ── Hoja: Consultas Veterinarias ──────────────────────────
            ws_consultas = wb.create_sheet("Veterinaria - Consultas")
            cursor.execute("""
                SELECT c.id, COALESCE(m.nombre, '—'), c.fecha, c.motivo,
                       c.diagnostico, c.costo
                FROM consultas_veterinarias c
                LEFT JOIN mascotas m ON c.mascota_id = m.id
                ORDER BY c.fecha DESC, c.id DESC
            """)
            cols_consultas = ["ID", "Mascota", "Fecha", "Motivo", "Diagnóstico", "Costo"]
            _escribir_hoja(ws_consultas, cols_consultas, cursor.fetchall())

            # ── Hoja: Comandas (Restaurante/Comedor) ──────────────────
            ws_comandas = wb.create_sheet("Restaurante - Comandas")
            cursor.execute("""
                SELECT co.id, COALESCE(me.numero, '—'), co.tipo,
                       COALESCE(cl.nombre, 'Ocasional'), co.estado,
                       co.fecha_apertura, co.fecha_cierre, co.venta_id
                FROM rest_comandas co
                LEFT JOIN rest_mesas me ON co.mesa_id = me.id
                LEFT JOIN clientes cl ON co.cliente_id = cl.id
                ORDER BY co.id DESC
            """)
            cols_comandas = ["ID", "Mesa", "Tipo", "Cliente", "Estado",
                             "Fecha Apertura", "Fecha Cierre", "ID Venta"]
            _escribir_hoja(ws_comandas, cols_comandas, cursor.fetchall())

            # ── Hoja: Suscripciones de Streaming ──────────────────────
            ws_stream = wb.create_sheet("Streaming - Suscripciones")
            cursor.execute("""
                SELECT s.id, COALESCE(cl.nombre, '—'), s.modalidad, s.fecha_inicio,
                       s.fecha_vencimiento, s.precio_mensual, s.estado, s.forma_pago
                FROM stream_suscripciones s
                LEFT JOIN clientes cl ON s.cliente_id = cl.id
                ORDER BY s.id DESC
            """)
            cols_stream = ["ID", "Cliente", "Modalidad", "Fecha Inicio",
                          "Fecha Vencimiento", "Precio Mensual", "Estado", "Forma de Pago"]
            _escribir_hoja(ws_stream, cols_stream, cursor.fetchall())

            # ── Hoja: Importaciones - Compras ─────────────────────────
            ws_import = wb.create_sheet("Importaciones - Compras")
            cursor.execute("""
                SELECT c.id, c.plataforma, c.referencia, COALESCE(co.nombre, '—'),
                       c.casillero, c.tipo_envio, c.peso_caja_kg, c.costo_envio_total,
                       c.estado, c.fecha_compra, c.fecha_envio_casillero, c.fecha_recepcion, c.notas
                FROM import_compras c
                LEFT JOIN import_couriers co ON c.courier_id = co.id
                ORDER BY c.id DESC
            """)
            cols_import = ["ID", "Plataforma", "Referencia", "Courier", "Casillero",
                           "Tipo de Envío", "Peso Caja (kg)", "Costo Envío Total (US$)",
                           "Estado", "Fecha Compra", "Fecha Envío a Casillero",
                           "Fecha Recepción", "Notas"]
            _escribir_hoja(ws_import, cols_import, cursor.fetchall())

            # ── Hoja: Importaciones - Detalle de Productos ────────────
            ws_import_det = wb.create_sheet("Importaciones - Detalle")
            cursor.execute("""
                SELECT d.id, d.compra_id, d.producto_nombre, d.cantidad,
                       d.costo_unitario_compra, d.costo_envio_unitario, d.costo_total_unitario,
                       d.precio_venta_publico,
                       CASE WHEN d.enviado_inventario = 1 THEN 'Sí' ELSE 'No' END
                FROM import_detalle d
                ORDER BY d.compra_id DESC, d.id
            """)
            cols_import_det = ["ID", "ID Compra", "Producto", "Cantidad",
                               "Costo Compra/U (US$)", "Costo Envío/U (US$)",
                               "Costo Total/U (US$)", "Precio Venta Público (US$)",
                               "Enviado a Inventario"]
            _escribir_hoja(ws_import_det, cols_import_det, cursor.fetchall())

            # ── Hojas: Recursos Humanos ────────────────────────────
            # Se garantiza que las tablas de RRHH existan aunque el
            # administrador todavía no haya abierto ese módulo.
            inicializar_tablas_rrhh()

            ws_rrhh_emp = wb.create_sheet("RRHH - Empleados")
            cursor.execute("""
                SELECT id, nombre, cargo, departamento, telefono, email,
                       fecha_ingreso, sueldo_mensual, horas_dia,
                       CASE WHEN activo = 1 THEN 'Activo' ELSE 'Inactivo' END,
                       observaciones
                FROM rrhh_empleados ORDER BY nombre
            """)
            cols_rrhh_emp = ["ID", "Nombre", "Cargo", "Departamento", "Teléfono",
                             "Email", "Fecha Ingreso", "Sueldo Mensual", "Horas/Día",
                             "Estado", "Observaciones"]
            _escribir_hoja(ws_rrhh_emp, cols_rrhh_emp, cursor.fetchall())

            ws_rrhh_asis = wb.create_sheet("RRHH - Asistencia")
            cursor.execute("""
                SELECT ra.id, COALESCE(e.nombre, '—'), ra.fecha, ra.estado,
                       ra.hora_entrada, ra.hora_salida, ra.observaciones
                FROM rrhh_asistencia ra
                LEFT JOIN rrhh_empleados e ON ra.empleado_id = e.id
                ORDER BY ra.fecha DESC, ra.id DESC
            """)
            cols_rrhh_asis = ["ID", "Empleado", "Fecha", "Estado", "Hora Entrada",
                              "Hora Salida", "Observaciones"]
            _escribir_hoja(ws_rrhh_asis, cols_rrhh_asis, cursor.fetchall())

            ws_rrhh_adel = wb.create_sheet("RRHH - Adelantos")
            cursor.execute("""
                SELECT ad.id, COALESCE(e.nombre, '—'), ad.fecha, ad.monto,
                       ad.descripcion, ad.estado, ad.registrado_por, ad.fecha_descuento
                FROM rrhh_adelantos ad
                LEFT JOIN rrhh_empleados e ON ad.empleado_id = e.id
                ORDER BY ad.fecha DESC, ad.id DESC
            """)
            cols_rrhh_adel = ["ID", "Empleado", "Fecha", "Monto", "Descripción",
                              "Estado", "Registrado por", "Fecha de Descuento"]
            _escribir_hoja(ws_rrhh_adel, cols_rrhh_adel, cursor.fetchall())

            conn.close()
            wb.save(destino)
            nombres_hojas = ", ".join(wb.sheetnames)
            messagebox.showinfo(
                "Exportación exitosa",
                f"✔ Archivo Excel generado correctamente.\n"
                f"Hojas: {nombres_hojas}\n\n{destino}",
            )
        except Exception as e:
            messagebox.showerror("Error al exportar", f"No se pudo generar el Excel:\n{e}", parent=self)

    # ── 4. IMPORTAR EXCEL ───────────────────────────────────────
    def _importar_excel(self):
        try:
            import openpyxl as _openpyxl_check  # noqa: verificar disponibilidad
            del _openpyxl_check
        except ImportError:
            messagebox.showerror(
                "Librería faltante",
                "Se necesita 'openpyxl' para importar desde Excel.\n\n"
                "Instálala con:\n  pip install openpyxl",
            )
            return

        origen = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Libro Excel", "*.xlsx *.xlsm"), ("Todos los archivos", "*.*")],
        )
        if not origen:
            return

        # Mostrar ventana de opciones de importación
        _VentanaOpcionesImport(self, origen)

    # ── 5. EXPORTAR CSV ─────────────────────────────────────────
    def _exportar_csv(self):
        _VentanaExportarCSV(self)

    # ── 6. IMPORTAR CSV ─────────────────────────────────────────
    def _importar_csv(self):
        origen = filedialog.askopenfilename(
            title="Seleccionar archivo CSV",
            filetypes=[("Archivo CSV", "*.csv"), ("Todos los archivos", "*.*")],
        )
        if not origen:
            return
        _VentanaImportarCSV(self, origen)


# ─────────────────────────────────────────────────────────────
#  VENTANA DE OPCIONES DE IMPORTACIÓN EXCEL
# ─────────────────────────────────────────────────────────────
class _VentanaOpcionesImport(tk.Toplevel):
    """Permite al usuario elegir qué hojas importar y qué hacer con duplicados."""

    def __init__(self, parent, ruta_excel):
        super().__init__(parent)
        self.ruta_excel = ruta_excel
        self.title("Importar desde Excel")
        self.minsize(480, 360)
        self.resizable(True, True)
        self.configure(bg=BLANCO)
        self.grab_set()

        try:
            import openpyxl
            self.wb = openpyxl.load_workbook(ruta_excel, read_only=True, data_only=True)
            self.hojas_disponibles = self.wb.sheetnames
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{e}", parent=self)
            self.destroy()
            return

        self._construir_ui()

    def _construir_ui(self):
        # Barra
        barra = tk.Frame(self, bg=AZUL_RIBBON, height=36)
        barra.pack(fill="x")
        barra.pack_propagate(False)
        tk.Label(barra, text="Opciones de importación",
                 font=("Segoe UI", 11, "bold"), bg=AZUL_RIBBON, fg=BLANCO
                 ).pack(side="left", padx=16, pady=7)

        cuerpo = tk.Frame(self, bg=BLANCO, padx=20, pady=16)
        cuerpo.pack(fill="both", expand=True)

        tk.Label(cuerpo, text="Selecciona qué hojas importar:",
                 font=("Segoe UI", 10, "bold"), bg=BLANCO).pack(anchor="w", pady=(0, 10))

        HOJAS_SOPORTADAS = {
            "Productos": "Importa productos (nombre, código, precios, stock).",
            "Clientes":  "Importa clientes (nombre, documento, teléfono, email).",
        }

        self.vars_hojas = {}
        for hoja, desc in HOJAS_SOPORTADAS.items():
            disponible = hoja in self.hojas_disponibles
            var = tk.BooleanVar(value=disponible)
            self.vars_hojas[hoja] = var
            f = tk.Frame(cuerpo, bg=BLANCO)
            f.pack(fill="x", pady=3)
            cb = tk.Checkbutton(f, variable=var, bg=BLANCO,
                                state="normal" if disponible else "disabled")
            cb.pack(side="left")
            estado = hoja if disponible else f"{hoja} (no encontrada en el archivo)"
            tk.Label(f, text=estado, font=("Segoe UI", 10, "bold"),
                     bg=BLANCO, fg="#1e293b" if disponible else GRIS_TEXTO).pack(side="left")
            tk.Label(cuerpo, text=f"   {desc}", font=("Segoe UI", 9),
                     bg=BLANCO, fg=GRIS_TEXTO).pack(anchor="w")

        tk.Frame(cuerpo, bg=GRIS_BORDE, height=1).pack(fill="x", pady=14)

        tk.Label(cuerpo, text="Si ya existe un registro con el mismo nombre:",
                 font=("Segoe UI", 10, "bold"), bg=BLANCO).pack(anchor="w")
        self.var_duplicados = tk.StringVar(value="omitir")
        for valor, etiqueta in [("omitir", "Omitir (no modificar el existente)"),
                                  ("actualizar", "Actualizar con los datos del Excel")]:
            tk.Radiobutton(cuerpo, text=etiqueta, variable=self.var_duplicados,
                           value=valor, bg=BLANCO, font=("Segoe UI", 9)).pack(anchor="w")

        tk.Frame(cuerpo, bg=GRIS_BORDE, height=1).pack(fill="x", pady=14)

        frame_botones = tk.Frame(cuerpo, bg=BLANCO)
        frame_botones.pack(fill="x")
        tk.Button(frame_botones, text="✔ Importar", font=("Segoe UI", 10, "bold"),
                  bg=VERDE, fg=BLANCO, relief="flat", padx=18, pady=8,
                  cursor="hand2", command=self._ejecutar_importacion).pack(side="left", padx=(0, 10))
        tk.Button(frame_botones, text="✕ Cancelar", font=("Segoe UI", 10, "bold"),
                  bg=BLANCO, fg=ROJO, relief="solid", bd=1, padx=18, pady=8,
                  cursor="hand2", command=self.destroy).pack(side="left")

    def _ejecutar_importacion(self):
        hojas_a_importar = [h for h, v in self.vars_hojas.items() if v.get()]
        if not hojas_a_importar:
            messagebox.showwarning("Sin selección", "Selecciona al menos una hoja para importar.",
                                   parent=self)
            return

        modo = self.var_duplicados.get()
        resumen = []
        errores = []

        conn = conectar()
        try:
            for hoja in hojas_a_importar:
                if hoja == "Productos":
                    ins, act, err = _importar_hoja_productos(conn, self.wb["Productos"], modo)
                    resumen.append(f"Productos: {ins} insertados, {act} actualizados.")
                    errores.extend(err)
                elif hoja == "Clientes":
                    ins, act, err = _importar_hoja_clientes(conn, self.wb["Clientes"], modo)
                    resumen.append(f"Clientes: {ins} insertados, {act} actualizados.")
                    errores.extend(err)
            conn.commit()
        except Exception as e:
            conn.rollback()
            messagebox.showerror("Error en importación", str(e), parent=self)
            conn.close()
            return
        finally:
            conn.close()

        msg = "✔ Importación completada.\n\n" + "\n".join(resumen)
        if errores:
            msg += f"\n\n⚠ {len(errores)} filas con errores (se omitieron)."
        messagebox.showinfo("Resultado", msg, parent=self)
        self.destroy()


# ─────────────────────────────────────────────────────────────
#  HELPERS DE ESCRITURA / LECTURA EXCEL
# ─────────────────────────────────────────────────────────────
def _estilo_encabezado(ws, fila_cols):
    """Aplica negrita y fondo azul a la fila de encabezados."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill("solid", fgColor="1D5FD6")
        font = Font(bold=True, color="FFFFFF")
        alin = Alignment(horizontal="center")
        for cell in ws[1]:
            cell.font = font
            cell.fill = fill
            cell.alignment = alin
    except Exception:
        pass  # si falla el estilo, no es crítico


def _escribir_hoja(ws, columnas, filas):
    """Escribe encabezados y datos en una hoja de openpyxl."""
    ws.append(columnas)
    _estilo_encabezado(ws, columnas)
    for fila in filas:
        ws.append(list(fila))
    # Ajustar ancho de columnas automáticamente
    try:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    except Exception:
        pass


def _celda(ws, fila, col):
    """Devuelve el valor de una celda de openpyxl (str limpio o None)."""
    val = ws.cell(row=fila, column=col).value
    if val is None:
        return None
    return str(val).strip() if isinstance(val, str) else val


def _float_o(val, default=0.0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _importar_hoja_productos(conn, ws, modo):
    """Lee la hoja 'Productos' e inserta/actualiza en la BD.
    Columnas esperadas (fila 1 = encabezados):
      Nombre | Código de Barras | Precio Compra | Precio Venta |
      Precio Mayorista | Stock | Tipo | Control Stock
    """
    cursor = conn.cursor()
    insertados = actualizados = 0
    errores = []

    encabezados = {str(ws.cell(1, c).value).strip().lower(): c
                   for c in range(1, ws.max_column + 1)
                   if ws.cell(1, c).value}

    col_nombre   = encabezados.get("nombre", 2)
    col_codigo   = encabezados.get("código de barras", 3)
    col_compra   = encabezados.get("precio compra", 4)
    col_venta    = encabezados.get("precio venta", 5)
    col_mayor    = encabezados.get("precio mayorista", 6)
    col_stock    = encabezados.get("stock", 7)
    col_tipo     = encabezados.get("tipo", 9)
    col_control  = encabezados.get("control stock", 10)

    for row in range(2, ws.max_row + 1):
        nombre = _celda(ws, row, col_nombre)
        if not nombre:
            continue
        try:
            codigo   = _celda(ws, row, col_codigo) or ""
            compra   = _float_o(_celda(ws, row, col_compra))
            venta    = _float_o(_celda(ws, row, col_venta))
            mayor    = _float_o(_celda(ws, row, col_mayor)) or venta
            stock    = int(_float_o(_celda(ws, row, col_stock)))
            tipo     = _celda(ws, row, col_tipo) or "Producto"
            control  = _celda(ws, row, col_control) or "Normal"

            cursor.execute("SELECT id FROM productos WHERE nombre = ?", (nombre,))
            existente = cursor.fetchone()

            if existente:
                if modo == "actualizar":
                    cursor.execute("""
                        UPDATE productos SET
                            codigo_barras=?, precio_compra=?, precio=?,
                            precio_mayorista=?, stock=?, tipo_producto=?, control_stock=?
                        WHERE nombre=?
                    """, (codigo, compra, venta, mayor, stock, tipo, control, nombre))
                    actualizados += 1
                # else: omitir
            else:
                cursor.execute("""
                    INSERT INTO productos
                        (nombre, codigo_barras, precio_compra, precio,
                         precio_mayorista, stock, tipo_producto, control_stock, activo)
                    VALUES (?,?,?,?,?,?,?,?,1)
                """, (nombre, codigo, compra, venta, mayor, stock, tipo, control))
                insertados += 1
        except Exception as e:
            errores.append(f"Fila {row}: {e}")

    return insertados, actualizados, errores


def _importar_hoja_clientes(conn, ws, modo):
    """Lee la hoja 'Clientes' e inserta/actualiza en la BD.
    Columnas esperadas:
      Nombre | Razón Social | Nro. Documento | RUC | Teléfono | Email | Dirección | Observaciones
    """
    cursor = conn.cursor()
    insertados = actualizados = 0
    errores = []

    encabezados = {str(ws.cell(1, c).value).strip().lower(): c
                   for c in range(1, ws.max_column + 1)
                   if ws.cell(1, c).value}

    col_nombre  = encabezados.get("nombre", 2)
    col_razon   = encabezados.get("razón social", 3)
    col_doc     = encabezados.get("nro. documento", 4)
    col_ruc     = encabezados.get("ruc", 5)
    col_tel     = encabezados.get("teléfono", 6)
    col_email   = encabezados.get("email", 7)
    col_dir     = encabezados.get("dirección", 8)
    col_obs     = encabezados.get("observaciones", 9)

    for row in range(2, ws.max_row + 1):
        nombre = _celda(ws, row, col_nombre)
        if not nombre:
            continue
        try:
            razon  = _celda(ws, row, col_razon) or nombre
            doc    = _celda(ws, row, col_doc) or ""
            ruc    = _celda(ws, row, col_ruc) or ""
            tel    = _celda(ws, row, col_tel) or ""
            email  = _celda(ws, row, col_email) or ""
            direc  = _celda(ws, row, col_dir) or ""
            obs    = _celda(ws, row, col_obs) or ""

            cursor.execute("SELECT id FROM clientes WHERE nombre = ?", (nombre,))
            existente = cursor.fetchone()

            if existente:
                if modo == "actualizar":
                    cursor.execute("""
                        UPDATE clientes SET
                            razon_social=?, nro_documento=?, ruc=?,
                            telefono=?, email=?, direccion=?, observaciones=?
                        WHERE nombre=?
                    """, (razon, doc, ruc, tel, email, direc, obs, nombre))
                    actualizados += 1
            else:
                cursor.execute("""
                    INSERT INTO clientes
                        (nombre, razon_social, nro_documento, ruc,
                         telefono, email, direccion, observaciones)
                    VALUES (?,?,?,?,?,?,?,?)
                """, (nombre, razon, doc, ruc, tel, email, direc, obs))
                insertados += 1
        except Exception as e:
            errores.append(f"Fila {row}: {e}")

    return insertados, actualizados, errores

# ─────────────────────────────────────────────────────────────
#  VENTANA EXPORTAR CSV
# ─────────────────────────────────────────────────────────────
class _VentanaExportarCSV(tk.Toplevel):
    """Permite elegir qué tabla exportar a CSV y la carpeta destino."""

    TABLAS = {
        "Productos": {
            "sql": """
                SELECT p.id, p.nombre, p.codigo_barras, p.precio_compra,
                       p.precio AS precio_venta, p.precio_mayorista, p.stock,
                       p.comprometido, p.tipo_producto, p.control_stock,
                       p.activo, c.nombre, m.nombre
                FROM productos p
                LEFT JOIN categorias c ON p.categoria_id = c.id
                LEFT JOIN marcas m     ON p.marca_id     = m.id
                ORDER BY p.id
            """,
            "columnas": ["ID", "Nombre", "Código de Barras", "Precio Compra",
                         "Precio Venta", "Precio Mayorista", "Stock",
                         "Comprometido", "Tipo", "Control Stock",
                         "Activo", "Categoría", "Marca"],
            "archivo": "productos",
        },
        "Clientes": {
            "sql": """
                SELECT id, nombre, razon_social, nro_documento, ruc,
                       telefono, email, direccion,
                       tipo_persona, nacionalidad,
                       credito_permitido, observaciones, fecha_creacion
                FROM clientes ORDER BY id
            """,
            "columnas": ["ID", "Nombre", "Razón Social", "Nro. Documento", "RUC",
                         "Teléfono", "Email", "Dirección",
                         "Tipo Persona", "Nacionalidad",
                         "Crédito Permitido", "Observaciones", "Fecha Alta"],
            "archivo": "clientes",
        },
        "Ventas": {
            "sql": """
                SELECT v.id, v.fecha,
                       COALESCE(cl.nombre, 'Ocasional'),
                       u.nombre_completo,
                       v.total, v.condicion, v.forma_pago, v.estado,
                       f.nro_factura
                FROM ventas v
                LEFT JOIN clientes cl ON v.cliente_id = cl.id
                LEFT JOIN usuarios u  ON v.usuario_id  = u.id
                LEFT JOIN facturas f  ON f.venta_id    = v.id
                ORDER BY v.id DESC
            """,
            "columnas": ["ID Venta", "Fecha", "Cliente", "Vendedor",
                         "Total", "Condición", "Forma de Pago", "Estado", "Factura"],
            "archivo": "ventas",
        },
        "Detalle de Ventas": {
            "sql": """
                SELECT dv.id, dv.venta_id,
                       COALESCE(p.nombre, dv.descripcion_libre, '(Artículo libre)'),
                       dv.cantidad, dv.precio_unitario,
                       (dv.cantidad * dv.precio_unitario)
                FROM detalle_ventas dv
                LEFT JOIN productos p ON dv.producto_id = p.id
                ORDER BY dv.venta_id DESC, dv.id
            """,
            "columnas": ["ID", "ID Venta", "Producto", "Cantidad",
                         "Precio Unitario", "Importe"],
            "archivo": "detalle_ventas",
        },
        "Inventario": {
            "sql": """
                SELECT p.id,
                       p.nombre,
                       COALESCE(c.nombre, '—') AS categoria,
                       COALESCE(m.nombre, '—') AS marca,
                       p.stock,
                       p.comprometido,
                       (p.stock - p.comprometido) AS disponible,
                       p.stock_minimo,
                       CASE WHEN (p.stock - p.comprometido) < p.stock_minimo
                            THEN 'Bajo stock' ELSE 'OK' END AS estado_stock,
                       p.precio_compra,
                       p.precio AS precio_venta,
                       (p.stock * p.precio_compra) AS valor_inventario,
                       p.unidad_medida,
                       p.control_stock
                FROM productos p
                LEFT JOIN categorias c ON p.categoria_id = c.id
                LEFT JOIN marcas m     ON p.marca_id     = m.id
                WHERE p.tipo_producto != 'Servicio'
                  AND p.control_stock != 'Ilimitado'
                ORDER BY categoria, p.nombre
            """,
            "columnas": ["ID", "Producto", "Categoría", "Marca",
                         "Stock", "Comprometido", "Disponible", "Stock Mínimo",
                         "Estado Stock", "Precio Compra", "Precio Venta",
                         "Valor Inventario", "Unidad de Medida", "Control Stock"],
            "archivo": "inventario",
        },
        "Movimientos Inventario": {
            "sql": """
                SELECT mi.id,
                       mi.fecha,
                       COALESCE(p.nombre, mi.producto_nombre_historico, '(Producto eliminado)') AS producto,
                       mi.tipo,
                       CASE WHEN mi.es_ilimitado = 1 THEN 'Ilimitado'
                            ELSE mi.cantidad END AS cantidad,
                       CASE WHEN mi.es_ilimitado = 1 THEN 'Ilimitado'
                            ELSE mi.stock_resultante END AS stock_resultante,
                       mi.motivo,
                       mi.nro_comprobante,
                       mi.observaciones,
                       COALESCE(u.nombre_completo, '—') AS usuario
                FROM movimientos_inventario mi
                LEFT JOIN productos p ON mi.producto_id = p.id
                LEFT JOIN usuarios u ON mi.usuario_id = u.id
                ORDER BY mi.fecha DESC, mi.id DESC
            """,
            "columnas": ["ID", "Fecha", "Producto", "Tipo",
                         "Cantidad", "Stock Resultante", "Motivo",
                         "Nro. Comprobante", "Observaciones", "Registrado por"],
            "archivo": "movimientos_inventario",
        },
        "Presupuestos": {
            "sql": """
                SELECT pr.id, pr.fecha, pr.cliente_nombre, pr.fecha_validez,
                       pr.estado, pr.total, COALESCE(u.nombre_completo, '—'),
                       pr.venta_id, pr.observaciones
                FROM presupuestos pr
                LEFT JOIN usuarios u ON pr.usuario_id = u.id
                ORDER BY pr.id DESC
            """,
            "columnas": ["ID", "Fecha", "Cliente", "Fecha Validez", "Estado",
                         "Total", "Vendedor", "ID Venta (si se convirtió)", "Observaciones"],
            "archivo": "presupuestos",
        },
        "Detalle Presupuestos": {
            "sql": """
                SELECT dp.id, dp.presupuesto_id,
                       COALESCE(p.nombre, dp.descripcion_libre, '(Artículo libre)'),
                       dp.cantidad, dp.precio_unitario,
                       (dp.cantidad * dp.precio_unitario)
                FROM detalle_presupuestos dp
                LEFT JOIN productos p ON dp.producto_id = p.id
                ORDER BY dp.presupuesto_id DESC, dp.id
            """,
            "columnas": ["ID", "ID Presupuesto", "Producto", "Cantidad",
                         "Precio Unitario", "Importe"],
            "archivo": "detalle_presupuestos",
        },
        "Créditos": {
            "sql": """
                SELECT c.id, c.venta_id, COALESCE(cl.nombre, '—'), c.fecha,
                       c.fecha_vencimiento, c.descripcion, c.deuda_total,
                       c.pagado, (c.deuda_total - c.pagado)
                FROM creditos c
                LEFT JOIN clientes cl ON c.cliente_id = cl.id
                ORDER BY c.id DESC
            """,
            "columnas": ["ID", "ID Venta", "Cliente", "Fecha", "Fecha Vencimiento",
                         "Descripción", "Deuda Total", "Pagado", "Saldo Pendiente"],
            "archivo": "creditos",
        },
        "Pagos de Créditos": {
            "sql": """
                SELECT id, credito_id, fecha, monto
                FROM pagos_credito ORDER BY credito_id DESC, id
            """,
            "columnas": ["ID", "ID Crédito", "Fecha", "Monto Pagado"],
            "archivo": "pagos_credito",
        },
        "Compras": {
            "sql": """
                SELECT co.id, co.fecha_y_hora, co.fecha_compra,
                       COALESCE(pv.nombre, '—'), co.nro_comprobante, co.importe
                FROM compras co
                LEFT JOIN proveedores pv ON co.proveedor_id = pv.id
                ORDER BY co.id DESC
            """,
            "columnas": ["ID", "Fecha y Hora", "Fecha Compra", "Proveedor",
                         "Nro. Comprobante", "Importe"],
            "archivo": "compras",
        },
        "Detalle de Compras": {
            "sql": """
                SELECT dc.id, dc.compra_id,
                       COALESCE(p.nombre, dc.producto_nombre_historico, '(Producto eliminado)'),
                       dc.cantidad, dc.precio_unitario,
                       (dc.cantidad * dc.precio_unitario)
                FROM detalle_compras dc
                LEFT JOIN productos p ON dc.producto_id = p.id
                ORDER BY dc.compra_id DESC, dc.id
            """,
            "columnas": ["ID", "ID Compra", "Producto", "Cantidad",
                         "Precio Unitario", "Importe"],
            "archivo": "detalle_compras",
        },
        "Proveedores": {
            "sql": """
                SELECT id, nombre, contacto, telefono, direccion, ruc
                FROM proveedores ORDER BY nombre
            """,
            "columnas": ["ID", "Nombre", "Contacto", "Teléfono", "Dirección", "RUC"],
            "archivo": "proveedores",
        },
        "Movimientos de Caja": {
            "sql": """
                SELECT cm.id, cm.fecha, cm.tipo, cm.monto, cm.descripcion,
                       COALESCE(u.nombre_completo, '—')
                FROM caja_movimientos cm
                LEFT JOIN usuarios u ON cm.usuario_id = u.id
                ORDER BY cm.fecha DESC, cm.id DESC
            """,
            "columnas": ["ID", "Fecha", "Tipo", "Monto", "Descripción", "Registrado por"],
            "archivo": "movimientos_caja",
        },
        "Devoluciones": {
            "sql": """
                SELECT d.id, d.venta_id,
                       COALESCE(p.nombre, d.producto_nombre_historico, '(Producto eliminado)'),
                       d.cantidad, d.importe, d.fecha, COALESCE(u.nombre_completo, '—')
                FROM devoluciones d
                LEFT JOIN productos p ON d.producto_id = p.id
                LEFT JOIN usuarios u ON d.usuario_id = u.id
                ORDER BY d.fecha DESC, d.id DESC
            """,
            "columnas": ["ID", "ID Venta", "Producto", "Cantidad", "Importe",
                         "Fecha", "Registrado por"],
            "archivo": "devoluciones",
        },
        "Facturas": {
            "sql": """
                SELECT id, venta_id, nro_factura, fecha, razon_social, ruc,
                       valor_total, estado
                FROM facturas ORDER BY id DESC
            """,
            "columnas": ["ID", "ID Venta", "Nro. Factura", "Fecha", "Razón Social",
                         "RUC", "Valor Total", "Estado"],
            "archivo": "facturas",
        },
        "Asistencia Técnica": {
            "sql": """
                SELECT ct.id, ct.fecha_entrada, ct.cliente_nombre, ct.cliente_telefono,
                       COALESCE(te.nombre, ct.tipo_equipo_texto, '—'), ct.nro_serie,
                       ct.descripcion_equipo, ct.prioridad, ct.estado,
                       ct.fecha_estado, ct.fecha_retiro,
                       CASE WHEN ct.anulado = 1 THEN 'Sí' ELSE 'No' END,
                       COALESCE(u.nombre_completo, '—')
                FROM casos_tecnicos ct
                LEFT JOIN tipos_equipo te ON ct.tipo_equipo_id = te.id
                LEFT JOIN usuarios u ON ct.usuario_id = u.id
                ORDER BY ct.id DESC
            """,
            "columnas": ["ID", "Fecha Entrada", "Cliente", "Teléfono", "Tipo de Equipo",
                         "Nro. Serie", "Descripción", "Prioridad", "Estado",
                         "Fecha de Estado", "Fecha de Retiro", "Anulado", "Registrado por"],
            "archivo": "asistencia_tecnica",
        },
        "RRHH - Empleados": {
            "sql": """
                SELECT id, nombre, cargo, departamento, telefono, email,
                       fecha_ingreso, sueldo_mensual, horas_dia,
                       CASE WHEN activo = 1 THEN 'Activo' ELSE 'Inactivo' END,
                       observaciones
                FROM rrhh_empleados ORDER BY nombre
            """,
            "columnas": ["ID", "Nombre", "Cargo", "Departamento", "Teléfono",
                         "Email", "Fecha Ingreso", "Sueldo Mensual", "Horas/Día",
                         "Estado", "Observaciones"],
            "archivo": "rrhh_empleados",
        },
        "RRHH - Asistencia": {
            "sql": """
                SELECT ra.id, COALESCE(e.nombre, '—'), ra.fecha, ra.estado,
                       ra.hora_entrada, ra.hora_salida, ra.observaciones
                FROM rrhh_asistencia ra
                LEFT JOIN rrhh_empleados e ON ra.empleado_id = e.id
                ORDER BY ra.fecha DESC, ra.id DESC
            """,
            "columnas": ["ID", "Empleado", "Fecha", "Estado", "Hora Entrada",
                         "Hora Salida", "Observaciones"],
            "archivo": "rrhh_asistencia",
        },
        "RRHH - Adelantos": {
            "sql": """
                SELECT ad.id, COALESCE(e.nombre, '—'), ad.fecha, ad.monto,
                       ad.descripcion, ad.estado, ad.registrado_por, ad.fecha_descuento
                FROM rrhh_adelantos ad
                LEFT JOIN rrhh_empleados e ON ad.empleado_id = e.id
                ORDER BY ad.fecha DESC, ad.id DESC
            """,
            "columnas": ["ID", "Empleado", "Fecha", "Monto", "Descripción",
                         "Estado", "Registrado por", "Fecha de Descuento"],
            "archivo": "rrhh_adelantos",
        },
        "Mascotas (Veterinaria)": {
            "sql": """
                SELECT m.id, COALESCE(cl.nombre, m.dueño_nombre, '—'), m.dueño_telefono,
                       m.nombre, COALESCE(es.nombre, m.especie_texto, '—'), m.raza,
                       m.sexo, m.fecha_nacimiento, m.peso_kg,
                       CASE WHEN m.esterilizado = 1 THEN 'Sí' ELSE 'No' END,
                       CASE WHEN m.fallecido = 1 THEN 'Sí' ELSE 'No' END,
                       m.observaciones
                FROM mascotas m
                LEFT JOIN clientes cl ON m.cliente_id = cl.id
                LEFT JOIN especies_mascota es ON m.especie_id = es.id
                ORDER BY m.nombre
            """,
            "columnas": ["ID", "Dueño", "Teléfono Dueño", "Nombre Mascota",
                         "Especie", "Raza", "Sexo", "Fecha Nacimiento",
                         "Peso (kg)", "Esterilizado", "Fallecido", "Observaciones"],
            "archivo": "veterinaria_mascotas",
        },
        "Consultas Veterinarias": {
            "sql": """
                SELECT cv.id, m.nombre, cv.fecha, cv.motivo, cv.diagnostico,
                       cv.tratamiento_indicado, cv.peso_kg, cv.temperatura,
                       cv.proxima_visita, cv.costo, COALESCE(u.nombre_completo, '—')
                FROM consultas_veterinarias cv
                LEFT JOIN mascotas m ON cv.mascota_id = m.id
                LEFT JOIN usuarios u ON cv.usuario_id = u.id
                ORDER BY cv.fecha DESC, cv.id DESC
            """,
            "columnas": ["ID", "Mascota", "Fecha", "Motivo", "Diagnóstico",
                         "Tratamiento Indicado", "Peso (kg)", "Temperatura",
                         "Próxima Visita", "Costo", "Atendido por"],
            "archivo": "veterinaria_consultas",
        },
        "Vacunas (Veterinaria)": {
            "sql": """
                SELECT v.id, m.nombre, v.vacuna, v.fecha_aplicacion,
                       v.proxima_dosis, v.lote, v.veterinario, v.observaciones
                FROM vacunas_mascota v
                LEFT JOIN mascotas m ON v.mascota_id = m.id
                ORDER BY v.fecha_aplicacion DESC, v.id DESC
            """,
            "columnas": ["ID", "Mascota", "Vacuna", "Fecha Aplicación",
                         "Próxima Dosis", "Lote", "Veterinario", "Observaciones"],
            "archivo": "veterinaria_vacunas",
        },
        "Préstamos": {
            "sql": """
                SELECT p.id, COALESCE(cl.nombre, '—'), p.fecha_desembolso,
                       p.capital, p.tasa_interes, p.frecuencia, p.cantidad_cuotas,
                       p.sistema, p.tasa_mora_diaria, p.estado, p.observaciones,
                       COALESCE(u.nombre_completo, '—')
                FROM prestamos p
                LEFT JOIN clientes cl ON p.cliente_id = cl.id
                LEFT JOIN usuarios u ON p.usuario_id = u.id
                ORDER BY p.id DESC
            """,
            "columnas": ["ID", "Cliente", "Fecha Desembolso", "Capital",
                         "Tasa de Interés", "Frecuencia", "Cant. Cuotas",
                         "Sistema", "Tasa Mora Diaria", "Estado",
                         "Observaciones", "Otorgado por"],
            "archivo": "prestamos",
        },
        "Cuotas de Préstamos": {
            "sql": """
                SELECT c.id, c.prestamo_id, c.nro_cuota, c.fecha_vencimiento,
                       c.capital, c.interes, c.pagado_capital, c.pagado_interes,
                       c.pagado_mora, c.fecha_pago_completo
                FROM cuotas_prestamo c
                ORDER BY c.prestamo_id DESC, c.nro_cuota
            """,
            "columnas": ["ID", "ID Préstamo", "Nro. Cuota", "Fecha Vencimiento",
                         "Capital", "Interés", "Capital Pagado", "Interés Pagado",
                         "Mora Pagada", "Fecha Pago Completo"],
            "archivo": "cuotas_prestamo",
        },
        "Movimientos del Fondo de Préstamos": {
            "sql": """
                SELECT f.id, f.fecha, f.tipo, f.monto, f.descripcion,
                       f.prestamo_id, COALESCE(u.nombre_completo, '—'), f.saldo_resultante
                FROM fondo_prestamos_movimientos f
                LEFT JOIN usuarios u ON f.usuario_id = u.id
                ORDER BY f.id DESC
            """,
            "columnas": ["ID", "Fecha", "Tipo", "Monto", "Descripción",
                         "ID Préstamo", "Registrado por", "Saldo Resultante"],
            "archivo": "fondo_prestamos_movimientos",
        },
        "Comandas (Restaurante)": {
            "sql": """
                SELECT c.id, COALESCE(mesa.numero, '—'), c.tipo,
                       COALESCE(cl.nombre, '—'), COALESCE(u.nombre_completo, '—'),
                       c.estado, c.turno, c.fecha_apertura, c.fecha_cierre,
                       c.venta_id, c.observaciones
                FROM rest_comandas c
                LEFT JOIN rest_mesas mesa ON c.mesa_id = mesa.id
                LEFT JOIN clientes cl ON c.cliente_id = cl.id
                LEFT JOIN usuarios u ON c.mozo_usuario_id = u.id
                ORDER BY c.id DESC
            """,
            "columnas": ["ID", "Mesa", "Tipo", "Cliente", "Mozo/Atendió",
                         "Estado", "Turno", "Fecha Apertura", "Fecha Cierre",
                         "ID Venta", "Observaciones"],
            "archivo": "restaurante_comandas",
        },
        "Platos (Restaurante)": {
            "sql": """
                SELECT id, nombre, categoria, precio_venta, descripcion,
                       tiempo_preparacion_min,
                       CASE WHEN activo = 1 THEN 'Activo' ELSE 'Inactivo' END
                FROM rest_platos ORDER BY nombre
            """,
            "columnas": ["ID", "Nombre", "Categoría", "Precio Venta",
                         "Descripción", "Tiempo Preparación (min)", "Estado"],
            "archivo": "restaurante_platos",
        },
        "Cuentas de Streaming": {
            "sql": """
                SELECT c.id, pl.nombre, c.email, c.plan_nombre, c.max_perfiles,
                       c.costo_mensual, c.fecha_compra, c.fecha_ultimo_cambio_password,
                       c.estado, c.notas
                FROM stream_cuentas c
                LEFT JOIN stream_plataformas pl ON c.plataforma_id = pl.id
                ORDER BY pl.nombre, c.id
            """,
            "columnas": ["ID", "Plataforma", "Email", "Plan", "Máx. Perfiles",
                         "Costo Mensual", "Fecha Compra",
                         "Último Cambio Contraseña", "Estado", "Notas"],
            "archivo": "streaming_cuentas",
        },
        "Suscripciones de Streaming": {
            "sql": """
                SELECT s.id, COALESCE(cl.nombre, '—'), COALESCE(co.nombre, '—'),
                       s.modalidad, s.fecha_inicio, s.fecha_vencimiento,
                       s.precio_mensual, s.estado, s.forma_pago,
                       s.dispositivos_conectados, s.max_dispositivos, s.notas
                FROM stream_suscripciones s
                LEFT JOIN clientes cl ON s.cliente_id = cl.id
                LEFT JOIN stream_combos co ON s.combo_id = co.id
                ORDER BY s.id DESC
            """,
            "columnas": ["ID", "Cliente", "Combo", "Modalidad", "Fecha Inicio",
                         "Fecha Vencimiento", "Precio Mensual", "Estado",
                         "Forma de Pago", "Dispositivos Conectados",
                         "Máx. Dispositivos", "Notas"],
            "archivo": "streaming_suscripciones",
        },
        "Importaciones - Compras": {
            "sql": """
                SELECT c.id, c.plataforma, c.referencia, COALESCE(co.nombre, '—'),
                       c.casillero, c.tipo_envio, c.peso_caja_kg, c.costo_envio_total,
                       c.estado, c.fecha_compra, c.fecha_envio_casillero, c.fecha_recepcion, c.notas
                FROM import_compras c
                LEFT JOIN import_couriers co ON c.courier_id = co.id
                ORDER BY c.id DESC
            """,
            "columnas": ["ID", "Plataforma", "Referencia", "Courier", "Casillero",
                         "Tipo de Envío", "Peso Caja (kg)", "Costo Envío Total (US$)",
                         "Estado", "Fecha Compra", "Fecha Envío a Casillero",
                         "Fecha Recepción", "Notas"],
            "archivo": "importaciones_compras",
        },
        "Importaciones - Detalle de Productos": {
            "sql": """
                SELECT d.id, d.compra_id, d.producto_nombre, d.cantidad,
                       d.costo_unitario_compra, d.costo_envio_unitario, d.costo_total_unitario,
                       d.precio_venta_publico,
                       CASE WHEN d.enviado_inventario = 1 THEN 'Sí' ELSE 'No' END
                FROM import_detalle d
                ORDER BY d.compra_id DESC, d.id
            """,
            "columnas": ["ID", "ID Compra", "Producto", "Cantidad",
                         "Costo Compra/U (US$)", "Costo Envío/U (US$)",
                         "Costo Total/U (US$)", "Precio Venta Público (US$)",
                         "Enviado a Inventario"],
            "archivo": "importaciones_detalle",
        },
        "Couriers de Importación": {
            "sql": """
                SELECT id, nombre, email, ruc, telefono, costo_kg_aereo, costo_kg_maritimo,
                       direccion_casillero_miami, direccion_casillero_shenzhen,
                       CASE WHEN activo = 1 THEN 'Activo' ELSE 'Inactivo' END, notas
                FROM import_couriers ORDER BY nombre
            """,
            "columnas": ["ID", "Nombre", "Email", "RUC", "Teléfono",
                         "Costo/kg Aéreo (US$)", "Costo/kg Marítimo (US$)",
                         "Dirección Casillero Miami", "Dirección Casillero Shenzhen",
                         "Estado", "Notas"],
            "archivo": "importaciones_couriers",
        },
    }

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Exportar a CSV")
        self.minsize(460, 420)
        self.geometry("460x560")
        self.resizable(True, True)
        self.configure(bg=BLANCO)
        self.grab_set()
        self._construir_ui()

    def _construir_ui(self):
        barra = tk.Frame(self, bg=AZUL_RIBBON, height=36)
        barra.pack(fill="x")
        barra.pack_propagate(False)
        tk.Label(barra, text="Exportar a CSV",
                 font=("Segoe UI", 11, "bold"), bg=AZUL_RIBBON, fg=BLANCO
                 ).pack(side="left", padx=16, pady=7)

        cuerpo = tk.Frame(self, bg=BLANCO, padx=20, pady=16)
        cuerpo.pack(fill="both", expand=True)

        fila_titulo = tk.Frame(cuerpo, bg=BLANCO)
        fila_titulo.pack(fill="x", pady=(0, 6))
        tk.Label(fila_titulo, text="Selecciona las tablas a exportar:",
                 font=("Segoe UI", 10, "bold"), bg=BLANCO).pack(side="left")
        tk.Label(fila_titulo, text="Todas", font=("Segoe UI", 9, "underline"),
                 bg=BLANCO, fg=AZUL_RIBBON, cursor="hand2").pack(side="right", padx=(8, 0))
        tk.Label(fila_titulo, text="Ninguna", font=("Segoe UI", 9, "underline"),
                 bg=BLANCO, fg=AZUL_RIBBON, cursor="hand2").pack(side="right")

        # Lista de tablas con scroll propio, porque ahora son bastantes
        # (todos los módulos del sistema) y no siempre entran de una.
        contenedor_lista = tk.Frame(cuerpo, bg=BLANCO, relief="solid", bd=1, height=200)
        contenedor_lista.pack(fill="both", expand=True, pady=(0, 10))
        contenedor_lista.pack_propagate(False)

        canvas_tablas = tk.Canvas(contenedor_lista, bg=BLANCO, highlightthickness=0, bd=0)
        sb_tablas = tk.Scrollbar(contenedor_lista, orient="vertical", command=canvas_tablas.yview)
        canvas_tablas.configure(yscrollcommand=sb_tablas.set)
        sb_tablas.pack(side="right", fill="y")
        canvas_tablas.pack(side="left", fill="both", expand=True)

        frame_lista = tk.Frame(canvas_tablas, bg=BLANCO)
        canvas_tablas.create_window((0, 0), window=frame_lista, anchor="nw")
        frame_lista.bind("<Configure>",
                         lambda e: canvas_tablas.configure(scrollregion=canvas_tablas.bbox("all")))

        def _scroll_lista(event):
            if event.num == 4:
                canvas_tablas.yview_scroll(-1, "units")
            elif event.num == 5:
                canvas_tablas.yview_scroll(1, "units")
            else:
                canvas_tablas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        for secuencia in ("<MouseWheel>", "<Button-4>", "<Button-5>"):
            contenedor_lista.bind(secuencia, _scroll_lista)
            canvas_tablas.bind(secuencia, _scroll_lista)
            frame_lista.bind(secuencia, _scroll_lista)

        self.vars_tablas = {}
        for nombre in self.TABLAS:
            var = tk.BooleanVar(value=True)
            self.vars_tablas[nombre] = var
            tk.Checkbutton(frame_lista, text=nombre, variable=var,
                           font=("Segoe UI", 10), bg=BLANCO, anchor="w"
                           ).pack(anchor="w", pady=2, fill="x")

        fila_titulo.winfo_children()[1].bind(
            "<Button-1>", lambda e: [v.set(True) for v in self.vars_tablas.values()])
        fila_titulo.winfo_children()[2].bind(
            "<Button-1>", lambda e: [v.set(False) for v in self.vars_tablas.values()])

        tk.Label(cuerpo, text="Separador de columnas:",
                 font=("Segoe UI", 10, "bold"), bg=BLANCO).pack(anchor="w")
        self.var_sep = tk.StringVar(value=",")
        f_sep = tk.Frame(cuerpo, bg=BLANCO)
        f_sep.pack(anchor="w", pady=4)
        for etiq, val in [("Coma  ( , )", ","), ("Punto y coma  ( ; )", ";"),
                           ("Tabulación  ( TAB )", "\t")]:
            tk.Radiobutton(f_sep, text=etiq, variable=self.var_sep, value=val,
                           font=("Segoe UI", 9), bg=BLANCO).pack(anchor="w")

        tk.Frame(cuerpo, bg=GRIS_BORDE, height=1).pack(fill="x", pady=12)

        f_bot = tk.Frame(cuerpo, bg=BLANCO)
        f_bot.pack(fill="x")
        tk.Button(f_bot, text="📁 Elegir carpeta y exportar",
                  font=("Segoe UI", 10, "bold"),
                  bg=VERDE, fg=BLANCO, relief="flat", padx=14, pady=8,
                  cursor="hand2", command=self._ejecutar).pack(side="left", padx=(0, 10))
        tk.Button(f_bot, text="Cancelar", font=("Segoe UI", 10),
                  bg=BLANCO, fg=ROJO, relief="solid", bd=1, padx=14, pady=8,
                  cursor="hand2", command=self.destroy).pack(side="left")

    def _ejecutar(self):

        tablas = [n for n, v in self.vars_tablas.items() if v.get()]
        if not tablas:
            messagebox.showwarning("Sin selección",
                                   "Selecciona al menos una tabla.", parent=self)
            return

        carpeta = filedialog.askdirectory(title="Elegir carpeta de destino")
        if not carpeta:
            return

        sep = self.var_sep.get()
        ahora = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        archivos_generados = []
        errores = []

        # Garantiza que existan las tablas de RRHH aunque el administrador
        # todavía no haya abierto ese módulo, para que su exportación no
        # falle con un error de "tabla no encontrada".
        inicializar_tablas_rrhh()

        conn = conectar()
        cursor = conn.cursor()

        for nombre in tablas:
            cfg = self.TABLAS[nombre]
            nombre_archivo = f"{cfg['archivo']}_{ahora}.csv"
            ruta = os.path.join(carpeta, nombre_archivo)
            try:
                cursor.execute(cfg["sql"])
                filas = cursor.fetchall()
                with open(ruta, "w", newline="", encoding="utf-8-sig") as f:
                    # utf-8-sig agrega BOM para que Excel abra correctamente
                    writer = csv.writer(f, delimiter=sep)
                    writer.writerow(cfg["columnas"])
                    writer.writerows(filas)
                archivos_generados.append(nombre_archivo)
            except Exception as e:
                errores.append(f"{nombre}: {e}")

        conn.close()

        if archivos_generados:
            msg = f"✔ {len(archivos_generados)} archivo(s) generado(s) en:\n{carpeta}\n\n"
            msg += "\n".join(f"  • {a}" for a in archivos_generados)
            if errores:
                msg += "\n\n⚠ Errores:\n" + "\n".join(errores)
            messagebox.showinfo("Exportación completada", msg, parent=self)
            self.destroy()
        else:
            messagebox.showerror("Error", "No se pudo generar ningún archivo.\n\n" +
                                 "\n".join(errores), parent=self)


# ─────────────────────────────────────────────────────────────
#  VENTANA IMPORTAR CSV
# ─────────────────────────────────────────────────────────────
class _VentanaImportarCSV(tk.Toplevel):
    """Detecta el separador, muestra una vista previa de las primeras
    filas y permite elegir si importar Productos o Clientes."""

    def __init__(self, parent, ruta_csv):
        super().__init__(parent)
        self.ruta_csv = ruta_csv
        self.title("Importar desde CSV")
        self.minsize(520, 400)
        self.resizable(True, True)
        self.configure(bg=BLANCO)
        self.grab_set()

        self.sep_detectado, self.filas_preview, self.encabezados = \
            self._detectar_y_previsualizar(ruta_csv)

        if self.encabezados is None:
            messagebox.showerror("Error",
                                 "No se pudo leer el archivo CSV.", parent=self)
            self.destroy()
            return

        self._construir_ui()

    def _detectar_y_previsualizar(self, ruta):
        """Detecta el separador (coma o punto y coma) leyendo las primeras líneas."""
        for sep in (",", ";", "\t"):
            try:
                with open(ruta, newline="", encoding="utf-8-sig") as f:
                    reader = csv.reader(f, delimiter=sep)
                    filas = [row for row in reader if any(c.strip() for c in row)]
                if len(filas) > 0 and len(filas[0]) > 1:
                    return sep, filas[1:6], filas[0]  # sep, preview, encabezados
            except Exception:
                continue
        # Fallback: una sola columna
        try:
            with open(ruta, newline="", encoding="utf-8-sig") as f:
                import csv
                reader = csv.reader(f, delimiter=",")
                filas = list(reader)
            return ",", filas[1:6], filas[0] if filas else None
        except Exception:
            return ",", [], None

    def _construir_ui(self):

        barra = tk.Frame(self, bg=AZUL_RIBBON, height=36)
        barra.pack(fill="x")
        barra.pack_propagate(False)
        tk.Label(barra, text="Importar desde CSV",
                 font=("Segoe UI", 11, "bold"), bg=AZUL_RIBBON, fg=BLANCO
                 ).pack(side="left", padx=16, pady=7)

        cuerpo = tk.Frame(self, bg=BLANCO, padx=20, pady=14)
        cuerpo.pack(fill="both", expand=True)

        # Archivo seleccionado
        nombre_archivo = os.path.basename(self.ruta_csv)
        tk.Label(cuerpo, text=f"📄  {nombre_archivo}",
                 font=("Segoe UI", 9), bg=BLANCO,
                 fg="#475569").pack(anchor="w", pady=(0, 10))

        # Separador detectado
        sep_txt = {"," : "Coma ( , )",
                   ";" : "Punto y coma ( ; )",
                   "\t": "Tabulación ( TAB )"}.get(self.sep_detectado, self.sep_detectado)
        tk.Label(cuerpo, text=f"Separador detectado: {sep_txt}",
                 font=("Segoe UI", 9), bg=BLANCO,
                 fg="#16a34a").pack(anchor="w", pady=(0, 8))

        # Vista previa
        tk.Label(cuerpo, text="Vista previa (primeras 5 filas):",
                 font=("Segoe UI", 9, "bold"), bg=BLANCO).pack(anchor="w")

        from tkinter import ttk
        frame_prev = tk.Frame(cuerpo, bg=BLANCO)
        frame_prev.pack(fill="x", pady=(4, 12))

        cols_id = [f"c{i}" for i in range(len(self.encabezados))]
        tree_prev = ttk.Treeview(frame_prev, columns=cols_id,
                                  show="headings", height=5)
        habilitar_deseleccion_treeview(tree_prev)
        for cid, enc in zip(cols_id, self.encabezados):
            tree_prev.heading(cid, text=enc)
            tree_prev.column(cid, width=max(80, len(enc) * 8), minwidth=50)
        for fila in self.filas_preview:
            tree_prev.insert("", "end", values=fila)

        sb_prev = ttk.Scrollbar(frame_prev, orient="horizontal",
                                 command=tree_prev.xview)
        tree_prev.configure(xscrollcommand=sb_prev.set)
        tree_prev.pack(fill="x")
        sb_prev.pack(fill="x")

        tk.Frame(cuerpo, bg=GRIS_BORDE, height=1).pack(fill="x", pady=8)

        # Tipo de dato
        tk.Label(cuerpo, text="¿Qué tipo de datos contiene este archivo?",
                 font=("Segoe UI", 10, "bold"), bg=BLANCO).pack(anchor="w")
        self.var_tipo = tk.StringVar(value="Productos")
        for val in ("Productos", "Clientes"):
            tk.Radiobutton(cuerpo, text=val, variable=self.var_tipo,
                           value=val, font=("Segoe UI", 9),
                           bg=BLANCO).pack(anchor="w")

        tk.Frame(cuerpo, bg=GRIS_BORDE, height=1).pack(fill="x", pady=8)

        # Duplicados
        tk.Label(cuerpo, text="Si ya existe un registro con el mismo nombre:",
                 font=("Segoe UI", 10, "bold"), bg=BLANCO).pack(anchor="w")
        self.var_dup = tk.StringVar(value="omitir")
        for val, etiq in [("omitir",     "Omitir (no modificar el existente)"),
                           ("actualizar", "Actualizar con los datos del CSV")]:
            tk.Radiobutton(cuerpo, text=etiq, variable=self.var_dup,
                           value=val, font=("Segoe UI", 9),
                           bg=BLANCO).pack(anchor="w")

        tk.Frame(cuerpo, bg=GRIS_BORDE, height=1).pack(fill="x", pady=10)

        f_bot = tk.Frame(cuerpo, bg=BLANCO)
        f_bot.pack(fill="x")
        tk.Button(f_bot, text="✔ Importar",
                  font=("Segoe UI", 10, "bold"),
                  bg=VERDE, fg=BLANCO, relief="flat", padx=14, pady=8,
                  cursor="hand2", command=self._ejecutar).pack(side="left", padx=(0, 10))
        tk.Button(f_bot, text="Cancelar",
                  font=("Segoe UI", 10),
                  bg=BLANCO, fg=ROJO, relief="solid", bd=1, padx=14, pady=8,
                  cursor="hand2", command=self.destroy).pack(side="left")

    def _ejecutar(self):
        tipo = self.var_tipo.get()
        modo = self.var_dup.get()

        try:
            with open(self.ruta_csv, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f, delimiter=self.sep_detectado)
                filas = list(reader)
        except Exception as e:
            messagebox.showerror("Error al leer el archivo", str(e), parent=self)
            return

        conn = conectar()
        try:
            if tipo == "Productos":
                ins, act, errores = _importar_csv_productos(conn, filas, modo)
            else:
                ins, act, errores = _importar_csv_clientes(conn, filas, modo)
            conn.commit()
        except Exception as e:
            conn.rollback()
            messagebox.showerror("Error en importación", str(e), parent=self)
            conn.close()
            return
        finally:
            conn.close()

        msg = f"✔ Importación completada.\n\n{tipo}: {ins} insertados, {act} actualizados."
        if errores:
            msg += f"\n\n⚠ {len(errores)} fila(s) con errores (omitidas)."
        messagebox.showinfo("Resultado", msg, parent=self)
        self.destroy()


# ─────────────────────────────────────────────────────────────
#  HELPERS CSV
# ─────────────────────────────────────────────────────────────
def _norm(encabezados: dict, *posibles) -> str | None:
    """Busca la primera clave que coincida (sin importar mayúsculas/tildes)."""
    for clave in encabezados:
        clave_n = clave.lower().strip()
        for p in posibles:
            if p.lower() in clave_n:
                return encabezados[clave]
    return None


def _val(fila: dict, *posibles):
    """Devuelve el valor de la primera columna que coincida."""
    for clave in fila:
        clave_n = clave.lower().strip()
        for p in posibles:
            if p.lower() in clave_n:
                v = fila[clave]
                return str(v).strip() if v is not None else ""
    return ""


def _float_csv(fila, *posibles, default=0.0):
    try:
        return float(_val(fila, *posibles).replace(",", ".").replace("Gs.", "").strip() or default)
    except (ValueError, TypeError):
        return default


def _importar_csv_productos(conn, filas: list[dict], modo: str):
    cursor = conn.cursor()
    insertados = actualizados = 0
    errores = []

    for i, fila in enumerate(filas, start=2):
        nombre = _val(fila, "nombre", "name", "descripcion", "producto")
        if not nombre:
            continue
        try:
            codigo   = _val(fila, "código", "codigo", "barras", "code")
            compra   = _float_csv(fila, "compra", "costo", "cost")
            venta    = _float_csv(fila, "venta", "precio", "price", "sale")
            mayor    = _float_csv(fila, "mayorista", "mayor", "wholesale") or venta
            stock    = int(_float_csv(fila, "stock", "cantidad", "qty"))
            tipo     = _val(fila, "tipo", "type") or "Producto"
            control  = _val(fila, "control", "stock control") or "Normal"

            cursor.execute("SELECT id FROM productos WHERE nombre = ?", (nombre,))
            existe = cursor.fetchone()

            if existe:
                if modo == "actualizar":
                    cursor.execute("""
                        UPDATE productos SET
                            codigo_barras=?, precio_compra=?, precio=?,
                            precio_mayorista=?, stock=?, tipo_producto=?, control_stock=?
                        WHERE nombre=?
                    """, (codigo, compra, venta, mayor, stock, tipo, control, nombre))
                    actualizados += 1
            else:
                cursor.execute("""
                    INSERT INTO productos
                        (nombre, codigo_barras, precio_compra, precio,
                         precio_mayorista, stock, tipo_producto, control_stock, activo)
                    VALUES (?,?,?,?,?,?,?,?,1)
                """, (nombre, codigo, compra, venta, mayor, stock, tipo, control))
                insertados += 1
        except Exception as e:
            errores.append(f"Fila {i}: {e}")

    return insertados, actualizados, errores


def _importar_csv_clientes(conn, filas: list[dict], modo: str):
    cursor = conn.cursor()
    insertados = actualizados = 0
    errores = []

    for i, fila in enumerate(filas, start=2):
        nombre = _val(fila, "nombre", "name", "cliente")
        if not nombre:
            continue
        try:
            razon  = _val(fila, "razón", "razon", "social", "empresa") or nombre
            doc    = _val(fila, "documento", "ci", "doc", "nro. documento")
            ruc    = _val(fila, "ruc")
            tel    = _val(fila, "teléfono", "telefono", "phone", "cel")
            email  = _val(fila, "email", "correo", "mail")
            direc  = _val(fila, "dirección", "direccion", "address")
            obs    = _val(fila, "observaciones", "obs", "notas", "notes")

            cursor.execute("SELECT id FROM clientes WHERE nombre = ?", (nombre,))
            existe = cursor.fetchone()

            if existe:
                if modo == "actualizar":
                    cursor.execute("""
                        UPDATE clientes SET
                            razon_social=?, nro_documento=?, ruc=?,
                            telefono=?, email=?, direccion=?, observaciones=?
                        WHERE nombre=?
                    """, (razon, doc, ruc, tel, email, direc, obs, nombre))
                    actualizados += 1
            else:
                cursor.execute("""
                    INSERT INTO clientes
                        (nombre, razon_social, nro_documento, ruc,
                         telefono, email, direccion, observaciones)
                    VALUES (?,?,?,?,?,?,?,?)
                """, (nombre, razon, doc, ruc, tel, email, direc, obs))
                insertados += 1
        except Exception as e:
            errores.append(f"Fila {i}: {e}")

    return insertados, actualizados, errores