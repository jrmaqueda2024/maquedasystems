from utilidades_ui import formatear_gs, habilitar_deseleccion_treeview
from traducciones import t
"""
ventana_reportes.py
Módulo de Reportes: registro completo de todas las ventas de todos los
usuarios con filtros por rango de fechas, vendedor, cliente y estado.
Incluye tarjetas de resumen financiero, tabla de ventas con búsqueda y
panel lateral de detalle de cada venta.
"""
import tkinter as tk
from tkinter import ttk
import datetime

from database import conectar

from widget_calendario import abrir_selector_fecha, abrir_selector_rango_fechas
from menu_reporte_general import BotonReporteGeneral
from auth import filtro_usuario_ventas

AZUL        = "#1d5fd6"
AZUL_OSC    = "#163d8c"
GRIS_FONDO  = "#f4f5f7"
GRIS_BORDE  = "#e2e8f0"
BLANCO      = "#ffffff"
VERDE       = "#16a34a"
ROJO        = "#dc2626"
NARANJA     = "#d97706"
GRIS_TEXTO  = "#6b7280"
NEGRO       = "#1e293b"


# ─────────────────────────────────────────────────────────────
#  HELPERS DE BD
# ─────────────────────────────────────────────────────────────
def _listar_todos_los_vendedores() -> list[dict]:
    conn = conectar()
    c = conn.cursor()
    c.execute("SELECT id, nombre_completo FROM usuarios WHERE activo=1 ORDER BY nombre_completo")
    filas = c.fetchall()
    conn.close()
    return [{"id": f[0], "nombre": f[1]} for f in filas]


def _listar_ventas_filtradas(fecha_desde: str, fecha_hasta: str,
                              vendedor_id: int | None = None,
                              estado: str | None = None,
                              busqueda: str = "") -> list[dict]:
    conn = conectar()
    c = conn.cursor()
    sql = """
        SELECT v.id, v.fecha,
               COALESCE(cl.nombre, 'Ocasional') AS cliente,
               u.nombre_completo AS vendedor,
               v.total, v.condicion, v.forma_pago, v.estado,
               f.nro_factura
        FROM ventas v
        LEFT JOIN clientes cl ON v.cliente_id = cl.id
        LEFT JOIN usuarios  u  ON v.usuario_id  = u.id
        LEFT JOIN facturas  f  ON f.venta_id    = v.id
        WHERE date(v.fecha) BETWEEN ? AND ?
    """
    params: list = [fecha_desde, fecha_hasta]

    if vendedor_id:
        sql += " AND v.usuario_id = ?"
        params.append(vendedor_id)
    if estado and estado != "Todos":
        sql += " AND v.estado = ?"
        params.append(estado)
    if busqueda:
        sql += """ AND (
            COALESCE(cl.nombre,'') LIKE ?
            OR u.nombre_completo LIKE ?
            OR CAST(v.id AS TEXT) LIKE ?
            OR COALESCE(f.nro_factura,'') LIKE ?
        )"""
        q = f"%{busqueda}%"
        params += [q, q, q, q]

    sql += " ORDER BY v.fecha DESC"
    c.execute(sql, params)
    filas = c.fetchall()
    conn.close()
    return [
        {
            "id": f[0], "fecha": f[1], "cliente": f[2], "vendedor": f[3],
            "total": f[4] or 0, "condicion": f[5], "forma_pago": f[6],
            "estado": f[7], "factura": f[8] or "",
        }
        for f in filas
    ]


def _resumen_de_ventas(ventas: list[dict]) -> dict:
    validas   = [v for v in ventas if v["estado"] != "Cancelado"]
    canceladas = [v for v in ventas if v["estado"] == "Cancelado"]
    total      = sum(v["total"] for v in validas)
    efectivo   = sum(v["total"] for v in validas if v["forma_pago"] == "Efectivo")
    transferencia = sum(v["total"] for v in validas if v["forma_pago"] == "Transferencia Bancaria")
    cripto     = sum(v["total"] for v in validas if v["forma_pago"] == "Criptomonedas")
    tarjeta    = sum(v["total"] for v in validas
                     if v["forma_pago"] not in ("Efectivo", "Transferencia Bancaria", "Crédito", "Criptomonedas"))
    credito    = sum(v["total"] for v in validas if v["forma_pago"] == "Crédito")
    return {
        "cantidad": len(validas),
        "canceladas": len(canceladas),
        "total": total,
        "efectivo": efectivo,
        "transferencia": transferencia,
        "tarjeta": tarjeta,
        "credito": credito,
        "cripto": cripto,
    }


def _obtener_detalle_venta_con_libres(venta_id: int) -> dict | None:
    """Como obtener_detalle_venta pero también incluye líneas de venta libre."""
    conn = conectar()
    c = conn.cursor()
    c.execute("""
        SELECT v.id, v.fecha, v.total, v.condicion, v.forma_pago, v.estado,
               COALESCE(cl.nombre, 'Ocasional'), COALESCE(cl.nro_documento, ''),
               f.nro_factura, u.nombre_completo
        FROM ventas v
        LEFT JOIN clientes cl ON v.cliente_id = cl.id
        LEFT JOIN facturas f  ON f.venta_id   = v.id
        LEFT JOIN usuarios u  ON v.usuario_id  = u.id
        WHERE v.id = ?
    """, (venta_id,))
    fila = c.fetchone()
    if not fila:
        conn.close()
        return None

    c.execute("""
        SELECT dv.id, dv.producto_id,
               COALESCE(p.nombre, dv.descripcion_libre, '(Artículo libre)') AS nombre,
               dv.cantidad, dv.precio_unitario,
               COALESCE(dv.cantidad_devuelta, 0)
        FROM detalle_ventas dv
        LEFT JOIN productos p ON dv.producto_id = p.id
        WHERE dv.venta_id = ?
        ORDER BY dv.id
    """, (venta_id,))
    lineas = []
    for f in c.fetchall():
        cant_activa = f[3] - f[5]
        lineas.append({
            "detalle_id": f[0], "producto_id": f[1], "nombre_producto": f[2],
            "cantidad": f[3], "precio_unitario": f[4], "cantidad_devuelta": f[5],
            "cantidad_activa": cant_activa,
            "importe": cant_activa * f[4],
            "es_libre": f[1] is None,
        })
    conn.close()
    return {
        "id": fila[0], "fecha": fila[1], "total": fila[2], "condicion": fila[3],
        "forma_pago": fila[4], "estado": fila[5], "cliente_nombre": fila[6],
        "cliente_documento": fila[7], "nro_factura": fila[8] or "",
        "vendedor": fila[9] or "", "lineas": lineas,
    }


# ─────────────────────────────────────────────────────────────
#  PANEL PRINCIPAL
# ─────────────────────────────────────────────────────────────
class _PanelReportesVentas(tk.Frame):
    def __init__(self, parent, usuario_actual):
        super().__init__(parent, bg=BLANCO)
        self.usuario_actual = usuario_actual

        hoy = datetime.date.today()
        self.fecha_desde = hoy.replace(day=1)   # primer día del mes
        self.fecha_hasta = hoy

        self.ventas_actuales: list[dict] = []
        self.venta_detalle_id: int | None = None

        self._construir_encabezado()
        self._construir_filtros()
        self._construir_tarjetas()
        self._construir_cuerpo()
        self._cargar()

    # ── Encabezado ───────────────────────────────────────────
    def _construir_encabezado(self):
        enc = tk.Frame(self, bg=AZUL, height=54)
        enc.pack(fill="x")
        enc.pack_propagate(False)
        tk.Label(enc, text=t("reportes_titulo"),
                 font=("Segoe UI", 15, "bold"), bg=AZUL, fg=BLANCO
                 ).pack(side="left", padx=20, pady=12)

        self.boton_reporte = BotonReporteGeneral(
            enc, obtener_datos_callback=self._obtener_datos_reporte,
            nombre_archivo_base="Reporte_General_Ventas",
        )
        self.boton_reporte.generador_pdf_dashboard = self._generar_pdf_dashboard
        self.boton_reporte.generador_excel = self._generar_excel
        self.boton_reporte.confirmar_antes_de_dashboard = self._confirmar_dashboard_sin_filtros
        self.boton_reporte.pack(side="right", padx=(0, 12), pady=12)

        tk.Button(enc, text=f"🔄 {t('actualizar')}", font=("Segoe UI", 9),
                  bg=AZUL_OSC, fg=BLANCO, relief="flat", padx=12, pady=4,
                  cursor="hand2", command=self._cargar
                  ).pack(side="right", padx=16, pady=12)

    # ── Barra de filtros ─────────────────────────────────────
    def _construir_filtros(self):
        barra = tk.Frame(self, bg=GRIS_FONDO, pady=10)
        barra.pack(fill="x", padx=0)

        interior = tk.Frame(barra, bg=GRIS_FONDO)
        interior.pack(fill="x", padx=16)

        # Fecha desde
        tk.Label(interior, text=t("desde_label"), font=("Segoe UI", 9, "bold"),
                 bg=GRIS_FONDO).pack(side="left")
        self.lbl_desde = tk.Label(interior,
                                   text=self.fecha_desde.strftime("%d/%m/%Y"),
                                   font=("Segoe UI", 9), bg=BLANCO, fg=AZUL,
                                   relief="solid", bd=1, padx=8, pady=4,
                                   cursor="hand2")
        self.lbl_desde.pack(side="left", padx=(4, 12))
        self.lbl_desde.bind("<Button-1>", lambda e: self._abrir_cal_desde())

        # Fecha hasta
        tk.Label(interior, text=t("hasta_label"), font=("Segoe UI", 9, "bold"),
                 bg=GRIS_FONDO).pack(side="left")
        self.lbl_hasta = tk.Label(interior,
                                   text=self.fecha_hasta.strftime("%d/%m/%Y"),
                                   font=("Segoe UI", 9), bg=BLANCO, fg=AZUL,
                                   relief="solid", bd=1, padx=8, pady=4,
                                   cursor="hand2")
        self.lbl_hasta.pack(side="left", padx=(4, 6))
        self.lbl_hasta.bind("<Button-1>", lambda e: self._abrir_cal_hasta())

        tk.Button(interior, text=t("reportes_elegir_rango"), font=("Segoe UI", 8), bg=BLANCO,
                  relief="solid", bd=1, padx=8, pady=3, cursor="hand2",
                  command=self._abrir_selector_rango).pack(side="left", padx=(0, 10))

        # Accesos rápidos de período
        for texto, accion in [("Hoy", self._periodo_hoy),
                               ("Esta semana", self._periodo_semana),
                               ("Este mes", self._periodo_mes),
                               ("Este año", self._periodo_anio)]:
            tk.Button(interior, text=texto, font=("Segoe UI", 8),
                      bg=BLANCO, relief="solid", bd=1, padx=8, pady=3,
                      cursor="hand2", command=accion
                      ).pack(side="left", padx=2)

        # Separador
        tk.Frame(interior, bg=GRIS_BORDE, width=1).pack(
            side="left", fill="y", padx=12, pady=2)

        # Filtro vendedor — un usuario con rol Vendedor NUNCA puede elegir
        # ver las ventas de otro: el combo queda bloqueado mostrando
        # únicamente su propio nombre. Gerente y Administrador sí pueden
        # elegir "Todos" o cualquier vendedor puntual.
        self._es_vendedor = bool(self.usuario_actual) and self.usuario_actual.get("rol") == "vendedor"
        nombre_propio = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""

        tk.Label(interior, text=t("vendedor_label"), font=("Segoe UI", 9, "bold"),
                 bg=GRIS_FONDO).pack(side="left")
        self.vendedores = _listar_todos_los_vendedores()
        if self._es_vendedor:
            self.var_vendedor = tk.StringVar(value=nombre_propio)
            opciones_vend = [nombre_propio]
        else:
            self.var_vendedor = tk.StringVar(value="Todos")
            opciones_vend = ["Todos"] + [v["nombre"] for v in self.vendedores]
        self.combo_vendedor = ttk.Combobox(interior, textvariable=self.var_vendedor,
                                            values=opciones_vend,
                                            state="disabled" if self._es_vendedor else "readonly",
                                            width=18, font=("Segoe UI", 9))
        self.combo_vendedor.pack(side="left", padx=(4, 12))
        self.combo_vendedor.bind("<<ComboboxSelected>>", lambda e: self.after(50, self._cargar))
        if self._es_vendedor:
            tk.Label(interior, text=t("reportes_solo_tus_ventas"),
                     font=("Segoe UI", 8, "italic"), bg=GRIS_FONDO, fg=GRIS_TEXTO
                     ).pack(side="left", padx=(0, 12))

        # Filtro estado
        tk.Label(interior, text=t("estado_label"), font=("Segoe UI", 9, "bold"),
                 bg=GRIS_FONDO).pack(side="left")
        self.var_estado = tk.StringVar(value="Todos")
        self.combo_estado = ttk.Combobox(interior, textvariable=self.var_estado,
                                          values=["Todos", "Pagado", "Cancelado"],
                                          state="readonly", width=10,
                                          font=("Segoe UI", 9))
        self.combo_estado.pack(side="left", padx=(4, 16))
        self.combo_estado.bind("<<ComboboxSelected>>", lambda e: self.after(50, self._cargar))

        # Buscar
        tk.Label(interior, text="🔍", font=("Segoe UI", 11),
                 bg=GRIS_FONDO).pack(side="left")
        self.var_busqueda = tk.StringVar()
        entry_busq = tk.Entry(interior, textvariable=self.var_busqueda,
                               font=("Segoe UI", 9), relief="solid", bd=1,
                               width=20)
        entry_busq.pack(side="left", padx=(2, 4), ipady=3)
        entry_busq.bind("<Return>", lambda e: self._cargar())
        entry_busq.insert(0, "Buscar cliente, factura, ID…")
        entry_busq.config(fg=GRIS_TEXTO)

        def _on_focus_in(e):
            if entry_busq.get() == "Buscar cliente, factura, ID…":
                entry_busq.delete(0, tk.END)
                entry_busq.config(fg=NEGRO)

        def _on_focus_out(e):
            if not entry_busq.get():
                entry_busq.insert(0, "Buscar cliente, factura, ID…")
                entry_busq.config(fg=GRIS_TEXTO)

        entry_busq.bind("<FocusIn>", _on_focus_in)
        entry_busq.bind("<FocusOut>", _on_focus_out)

        tk.Button(interior, text=t("buscar"), font=("Segoe UI", 9),
                  bg=AZUL, fg=BLANCO, relief="flat", padx=10, pady=3,
                  cursor="hand2", command=self._cargar
                  ).pack(side="left")

    # ── Tarjetas de resumen ──────────────────────────────────
    def _construir_tarjetas(self):
        self.frame_tarjetas = tk.Frame(self, bg=BLANCO)
        self.frame_tarjetas.pack(fill="x", padx=16, pady=(12, 4))

        definiciones = [
            ("cantidad",      t("modulo_ventas"),  AZUL,   "🧾"),
            ("total",         t("reportes_total_vendido"), VERDE,  "💰"),
            ("efectivo",      t("forma_pago_efectivo"), NEGRO,  "💵"),
            ("transferencia", t("forma_pago_transferencia"), NARANJA,"🏦"),
            ("cripto",        t("forma_pago_cripto"), "#7c3aed", "🪙"),
            ("credito",       t("credito_label"),  AZUL_OSC,"📋"),
            ("canceladas",    t("reportes_canceladas"), ROJO,   "✕"),
        ]
        self.labels_tarjetas: dict[str, tk.Label] = {}
        for clave, titulo, color, icono in definiciones:
            card = tk.Frame(self.frame_tarjetas, bg=BLANCO,
                            highlightthickness=1, highlightbackground=GRIS_BORDE)
            card.pack(side="left", fill="both", expand=True, padx=(0, 8), ipady=8)

            tk.Label(card, text=icono, font=("Segoe UI", 16),
                     bg=BLANCO).pack(anchor="w", padx=12, pady=(10, 0))
            lbl_val = tk.Label(card, text="—", font=("Segoe UI", 14, "bold"),
                               bg=BLANCO, fg=color)
            lbl_val.pack(anchor="w", padx=12)
            tk.Label(card, text=titulo, font=("Segoe UI", 8),
                     bg=BLANCO, fg=GRIS_TEXTO).pack(anchor="w", padx=12, pady=(0, 8))
            self.labels_tarjetas[clave] = lbl_val

    # ── Cuerpo: tabla + panel detalle ───────────────────────
    def _construir_cuerpo(self):
        self.frame_cuerpo = tk.Frame(self, bg=BLANCO)
        self.frame_cuerpo.pack(fill="both", expand=True, padx=16, pady=(4, 12))
        self.frame_cuerpo.grid_columnconfigure(0, weight=1)
        self.frame_cuerpo.grid_columnconfigure(1, weight=0)
        self.frame_cuerpo.grid_rowconfigure(0, weight=1)

        self._construir_tabla()
        self._construir_panel_detalle()

    def _construir_tabla(self):
        frame_tabla = tk.Frame(self.frame_cuerpo, bg=BLANCO)
        frame_tabla.grid(row=0, column=0, sticky="nsew")

        cols = ("id", "fecha", "cliente", "vendedor",
                "total", "condicion", "forma_pago", "estado", "factura")
        encabs = (t("col_id"), t("col_fecha_hora"), t("col_cliente_mayus").title(), t("col_vendedor_mayus").title(),
                  t("col_total"), t("col_condicion"), t("col_forma_pago"), t("col_estado"), t("col_factura"))
        anchos = (50, 140, 160, 140, 100, 80, 120, 80, 130)

        self.tabla = ttk.Treeview(frame_tabla, columns=cols,
                                   show="headings", selectmode="browse")
        habilitar_deseleccion_treeview(self.tabla)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        style.configure("Treeview", font=("Segoe UI", 9), rowheight=26)

        for col, enc, ancho in zip(cols, encabs, anchos):
            self.tabla.heading(col, text=enc,
                               command=lambda c=col: self._ordenar_por(c))
            self.tabla.column(col, width=ancho, anchor="center", minwidth=40)
        self.tabla.column("cliente",  anchor="w")
        self.tabla.column("vendedor", anchor="w")
        self.tabla.column("factura",  anchor="w")

        self.tabla.tag_configure("cancelado",  foreground="#9ca3af")
        self.tabla.tag_configure("par",        background="#f8fafc")
        self.tabla.tag_configure("impar",      background=BLANCO)

        sb_y = ttk.Scrollbar(frame_tabla, orient="vertical",
                              command=self.tabla.yview)
        sb_x = ttk.Scrollbar(frame_tabla, orient="horizontal",
                              command=self.tabla.xview)
        self.tabla.configure(yscrollcommand=sb_y.set,
                             xscrollcommand=sb_x.set)

        self.tabla.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_x.grid(row=1, column=0, sticky="ew")
        frame_tabla.grid_rowconfigure(0, weight=1)
        frame_tabla.grid_columnconfigure(0, weight=1)

        self.tabla.bind("<<TreeviewSelect>>", self._al_seleccionar)

        # Barra inferior con contador
        self.lbl_contador = tk.Label(frame_tabla, text="",
                                      font=("Segoe UI", 8), bg=GRIS_FONDO,
                                      fg=GRIS_TEXTO, anchor="w")
        self.lbl_contador.grid(row=2, column=0, columnspan=2,
                                sticky="ew", pady=(2, 0))

        self._orden_col = None
        self._orden_asc = True

    def _construir_panel_detalle(self):
        self.frame_detalle = tk.Frame(self.frame_cuerpo, bg=GRIS_FONDO,
                                       width=310)
        self.frame_detalle.grid(row=0, column=1, sticky="ns",
                                 padx=(10, 0))
        self.frame_detalle.grid_propagate(False)
        self.frame_detalle.grid_remove()   # oculto hasta seleccionar

    # ── Carga y refresco ────────────────────────────────────
    def _cargar(self):
        busq = self.var_busqueda.get()
        if busq == "Buscar cliente, factura, ID…":
            busq = ""

        # Segunda barrera de seguridad (además del combo bloqueado en la
        # UI): un usuario con rol Vendedor siempre consulta con su propio
        # ID, sin importar qué tenga seleccionado self.var_vendedor.
        if self._es_vendedor:
            vendedor_id = filtro_usuario_ventas(self.usuario_actual)
        else:
            vendedor_id = None
            sel_vend = self.var_vendedor.get()
            if sel_vend != "Todos":
                for v in self.vendedores:
                    if v["nombre"] == sel_vend:
                        vendedor_id = v["id"]
                        break

        self.ventas_actuales = _listar_ventas_filtradas(
            self.fecha_desde.isoformat(),
            self.fecha_hasta.isoformat(),
            vendedor_id=vendedor_id,
            estado=self.var_estado.get(),
            busqueda=busq,
        )
        self._poblar_tabla()
        self._actualizar_tarjetas()

    def _poblar_tabla(self):
        for fila in self.tabla.get_children():
            self.tabla.delete(fila)

        for i, v in enumerate(self.ventas_actuales):
            fila_hora = v["fecha"][:16] if v["fecha"] else ""
            tags = []
            if v["estado"] == "Cancelado":
                tags.append("cancelado")
            tags.append("par" if i % 2 == 0 else "impar")

            self.tabla.insert("", "end", iid=str(v["id"]),
                              values=(
                                  v["id"], fila_hora, v["cliente"],
                                  v["vendedor"],
                                  formatear_gs(v['total']),
                                  v["condicion"].capitalize(),
                                  v["forma_pago"], v["estado"],
                                  v["factura"],
                              ), tags=tuple(tags))

        n = len(self.ventas_actuales)
        self.lbl_contador.config(
            text=f"  {n} venta{'s' if n != 1 else ''} encontrada{'s' if n != 1 else ''}"
                 f"  —  {self.fecha_desde.strftime('%d/%m/%Y')} al "
                 f"{self.fecha_hasta.strftime('%d/%m/%Y')}"
        )

    def _actualizar_tarjetas(self):
        res = _resumen_de_ventas(self.ventas_actuales)
        self.labels_tarjetas["cantidad"].config(text=str(res["cantidad"]))
        self.labels_tarjetas["total"].config(
            text=formatear_gs(res['total']))
        self.labels_tarjetas["efectivo"].config(
            text=formatear_gs(res['efectivo']))
        self.labels_tarjetas["transferencia"].config(
            text=formatear_gs(res['transferencia']))
        self.labels_tarjetas["credito"].config(
            text=formatear_gs(res['credito']))
        self.labels_tarjetas["canceladas"].config(
            text=str(res["canceladas"]))

    # ── Selección de venta → panel detalle ──────────────────
    def _al_seleccionar(self, event=None):
        sel = self.tabla.selection()
        if not sel:
            return
        venta_id = int(sel[0])
        if venta_id == self.venta_detalle_id:
            return
        self.venta_detalle_id = venta_id
        self._mostrar_detalle(venta_id)

    def _mostrar_detalle(self, venta_id: int):
        # Limpiar panel
        for w in self.frame_detalle.winfo_children():
            w.destroy()
        self.frame_detalle.grid()

        det = _obtener_detalle_venta_con_libres(venta_id)
        if not det:
            tk.Label(self.frame_detalle, text=t("reportes_no_detalle"),
                     bg=GRIS_FONDO, font=("Segoe UI", 9)).pack(pady=20)
            return

        # ── Encabezado del panel ─────────────────────────────
        hdr = tk.Frame(self.frame_detalle, bg=AZUL)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"Venta #{det['id']}",
                 font=("Segoe UI", 11, "bold"), bg=AZUL, fg=BLANCO
                 ).pack(side="left", padx=12, pady=8)
        tk.Button(hdr, text="✕", font=("Segoe UI", 9), bg=AZUL, fg=BLANCO,
                  relief="flat", cursor="hand2",
                  command=self._cerrar_detalle
                  ).pack(side="right", padx=8)

        # ── Cuerpo scrolleable ───────────────────────────────
        canvas = tk.Canvas(self.frame_detalle, bg=GRIS_FONDO,
                           highlightthickness=0)
        sb = tk.Scrollbar(self.frame_detalle, orient="vertical",
                          command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)

        inner = tk.Frame(canvas, bg=GRIS_FONDO)
        canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>",
                   lambda e: canvas.configure(
                       scrollregion=canvas.bbox("all")))

        def _scroll(event):
            if event.num == 4:
                canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                canvas.yview_scroll(1, "units")
            else:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind("<MouseWheel>", _scroll)
        canvas.bind("<Button-4>", _scroll)
        canvas.bind("<Button-5>", _scroll)
        inner.bind("<MouseWheel>", _scroll)

        pad = {"padx": 12, "pady": 3}

        # Estado badge
        color_est = VERDE if det["estado"] == "Pagado" else ROJO
        tk.Label(inner, text=f"  {det['estado']}  ",
                 font=("Segoe UI", 9, "bold"), bg=color_est, fg=BLANCO,
                 ).pack(anchor="w", padx=12, pady=(10, 4))

        # Datos cabecera
        def fila_info(etiqueta, valor, color_val=NEGRO):
            f = tk.Frame(inner, bg=GRIS_FONDO)
            f.pack(fill="x", **pad)
            tk.Label(f, text=etiqueta, font=("Segoe UI", 8),
                     bg=GRIS_FONDO, fg=GRIS_TEXTO, width=12,
                     anchor="w").pack(side="left")
            tk.Label(f, text=valor, font=("Segoe UI", 9, "bold"),
                     bg=GRIS_FONDO, fg=color_val,
                     wraplength=170, justify="left").pack(side="left")

        fecha_fmt = det["fecha"][:16] if det["fecha"] else ""
        fila_info("Fecha:",    fecha_fmt)
        fila_info("Factura:",  det["nro_factura"] or "—")
        fila_info("Cliente:",  det["cliente_nombre"])
        if det["cliente_documento"]:
            fila_info("Doc.:",    det["cliente_documento"])
        fila_info("Vendedor:", det["vendedor"])
        fila_info("Condición:",det["condicion"].capitalize())
        fila_info("Pago:",     det["forma_pago"])

        # Separador
        tk.Frame(inner, bg=GRIS_BORDE, height=1).pack(
            fill="x", padx=12, pady=8)

        # Líneas de la venta
        tk.Label(inner, text=t("reportes_articulos"), font=("Segoe UI", 9, "bold"),
                 bg=GRIS_FONDO, fg=NEGRO).pack(anchor="w", padx=12, pady=(0, 4))

        for linea in det["lineas"]:
            f_lin = tk.Frame(inner, bg=BLANCO,
                             highlightthickness=1,
                             highlightbackground=GRIS_BORDE)
            f_lin.pack(fill="x", padx=12, pady=2)

            nombre = linea["nombre_producto"]
            if linea.get("es_libre"):
                nombre = f"✏ {nombre}"
            tk.Label(f_lin, text=nombre, font=("Segoe UI", 9, "bold"),
                     bg=BLANCO, fg=NEGRO, wraplength=240,
                     justify="left").pack(anchor="w", padx=8, pady=(5, 0))

            f_detlin = tk.Frame(f_lin, bg=BLANCO)
            f_detlin.pack(fill="x", padx=8, pady=(0, 5))
            cant_txt = f"{linea['cantidad']:g}"
            if linea["cantidad_devuelta"] > 0:
                cant_txt += f"  (dev. {linea['cantidad_devuelta']:g})"
            tk.Label(f_detlin, text=cant_txt, font=("Segoe UI", 8),
                     bg=BLANCO, fg=GRIS_TEXTO).pack(side="left")
            tk.Label(f_detlin,
                     text=formatear_gs(linea['importe']),
                     font=("Segoe UI", 9, "bold"),
                     bg=BLANCO, fg=AZUL).pack(side="right")

        # Separador
        tk.Frame(inner, bg=GRIS_BORDE, height=1).pack(
            fill="x", padx=12, pady=8)

        # Total
        f_tot = tk.Frame(inner, bg=GRIS_FONDO)
        f_tot.pack(fill="x", padx=12, pady=(0, 12))
        tk.Label(f_tot, text=t("col_total_mayus"),
                 font=("Segoe UI", 11, "bold"),
                 bg=GRIS_FONDO, fg=NEGRO).pack(side="left")
        tk.Label(f_tot, text=formatear_gs(det['total']),
                 font=("Segoe UI", 13, "bold"),
                 bg=GRIS_FONDO, fg=VERDE).pack(side="right")

    def _cerrar_detalle(self):
        self.venta_detalle_id = None
        self.frame_detalle.grid_remove()
        self.tabla.selection_remove(*self.tabla.selection())

    # ── Ordenamiento de columnas ─────────────────────────────
    def _ordenar_por(self, col):
        if self._orden_col == col:
            self._orden_asc = not self._orden_asc
        else:
            self._orden_col = col
            self._orden_asc = True

        reverse = not self._orden_asc
        col_map = {
            "id": "id", "fecha": "fecha", "cliente": "cliente",
            "vendedor": "vendedor", "total": "total",
            "condicion": "condicion", "forma_pago": "forma_pago",
            "estado": "estado", "factura": "factura",
        }
        clave = col_map.get(col, "fecha")
        try:
            self.ventas_actuales.sort(
                key=lambda v: (v.get(clave) or 0)
                if clave in ("id", "total")
                else str(v.get(clave) or "").lower(),
                reverse=reverse,
            )
        except Exception:
            pass
        self._poblar_tabla()

    # ── Selectores de fecha ──────────────────────────────────
    def _abrir_cal_desde(self):
        abrir_selector_fecha(self, self.fecha_desde, self._set_desde)

    def _abrir_cal_hasta(self):
        abrir_selector_fecha(self, self.fecha_hasta, self._set_hasta)

    def _abrir_selector_rango(self):
        abrir_selector_rango_fechas(self, self.fecha_desde, self.fecha_hasta, self._set_rango)

    def _set_rango(self, fecha_desde: datetime.date, fecha_hasta: datetime.date):
        self.fecha_desde = fecha_desde
        self.fecha_hasta = fecha_hasta
        self.lbl_desde.config(text=fecha_desde.strftime("%d/%m/%Y"))
        self.lbl_hasta.config(text=fecha_hasta.strftime("%d/%m/%Y"))
        self._cargar()

    def _set_desde(self, fecha: datetime.date):
        self.fecha_desde = fecha
        self.lbl_desde.config(text=fecha.strftime("%d/%m/%Y"))
        self._cargar()

    def _set_hasta(self, fecha: datetime.date):
        self.fecha_hasta = fecha
        self.lbl_hasta.config(text=fecha.strftime("%d/%m/%Y"))
        self._cargar()

    # ── Atajos de período ────────────────────────────────────
    def _periodo_hoy(self):
        hoy = datetime.date.today()
        self.fecha_desde = hoy
        self.fecha_hasta = hoy
        self._sync_labels()
        self._cargar()

    def _periodo_semana(self):
        hoy = datetime.date.today()
        self.fecha_desde = hoy - datetime.timedelta(days=hoy.weekday())
        self.fecha_hasta = hoy
        self._sync_labels()
        self._cargar()

    def _periodo_mes(self):
        hoy = datetime.date.today()
        self.fecha_desde = hoy.replace(day=1)
        self.fecha_hasta = hoy
        self._sync_labels()
        self._cargar()

    def _periodo_anio(self):
        hoy = datetime.date.today()
        self.fecha_desde = hoy.replace(month=1, day=1)
        self.fecha_hasta = hoy
        self._sync_labels()
        self._cargar()

    def _sync_labels(self):
        self.lbl_desde.config(text=self.fecha_desde.strftime("%d/%m/%Y"))
        self.lbl_hasta.config(text=self.fecha_hasta.strftime("%d/%m/%Y"))
    # ---------------- REPORTE GENERAL: puentes hacia el controlador genérico ----------------
    def _obtener_datos_reporte(self) -> dict:
        """Estructura neutral (consumida por PDF simple, Word, ODT, CSV y
        JSON) construida a partir de self.ventas_actuales: exactamente las
        ventas que el usuario está viendo en pantalla en este momento, con
        TODOS los filtros activos aplicados (rango, vendedor, estado y
        búsqueda libre), no solo el rango de fechas."""
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""
        res = _resumen_de_ventas(self.ventas_actuales)

        filtros_aplicados = []
        sel_vend = self.var_vendedor.get()
        if sel_vend != "Todos":
            filtros_aplicados.append(f"Vendedor: {sel_vend}")
        if self.var_estado.get() != "Todos":
            filtros_aplicados.append(f"Estado: {self.var_estado.get()}")
        busq = self.var_busqueda.get()
        if busq and busq != "Buscar cliente, factura, ID…":
            filtros_aplicados.append(f"Búsqueda: \"{busq}\"")
        texto_filtros = "  •  ".join(filtros_aplicados) if filtros_aplicados else "Sin filtros adicionales"

        periodo_texto = (
            f"Período: {self.fecha_desde.strftime('%d/%m/%Y')} al "
            f"{self.fecha_hasta.strftime('%d/%m/%Y')}   ({texto_filtros})"
        )

        secciones = [
            {
                "tipo": "resumen",
                "titulo": "RESUMEN GENERAL",
                "filas": [
                    ("Cantidad de Ventas", str(res["cantidad"])),
                    ("Ventas Canceladas", str(res["canceladas"])),
                    ("Total Vendido", formatear_gs(res['total'])),
                    ("Efectivo", formatear_gs(res['efectivo'])),
                    ("Transferencia", formatear_gs(res['transferencia'])),
                    ("Criptomonedas", formatear_gs(res['cripto'])),
                    ("Crédito", formatear_gs(res['credito'])),
                ],
            },
            {
                "tipo": "tabla",
                "titulo": "DETALLE DE VENTAS",
                "encabezados": ["ID", "Fecha y Hora", "Cliente", "Vendedor",
                                "Total", "Condición", "Forma de Pago", "Estado", "Factura"],
                "filas": [
                    [str(v["id"]), v["fecha"], v["cliente"], v["vendedor"],
                     formatear_gs(v['total']), v["condicion"].capitalize(),
                     v["forma_pago"], v["estado"], v["factura"]]
                    for v in self.ventas_actuales
                ],
            },
        ]

        return {
            "titulo": "Reporte General de Ventas",
            "subtitulo": periodo_texto,
            "generado_por": nombre_usuario,
            "secciones": secciones,
        }

    def _confirmar_dashboard_sin_filtros(self) -> bool:
        """Se llama ANTES de pedir la ruta del PDF con dashboard. Si hay
        filtros de Estado/Búsqueda activos (que ese formato no respeta, ya
        que trabaja por rango de fechas puro), pide confirmación. Devuelve
        False para cancelar sin mostrar error.

        El filtro de Vendedor NO entra en este chequeo: para un usuario
        con rol Vendedor es obligatorio y SIEMPRE se respeta (ver
        _generar_pdf_dashboard), no es algo que se esté ignorando. Para
        Gerente/Administrador eligiendo un vendedor puntual, en cambio, sí
        es un filtro opcional que este formato no aplica, así que se
        avisa igual que antes.
        """
        from tkinter import messagebox
        sel_vend = self.var_vendedor.get()
        vendedor_filtro_opcional_ignorado = (not self._es_vendedor) and sel_vend != "Todos"
        hay_filtros_extra = (
            vendedor_filtro_opcional_ignorado
            or self.var_estado.get() != "Todos"
            or (self.var_busqueda.get() and self.var_busqueda.get() != "Buscar cliente, factura, ID…")
        )
        if not hay_filtros_extra:
            return True
        return messagebox.askyesno(
            "Filtros activos",
            "El 'PDF con dashboard' siempre muestra TODAS las ventas del "
            "rango de fechas, sin aplicar los filtros de Vendedor, Estado "
            "o Búsqueda que tenés activos ahora.\n\n"
            "¿Querés continuar de todas formas? (Para un reporte que sí "
            "respete los filtros, usá 'PDF simple' o 'Excel'.)",
            parent=self,
        )

    def _generar_pdf_dashboard(self, ruta: str):
        """El PDF 'con dashboard' (gráfico de barras + KPIs) reutiliza el
        generador existente de reporte_pdf.py, que trabaja por rango de
        fechas puro y siempre incluye TODAS las ventas del rango, sin
        aplicar Estado/Búsqueda (ver advertencia en
        _confirmar_dashboard_sin_filtros).

        El filtro de Vendedor es la única excepción: para un usuario con
        rol Vendedor SIEMPRE se aplica (ve solo sus propias ventas, nunca
        las de otro), independientemente de cualquier otra cosa — por eso
        se pasa acá explícitamente, sin depender del combo de la UI.
        """
        from reporte_pdf import generar_reporte_pdf
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""
        generar_reporte_pdf(
            ruta, self.fecha_desde.isoformat(), self.fecha_hasta.isoformat(),
            generado_por=nombre_usuario,
            usuario_id=filtro_usuario_ventas(self.usuario_actual),
        )

    def _generar_excel(self, ruta: str):
        """Excel propio del módulo Reportes: vuelca exactamente
        self.ventas_actuales (con todos los filtros activos aplicados) en
        una hoja de resumen y otra de detalle."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        res = _resumen_de_ventas(self.ventas_actuales)
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""

        wb = openpyxl.Workbook()

        # ── Hoja Resumen ──────────────────────────────────────
        ws = wb.active
        ws.title = "Resumen"
        ws.append(["Reporte General de Ventas"])
        ws.append([f"Período: {self.fecha_desde.strftime('%d/%m/%Y')} al "
                    f"{self.fecha_hasta.strftime('%d/%m/%Y')}"])
        ws.append([f"Generado por: {nombre_usuario}"])
        ws.append([])
        filas_resumen = [
            ("Cantidad de Ventas", res["cantidad"]),
            ("Ventas Canceladas", res["canceladas"]),
            ("Total Vendido", res["total"]),
            ("Efectivo", res["efectivo"]),
            ("Transferencia", res["transferencia"]),
            ("Criptomonedas", res["cripto"]),
            ("Crédito", res["credito"]),
        ]
        ws.append(["Indicador", "Valor"])
        for etiqueta, valor in filas_resumen:
            ws.append([etiqueta, valor])

        for cell in ws[6]:  # fila de encabezados "Indicador" / "Valor"
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D5FD6")
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 20

        # ── Hoja Detalle de Ventas ───────────────────────────────
        ws2 = wb.create_sheet("Detalle de Ventas")
        cols_det = ["ID", "Fecha y Hora", "Cliente", "Vendedor", "Total",
                    "Condición", "Forma de Pago", "Estado", "Factura"]
        ws2.append(cols_det)
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D5FD6")
            cell.alignment = Alignment(horizontal="center")
        for v in self.ventas_actuales:
            ws2.append([
                v["id"], v["fecha"], v["cliente"], v["vendedor"], v["total"],
                v["condicion"].capitalize(), v["forma_pago"], v["estado"], v["factura"],
            ])
        for col in ws2.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(ruta)


# ─────────────────────────────────────────────────────────────
#  PANEL PÚBLICO: pestañas Ventas / Compras / Presupuestos
# ─────────────────────────────────────────────────────────────
class PanelReportes(tk.Frame):
    """Contenedor del módulo Reportes con tres pestañas: Ventas, Compras
    y Presupuestos. Cada una es un panel independiente y completo
    (filtros, tarjetas, tabla + detalle, y exportación en los 7 formatos)."""

    def __init__(self, parent, usuario_actual):
        super().__init__(parent, bg=BLANCO)
        self.usuario_actual = usuario_actual

        style = ttk.Style()
        style.configure("Reportes.TNotebook.Tab", font=("Segoe UI", 10, "bold"), padding=(16, 8))

        notebook = ttk.Notebook(self, style="Reportes.TNotebook")
        notebook.pack(fill="both", expand=True)

        from ventana_reporte_compras import PanelReporteCompras
        from ventana_reporte_presupuestos import PanelReportePresupuestos

        tab_ventas = _PanelReportesVentas(notebook, usuario_actual)
        tab_compras = PanelReporteCompras(notebook, usuario_actual)
        tab_presupuestos = PanelReportePresupuestos(notebook, usuario_actual)

        notebook.add(tab_ventas, text=t("reportes_tab_ventas"))
        notebook.add(tab_compras, text=t("reportes_tab_compras"))
        notebook.add(tab_presupuestos, text=t("reportes_tab_presupuestos"))

