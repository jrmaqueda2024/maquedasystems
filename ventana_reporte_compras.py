"""
ventana_reporte_compras.py
Pestaña "Compras" del módulo Reportes: registro completo de todas las
compras con filtros por rango de fechas, proveedor y búsqueda. Incluye
tarjetas de resumen, tabla con panel lateral de detalle (líneas de cada
compra) y exportación en los 7 formatos vía BotonReporteGeneral (PDF con
dashboard, PDF simple, Word, ODT, Excel, CSV, JSON).
"""
import tkinter as tk
from tkinter import ttk
import datetime

from utilidades_ui import formatear_gs, habilitar_deseleccion_treeview
from widget_calendario import abrir_selector_fecha, abrir_selector_rango_fechas
from menu_reporte_general import BotonReporteGeneral
from models_catalogo import listar_proveedores
from models_compras import listar_compras_en_rango, resumen_compras_en_rango, obtener_detalle_compra

AZUL        = "#1d5fd6"
AZUL_OSC    = "#163d8c"
GRIS_FONDO  = "#f4f5f7"
GRIS_BORDE  = "#e2e8f0"
BLANCO      = "#ffffff"
VERDE       = "#16a34a"
ROJO        = "#dc2626"
NARANJA     = "#d97706"
GRIS_TEXTO  = "#6b7280"
NEGRO       = "#1e293b"


class PanelReporteCompras(tk.Frame):
    def __init__(self, parent, usuario_actual):
        super().__init__(parent, bg=BLANCO)
        self.usuario_actual = usuario_actual

        hoy = datetime.date.today()
        self.fecha_desde = hoy.replace(day=1)   # primer día del mes
        self.fecha_hasta = hoy

        self.compras_actuales: list[dict] = []
        self.compra_detalle_id: int | None = None
        self.proveedores = listar_proveedores()

        self._construir_encabezado()
        self._construir_filtros()
        self._construir_tarjetas()
        self._construir_cuerpo()
        self._cargar()

    # ── Encabezado ───────────────────────────────────────────
    def _construir_encabezado(self):
        enc = tk.Frame(self, bg=AZUL, height=54)
        enc.pack(fill="x")
        enc.pack_propagate(False)
        tk.Label(enc, text="📊  Reportes de Compras",
                 font=("Segoe UI", 15, "bold"), bg=AZUL, fg=BLANCO
                 ).pack(side="left", padx=20, pady=12)

        self.boton_reporte = BotonReporteGeneral(
            enc, obtener_datos_callback=self._obtener_datos_reporte,
            nombre_archivo_base="Reporte_General_Compras",
        )
        self.boton_reporte.generador_pdf_dashboard = self._generar_pdf_dashboard
        self.boton_reporte.generador_excel = self._generar_excel
        self.boton_reporte.pack(side="right", padx=(0, 12), pady=12)

        tk.Button(enc, text="🔄 Actualizar", font=("Segoe UI", 9),
                  bg=AZUL_OSC, fg=BLANCO, relief="flat", padx=12, pady=4,
                  cursor="hand2", command=self._cargar
                  ).pack(side="right", padx=16, pady=12)

    # ── Barra de filtros ─────────────────────────────────────
    def _construir_filtros(self):
        barra = tk.Frame(self, bg=GRIS_FONDO, pady=10)
        barra.pack(fill="x", padx=0)

        interior = tk.Frame(barra, bg=GRIS_FONDO)
        interior.pack(fill="x", padx=16)

        tk.Label(interior, text="Desde:", font=("Segoe UI", 9, "bold"),
                 bg=GRIS_FONDO).pack(side="left")
        self.lbl_desde = tk.Label(interior,
                                   text=self.fecha_desde.strftime("%d/%m/%Y"),
                                   font=("Segoe UI", 9), bg=BLANCO, fg=AZUL,
                                   relief="solid", bd=1, padx=8, pady=4,
                                   cursor="hand2")
        self.lbl_desde.pack(side="left", padx=(4, 12))
        self.lbl_desde.bind("<Button-1>", lambda e: self._abrir_cal_desde())

        tk.Label(interior, text="Hasta:", font=("Segoe UI", 9, "bold"),
                 bg=GRIS_FONDO).pack(side="left")
        self.lbl_hasta = tk.Label(interior,
                                   text=self.fecha_hasta.strftime("%d/%m/%Y"),
                                   font=("Segoe UI", 9), bg=BLANCO, fg=AZUL,
                                   relief="solid", bd=1, padx=8, pady=4,
                                   cursor="hand2")
        self.lbl_hasta.pack(side="left", padx=(4, 6))
        self.lbl_hasta.bind("<Button-1>", lambda e: self._abrir_cal_hasta())

        tk.Button(interior, text="📅 Elegir Rango", font=("Segoe UI", 8), bg=BLANCO,
                  relief="solid", bd=1, padx=8, pady=3, cursor="hand2",
                  command=self._abrir_selector_rango).pack(side="left", padx=(0, 10))

        for texto, accion in [("Hoy", self._periodo_hoy),
                               ("Esta semana", self._periodo_semana),
                               ("Este mes", self._periodo_mes),
                               ("Este año", self._periodo_anio)]:
            tk.Button(interior, text=texto, font=("Segoe UI", 8),
                      bg=BLANCO, relief="solid", bd=1, padx=8, pady=3,
                      cursor="hand2", command=accion
                      ).pack(side="left", padx=2)

        tk.Frame(interior, bg=GRIS_BORDE, width=1).pack(
            side="left", fill="y", padx=12, pady=2)

        # Filtro proveedor
        tk.Label(interior, text="Proveedor:", font=("Segoe UI", 9, "bold"),
                 bg=GRIS_FONDO).pack(side="left")
        self.var_proveedor = tk.StringVar(value="Todos")
        opciones_prov = ["Todos"] + [p["nombre"] for p in self.proveedores]
        self.combo_proveedor = ttk.Combobox(interior, textvariable=self.var_proveedor,
                                             values=opciones_prov, state="readonly",
                                             width=18, font=("Segoe UI", 9))
        self.combo_proveedor.pack(side="left", padx=(4, 16))
        self.combo_proveedor.bind("<<ComboboxSelected>>", lambda e: self.after(50, self._cargar))

        # Buscar
        tk.Label(interior, text="🔍", font=("Segoe UI", 11),
                 bg=GRIS_FONDO).pack(side="left")
        self.var_busqueda = tk.StringVar()
        entry_busq = tk.Entry(interior, textvariable=self.var_busqueda,
                               font=("Segoe UI", 9), relief="solid", bd=1,
                               width=20)
        entry_busq.pack(side="left", padx=(2, 4), ipady=3)
        entry_busq.bind("<Return>", lambda e: self._cargar())
        entry_busq.insert(0, "Buscar proveedor, comprobante, ID…")
        entry_busq.config(fg=GRIS_TEXTO)

        def _on_focus_in(e):
            if entry_busq.get() == "Buscar proveedor, comprobante, ID…":
                entry_busq.delete(0, tk.END)
                entry_busq.config(fg=NEGRO)

        def _on_focus_out(e):
            if not entry_busq.get():
                entry_busq.insert(0, "Buscar proveedor, comprobante, ID…")
                entry_busq.config(fg=GRIS_TEXTO)

        entry_busq.bind("<FocusIn>", _on_focus_in)
        entry_busq.bind("<FocusOut>", _on_focus_out)

        tk.Button(interior, text="Buscar", font=("Segoe UI", 9),
                  bg=AZUL, fg=BLANCO, relief="flat", padx=10, pady=3,
                  cursor="hand2", command=self._cargar
                  ).pack(side="left")

    # ── Tarjetas de resumen ──────────────────────────────────
    def _construir_tarjetas(self):
        self.frame_tarjetas = tk.Frame(self, bg=BLANCO)
        self.frame_tarjetas.pack(fill="x", padx=16, pady=(12, 4))

        definiciones = [
            ("cantidad",    "Compras",             AZUL,    "🛒"),
            ("total",       "Total Comprado",      VERDE,   "💰"),
            ("promedio",    "Promedio por Compra", NARANJA, "📈"),
            ("proveedores", "Proveedores Distintos", AZUL_OSC, "🏭"),
        ]
        self.labels_tarjetas: dict[str, tk.Label] = {}
        for clave, titulo, color, icono in definiciones:
            card = tk.Frame(self.frame_tarjetas, bg=BLANCO,
                            highlightthickness=1, highlightbackground=GRIS_BORDE)
            card.pack(side="left", fill="both", expand=True, padx=(0, 8), ipady=8)

            tk.Label(card, text=icono, font=("Segoe UI", 16),
                     bg=BLANCO).pack(anchor="w", padx=12, pady=(10, 0))
            lbl_val = tk.Label(card, text="—", font=("Segoe UI", 14, "bold"),
                               bg=BLANCO, fg=color)
            lbl_val.pack(anchor="w", padx=12)
            tk.Label(card, text=titulo, font=("Segoe UI", 8),
                     bg=BLANCO, fg=GRIS_TEXTO).pack(anchor="w", padx=12, pady=(0, 8))
            self.labels_tarjetas[clave] = lbl_val

    # ── Cuerpo: tabla + panel detalle ───────────────────────
    def _construir_cuerpo(self):
        self.frame_cuerpo = tk.Frame(self, bg=BLANCO)
        self.frame_cuerpo.pack(fill="both", expand=True, padx=16, pady=(4, 12))
        self.frame_cuerpo.grid_columnconfigure(0, weight=1)
        self.frame_cuerpo.grid_columnconfigure(1, weight=0)
        self.frame_cuerpo.grid_rowconfigure(0, weight=1)

        self._construir_tabla()
        self._construir_panel_detalle()

    def _construir_tabla(self):
        frame_tabla = tk.Frame(self.frame_cuerpo, bg=BLANCO)
        frame_tabla.grid(row=0, column=0, sticky="nsew")

        cols = ("id", "fecha_hora", "fecha_compra", "proveedor", "nro_comprobante", "importe")
        encabs = ("ID", "Fecha y Hora", "Fecha Compra", "Proveedor", "N° Comprobante", "Importe")
        anchos = (50, 140, 110, 200, 140, 110)

        self.tabla = ttk.Treeview(frame_tabla, columns=cols,
                                   show="headings", selectmode="browse")
        habilitar_deseleccion_treeview(self.tabla)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        style.configure("Treeview", font=("Segoe UI", 9), rowheight=26)

        for col, enc, ancho in zip(cols, encabs, anchos):
            self.tabla.heading(col, text=enc)
            self.tabla.column(col, width=ancho, anchor="center", minwidth=40)
        self.tabla.column("proveedor", anchor="w")

        self.tabla.tag_configure("par",   background="#f8fafc")
        self.tabla.tag_configure("impar", background=BLANCO)

        sb_y = ttk.Scrollbar(frame_tabla, orient="vertical", command=self.tabla.yview)
        sb_x = ttk.Scrollbar(frame_tabla, orient="horizontal", command=self.tabla.xview)
        self.tabla.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)

        self.tabla.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_x.grid(row=1, column=0, sticky="ew")
        frame_tabla.grid_rowconfigure(0, weight=1)
        frame_tabla.grid_columnconfigure(0, weight=1)

        self.tabla.bind("<<TreeviewSelect>>", self._al_seleccionar)

        self.lbl_contador = tk.Label(frame_tabla, text="",
                                      font=("Segoe UI", 8), bg=GRIS_FONDO,
                                      fg=GRIS_TEXTO, anchor="w")
        self.lbl_contador.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(2, 0))

    def _construir_panel_detalle(self):
        self.frame_detalle = tk.Frame(self.frame_cuerpo, bg=GRIS_FONDO, width=320)
        self.frame_detalle.grid(row=0, column=1, sticky="ns", padx=(10, 0))
        self.frame_detalle.grid_propagate(False)
        self.frame_detalle.grid_remove()   # oculto hasta seleccionar

    # ── Accesos rápidos de período ──────────────────────────
    def _abrir_cal_desde(self):
        abrir_selector_fecha(self, self.fecha_desde, self._al_elegir_desde)

    def _abrir_cal_hasta(self):
        abrir_selector_fecha(self, self.fecha_hasta, self._al_elegir_hasta)

    def _abrir_selector_rango(self):
        abrir_selector_rango_fechas(self, self.fecha_desde, self.fecha_hasta, self._al_elegir_rango)

    def _al_elegir_rango(self, fecha_desde, fecha_hasta):
        self.fecha_desde = fecha_desde
        self.fecha_hasta = fecha_hasta
        self._sync_labels()
        self._cargar()

    def _al_elegir_desde(self, fecha):
        self.fecha_desde = fecha
        self._sync_labels()
        self._cargar()

    def _al_elegir_hasta(self, fecha):
        self.fecha_hasta = fecha
        self._sync_labels()
        self._cargar()

    def _periodo_hoy(self):
        hoy = datetime.date.today()
        self.fecha_desde = self.fecha_hasta = hoy
        self._sync_labels()
        self._cargar()

    def _periodo_semana(self):
        hoy = datetime.date.today()
        self.fecha_desde = hoy - datetime.timedelta(days=hoy.weekday())
        self.fecha_hasta = hoy
        self._sync_labels()
        self._cargar()

    def _periodo_mes(self):
        hoy = datetime.date.today()
        self.fecha_desde = hoy.replace(day=1)
        self.fecha_hasta = hoy
        self._sync_labels()
        self._cargar()

    def _periodo_anio(self):
        hoy = datetime.date.today()
        self.fecha_desde = hoy.replace(month=1, day=1)
        self.fecha_hasta = hoy
        self._sync_labels()
        self._cargar()

    def _sync_labels(self):
        self.lbl_desde.config(text=self.fecha_desde.strftime("%d/%m/%Y"))
        self.lbl_hasta.config(text=self.fecha_hasta.strftime("%d/%m/%Y"))

    # ── Carga y refresco ────────────────────────────────────
    def _proveedor_id_seleccionado(self):
        sel = self.var_proveedor.get()
        if sel == "Todos":
            return None
        for p in self.proveedores:
            if p["nombre"] == sel:
                return p["id"]
        return None

    def _cargar(self):
        busq = self.var_busqueda.get()
        if busq == "Buscar proveedor, comprobante, ID…":
            busq = ""

        self.compras_actuales = listar_compras_en_rango(
            self.fecha_desde.isoformat(),
            self.fecha_hasta.isoformat(),
            proveedor_id=self._proveedor_id_seleccionado(),
            busqueda=busq,
        )
        self._poblar_tabla()
        self._actualizar_tarjetas()

    def _poblar_tabla(self):
        for fila in self.tabla.get_children():
            self.tabla.delete(fila)

        for i, c in enumerate(self.compras_actuales):
            fila_hora = _formatear_fecha_hora_iso(c["fecha_y_hora"])
            fecha_compra_fmt = _formatear_fecha_iso(c["fecha_compra"])
            tags = ("par" if i % 2 == 0 else "impar",)
            self.tabla.insert("", "end", iid=str(c["id"]), values=(
                c["id"], fila_hora, fecha_compra_fmt, c["proveedor"],
                c["nro_comprobante"] or "—", formatear_gs(c["importe"]),
            ), tags=tags)

        n = len(self.compras_actuales)
        self.lbl_contador.config(
            text=f"  {n} compra{'s' if n != 1 else ''} encontrada{'s' if n != 1 else ''}"
                 f"  —  {self.fecha_desde.strftime('%d/%m/%Y')} al "
                 f"{self.fecha_hasta.strftime('%d/%m/%Y')}"
        )

    def _actualizar_tarjetas(self):
        total = sum(c["importe"] for c in self.compras_actuales)
        cantidad = len(self.compras_actuales)
        promedio = (total / cantidad) if cantidad else 0
        proveedores_distintos = len({c["proveedor_id"] for c in self.compras_actuales if c["proveedor_id"]})

        self.labels_tarjetas["cantidad"].config(text=str(cantidad))
        self.labels_tarjetas["total"].config(text=formatear_gs(total))
        self.labels_tarjetas["promedio"].config(text=formatear_gs(promedio))
        self.labels_tarjetas["proveedores"].config(text=str(proveedores_distintos))

    # ── Selección de compra → panel detalle ─────────────────
    def _al_seleccionar(self, event=None):
        sel = self.tabla.selection()
        if not sel:
            return
        compra_id = int(sel[0])
        if compra_id == self.compra_detalle_id:
            return
        self.compra_detalle_id = compra_id
        self._mostrar_detalle(compra_id)

    def _mostrar_detalle(self, compra_id: int):
        for w in self.frame_detalle.winfo_children():
            w.destroy()
        self.frame_detalle.grid()

        det = obtener_detalle_compra(compra_id)
        if not det:
            tk.Label(self.frame_detalle, text="No se pudo cargar el detalle.",
                     bg=GRIS_FONDO, font=("Segoe UI", 9)).pack(pady=20)
            return

        hdr = tk.Frame(self.frame_detalle, bg=AZUL)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"Compra #{det['id']}",
                 font=("Segoe UI", 11, "bold"), bg=AZUL, fg=BLANCO
                 ).pack(side="left", padx=12, pady=8)
        tk.Button(hdr, text="✕", font=("Segoe UI", 9), bg=AZUL, fg=BLANCO,
                  relief="flat", cursor="hand2",
                  command=self._cerrar_detalle
                  ).pack(side="right", padx=8)

        canvas = tk.Canvas(self.frame_detalle, bg=GRIS_FONDO, highlightthickness=0)
        sb = tk.Scrollbar(self.frame_detalle, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)

        inner = tk.Frame(canvas, bg=GRIS_FONDO)
        canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        def _scroll(event):
            if event.num == 4:
                canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                canvas.yview_scroll(1, "units")
            else:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind("<MouseWheel>", _scroll)
        canvas.bind("<Button-4>", _scroll)
        canvas.bind("<Button-5>", _scroll)
        inner.bind("<MouseWheel>", _scroll)

        pad = {"padx": 12, "pady": 3}

        def fila_info(etiqueta, valor, color_val=NEGRO):
            f = tk.Frame(inner, bg=GRIS_FONDO)
            f.pack(fill="x", **pad)
            tk.Label(f, text=etiqueta, font=("Segoe UI", 8),
                     bg=GRIS_FONDO, fg=GRIS_TEXTO, width=13,
                     anchor="w").pack(side="left")
            tk.Label(f, text=valor, font=("Segoe UI", 9, "bold"),
                     bg=GRIS_FONDO, fg=color_val,
                     wraplength=180, justify="left").pack(side="left")

        tk.Label(inner, text="", bg=GRIS_FONDO).pack(pady=(6, 0))
        fila_info("Fecha y Hora:", _formatear_fecha_hora_iso(det["fecha_y_hora"]))
        fila_info("Fecha Compra:", _formatear_fecha_iso(det["fecha_compra"]))
        fila_info("Comprobante:", det["nro_comprobante"] or "—")
        fila_info("Proveedor:", det["proveedor"])

        tk.Frame(inner, bg=GRIS_BORDE, height=1).pack(fill="x", padx=12, pady=(8, 8))

        tk.Label(inner, text="ARTÍCULOS", font=("Segoe UI", 8, "bold"),
                 bg=GRIS_FONDO, fg=GRIS_TEXTO).pack(anchor="w", padx=12)

        for item in det["items"]:
            f = tk.Frame(inner, bg=BLANCO, highlightthickness=1, highlightbackground=GRIS_BORDE)
            f.pack(fill="x", padx=12, pady=4)
            tk.Label(f, text=item["nombre"], font=("Segoe UI", 9, "bold"),
                     bg=BLANCO, wraplength=260, justify="left").pack(anchor="w", padx=8, pady=(6, 0))
            sub = tk.Frame(f, bg=BLANCO)
            sub.pack(fill="x", padx=8, pady=(2, 6))
            tk.Label(sub, text=f"{item['cantidad']:g} x {formatear_gs(item['precio_unitario'])}",
                     font=("Segoe UI", 8), bg=BLANCO, fg=GRIS_TEXTO).pack(side="left")
            tk.Label(sub, text=formatear_gs(item["importe"]), font=("Segoe UI", 9, "bold"),
                     bg=BLANCO, fg=VERDE).pack(side="right")

        tk.Frame(inner, bg=GRIS_BORDE, height=1).pack(fill="x", padx=12, pady=(8, 8))
        f_total = tk.Frame(inner, bg=GRIS_FONDO)
        f_total.pack(fill="x", padx=12, pady=(0, 16))
        tk.Label(f_total, text="TOTAL", font=("Segoe UI", 9, "bold"),
                 bg=GRIS_FONDO, fg=NEGRO).pack(side="left")
        tk.Label(f_total, text=formatear_gs(det["importe"]), font=("Segoe UI", 12, "bold"),
                 bg=GRIS_FONDO, fg=AZUL).pack(side="right")

    def _cerrar_detalle(self):
        self.frame_detalle.grid_remove()
        self.compra_detalle_id = None
        if self.tabla.selection():
            self.tabla.selection_remove(*self.tabla.selection())

    # ---------------- REPORTE GENERAL: puentes hacia el controlador genérico ----------------
    def _obtener_datos_reporte(self) -> dict:
        """Estructura neutral (consumida por PDF simple, Word, ODT, CSV y
        JSON) construida a partir de self.compras_actuales: exactamente lo
        que el usuario está viendo en pantalla, con todos los filtros
        activos aplicados (rango, proveedor y búsqueda)."""
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""
        total = sum(c["importe"] for c in self.compras_actuales)
        cantidad = len(self.compras_actuales)
        promedio = (total / cantidad) if cantidad else 0
        proveedores_distintos = len({c["proveedor_id"] for c in self.compras_actuales if c["proveedor_id"]})

        filtros_aplicados = []
        if self.var_proveedor.get() != "Todos":
            filtros_aplicados.append(f"Proveedor: {self.var_proveedor.get()}")
        busq = self.var_busqueda.get()
        if busq and busq != "Buscar proveedor, comprobante, ID…":
            filtros_aplicados.append(f"Búsqueda: \"{busq}\"")
        texto_filtros = "  •  ".join(filtros_aplicados) if filtros_aplicados else "Sin filtros adicionales"

        periodo_texto = (
            f"Período: {self.fecha_desde.strftime('%d/%m/%Y')} al "
            f"{self.fecha_hasta.strftime('%d/%m/%Y')}   ({texto_filtros})"
        )

        secciones = [
            {
                "tipo": "resumen",
                "titulo": "RESUMEN GENERAL",
                "filas": [
                    ("Cantidad de Compras", str(cantidad)),
                    ("Total Comprado", formatear_gs(total)),
                    ("Promedio por Compra", formatear_gs(promedio)),
                    ("Proveedores Distintos", str(proveedores_distintos)),
                ],
            },
            {
                "tipo": "tabla",
                "titulo": "DETALLE DE COMPRAS",
                "encabezados": ["ID", "Fecha y Hora", "Fecha Compra", "Proveedor",
                                "N° Comprobante", "Importe"],
                "filas": [
                    [str(c["id"]), _formatear_fecha_hora_iso(c["fecha_y_hora"]),
                     _formatear_fecha_iso(c["fecha_compra"]), c["proveedor"],
                     c["nro_comprobante"] or "—", formatear_gs(c["importe"])]
                    for c in self.compras_actuales
                ],
            },
        ]

        return {
            "titulo": "Reporte General de Compras",
            "subtitulo": periodo_texto,
            "generado_por": nombre_usuario,
            "secciones": secciones,
        }

    def _generar_pdf_dashboard(self, ruta: str):
        """El PDF 'con dashboard' (gráfico de barras + resumen + top
        proveedores/productos) siempre trabaja por rango de fechas puro,
        sin aplicar el filtro de Proveedor/Búsqueda activo en pantalla."""
        from reporte_compras_pdf import generar_reporte_compras_pdf
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""
        generar_reporte_compras_pdf(
            ruta, self.fecha_desde.isoformat(), self.fecha_hasta.isoformat(),
            generado_por=nombre_usuario,
        )

    def _generar_excel(self, ruta: str):
        """Excel propio del reporte: vuelca exactamente self.compras_actuales
        (con todos los filtros activos aplicados) en una hoja de resumen,
        una de detalle de compras y una de detalle de artículos."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from models_compras import obtener_detalle_compra as _detalle

        total = sum(c["importe"] for c in self.compras_actuales)
        cantidad = len(self.compras_actuales)
        promedio = (total / cantidad) if cantidad else 0
        proveedores_distintos = len({c["proveedor_id"] for c in self.compras_actuales if c["proveedor_id"]})
        nombre_usuario = self.usuario_actual.get("nombre_completo", "") if self.usuario_actual else ""

        wb = openpyxl.Workbook()

        # ── Hoja Resumen ──────────────────────────────────────
        ws = wb.active
        ws.title = "Resumen"
        ws.append(["Reporte General de Compras"])
        ws.append([f"Período: {self.fecha_desde.strftime('%d/%m/%Y')} al "
                    f"{self.fecha_hasta.strftime('%d/%m/%Y')}"])
        ws.append([f"Generado por: {nombre_usuario}"])
        ws.append([])
        filas_resumen = [
            ("Cantidad de Compras", cantidad),
            ("Total Comprado", total),
            ("Promedio por Compra", round(promedio, 2)),
            ("Proveedores Distintos", proveedores_distintos),
        ]
        ws.append(["Indicador", "Valor"])
        for etiqueta, valor in filas_resumen:
            ws.append([etiqueta, valor])

        for cell in ws[6]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D5FD6")
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 20

        # ── Hoja Detalle de Compras ───────────────────────────
        ws2 = wb.create_sheet("Detalle de Compras")
        cols_det = ["ID", "Fecha y Hora", "Fecha Compra", "Proveedor",
                    "N° Comprobante", "Importe"]
        ws2.append(cols_det)
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D5FD6")
            cell.alignment = Alignment(horizontal="center")
        for c in self.compras_actuales:
            ws2.append([
                c["id"], _formatear_fecha_hora_iso(c["fecha_y_hora"]),
                _formatear_fecha_iso(c["fecha_compra"]), c["proveedor"],
                c["nro_comprobante"] or "—", c["importe"],
            ])
        for col in ws2.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # ── Hoja Detalle de Artículos ──────────────────────────
        ws3 = wb.create_sheet("Detalle de Artículos")
        cols_art = ["Compra Nro.", "Producto", "Cantidad", "Precio Unit.", "Importe"]
        ws3.append(cols_art)
        for cell in ws3[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D5FD6")
            cell.alignment = Alignment(horizontal="center")
        for c in self.compras_actuales:
            det = _detalle(c["id"])
            if not det:
                continue
            for item in det["items"]:
                ws3.append([
                    c["id"], item["nombre"], item["cantidad"],
                    item["precio_unitario"], item["importe"],
                ])
        for col in ws3.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws3.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(ruta)


def _formatear_fecha_iso(fecha_iso: str) -> str:
    if not fecha_iso:
        return "—"
    try:
        return datetime.date.fromisoformat(fecha_iso[:10]).strftime("%d/%m/%Y")
    except ValueError:
        return fecha_iso


def _formatear_fecha_hora_iso(fecha_hora: str) -> str:
    if not fecha_hora:
        return "—"
    try:
        dt = datetime.datetime.fromisoformat(fecha_hora)
        return dt.strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return fecha_hora
